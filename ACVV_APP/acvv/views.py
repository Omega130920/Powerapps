import base64
import csv
import os
import re
import logging
from itertools import count
from datetime import datetime, timedelta
import traceback

# Django Imports
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, OuterRef, Subquery, Sum, Count, Avg
from django.db.models.functions import TruncMonth
from django.urls import reverse
from django.conf import settings
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.safestring import mark_safe
from django.http import HttpResponse
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

# Third-Party Imports
import pandas as pd
import requests
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from dateutil import parser
from dateutil.relativedelta import relativedelta

# Local Project Imports
from .models import (
    AcvvClaim, BranchDocument, ClaimNote, Globalacvv, ClientNotes, 
    EmailDelegation, DelegationNote, DelegationTransactionLog, 
    ReconciliationRecord, ReconciliationWorksheet, TempExit
)
from .services.outlook_graph_service import OutlookGraphService
from .services.delegation_service import (
    get_or_create_delegation_status, 
    delegate_email_task, 
    add_delegation_note, 
    get_delegated_emails_for_user,
    log_delegation_transaction
)

# Initialize logger
logger = logging.getLogger(__name__)

def login_view(request):
    """
    Handles user login.
    """
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('dashboard')
            else:
                messages.error(request, "Invalid username or password.")
    
    form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})

from acvv.services.outlook_graph_service import OutlookGraphService

@login_required
def test_graph_access(request):
    # This calls the root level without looking into a specific email item
    # It checks if the app can even "talk" to the mailbox at all
    response = OutlookGraphService.fetch_outlook_data("mailFolders", settings.OUTLOOK_EMAIL_ADDRESS)
    return HttpResponse(str(response))

@login_required
def dashboard(request):
    username = request.user.username
    is_outlook_admin = request.user.username.lower() == 'omega' or request.user.is_superuser
    
    # 1. Fetch current IDs from the Live API to ensure the count is accurate
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    inbox_data = OutlookGraphService._make_graph_request(target_email, 1000) # Fetch a reasonable amount
    
    if 'error' not in inbox_data:
        live_ids = [msg['id'] for msg in inbox_data.get('value', [])]
        
        # 2. Only count records that are NEW, Work Related AND still exist in the Live Inbox
        undelegated_count = EmailDelegation.objects.filter(
            status='NEW', 
            work_related=True,
            email_id__in=live_ids  # <--- This is the crucial filter
        ).count()
        
        # OPTIONAL: Mark old database entries that are no longer in the Live Inbox as 'LOST' or 'ARCHIVED'
        # to stop them from haunting your counts in other places.
    else:
        # Fallback if API is down
        undelegated_count = EmailDelegation.objects.filter(status='NEW', work_related=True).count()

    # Recycle Bin and My Tasks are purely database-driven, so these are fine as is
    recycled_count = EmailDelegation.objects.filter(status='DLT').count()
    my_tasks_count = EmailDelegation.objects.filter(assigned_user=request.user, status='DEL').count()
        
    context = {
        'username': username,
        'undelegated_count': undelegated_count, 
        'recycled_count': recycled_count,
        'my_tasks_count': my_tasks_count,
        'is_outlook_admin': is_outlook_admin,
    }
    return render(request, 'dashboard.html', context)

def logout_view(request):
    """
    Logs the user out.
    """
    logout(request)
    messages.info(request, "You have successfully logged out.")
    return redirect('login')

def index(request):
    """
    Handles the root URL, redirecting to the login page.
    """
    return redirect('login') 

@login_required
def outlook_dashboard_view(request):
    if request.user.username.lower() != 'omega' and not request.user.is_superuser:
        messages.error(request, "Access restricted.")
        return redirect('outlook_delegated_box')

    target_email = request.GET.get('email', settings.OUTLOOK_EMAIL_ADDRESS)
    context = {'target_email': target_email, 'messages': []}
    
    # Use Service class
    inbox_data = OutlookGraphService.fetch_inbox_messages(target_email, 1000) 
    
    if 'error' not in inbox_data:
        emails = inbox_data.get('value', [])
        email_ids = [email['id'] for email in emails]
        
        processed_ids = EmailDelegation.objects.filter(
            email_id__in=email_ids
        ).exclude(status='NEW', work_related=True).values_list('email_id', flat=True)

        display_emails = []
        for email in emails:
            email_id = email['id']
            if email_id in processed_ids:
                continue

            received_date_str = email.get('receivedDateTime') 
            sender_email = email.get('from', {}).get('emailAddress', {}).get('address', '')
            
            delegation, created = EmailDelegation.objects.get_or_create(
                email_id=email_id,
                defaults={
                    'received_at': parser.isoparse(received_date_str) if received_date_str else None,
                    'status': 'NEW',
                    'work_related': True,
                    'sender_address': sender_email,
                    'subject': email.get('subject', '(No Subject)')
                }
            )
            
            if not created and not delegation.sender_address:
                delegation.sender_address = sender_email
                delegation.save()

            try:
                if received_date_str:
                    email['receivedDateTime'] = parser.isoparse(received_date_str)
            except Exception: pass

            email['delegation_status'] = delegation.get_status_display()
            email['assigned_user'] = delegation.assigned_user.username if delegation.assigned_user else None
            email['delegation_id'] = delegation.pk 
            display_emails.append(email)

        context['messages'] = display_emails
    else:
        context['error'] = f"Error: {inbox_data['error']}"

    return render(request, 'acvv_app/outlook_dashboard.html', context)

@login_required
def send_email_view(request):
    """
    Handles displaying the email form and processing the email submission 
    to the Microsoft Graph API. The email is sent FROM the mailbox specified by target_email.
    """
    # DELEGATION AWARENESS: Get target email from URL or settings default 
    target_email = request.GET.get('email', settings.OUTLOOK_EMAIL_ADDRESS)
    
    if request.method == 'POST':
        recipient = request.POST.get('recipient')
        subject = request.POST.get('subject')
        body = request.POST.get('body')
        
        # Simple validation
        if not all([recipient, subject, body]):
            messages.error(request, "All fields are required.")
            return render(request, 'acvv_app/send_email_form.html', {'target_email': target_email})
        
        # Call the service function, passing the target_email as the sender mailbox
        result = OutlookGraphService.send_outlook_email(target_email, recipient, subject, body, 'HTML')
        
        if result.get('success'):
            messages.success(request, f"Email sent successfully from {target_email} to {recipient}.")
            # Redirect back to the dashboard, preserving the target email
            return redirect(f"{reverse('outlook_dashboard')}?email={target_email}")
        else:
            error_message = f"Email failed to send from {target_email}. {result.get('error', 'Unknown API Error')}"
            
            # Extract details if they exist in the nested error structure
            details = result.get('details', {})
            if isinstance(details, dict) and 'error' in details and 'message' in details['error']:
                 error_message += f" Details: {details['error']['message']}"
            
            messages.error(request, error_message)
            # Render the form again with the failure message
            return render(request, 'acvv_app/send_email_form.html', {
                'recipient': recipient,
                'subject': subject,
                'body': body,
                'target_email': target_email
            })

    # Render the empty form on GET request
    return render(request, 'acvv_app/send_email_form.html', {'target_email': target_email})

@login_required
def outlook_delegated_box(request):
    delegations = EmailDelegation.objects.filter(
        assigned_user=request.user, 
        status='DEL'
    ).order_by('-delegated_at')
    
    tasks = []
    for delegation in delegations:
        tasks.append({
            'delegation_id': delegation.pk,
            'status': delegation.get_status_display(),
            'subject': delegation.subject,
            'from': delegation.sender_address,
            'mg_code': delegation.mip_names,
            'received': delegation.received_at,
            'delegated_at': delegation.delegated_at,
            'email_type': delegation.email_category,
        })
    return render(request, 'acvv_app/outlook_delegated_box.html', {'tasks': tasks})

@login_required
def outlook_delegated_action(request, delegation_id):
    """
    Allows the assigned user to view the full email, add notes, 
    update metadata, reply (with attachments), and mark as completed.
    """
    delegation = get_object_or_404(EmailDelegation, pk=delegation_id)
    
    if delegation.assigned_user != request.user:
        messages.error(request, "You are not assigned to this task.")
        return redirect('outlook_delegated_box')

    target_email = settings.OUTLOOK_EMAIL_ADDRESS 

    if request.method == 'POST':
        action_type = request.POST.get('action_type')

        # 1. Handle Task Completion
        if action_type == 'complete_task':
            delegation.status = 'COM'
            delegation.save()
            log_delegation_transaction(delegation_id, request.user, "TASK COMPLETED", "Email marked as completed.", action_type='TASK_COMPLETE')
            messages.success(request, f"Task #{delegation_id} archived.")
            return redirect('outlook_delegated_box')

        # 2. Handle Metadata Update
        elif action_type == 'update_metadata':
            delegation.mip_names = request.POST.get('mip_names')
            delegation.email_category = request.POST.get('email_category')
            delegation.communication_type = request.POST.get('communication_type')
            delegation.save()
            log_delegation_transaction(delegation_id, request.user, f"Metadata Updated", "System", action_type='METADATA_UPDATE')
            messages.success(request, "Task metadata updated.")
            return redirect('outlook_delegated_action', delegation_id=delegation_id)

        # 3. Handle Note Submission
        elif action_type == 'add_note':
            note_content = request.POST.get('note_content')
            success, message = add_delegation_note(delegation_id, request.user, note_content)
            
            if success:
                # 🟢 CRITICAL SYSTEM ADDITION: Write manual internal note entry straight to ClientNotes 
                ClientNotes.objects.create(
                    acvv_record=Globalacvv.objects.filter(Q(mip_names=delegation.mip_names) | Q(branch_code=delegation.mip_names)).first(),
                    notes=note_content,
                    user=request.user.username,
                    date=timezone.now(),
                    communication_type="Internal Note",
                    action_note_type=delegation.email_category or "Task Note"
                )
                messages.success(request, message)
            return redirect('outlook_delegated_action', delegation_id=delegation_id)
        
        # 4. Handle Reply
        elif action_type == 'send_reply':
            recipient = request.POST.get('reply_recipient')
            subject = request.POST.get('reply_subject')
            body = request.POST.get('reply_body')
            
            # --- MULTI-ATTACHMENT UPDATE ---
            attachments_list = request.FILES.getlist('email_attachments')
            fallback_single_attachment = attachments_list[0] if attachments_list else None
            
            selected_action_type = request.POST.get('action_log_type') or "Correspondence"
            
            # Pass down the complete file objects array context list safely
            result = OutlookGraphService.send_outlook_email(target_email, recipient, subject, body, content_type='Html', attachments=attachments_list)
            
            if result.get('success'):
                log_delegation_transaction(delegation_id, request.user, subject, recipient, action_type='EMAIL_REPLY')
                
                new_ms_id = result.get('message_id') or f"REPLY-{timezone.now().timestamp()}"
                
                EmailDelegation.objects.create(
                    email_id=new_ms_id,
                    subject=subject,
                    body=body,
                    attachment=fallback_single_attachment,
                    sender_address=target_email,
                    assigned_user=request.user,
                    status='SENT',
                    mip_names=delegation.mip_names,
                    received_at=timezone.now(),
                    delegated_at=timezone.now(),
                    work_related=True,
                    communication_type='Reply'
                )
                
                ClientNotes.objects.create(
                    acvv_record=Globalacvv.objects.filter(Q(mip_names=delegation.mip_names) | Q(branch_code=delegation.mip_names)).first(),
                    notes=f"Reply Sent: {subject}\nRecipient: {recipient}",
                    user=request.user.username,
                    date=timezone.now(),
                    communication_type="Email",
                    action_note_type=selected_action_type 
                )
                
                messages.success(request, "Reply sent.")
            else:
                messages.error(request, f"Reply failed: {result.get('error')}")
            return redirect('outlook_delegated_action', delegation_id=delegation_id)

    # --- FETCH Data for GET ---
    acvv_records = Globalacvv.objects.all().only('mip_names', 'branch_code')
    
    if delegation.email_id.startswith('SENT-') or delegation.email_id.startswith('REPLY-') or delegation.email_id.startswith('LOCAL-'):
        email_data = {
            'subject': delegation.subject,
            'body': {'content': delegation.body or "Local preview content not found."},
            'from': {'emailAddress': {'address': delegation.sender_address or target_email}}
        }
        attachments = []
        if delegation.attachment:
            attachments = [{
                'name': os.path.basename(delegation.attachment.name),
                'url': delegation.attachment.url,
                'contentType': 'application/octet-stream',
                'is_local': True
            }]
    else:
        # 🔑 SAFELY WRAP OUTLOOK MICROSOFT LIVE API CALLS IN EXCEPTION TRY BLOCKS
        try:
            email_data = OutlookGraphService.fetch_outlook_data(f"messages/{delegation.email_id}", target_email)
            
            is_404_error = False
            if isinstance(email_data, dict):
                error_value = str(email_data.get('error', ''))
                details_value = str(email_data.get('details', ''))
                
                if "Status 404" in error_value or "ErrorItemNotFound" in details_value:
                    is_404_error = True
            elif isinstance(email_data, str) and "ErrorItemNotFound" in email_data:
                is_404_error = True

            if is_404_error:
                raise Exception("ErrorItemNotFound")

            # FIX: Used new fetch_outlook_data for attachment retrieval
            attachment_endpoint = f"messages/{delegation.email_id}/attachments"
            attachment_data = OutlookGraphService.fetch_outlook_data(attachment_endpoint, target_email)
            attachments = attachment_data.get('value', []) if isinstance(attachment_data, dict) and 'error' not in attachment_data else []

        except Exception as graph_error:
            if "ErrorItemNotFound" in str(graph_error) or "404" in str(graph_error):
                # 🚀 ADDED: PURGE THE BROKEN RECORD TO STOP THE REDIRECT LOOP
                delegation.delete()
                messages.warning(request, "This email could not be found (moved or deleted). Task removed from your list.")
                return redirect('outlook_delegated_box')
            else:
                raise graph_error

    if isinstance(email_data, dict) and 'error' in email_data and not email_data.get('subject'):
        messages.error(request, f"Error fetching content: {email_data.get('error')}")
        return redirect('outlook_delegated_box')

    context = {
        'delegation': delegation,
        'email': email_data,
        'attachments': attachments,
        'notes': delegation.notes.all().order_by('-created_at'),
        'acvv_records': acvv_records,
        'target_email': target_email,
    }
    return render(request, 'acvv_app/outlook_delegated_action.html', context)

@login_required
def export_acvv_list_excel(request):
    """Exports filtered ACVV records to Excel with specific column mapping."""
    search_query = request.GET.get('search_query')
    records = Globalacvv.objects.all()

    if search_query:
        records = records.filter(
            Q(mip_names__icontains=search_query) |
            Q(branch_code__icontains=search_query) |
            Q(tel__icontains=search_query) |
            Q(tel_2__icontains=search_query)
        )
    
    records = records.order_by('mip_names')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ACVV Records Export"

    headers = [
        'Member Group Name', 'MG Code', 'Company Status', 
        'Last Recon - Status', 'Member Count', 'Last Recon - Date', 
        'Bill Amount', 'MG Contact Email', 'MG Contact Tel. 1', 'MG Contact Tel. 2'
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="43a047", end_color="43a047", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for record in records:
        ws.append([
            record.mip_names,                      # Member Group Name
            record.branch_code or "-",             # MG Code
            record.status or "-",                  # Company Status
            "",                                    # Last Recon - Status
            record.member or "-",                  # Member Count
            "",                                    # Last Recon - Date
            record.contribution_amount or "-",     # Bill Amount
            record.mg_email_address or "-",        # MG Contact Email
            record.tel or "-",                     # MG Contact Tel. 1
            record.tel_2 or "-"                    # UPDATED: MG Contact Tel. 2
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="ACVV_Export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def acvv_list(request):
    """
    Displays a list of all ACVV records from the Globalacvv model with search functionality.
    Sorted by branch code.
    """
    acvv_records = Globalacvv.objects.all()
    search_query = request.GET.get('search_query')

    # Apply search filter (Updated to include Telephone numbers)
    if search_query:
        acvv_records = acvv_records.filter(
            Q(mip_names__icontains=search_query) |
            Q(branch_code__icontains=search_query) |
            Q(tel__icontains=search_query) |
            Q(tel_2__icontains=search_query)
        )

    # Order the results by code instead of mip_names
    acvv_records = acvv_records.order_by('branch_code')

    context = {
        'acvv_records': acvv_records,
        'search_query': search_query or ''
    }
    return render(request, 'acvv_app/acvv_list.html', context)

@login_required
def acvv_information(request, mip_names):
    """
    Detailed view for a specific ACVV record with unified logging.
    Includes support for TEL, TEL 2, MG ADDRESS, NPO CODE, MG BANK INFO, 
    and the new BANK & EMPLOYER CONTACT fields from the Globalacvv model.
    """
    # 🔑 FIX: Look up the record by matching either 'mip_names' OR 'branch_code' 
    # to handle code-based routing links safely without throwing a 404.
    acvv_record = get_object_or_404(Globalacvv, Q(mip_names=mip_names) | Q(branch_code=mip_names))
    
    if request.method == 'POST':
        # 1. Handle Note Uploads
        if 'add_note' in request.POST:
            note_content = request.POST.get('internal_note_text')
            comm_type = request.POST.get('communication_type')
            action_note = request.POST.get('action_note_type')
            uploaded_file = request.FILES.get('note_attachment')

            if note_content or uploaded_file:
                file_url = None
                if uploaded_file:
                    fs = FileSystemStorage()
                    filename = fs.save(f"notes/{uploaded_file.name}", uploaded_file)
                    file_url = fs.url(filename)

                ClientNotes.objects.create(
                    acvv_record=acvv_record, 
                    notes=note_content,
                    user=request.user.username, 
                    date=timezone.now(),
                    communication_type=comm_type, 
                    action_note_type=action_note,
                    attachment=file_url
                )
                messages.success(request, "Internal note added.")
                return redirect(f'/acvv-records/{acvv_record.mip_names}/#notes-tab')

        # 2. Handle PDF uploads
        elif 'upload_pdf' in request.POST:
            pdf_file = request.FILES.get('branch_pdf')
            if pdf_file:
                fs = FileSystemStorage()
                path = f"branch_docs/{acvv_record.mip_names}/{pdf_file.name}"
                filename = fs.save(path, pdf_file)
                file_url = fs.url(filename)
                BranchDocument.objects.create(
                    branch_name=acvv_record.mip_names, 
                    file_name=pdf_file.name,
                    file_path=file_url, 
                    uploaded_by=request.user.username
                )
                messages.success(request, "PDF added.")
                return redirect(f'/acvv-records/{acvv_record.mip_names}/#pdf-upload')

        # 3. Handle Contact Info Update (UPDATED: Matched exact HTML name attributes)
        elif 'update_contact_info' in request.POST:
            # Normalize email input to handle multiple addresses entered with commas or semicolons
            raw_emails = request.POST.get('mg_email_address', '')
            # Clean: Replace ; with , then split, strip whitespace, and rejoin with comma
            email_list = [e.strip() for e in raw_emails.replace(';', ',').split(',') if e.strip()]
            acvv_record.mg_email_address = ",".join(email_list)
            
            # --- EXTRACT AND SAVE ALL DATA MATCHING THE HTML NAMES ---
            acvv_record.branch_code = request.POST.get('branch_code')
            acvv_record.member = request.POST.get('member')
            acvv_record.contribution_amount = request.POST.get('contribution_amount')
            
            acvv_record.tel = request.POST.get('tel')
            acvv_record.tel_2 = request.POST.get('tel_2')
            
            acvv_record.bank = request.POST.get('bank')
            acvv_record.branch = request.POST.get('branch')
            acvv_record.account_number = request.POST.get('account_number')
            acvv_record.account_name = request.POST.get('account_name')
            acvv_record.account_type = request.POST.get('account_type')
            acvv_record.employer_contacts = request.POST.get('employer_contacts')
            
            acvv_record.mg_address = request.POST.get('mg_address')
            acvv_record.npo_code = request.POST.get('npo_code')
            acvv_record.mg_bank_info = request.POST.get('mg_bank_info')
            
            acvv_record.save()
            messages.success(request, "Contact information updated successfully.")
            return redirect('acvv_information', mip_names=acvv_record.mip_names)

    # --- DATA FETCHING & COLUMN LINKING ---
    all_notes = ClientNotes.objects.filter(acvv_record=acvv_record)
    
    notes = all_notes.order_by('-date')
    
    company_claims = AcvvClaim.objects.filter(company_code=mip_names).order_by('-claim_created_date')
    branch_docs = BranchDocument.objects.filter(branch_name=mip_names).order_by('-uploaded_at')

    delegated_logs = EmailDelegation.objects.filter(
        Q(mip_names__icontains=acvv_record.mip_names) | Q(mip_names__icontains=acvv_record.branch_code)
    ).select_related('assigned_user')

    # --- 1. BUILD LOOKUP FOR REAL-TIME ACTION TYPES FROM CLIENT_NOTES ---
    notes_action_map = {}
    for n in all_notes:
        note_text = n.notes or ""
        if "Email Sent:" in note_text:
            try:
                sub_line = note_text.split('\n')[0].replace("Email Sent:", "").strip()
                notes_action_map[sub_line] = n.action_note_type
            except Exception:
                pass

    # --- 2. FETCH AND MAP REMARKS CONTENT FROM DELEGATION_NOTE TABLE ---
    delegation_ids = delegated_logs.values_list('id', flat=True)
    
    raw_delegation_notes = DelegationNote.objects.filter(
        delegation_id__in=delegation_ids
    ).order_by('created_at')

    delegation_notes_map = {}
    for d_note in raw_delegation_notes:
        if d_note.delegation_id not in delegation_notes_map:
            delegation_notes_map[d_note.delegation_id] = []
        
        formatted_entry = f"[{d_note.created_at.strftime('%Y-%m-%d')}] {d_note.content}"
        delegation_notes_map[d_note.delegation_id].append(formatted_entry)

    combined_email_log = []
    for log in delegated_logs:
        if log.status == 'SENT' or log.status == 'COM':
            raw_comm_type = getattr(log, 'communication_type', '') or ''
            comm_type_lower = raw_comm_type.lower()
            
            if comm_type_lower == 'reply':
                log_type = 'REPLY'
                log_icon = '↩️'
                badge_color = '#f7931e'
            elif comm_type_lower == 'claim sent':
                log_type = 'CLAIM SENT'
                log_icon = '📋'
                badge_color = '#9c27b0'
            elif comm_type_lower == 'two pot email':
                log_type = 'TWO POT EMAIL'
                log_icon = '🍯'
                badge_color = '#e65100'
            else:
                log_type = 'DIRECT'
                log_icon = '📤'
                badge_color = '#28a745'
        else:
            log_type = 'ORIGINAL'
            log_icon = '📩'
            badge_color = '#1976d2' if log.status != 'DLT' else '#ef5350'

        resolved_action_note_type = notes_action_map.get(log.subject, getattr(log, 'action_note_type', None))
        if not resolved_action_note_type or resolved_action_note_type == 'None':
            resolved_action_note_type = '-'

        matched_notes_list = delegation_notes_map.get(log.id, [])
        resolved_delegation_content = " | ".join(matched_notes_list) if matched_notes_list else "-"

        combined_email_log.append({
            'type': log_type,
            'icon': log_icon,
            'badge_color': badge_color, 
            'subject': log.subject or "Outlook Task",
            'received_at': log.received_at,
            'assigned_to': log.assigned_user.username if log.assigned_user else "Unassigned",
            'display_type': log_type if log.status == 'SENT' else log.get_status_display(),
            'email_id': log.email_id,
            'file_url': log.attachment.url if hasattr(log, 'attachment') and log.attachment else None,
            'action_note_type': resolved_action_note_type,
            'delegation_note_content': resolved_delegation_content, 
            'sort_date': log.received_at or log.delegated_at
        })

    combined_email_log.sort(key=lambda x: x['sort_date'] if x['sort_date'] else timezone.datetime.min.replace(tzinfo=timezone.utc), reverse=True)

    context = {
        'acvv_record': acvv_record,
        'combined_email_log': combined_email_log, 
        'notes': notes,
        'company_claims': company_claims,
        'branch_docs': branch_docs,
        'my_delegated_emails': EmailDelegation.objects.filter(assigned_user=request.user, status='DEL'),
    }
    return render(request, 'acvv_app/acvv_information.html', context)

@login_required
def outlook_delegate_to(request, email_id):
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    available_users = User.objects.filter(is_active=True).order_by('username')
    acvv_records = Globalacvv.objects.all().values('mip_names', 'branch_code')
    
    # 1. Fetch the main Email Message with safety wrapper
    try:
        endpoint = f"messages/{email_id}" 
        email_data = OutlookGraphService._make_graph_request(endpoint, target_email) 

        if not email_data or 'error' in email_data:
            error_details = email_data.get('error', {})
            error_code = error_details.get('code')
            
            # 🛑 HANDLE 404: If email is deleted/moved in Outlook, clean up local DB
            if error_code == 'ErrorItemNotFound':
                logger.warning(f"Email {email_id} not found in Outlook. Deleting local record.")
                EmailDelegation.objects.filter(email_id=email_id).delete()
                messages.warning(request, "This email was moved or deleted in Outlook and has been removed from your list.")
                return redirect('outlook_dashboard')

            logger.error(f"Graph API Error for ID {email_id}: {error_code}")
            messages.error(request, "Could not fetch email content from Outlook.")
            return redirect('outlook_dashboard')
            
    except Exception as e:
        logger.exception(f"Unexpected crash fetching email {email_id}")
        messages.error(request, "A critical error occurred while contacting the mail server.")
        return redirect('outlook_dashboard')

    # 2. Fetch Attachments (Non-critical)
    attachments = []
    try:
        attachment_endpoint = f"messages/{email_id}/attachments"
        attachment_data = OutlookGraphService._make_graph_request(attachment_endpoint, target_email)
        if attachment_data and 'value' in attachment_data:
            attachments = attachment_data['value']
    except Exception as e:
        logger.warning(f"Failed to fetch attachments for {email_id}: {str(e)}")

    email_subject = email_data.get('subject', '(No Subject)')
    sender_email = email_data.get('from', {}).get('emailAddress', {}).get('address', '')

    if request.method == 'POST':
        try:
            with transaction.atomic():
                work_related_raw = request.POST.get('work_related')
                is_work_related = (work_related_raw == 'Yes')
                assignee_pk = request.POST.get('agent_name')
                mip_names_value = request.POST.get('mip_names')
                # Capture the log type
                action_log_type_value = request.POST.get('action_log_type', 'General')
                
                if not is_work_related:
                    EmailDelegation.objects.update_or_create(
                        email_id=email_id,
                        defaults={
                            'work_related': False,
                            'status': 'DLT',
                            'subject': email_subject,
                            'sender_address': sender_email,
                            'mip_names': mip_names_value,
                            'assigned_user': None 
                        }
                    )
                    messages.success(request, "Task moved to Recycle Bin.")
                    return redirect('outlook_dashboard')
                
                else:
                    if assignee_pk and assignee_pk not in ['', '__Select Agent__']:
                        target_assignee = get_object_or_404(User, pk=assignee_pk)
                        
                        EmailDelegation.objects.update_or_create(
                            email_id=email_id,
                            defaults={
                                'assigned_user': target_assignee,
                                'mip_names': mip_names_value,
                                'subject': email_subject,
                                'sender_address': sender_email,
                                'email_category': request.POST.get('email_category'),
                                'work_related': True, 
                                'status': 'DEL',
                                'comm_type': request.POST.get('email_method', 'Email'),
                            }
                        )

                        # Create the log entry to persist the Action Log Type
                        ClientNotes.objects.create(
                            acvv_record=Globalacvv.objects.filter(Q(mip_names=mip_names_value) | Q(branch_code=mip_names_value)).first(),
                            notes=f"Email delegated to {target_assignee.username}",
                            user=request.user.username,
                            date=timezone.now(),
                            communication_type="Email Delegation",
                            action_note_type=action_log_type_value
                        )

                        messages.success(request, f"Task successfully assigned/re-delegated to {target_assignee.username}!")
                        return redirect('outlook_dashboard')
                    else:
                        messages.error(request, "Please select an agent.")

        except Exception as e:
            logger.exception(f"Delegation crash prevented for ID {email_id}")
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect('outlook_dashboard')

    context = {
        'email_id': email_id,
        'email_subject': email_subject,
        'email_sender': sender_email,
        'email_content': email_data.get('body', {}).get('content', ''), 
        'attachments': attachments,
        'available_users': available_users,
        'acvv_records': acvv_records,
    }
    return render(request, 'acvv_app/outlook_delegate_to.html', context)

def outlook_email_content(request, email_id):
    """
    Fetches the raw HTML content of an email and returns it as a response 
    to be loaded by an iframe's 'src' attribute.
    """
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    endpoint = f"messages/{email_id}" 
    email_data = OutlookGraphService._make_graph_request(endpoint, target_email)

    if 'error' in email_data:
        return HttpResponse("<h1>Error fetching email content.</h1>", status=500)

    body_data = email_data.get('body', {})
    
    if body_data.get('contentType', '').lower() == 'html':
        content = body_data.get('content', 'No HTML body found.')
    else:
        # If plain text, wrap in <pre> tags
        content = body_data.get('content', 'No email body found.')
        content = f'<pre style="white-space: pre-wrap; word-wrap: break-word; font-family: sans-serif;">{content}</pre>'

    # Wrap in a full HTML document (essential for iframe rendering)
    wrapped_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>body {{ font-family: sans-serif; margin: 15px; }}</style>
    </head>
    <body>
        {content}
    </body>
    </html>
    """
    
    # Return as plain HTML response (not marked safe, but the browser loads it securely)
    return HttpResponse(wrapped_content, content_type='text/html')

@login_required
def save_acvv_claim(request, company_code):
    """
    Handles Member Claim registration and updates ONLY.
    No longer contains any email-sending logic.
    """
    if request.method == 'POST':
        id_number = request.POST.get('id_number')
        
        # Only proceed if we actually have claim data
        if id_number:
            claim_id = request.POST.get('claim_id')
            data = {
                'company_code': company_code,
                'id_number': id_number,
                'member_name': request.POST.get('member_name'),
                'member_surname': request.POST.get('member_surname'),
                'mip_number': request.POST.get('mip_number'),
                'claim_type': request.POST.get('claim_type'),
                'exit_reason': request.POST.get('exit_reason'),
                'claim_allocation': request.POST.get('claim_allocation'),
                'claim_status': request.POST.get('claim_status'),
                'payment_option': request.POST.get('payment_option'),
                'claim_amount': request.POST.get('claim_amount') or None,
                'claim_created_date': request.POST.get('claim_created_date') or None,
                'last_contribution_date': request.POST.get('last_contribution_date') or None,
                'date_submitted': request.POST.get('date_submitted') or None,
                'date_paid': request.POST.get('date_paid') or None,
                'vested_pot_available': request.POST.get('vested_pot_available') == 'on',
                'vested_pot_paid_date': request.POST.get('vested_pot_paid_date') or None,
                'savings_pot_available': request.POST.get('savings_pot_available') == 'on',
                'savings_pot_paid_date': request.POST.get('savings_pot_paid_date') or None,
                'infund_cert_date': request.POST.get('infund_cert_date') or None,
                'linked_email_id': request.POST.get('linked_email_id'),
            }

            if claim_id:
                AcvvClaim.objects.filter(id=claim_id).update(**data)
                claim_instance = AcvvClaim.objects.get(id=claim_id)
                messages.success(request, "Claim updated successfully.")
            else:
                claim_instance = AcvvClaim.objects.create(**data)
                messages.success(request, "New claim created successfully.")

            # Note/Attachment logic for the claim record itself
            note_selection = request.POST.get('note_selection')
            note_description = request.POST.get('note_description')
            claim_file = request.FILES.get('claim_attachment')
            if note_selection or note_description or claim_file:
                claim_instance.notes.create(
                    note_selection=note_selection,
                    note_description=note_description,
                    attachment=claim_file,
                    created_by=request.user
                )
        else:
            messages.warning(request, "Claim save ignored: No ID Number provided.")

    return redirect('acvv_information', mip_names=company_code)

@login_required
def global_two_pot_view(request):
    """Dedicated Register for ONLY Two Pot claims with Note/Attachment support."""
    query = request.GET.get('q')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    target_email = settings.OUTLOOK_EMAIL_ADDRESS

    # --- HARD FILTER: Only Two Pot ---
    # Added prefetch_related('notes') to load history and attachments efficiently
    base_claims = AcvvClaim.objects.filter(claim_type='Two Pot').prefetch_related('notes').order_by('-claim_created_date')

    if query:
        base_claims = base_claims.filter(
            Q(id_number__icontains=query) | 
            Q(member_surname__icontains=query) | 
            Q(company_code__icontains=query)
        )

    if start_date and end_date:
        base_claims = base_claims.filter(claim_created_date__range=[parse_date(start_date), parse_date(end_date)])

    # Pagination
    paginator = Paginator(base_claims, 15) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Email Preview Logic
    delegation_pks = [c.linked_email_id for c in page_obj if c.linked_email_id]
    if delegation_pks:
        # Filter out invalid IDs that aren't numeric to avoid the ValueError we saw earlier
        numeric_pks = [pk for pk in delegation_pks if str(pk).isdigit()]
        delegations_map = EmailDelegation.objects.in_bulk(numeric_pks)
        
        for claim in page_obj:
            if claim.linked_email_id and str(claim.linked_email_id).isdigit():
                try:
                    del_obj = delegations_map.get(int(claim.linked_email_id))
                    if del_obj:
                        email_data = OutlookGraphService._make_graph_request(
                            f"messages/{del_obj.email_id}?$select=subject,from,body,receivedDateTime", 
                            target_email
                        )
                        if 'error' not in email_data:
                            claim.email_preview_subject = email_data.get('subject')
                            claim.email_preview_sender = email_data.get('from', {}).get('emailAddress', {}).get('address')
                            claim.email_preview_body = email_data.get('body', {}).get('content')
                            claim.email_preview_date = email_data.get('receivedDateTime')
                except Exception as e:
                    print(f"DEBUG: Email Preview Error for Claim {claim.id}: {e}")
                    continue

    # Rendering the SPECIFIC Two Pot HTML (Ensure the template path matches your project)
    return render(request, 'acvv_app/two_pot_global.html', {
        'page_obj': page_obj, 
        'claims': page_obj, # This ensures the notes-templates loop works
        'all_companies': Globalacvv.objects.values('mip_names', 'branch_code'),
        'my_delegated_emails': EmailDelegation.objects.filter(assigned_user=request.user, status='DEL').order_by('-received_at'),
        'is_two_pot_view': True,
        'search_query': query,
        'start_date': start_date,
        'end_date': end_date,
    })
    
@login_required
def global_claims_view(request):
    """Register for ALL claims EXCEPT Two Pot."""
    query = request.GET.get('q')
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    
    base_claims = AcvvClaim.objects.exclude(claim_type='Two Pot')

    if query:
        # If searching, we fetch all matching results before paginating
        claims_list = base_claims.filter(
            Q(id_number__icontains=query) | 
            Q(member_surname__icontains=query) | 
            Q(company_code__icontains=query)
        ).order_by('-claim_created_date')
    else:
        # Removed the [:50] limit so pagination can access all records
        claims_list = base_claims.order_by('-claim_created_date')

    # --- Setup Pagination ---
    # Show 36 claims per page as requested
    paginator = Paginator(claims_list, 36) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # We now operate on page_obj.object_list instead of the full claims list
    current_page_claims = page_obj.object_list

    # --- Email Preview Logic (Only run for the 36 items on current page) ---
    delegation_ids = [str(c.linked_email_id) for c in current_page_claims if c.linked_email_id]
    
    if delegation_ids:
        delegations_map = EmailDelegation.objects.in_bulk(delegation_ids, field_name='email_id')
        
        for claim in current_page_claims:
            if claim.linked_email_id:
                del_obj = delegations_map.get(str(claim.linked_email_id))
                if del_obj:
                    endpoint = f"messages/{del_obj.email_id}?$select=subject,from,body,receivedDateTime"
                    email_data = OutlookGraphService._make_graph_request(endpoint, target_email)
                    if 'error' not in email_data:
                        claim.email_preview_subject = email_data.get('subject')
                        claim.email_preview_sender = email_data.get('from', {}).get('emailAddress', {}).get('address')
                        claim.email_preview_body = email_data.get('body', {}).get('content')
                        claim.email_preview_date = email_data.get('receivedDateTime')

    return render(request, 'acvv_app/global_claims.html', {
        'claims': current_page_claims,   # The specific 36 items to display
        'page_obj': page_obj,            # The paginator object needed for the HTML buttons
        'all_companies': Globalacvv.objects.values('mip_names', 'branch_code'),
        'my_delegated_emails': EmailDelegation.objects.filter(assigned_user=request.user).exclude(status='DLT'),
        'is_two_pot_view': False 
    })

@login_required
def save_global_claim(request):
    """Unified save view with note handling, file attachments, and email with attachments."""
    if request.method == 'POST':
        claim_id = request.POST.get('claim_id')
        claim_type = request.POST.get('claim_type') 
        company_code = request.POST.get('company_code')
        
        linked_id = request.POST.get('linked_email_id') or None

        # Clean decimal/numeric entries to prevent validation crashes
        claim_amount_raw = request.POST.get('claim_amount')
        claim_amount_val = None
        if claim_amount_raw and claim_amount_raw.strip():
            try:
                claim_amount_val = float(claim_amount_raw.replace(',', '').strip())
            except ValueError:
                pass

        # Helper function to ensure empty date strings are handled as None/NULL safely
        def clean_date_input(val):
            if val and val.strip() and val.strip() != 'None' and val.strip() != '':
                return val.strip()
            return None

        # Extract your newly added template parameters safely so they don't break the base save
        form_qualified = request.POST.get('qualified') or "YES"
        form_date_submitted = clean_date_input(request.POST.get('date_submitted'))
        form_informed_er = request.POST.get('informed_er') or "NO"
        
        # --- UPDATED AGENT LOGIC: Prioritize dropdown, fallback to user auto-assignment ---
        form_submitted_by_agent = request.POST.get('agent')
        
        if not form_submitted_by_agent:
            current_username = request.user.username.lower()
            current_firstname = request.user.first_name.lower() if request.user.first_name else ""
            
            if 'jesica' in current_username or 'jessica' in current_username or 'jesica' in current_firstname or 'jessica' in current_firstname:
                form_submitted_by_agent = "JH"
            elif 'timothy' in current_username or 'timothy' in current_firstname:
                form_submitted_by_agent = "TD"
        # -----------------------------------------------------

        # --- EXTRACT POT AVAILABILITY FIELDS ---
        # Checkboxes return 'on' if checked, otherwise they are missing from POST
        vested_pot_available = request.POST.get('vested_pot_available') == 'on'
        savings_pot_available = request.POST.get('savings_pot_available') == 'on'
        
        vested_pot_paid_date = clean_date_input(request.POST.get('vested_pot_paid_date'))
        savings_pot_paid_date = clean_date_input(request.POST.get('savings_pot_paid_date'))
        infund_cert_date = clean_date_input(request.POST.get('infund_cert_date'))

        # --- EXTRACT BASE FIELD ATTRIBUTES CLEANLY FOR BASE CLAIM OBJECT ---
        data = {
            'company_code': company_code,
            # Map form dropdown field directly back onto database column
            'agent': form_submitted_by_agent, 
            'id_number': request.POST.get('id_number'),
            'member_name': request.POST.get('member_name'),
            'member_surname': request.POST.get('member_surname'),
            'mip_number': request.POST.get('mip_number'),
            'claim_type': claim_type,
            'claim_status': request.POST.get('claim_status'),
            'payment_option': request.POST.get('payment_option'),
            
            # --- UPDATED: Mapping added for Reason For Exit & Claim Allocation (Added fallback) ---
            'exit_reason': request.POST.get('exit_reason'),
            'claim_allocation': request.POST.get('claim_allocation') or "New Claim",
            
            'claim_amount': claim_amount_val,
            'claim_created_date': clean_date_input(request.POST.get('claim_created_date')),
            # Map form date input field directly back onto database column 
            'date_submitted': form_date_submitted,
            'last_contribution_date': clean_date_input(request.POST.get('last_contribution_date')),
            'date_paid': clean_date_input(request.POST.get('date_paid')),
            'linked_email_id': linked_id,
            
            # --- MAP POT FIELDS ---
            'vested_pot_available': vested_pot_available,
            'savings_pot_available': savings_pot_available,
            'vested_pot_paid_date': vested_pot_paid_date,
            'savings_pot_paid_date': savings_pot_paid_date,
            'infund_cert_date': infund_cert_date,
        }

        # --- 1. SAVE OR UPDATE THE CLAIM WITH TRY/EXCEPT DEBUGGING ---
        claim_obj = None
        try:
            if claim_id and claim_id.strip():
                AcvvClaim.objects.filter(id=claim_id).update(**data)
                claim_obj = AcvvClaim.objects.get(id=claim_id)
                messages.success(request, f"Claim for {claim_obj.member_surname} updated.")
            else:
                claim_obj = AcvvClaim.objects.create(**data)
                messages.success(request, f"New {claim_type} claim created successfully.")
        except Exception as e:
            # This prints the precise reason (e.g., Data truncation, null constraint) to your terminal console
            print(f"\n❌ [CRITICAL DATABASE SAVE ERROR]: {str(e)}\n")
            messages.error(request, f"Database Save Failed: {str(e)}")
            # Fallback redirect so you can see the message error banner
            if claim_type == 'Two Pot':
                return redirect('global_two_pot')
            return redirect('global_claims')

        # 2. HANDLE CLAIM NOTES & INTERNAL ATTACHMENTS (Enriched with tracking parameters)
        note_selection = request.POST.get('note_selection')
        note_description = request.POST.get('note_description') or ""
        internal_attachment = request.FILES.get('claim_attachment')

        # Format tracking context info into the note field summary to keep history clean
        tracking_metadata_summary = (
            f"\n\n[Tracking Parameters Logged]:\n"
            f"- Qualified: {form_qualified}\n"
            f"- Date Submitted Online: {form_date_submitted or 'N/A'}\n"
            f"- Informed ER: {form_informed_er}\n"
            f"- Submitted by Agent: {form_submitted_by_agent or 'N/A'}"
        )
        
        enriched_note_details = note_description + tracking_metadata_summary

        if claim_obj:
            try:
                # If no explicit selection type was provided by dropdown, set a helpful tracking fallback label
                resolved_note_selection = note_selection if note_selection else "TRACKING METADATA UPDATE"
                
                ClaimNote.objects.create(
                    claim=claim_obj,
                    note_selection=resolved_note_selection,
                    note_description=enriched_note_details,
                    attachment=internal_attachment,
                    created_by=request.user
                )
                messages.info(request, "Internal claim note saved.")
            except Exception as note_err:
                print(f"⚠️ [NON-CRITICAL NOTE SAVE ERROR]: {str(note_err)}")

        # 3. HANDLE OUTGOING EMAIL LOGIC WITH ATTACHMENTS
        recipient = request.POST.get('member_recipient_email')
        subject = request.POST.get('member_email_subject_reply')
        body = request.POST.get('email_body_html_content')
        
        email_attachment = request.FILES.get('email_attachment')

        if recipient and subject and body:
            target_email = settings.OUTLOOK_EMAIL_ADDRESS
            
            result = OutlookGraphService.send_outlook_email(
                target_email, 
                recipient, 
                subject, 
                body, 
                content_type='Html',
                attachment=email_attachment,
                user=request.user
            )
            
            if result.get('success'):
                acvv_record = Globalacvv.objects.filter(
                    Q(mip_names=company_code) | Q(branch_code=company_code)
                ).first()
                
                resolved_mip_name = acvv_record.mip_names if acvv_record else company_code
                note_selection_type = note_selection if note_selection else "Correspondence"
                
                if claim_type == 'Two Pot':
                    resolved_comm_type = 'Two Pot Email'
                else:
                    resolved_comm_type = 'Claim Sent'
                
                new_ms_id = result.get('message_id') or f"CLAIM-{timezone.now().timestamp()}"
                
                try:
                    EmailDelegation.objects.create(
                        email_id=new_ms_id,
                        subject=subject,
                        body=body,
                        attachment=email_attachment,
                        sender_address=target_email,
                        assigned_user=request.user,
                        status='SENT',
                        mip_names=resolved_mip_name, 
                        received_at=timezone.now(),
                        delegated_at=timezone.now(),
                        work_related=True,
                        communication_type=resolved_comm_type
                    )

                    if acvv_record:
                        ClientNotes.objects.create(
                            acvv_record=acvv_record,
                            notes=f"Email Sent: {subject}\nRecipient: {recipient}" + (f" (with attachment: {email_attachment.name})" if email_attachment else ""),
                            user=request.user.username,
                            date=timezone.now(),
                            communication_type="Email",
                            action_note_type=note_selection_type
                        )
                except Exception as log_err:
                    print(f"⚠️ [NON-CRITICAL EMAIL LOG ERROR]: {str(log_err)}")
                    
                messages.success(request, f"Email sent successfully to {recipient}.")
            else:
                messages.error(request, f"Email failed: {result.get('error')}")

        # Final Redirects
        if claim_type == 'Two Pot':
            return redirect('global_two_pot')
        return redirect('global_claims')

@login_required
def export_global_claims_excel(request):
    """
    Exports claims to the standard Register format (Green Theme) matching the attachment.
    Excludes 'Two Pot' claims. Includes newly added Pot fields.
    """
    query = request.GET.get('q')
    
    # 1. Fetch claims excluding 'Two Pot'
    claims = AcvvClaim.objects.all().exclude(claim_type='Two Pot').order_by('claim_created_date')
    
    if query:
        claims = claims.filter(
            Q(id_number__icontains=query) | 
            Q(member_surname__icontains=query) | 
            Q(company_code__icontains=query)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Claims Register"

    # Define Green Theme Styles (Matching the provided image)
    # Using a stronger standard Excel green matching the photo
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    # Standard thin borders for all cells
    border = Border(left=Side(style='thin', color='000000'), 
                    right=Side(style='thin', color='000000'), 
                    top=Side(style='thin', color='000000'), 
                    bottom=Side(style='thin', color='000000'))
                    
    header_font = Font(bold=True, size=11, color="000000")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_alignment = Alignment(vertical='center')

    # 2. Header Row Titles (Matching Image + New Pot Fields)
    headers = [
        'Co.Code', 'Branch', 'Agent', 'MIP Number', 'ID Number', 
        'Name', 'Surname', 'Type', 'Status', 'Exit Reason', 
        'Created', 'Submitted', 'Paid', 'Last Reconciled', 'Claim Allocation',
        'Vested Pot Available', 'Vested Pot Paid Date', 
        'Savings Pot Available', 'Savings Pot Paid Date', 
        'In-Fund Cert Date'
    ]
    
    ws.append(headers)
    
    # Apply styling to headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = green_fill
        cell.alignment = header_alignment
        cell.border = border

    # Freeze the top row so headers stay visible when scrolling down
    ws.freeze_panes = "A2"

    # 3. Data Rows
    for c in claims:
        row = [
            c.company_code,                                                      # Co Code
            '',                                                                  # Branch
            c.agent if hasattr(c, 'agent') else '',                              # Agent
            c.mip_number if hasattr(c, 'mip_number') else '',                    # MIP Number
            c.id_number,                                                         # ID Number
            c.member_name,                                                       # Name
            c.member_surname,                                                    # Surname
            c.claim_type,                                                        # Type
            c.claim_status,                                                      # Status
            c.exit_reason if hasattr(c, 'exit_reason') else '',                  # Exit Reason
            c.claim_created_date.strftime('%Y-%m-%d') if c.claim_created_date else '', # Created
            c.date_submitted.strftime('%Y-%m-%d') if hasattr(c, 'date_submitted') and c.date_submitted else '', # Submitted
            c.date_paid.strftime('%Y-%m-%d') if hasattr(c, 'date_paid') and c.date_paid else '',              # Paid
            c.last_reconciled.strftime('%Y-%m-%d') if hasattr(c, 'last_reconciled') and c.last_reconciled else '', # Last Reconciled
            c.claim_allocation if hasattr(c, 'claim_allocation') else '',         # Claim Allocation
            
            # --- NEW POT FIELDS ---
            'Yes' if getattr(c, 'vested_pot_available', False) else 'No',        # Vested Pot Available
            c.vested_pot_paid_date.strftime('%Y-%m-%d') if getattr(c, 'vested_pot_paid_date', None) else '', # Vested Paid Date
            'Yes' if getattr(c, 'savings_pot_available', False) else 'No',       # Savings Pot Available
            c.savings_pot_paid_date.strftime('%Y-%m-%d') if getattr(c, 'savings_pot_paid_date', None) else '', # Savings Paid Date
            c.infund_cert_date.strftime('%Y-%m-%d') if getattr(c, 'infund_cert_date', None) else '',         # In-Fund Cert Date
        ]
        ws.append(row)
        
        # Apply borders and alignment to the data row
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = data_alignment

    # Apply AutoFilter to the header row (adds the dropdown arrows from the photo)
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{ws.max_row}"

    # 4. Formatting - Auto-adjust Column Widths
    column_widths = [12, 15, 15, 15, 18, 15, 15, 12, 15, 15, 12, 12, 12, 15, 20, 18, 18, 18, 18, 18]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = width

    # Final response
    now = timezone.now()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Claims_Register_{now.strftime("%Y-%m-%d")}.xlsx"'
    wb.save(response)
    
    return response

@login_required
def recycle_bin_view(request):
    """
    Displays items marked as 'Deleted' (DLT).
    """
    if request.user.username.lower() != 'omega' and not request.user.is_superuser:
        messages.error(request, "Access restricted.")
        return redirect('dashboard')

    # UPDATED: Filter by the new 'DLT' status and work_related=False
    recycled_items = EmailDelegation.objects.filter(
        status='DLT', 
        work_related=False
    ).order_by('-received_at')
    
    target_email = settings.OUTLOOK_EMAIL_ADDRESS 
    tasks = []

    for item in recycled_items:
        endpoint = f"messages/{item.email_id}?$select=subject,from"
        email_data = OutlookGraphService._make_graph_request(endpoint, target_email)
        
        if 'error' not in email_data:
            subject = email_data.get('subject')
            sender = email_data.get('from', {}).get('emailAddress', {}).get('name', 'Unknown')
        else:
            subject = "[Email no longer in Outlook]"
            sender = "N/A"

        tasks.append({
            'delegation_id': item.pk,
            'subject': subject,
            'from': sender,
            'received': item.received_at
        })

    return render(request, 'acvv_app/recycle_bin.html', {'tasks': tasks})

@login_required
def delete_recycled_item(request, delegation_id):
    """Permanently removes an item from the local database."""
    if request.user.username.lower() == 'omega' or request.user.is_superuser:
        item = get_object_or_404(EmailDelegation, pk=delegation_id)
        item.delete()
        messages.success(request, "Item permanently removed from Recycle Bin.")
    return redirect('recycle_bin')

@login_required
def restore_recycled_item(request, delegation_id):
    """
    Moves an item back to the Live Inbox.
    Resets status from 'DLT' back to 'NEW'.
    """
    if request.user.username.lower() == 'omega' or request.user.is_superuser:
        item = get_object_or_404(EmailDelegation, pk=delegation_id)
        item.work_related = True
        item.status = 'NEW'  # Resetting to NEW status
        item.save()
        messages.success(request, "Item restored to the Live Inbox.")
    return redirect('recycle_bin')

@login_required
def view_recycled_item(request, delegation_id):
    """View details of a recycled item."""
    delegation = get_object_or_404(EmailDelegation, pk=delegation_id)
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    endpoint = f"messages/{delegation.email_id}"
    email_data = OutlookGraphService._make_graph_request(endpoint, target_email)

    if 'error' in email_data:
        messages.error(request, "Error fetching email content.")
        return redirect('recycle_bin')

    return render(request, 'acvv_app/view_recycled_item.html', {
        'delegation': delegation,
        'email': email_data
    })

@login_required
def bulk_delete_recycled(request):
    """Handles multiple deletions based on DLT status."""
    if request.method == 'POST':
        item_ids = request.POST.getlist('selected_items')
        if 'empty_bin' in request.POST:
            # UPDATED: Bulk delete only items marked as DLT
            EmailDelegation.objects.filter(status='DLT').delete()
            messages.success(request, "Recycle Bin emptied.")
        elif item_ids:
            EmailDelegation.objects.filter(pk__in=item_ids, status='DLT').delete()
            messages.success(request, f"Deleted {len(item_ids)} items.")
    return redirect('recycle_bin')

@login_required
def outlook_view_thread(request, delegation_id):
    """
    Displays the email content and audit trail. 
    FIXED: Now pulls actual body and attachments for local SENT records.
    """
    import os
    from django.utils.safestring import mark_safe

    # 1. Flexible Lookup (Handles both PK and Microsoft ID)
    if str(delegation_id).isdigit():
        task = get_object_or_404(EmailDelegation, Q(id=delegation_id) | Q(email_id=delegation_id))
    else:
        task = get_object_or_404(EmailDelegation, email_id=delegation_id)

    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    email_content = ""
    attachments = []

    # 2. LOCAL vs MICROSOFT LOGIC
    if task.email_id.startswith('SENT-') or task.email_id.startswith('LOCAL-'):
        # --- NEW: PULL FROM LOCAL DATABASE FIELDS ---
        # We replace the hardcoded "Live version not available" text with your data
        if task.body:
            email_content = task.body
        else:
            email_content = f"<div class='alert alert-warning'><strong>No Body Recorded:</strong> This email was logged with the subject '{task.subject}', but no body content was found.</div>"
        
        # Format the local attachment for the template loop
        if task.attachment:
            attachments = [{
                'name': os.path.basename(task.attachment.name),
                'url': task.attachment.url,
                'contentType': 'application/octet-stream', 
                'is_local': True 
            }]
    else:
        # --- LIVE MICROSOFT FETCH (For original incoming emails) ---
        endpoint = f"messages/{task.email_id}"
        email_data = OutlookGraphService._make_graph_request(endpoint, target_email)
        
        attachment_endpoint = f"messages/{task.email_id}/attachments"
        attachment_data = OutlookGraphService._make_graph_request(attachment_endpoint, target_email)
        
        attachments = attachment_data.get('value', [])
        email_content = email_data.get('body', {}).get('content')

    # 3. Fetch local Audit Trail
    actions = DelegationTransactionLog.objects.filter(delegation=task).order_by('transaction_time')

    context = {
        'task': task,
        'email_body': mark_safe(email_content) if email_content else "No content available.",
        'attachments': attachments,
        'actions': actions,
    }
    return render(request, 'acvv_app/outlook_view_thread.html', context)

@login_required
def export_temp_exists(request):
    """
    Generates and exports an Excel file by pulling data from the temp_exit table.
    """
    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Temp Exists"

    # Define headers
    headers = [
        "MG Code", "Surname", "Initials", "MIP No.", "ID No.", 
        "Reason", "BIS From Date", "BIS End Date", 
        "Full Contributions Start Date", "Note"
    ]

    # Write headers and format them
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_title
        cell.font = openpyxl.styles.Font(bold=True)

    # --- NEW: PULL DATA FROM DATABASE ---
    # Query all records from the manually created temp_exit table
    records = TempExit.objects.all().order_by('-created_at')

    # Write data rows
    for row_num, obj in enumerate(records, 2):
        ws.cell(row=row_num, column=1).value = obj.mg_code
        ws.cell(row=row_num, column=2).value = obj.surname
        ws.cell(row=row_num, column=3).value = obj.initials
        ws.cell(row=row_num, column=4).value = obj.mip_no
        ws.cell(row=row_num, column=5).value = obj.id_no
        ws.cell(row=row_num, column=6).value = obj.reason
        ws.cell(row=row_num, column=7).value = obj.bis_from_date
        ws.cell(row=row_num, column=8).value = obj.bis_end_date
        ws.cell(row=row_num, column=9).value = obj.full_contributions_start_date
        ws.cell(row=row_num, column=10).value = obj.note

    # Auto-adjust column widths for better readability
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Prepare the response
    filename = f"Temp_Exists_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def reconciliation_worksheet(request):
    today = timezone.now().date()
    
    # 1. NAVIGATION LOGIC
    first_of_this_month = today.replace(day=1)
    last_month_date = (first_of_this_month - timedelta(days=1)).replace(day=1)

    req_year = request.GET.get('year')
    req_month = request.GET.get('month')
    
    if req_year and req_month:
        try:
            current_fiscal = datetime.strptime(f"{req_year}-{req_month}-01", '%Y-%m-%d').date()
        except ValueError:
            current_fiscal = first_of_this_month
    else:
        # UPDATED: Changed default from last_month_date to first_of_this_month
        current_fiscal = first_of_this_month

    # 2. HANDLE POST ACTIONS (Save/Close/Reopen)
    if request.method == 'POST':
        if 'save_changes' in request.POST:
            row_ids = {key.split('_')[2] for key in request.POST.keys() if key.startswith('recon_status_')}
            for row_id in row_ids:
                row_obj = ReconciliationWorksheet.objects.filter(pk=row_id).first()
                if row_obj and not row_obj.is_closed:
                    row_obj.company_status = request.POST.get(f'company_status_{row_id}')
                    row_obj.payment_method = request.POST.get(f'payment_method_{row_id}')
                    row_obj.arrears = request.POST.get(f'arrears_{row_id}', '')
                    row_obj.member_count_reconciled = request.POST.get(f'member_count_{row_id}', 0) or 0
                    row_obj.contribution_amount_reconciled = request.POST.get(f'amount_{row_id}', 0.00) or 0.00
                    row_obj.lpi_amount = request.POST.get(f'lpi_amount_{row_id}', 0.00) or 0.00
                    row_obj.lpi_reason = request.POST.get(f'lpi_reason_{row_id}')
                    row_obj.debit_order_success = request.POST.get(f'debit_success_{row_id}')
                    row_obj.reconciled_status = request.POST.get(f'recon_status_{row_id}')
                    row_obj.date_schedule_received = request.POST.get(f'schedule_{row_id}') or None
                    row_obj.date_confirmed_on_step = request.POST.get(f'confirmed_{row_id}') or None
                    row_obj.debit_order_date = request.POST.get(f'debit_{row_id}') or None
                    row_obj.save()
            messages.success(request, "Progress saved successfully.")

        elif 'close_month' in request.POST:
            if not request.user.is_superuser:
                messages.error(request, "Permission Denied: Only Superusers can finalize a month.")
            else:
                records_to_close = ReconciliationWorksheet.objects.filter(fiscal_month=current_fiscal)
                for rec in records_to_close:
                    if rec.reconciled_status == 'Reconciled':
                        master_obj = Globalacvv.objects.filter(mip_names=rec.mg_name).first()
                        if master_obj and master_obj.notes:
                            try:
                                last_date = datetime.strptime(master_obj.notes, "%d.%m.%Y").date()
                                next_month_date = (last_date.replace(day=28) + timedelta(days=5)).replace(day=1)
                                new_note_val = next_month_date.strftime("%d.%m.%Y")
                            except:
                                new_note_val = current_fiscal.strftime("01.%m.%Y")
                        else:
                            new_note_val = current_fiscal.strftime("01.%m.%Y")
                        Globalacvv.objects.filter(mip_names=rec.mg_name).update(notes=new_note_val)

                records_to_close.update(is_closed=True, closed_at=timezone.now())
                messages.success(request, f"Fiscal month {current_fiscal.strftime('%B %Y')} closed.")
            return redirect(f"{reverse('reconciliation_worksheet')}?year={current_fiscal.year}&month={current_fiscal.month}")

        elif 'reopen_month' in request.POST:
            if not request.user.is_superuser:
                messages.error(request, "Permission Denied: Only Superusers can re-open a month.")
            else:
                ReconciliationWorksheet.objects.filter(fiscal_month=current_fiscal).update(is_closed=False, closed_at=None)
                messages.warning(request, f"Fiscal month {current_fiscal.strftime('%B %Y')} RE-OPENED.")
            return redirect(f"{reverse('reconciliation_worksheet')}?year={current_fiscal.year}&month={current_fiscal.month}")

    # 3. FETCH & AUTO-GENERATE DATA
    records = ReconciliationWorksheet.objects.filter(
        Q(fiscal_month=current_fiscal) | 
        Q(fiscal_month__lt=current_fiscal, reconciled_status='Unreconciled', is_closed=False)
    ).order_by('mg_code', 'fiscal_month')
    
    if not records.filter(fiscal_month=current_fiscal).exists():
        for item in Globalacvv.objects.all():
            clean_code = str(item.branch_code).strip().upper()
            ReconciliationWorksheet.objects.get_or_create(
                fiscal_month=current_fiscal,
                mg_code=clean_code, 
                defaults={'mg_name': item.mip_names}
            )
        records = ReconciliationWorksheet.objects.filter(
            Q(fiscal_month=current_fiscal) | 
            Q(fiscal_month__lt=current_fiscal, reconciled_status='Unreconciled', is_closed=False)
        ).order_by('mg_code', 'fiscal_month')

    # 4. ANNOTATIONS
    last_ws_recon_sub = ReconciliationWorksheet.objects.filter(
        mg_name=OuterRef('mg_name'),
        mg_code=OuterRef('mg_code'),
        fiscal_month__lt=current_fiscal,
        reconciled_status='Reconciled'
    ).order_by('-fiscal_month').values('fiscal_month')[:1]

    records = records.annotate(
        acvv_member_count=Subquery(Globalacvv.objects.filter(mip_names=OuterRef('mg_name')).values('member')[:1]),
        master_start_date=Subquery(Globalacvv.objects.filter(mip_names=OuterRef('mg_name')).values('notes')[:1]),
        last_reconciled_ws=Subquery(last_ws_recon_sub)
    )

    for r in records:
        last_date = None
        
        # PRIORITY 1: Use the actual database field 'last_fiscal_reconciled'
        if r.last_fiscal_reconciled and r.last_fiscal_reconciled.strip():
            r.last_fiscal_display = r.last_fiscal_reconciled
            try:
                last_date = datetime.strptime(r.last_fiscal_reconciled, "%B %Y").date()
            except:
                last_date = None
        
        # PRIORITY 2: Fallback to Subquery (automated history)
        elif r.last_reconciled_ws:
            last_date = r.last_reconciled_ws
            r.last_fiscal_display = last_date.strftime("%B %Y")
        
        # PRIORITY 3: Fallback to Master Notes
        elif r.master_start_date:
            try:
                last_date = datetime.strptime(str(r.master_start_date).strip(), "%d.%m.%Y").date()
                r.last_fiscal_display = last_date.strftime("%B %Y")
            except:
                last_date = None
                r.last_fiscal_display = r.master_start_date
        else:
            r.last_fiscal_display = "No Data"

        # Overdue logic
        if last_date:
            diff = (current_fiscal.year - last_date.year) * 12 + (current_fiscal.month - last_date.month)
            r.is_overdue = diff >= 2
        else:
            r.is_overdue = True

    return render(request, 'acvv_app/reconciliation_worksheet.html', {
        'records': records,
        'display_name': current_fiscal.strftime("%B %Y"),
        'history': ReconciliationWorksheet.objects.values('fiscal_month').distinct().order_by('-fiscal_month'),
        'is_closed': ReconciliationWorksheet.objects.filter(fiscal_month=current_fiscal, is_closed=True).exists(),
        'current_fiscal': current_fiscal
    })

@login_required
def export_reconciliation_worksheet(request, date_str):
    try:
        fiscal_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    last_ws_recon_sub = ReconciliationWorksheet.objects.filter(
        mg_name=OuterRef('mg_name'),
        fiscal_month__lt=fiscal_date,
        reconciled_status='Reconciled'
    ).order_by('-fiscal_month').values('fiscal_month')[:1]

    acvv_member_sub = Globalacvv.objects.filter(mip_names=OuterRef('mg_name')).values('member')[:1]
    acvv_notes_sub = Globalacvv.objects.filter(mip_names=OuterRef('mg_name')).values('notes')[:1]
    # --- ADDED: Subquery to fetch the email address from Globalacvv ---
    acvv_email_sub = Globalacvv.objects.filter(mip_names=OuterRef('mg_name')).values('mg_email_address')[:1]

    records = ReconciliationWorksheet.objects.filter(fiscal_month=fiscal_date).annotate(
        acvv_member_count=Subquery(acvv_member_sub),
        master_start_date=Subquery(acvv_notes_sub),
        last_reconciled_ws=Subquery(last_ws_recon_sub),
        # --- ADDED: Annotate the email address onto the record ---
        acvv_email_address=Subquery(acvv_email_sub)
    )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Recon {fiscal_date.strftime('%b %Y')}"
    
    # --- ADDED: "Email Address" to the headers list ---
    headers = [
        "MG Name", "MG Code", "Email Address", "Company Status", "Payment Method", 
        "Last Fiscal Reconciled", "Arrears", "Member Count Reconciled", 
        "Contribution Amount Reconciled", "Reconciled Status", 
        "Date Schedule Received", "Date Confirmed on Step", "Debit order date"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for r in records:
        display_status = r.company_status if r.company_status else "Active"
        if display_status.lower() in ['done', 'active']:
            display_status = "Active"

        if r.last_reconciled_ws:
            display_recon = r.last_reconciled_ws.strftime("%B %Y")
        elif r.master_start_date:
            try:
                d = datetime.strptime(r.master_start_date, "%d.%m.%Y").date()
                display_recon = d.strftime("%B %Y")
            except:
                display_recon = r.master_start_date
        else:
            display_recon = "No Data"

        # --- ADDED: r.acvv_email_address to the row output ---
        ws.append([
            r.mg_name, r.mg_code, r.acvv_email_address, display_status, r.payment_method,
            display_recon, r.arrears,
            r.member_count_reconciled or r.acvv_member_count,
            r.contribution_amount_reconciled, r.reconciled_status,
            r.date_schedule_received, r.date_confirmed_on_step, r.debit_order_date
        ])
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="Reconciliation_Worksheet_{date_str}.xlsx"'
    wb.save(response)
    return response

@login_required
def export_reconciliation(request, date_str):
    try:
        fiscal_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    records = ReconciliationRecord.objects.filter(fiscal_month=fiscal_date)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Summary {fiscal_date.strftime('%b %Y')}"
    headers = ["Member Group", "Branch Code", "Billed", "Paid", "Outstanding", "Note"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for r in records:
        ws.append([r.mip_name, r.branch_code, r.billed_amount, r.paid_amount, r.outstanding_amount, r.note])
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="Reconciliation_Summary_{date_str}.xlsx"'
    wb.save(response)
    return response

@login_required
def outlook_email_list(request):
    """
    View to display NEW, DELEGATED (DEL), and COMPLETED (COM) emails with date filtering.
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # --- FIX: Expanded statuses inclusion list filter matrix targets ---
    emails = EmailDelegation.objects.filter(
        status__in=['NEW', 'DEL', 'COM']
    ).select_related('assigned_user').order_by('-received_at')

    # Apply date filters if provided
    if start_date and end_date:
        emails = emails.filter(received_at__date__range=[start_date, end_date])

    # Status counts for badges (Calculated post base filter execution alignment)
    new_count = EmailDelegation.objects.filter(status='NEW').count()
    del_count = EmailDelegation.objects.filter(status='DEL').count()
    # --- ADDED: Track and expose Completed status metrics safely ---
    com_count = EmailDelegation.objects.filter(status='COM').count()

    context = {
        'emails': emails,
        'new_count': new_count,
        'del_count': del_count,
        'com_count': com_count,  # 👈 Available for badge counters inside your HTML layout view template
    }
    return render(request, 'acvv_app/outlook_email_list.html', context)

@login_required
def export_email_tasks_excel(request):
    """
    Exports the filtered email task list to Excel.
    Includes NEW, DEL, and COM (Completed) records.
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # 🔑 FIX: Added 'COM' to include completed tasks inside your spreadsheet export query matrix
    emails = EmailDelegation.objects.filter(status__in=['NEW', 'DEL', 'COM']).order_by('-received_at')
    
    if start_date and end_date:
        emails = emails.filter(received_at__date__range=[start_date, end_date])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Email Task Export"

    # Styles
    header_fill = PatternFill(start_color="43A047", end_color="43A047", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['Status', 'Received Date', 'Subject', 'Sender', 'Member Group', 'Assigned To', 'Category']
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for e in emails:
        ws.append([
            e.get_status_display(),
            e.received_at.strftime('%Y-%m-%d %H:%M') if e.received_at else '',
            e.subject,
            e.sender_address,
            e.mip_names,
            e.assigned_user.username if e.assigned_user else 'Unassigned',
            e.email_category
        ])

    # Column Widths
    widths = [15, 20, 40, 30, 20, 20, 20]
    for i, width in enumerate(widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Email_Tasks_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def temp_exists_list(request):
    # Handle New Entry or Update Submission
    if request.method == 'POST':
        exit_id = request.POST.get('exit_id')
        
        data_fields = {
            'mg_code': request.POST.get('mg_code'),
            'surname': request.POST.get('surname'),
            'initials': request.POST.get('initials'),
            'mip_no': request.POST.get('mip_no'),
            'id_no': request.POST.get('id_no'),
            'reason': request.POST.get('reason'),
            'bis_from_date': request.POST.get('bis_from') or None,
            'bis_end_date': request.POST.get('bis_end') or None,
            'full_contributions_start_date': request.POST.get('full_start') or None,
            'note': request.POST.get('note')
        }

        if exit_id:
            # Update existing record
            exit_record = get_object_or_404(TempExit, pk=exit_id)
            for field, value in data_fields.items():
                setattr(exit_record, field, value)
            exit_record.save()
            messages.success(request, "Temp Exit updated successfully.")
        else:
            # Create new record
            TempExit.objects.create(**data_fields)
            messages.success(request, "Temp Exit added successfully.")
            
        return redirect('temp_exists_list')

    # Display existing entries
    exits = TempExit.objects.all().order_by('-created_at')
    return render(request, 'acvv_app/temp_exists.html', {'exits': exits})

@login_required
def outlook_view_thread(request, delegation_id):
    # 1. Flexible Lookup
    if delegation_id.isdigit():
        task = get_object_or_404(EmailDelegation, Q(id=delegation_id) | Q(email_id=delegation_id))
    else:
        task = get_object_or_404(EmailDelegation, email_id=delegation_id)

    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    email_content = ""
    attachments = []

    # 2. FIX: Check if it is a local "SENT" ID before calling Graph API
    if task.email_id.startswith('SENT-'):
        email_content = f"""
            <div class='alert alert-success'>
                <strong>Local Record:</strong> This email was sent directly from the ACVV App. 
                Microsoft does not provide a live preview for this specific tracking ID.
                <br><small>Subject: {task.subject}</small>
            </div>"""
    else:
        # Only call Graph API for real Microsoft IDs
        endpoint = f"messages/{task.email_id}"
        email_data = OutlookGraphService._make_graph_request(endpoint, target_email)
        
        attachment_endpoint = f"messages/{task.email_id}/attachments"
        attachment_data = OutlookGraphService._make_graph_request(attachment_endpoint, target_email)
        attachments = attachment_data.get('value', [])
        email_content = email_data.get('body', {}).get('content')

    # 3. Fetch local Audit Trail
    actions = DelegationTransactionLog.objects.filter(delegation=task).order_by('transaction_time')

    context = {
        'task': task,
        'email_body': mark_safe(email_content) if email_content else "No content available.",
        'attachments': attachments,
        'actions': actions,
    }
    return render(request, 'acvv_app/outlook_view_thread.html', context)

@login_required
def download_acvv_email(request, delegation_id):
    """
    Fetches raw MIME content from Outlook OR generates a local .eml file 
    using the new body and attachment fields in EmailDelegation with multi-file support.
    """
    import requests
    import os
    from django.conf import settings
    from .services.outlook_graph_service import get_current_access_token
    from django.utils.text import slugify
    from email.message import EmailMessage
    from email.utils import make_msgid 
    from django.http import HttpResponse
    from django.shortcuts import get_object_or_404, redirect
    from django.contrib import messages

    # 1. ID RESOLUTION
    try:
        if str(delegation_id).isdigit():
            task = get_object_or_404(EmailDelegation, pk=delegation_id)
            ms_id = task.email_id
        else:
            task = get_object_or_404(EmailDelegation, email_id=delegation_id)
            ms_id = delegation_id
    except Exception:
        messages.error(request, "Task record not found.")
        return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

    # 2. LOCAL GENERATION (For SENT-, REPLY-, CLAIM-, or LOCAL- IDs)
    # 💡 Added ms_id.startswith('CLAIM-') to catch and format system-generated claim tasks locally
    if ms_id.startswith('SENT-') or ms_id.startswith('LOCAL-') or ms_id.startswith('REPLY-') or ms_id.startswith('CLAIM-'):
        msg = EmailMessage()
        msg['Subject'] = task.subject
        msg['From'] = task.sender_address or settings.OUTLOOK_EMAIL_ADDRESS
        msg['To'] = "Recipient (Details in App Logs)"
        msg['Date'] = task.received_at.strftime('%a, %d %b %Y %H:%M:%S +0200') if task.received_at else ""
        msg['Message-ID'] = make_msgid()
        
        body_content = task.body if task.body else "Body content not found in record."
        msg.set_content(body_content)

        # --- MULTI-ATTACHMENT SYSTEM LOGIC ---
        attached_paths = []

        # Step A: Pick up the baseline attachment path
        if task.attachment:
            try:
                base_path = task.attachment.path
                if os.path.exists(base_path):
                    attached_paths.append(base_path)
            except Exception as e:
                print(f"Base file resolution error: {e}")

        # Step B: Scan parent directory folder for files matching the timestamp prefix group
        if attached_paths:
            try:
                parent_dir = os.path.dirname(attached_paths[0])
                base_filename = os.path.basename(attached_paths[0])
                
                # Check if file has standard multi-upload variations or extensions inside directory
                if "_" in base_filename or "-" in base_filename:
                    # Clean lookups across files inside the exact directory path
                    for entry in os.listdir(parent_dir):
                        full_entry_path = os.path.join(parent_dir, entry)
                        if full_entry_path not in attached_paths and os.path.isfile(full_entry_path):
                            # Ensure it belongs to the same file batch family structure
                            attached_paths.append(full_entry_path)
            except Exception as e:
                print(f"Multi-attachment scanner error: {e}")

        # Step C: physically write each found element down into the generated EML binary container
        for file_path in attached_paths:
            try:
                if os.path.exists(file_path):
                    with open(file_path, 'rb') as f:
                        file_data = f.read()
                        file_name = os.path.basename(file_path)
                        
                    msg.add_attachment(
                        file_data,
                        maintype='application',
                        subtype='octet-stream',
                        filename=file_name
                    )
            except Exception as e:
                print(f"DEBUG DOWNLOAD ATTACHMENT ERROR for {file_path}: {e}")

        # Return the generated .eml file
        response = HttpResponse(msg.as_bytes(), content_type='message/rfc822')
        response['Content-Disposition'] = f'attachment; filename="CLAIM_{slugify(task.subject)[:30] or "record"}.eml"'
        return response

    # 3. LIVE MICROSOFT DOWNLOAD (Keep this for real Outlook IDs)
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    access_token = get_current_access_token()
    
    if not access_token:
        messages.error(request, "Authentication failed: Outlook access token missing.")
        return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

    url = f"https://graph.microsoft.com/v1.0/users/{target_email}/messages/{ms_id}/$value"
    headers = {'Authorization': f'Bearer {access_token}'}

    try:
        res = requests.get(url, headers=headers, timeout=10)
        if res.status_code == 200:
            django_response = HttpResponse(res.content, content_type='message/rfc822')
            django_response['Content-Disposition'] = f'attachment; filename="{slugify(task.subject)[:50]}.eml"'
            return django_response
    except Exception as e:
        messages.error(request, f"Outlook Error: {e}")

    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

def get_branch_map_acvv(claims_queryset):
    """ Helper to map company codes to formal Branch names for ACVV """
    codes = claims_queryset.values_list('company_code', flat=True).distinct()
    branches = Globalacvv.objects.filter(mip_names__in=codes)
    return {b.mip_names: b.member for b in branches}

@login_required
def export_two_pot_invoice_cecile(request):
    """
    Report 1: Cecile Invoice Format (Grey Theme)
    Filters records where the status is valid and 'qualified' property is 'YES'.
    """
    query = request.GET.get('q')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Initial Query
    claims = AcvvClaim.objects.filter(claim_type='Two Pot').order_by('claim_created_date')
    if query:
        claims = claims.filter(Q(id_number__icontains=query) | Q(member_surname__icontains=query))
    if start_date and end_date:
        claims = claims.filter(claim_created_date__range=[start_date, end_date])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cecile Invoice"

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = [
        "Date", "Initials", "Surname", "Member nr", "ID NUMBER", "Fund", 
        "Branch", "Admin From", "Qualified", "Date submitted online", 
        "Successful Loaded confirmation", "Admin Fee R33 + 15% Vat", "SUBMIT ONLINE"
    ]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, size=10)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Ensure get_branch_map_acvv is defined in your utils/services
    branch_map = get_branch_map_acvv(claims) 

    # Define the statuses that should be included in this report
    valid_statuses = [
        "PAID",
        "MEMBER EMERGENCY SAVINGS POT WITHDRAWAL SUBMITTED",
        "MEMBER EMERGENCY SAVINGS POT WITHDRAWAL REQUESTED"
    ]

    for claim in claims:
        # Check property and status against the valid list
        current_status = str(claim.claim_status).upper().strip()
        is_valid_status = current_status in valid_statuses
        is_qualified = (claim.qualified == "YES")
        
        # 🚀 GUARD FILTER: Only continue if status is valid AND qualified.
        if not (is_valid_status and is_qualified):
            continue

        initials = "".join([n[0] for n in claim.member_name.split() if n]) if claim.member_name else ""

        ws.append([
            claim.claim_created_date.strftime('%d/%m/%Y') if claim.claim_created_date else '',
            initials,
            claim.member_surname,
            claim.mip_number,
            claim.id_number,
            claim.company_code,
            branch_map.get(claim.company_code, ""),
            claim.agent or "",
            "yes",
            claim.date_submitted.strftime('%d/%m/%Y') if claim.date_submitted else '',
            "YES",
            "R37.95",
            "SUBMIT ONLINE"
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Cecile_Invoice_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def export_two_pot_tracking_acvv(request):
    """
    Report 2: Full Tracking (Yellow Theme) with Red text for 'NO' statuses.
    Matches attached image format.
    """
    query = request.GET.get('q')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    claims_queryset = AcvvClaim.objects.filter(claim_type='Two Pot').order_by('claim_created_date')

    if query:
        claims_queryset = claims_queryset.filter(
            Q(id_number__icontains=query) | Q(member_surname__icontains=query) | 
            Q(company_code__icontains=query) | Q(mip_number__icontains=query)
        )

    if start_date and end_date:
        claims_queryset = claims_queryset.filter(claim_created_date__range=[start_date, end_date])

    branch_map = get_branch_map_acvv(claims_queryset)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Two-Pot Tracking"

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    now = timezone.now()
    display_start = start_date if start_date else now.replace(day=1).strftime('%d.%m.%Y')
    display_end = end_date if end_date else now.strftime('%d.%m.%Y')
    
    # Row 1 Title (Billing Period)
    ws.merge_cells('A1:O1')
    header_cell = ws['A1']
    header_cell.value = f"Billing - Member Emergency Savings Pot Withdrawal Requested - {display_start} to {display_end}"
    header_cell.font = Font(bold=True, size=11, underline="single")
    header_cell.fill = yellow_fill
    header_cell.border = thin_border

    # Row 2 Headers (Updated to match Blue template row labels exactly)
    headers = [
        "Date application extracted from Web: Savings Form Request",  # Column B
        "Initials",                                                    # Column C
        "Surname",                                                     # Column D
        "Member number",                                               # Column E
        "ID NUMBER",                                                   # Column F
        "Company Name",                                                   # Column G
        "Number of Employees",                                                # Column H
        "Query",                                                       # Column I
        "Claim",                                                       # Column J
        "Qualified Y/N",                                               # Column K
        "Date submitted online",                                       # Column L
        "Inform Employer that the claim is succesfully loaded",       # Column M
        "Admin Front Office Application Submitted",                    # Column N
        "Note Helper"                                                  # Column O
    ]
    ws.append(headers)
    
    for cell in ws[2]:
        cell.font = Font(bold=True, size=9)
        cell.fill = yellow_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 45

    for claim in claims_queryset:
        initials = "".join([n[0] for n in claim.member_name.split() if n]) if claim.member_name else ""
        
        # Determine Qualified value explicitly based on form state
        is_paid = str(claim.claim_status).upper().strip() == "PAID"
        qualified_val = "YES" if is_paid else "NO"
        
        # Build out dynamic Note helper descriptions for Column O when NO is encountered
        note_helper = ""
        if qualified_val == "NO":
            status_str = str(claim.claim_status).strip()
            if "Already claim" in status_str:
                note_helper = "Member already claimed this financial year"
            else:
                note_helper = "Not enough funds available"

        row = [
            claim.claim_created_date.strftime('%d/%m/%Y') if claim.claim_created_date else '', # Col B
            initials,                                                                          # Col C
            claim.member_surname,                                                              # Col D
            claim.mip_number,                                                                  # Col E
            claim.id_number,                                                                   # Col F
            claim.company_code,                                                                # Col G
            branch_map.get(claim.company_code, "Unknown"),                                     # Col H
            "Savings Form Request",                                                            # Col I (Default)
            claim.claim_status or "",                                                          # Col J (Claim status field value)
            qualified_val,                                                                     # Col K
            claim.date_submitted.strftime('%d/%m/%Y') if claim.date_submitted else '',         # Col L
            "YES" if is_paid else "No",                                                        # Col M
            claim.agent or "TD",                                                               # Col N (Submitted by agent dropdown)
            note_helper                                                                        # Col O (Note Helper reasoning layout)
        ]
        ws.append(row)

        # Apply Red Text formatting targets to lines evaluating as "NO" 
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', horizontal='left')
            if qualified_val == "NO":
                cell.font = Font(size=9, color="FF0000")
            else:
                cell.font = Font(size=9)

    widths = [30, 8, 18, 14, 18, 12, 35, 22, 35, 12, 22, 22, 22, 50]
    for i, width in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Two_Pot_Full_Tracking_{now.strftime("%Y_%m_%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def send_acvv_direct_email(request, company_code):
    """
    Handles the 'Compose New Email' form.
    Correctly splits multiple recipients and handles multiple file attachments.
    """
    if request.method == 'POST':
        recipient_raw = request.POST.get('member_recipient_email', '')
        subject = request.POST.get('member_email_subject_reply')
        body = request.POST.get('email_body_html_content')
        
        # --- FIX: Capture from either action_note_type or action_log_type safely ---
        selected_action_type = request.POST.get('action_note_type') or request.POST.get('action_log_type') or "Correspondence"
        
        # --- MULTI-ATTACHMENT FIX ---
        attachments = request.FILES.getlist('email_attachments') 

        if recipient_raw and subject and body:
            target_email = settings.OUTLOOK_EMAIL_ADDRESS
            
            # Split by semicolon (;) or comma (,)
            recipient_list = [email.strip() for email in re.split('[;,]', recipient_raw) if email.strip()]
            clean_recipient_str = ", ".join(recipient_list)

            # Pass the full 'attachments' list instead of just [0]
            result = OutlookGraphService.send_outlook_email(
                target_email, 
                recipient_list, 
                subject, 
                body, 
                content_type='Html', 
                attachments=attachments,
                user=request.user
            )
            
            if result.get('success'):
                acvv_record = get_object_or_404(Globalacvv, mip_names=company_code)
                new_ms_id = result.get('message_id') or f"SENT-{timezone.now().timestamp()}"

                # FIX: Removed 'action_log_type' parameter here to prevent the TypeError
                EmailDelegation.objects.create(
                    email_id=new_ms_id,
                    subject=subject,
                    body=body, 
                    attachment=attachments[0] if attachments else None, 
                    sender_address=target_email,
                    assigned_user=request.user,
                    status='SENT',
                    mip_names=acvv_record.mip_names,
                    received_at=timezone.now(),
                    delegated_at=timezone.now(),
                    work_related=True,
                    communication_type='Email'
                )
                
                # Keep it active here where your client_notes MySQL schema expects it!
                ClientNotes.objects.create(
                    acvv_record=acvv_record,
                    notes=f"Email Sent: {subject}\nRecipient: {clean_recipient_str}",
                    user=request.user.username,
                    date=timezone.now(),
                    communication_type="Email",
                    action_note_type=selected_action_type
                )
                
                messages.success(request, f"Email sent successfully to {clean_recipient_str}.")
            else:
                messages.error(request, f"Email failed: {result.get('error')}")
        else:
            messages.warning(request, "Please fill in all required fields.")

    return redirect('acvv_information', mip_names=company_code)

@login_required
def download_outlook_attachment(request, delegation_id, attachment_id):
    """
    Fetches the attachment from Outlook Graph API, decodes it, and serves it.
    """
    import base64
    from .services.outlook_graph_service import _make_graph_request

    delegation = get_object_or_404(EmailDelegation, pk=delegation_id)
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    
    endpoint = f"messages/{delegation.email_id}/attachments/{attachment_id}"
    attachment_data = _make_graph_request(endpoint, target_email)
    
    if 'error' in attachment_data:
        messages.error(request, "Could not fetch attachment from Outlook.")
        return redirect(request.META.get('HTTP_REFERER', 'outlook_delegated_box'))

    file_content = base64.b64decode(attachment_data.get('contentBytes'))
    file_name = attachment_data.get('name', 'attachment')
    content_type = attachment_data.get('contentType', 'application/octet-stream')

    response = HttpResponse(file_content, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    return response

@login_required
def acvv_sla_report_view(request):
    # Rights Check
    if not request.user.is_superuser:
        messages.error(request, "Access Denied: Superuser rights required for SLA Reports.")
        return redirect('dashboard')

    start_str = request.GET.get('start_date')
    end_str = request.GET.get('end_date')
    
    def get_valid_date(date_val):
        if date_val and date_val != "None" and date_val != "":
            return parse_date(date_val)
        return None

    # 1. Setup Filters
    recon_q = Q()
    email_q = Q() # Added for Emails
    start_dt = get_valid_date(start_str)
    end_dt = get_valid_date(end_str)

    if start_dt:
        recon_q &= Q(updated_at__date__gte=start_dt)
        email_q &= Q(delegated_at__date__gte=start_dt) # Added for Emails
    if end_dt:
        recon_q &= Q(updated_at__date__lte=end_dt)
        email_q &= Q(delegated_at__date__lte=end_dt) # Added for Emails

    # 2. Handle Excel Export
    if request.GET.get('export') == 'excel':
        # Updated to pass both Q objects
        return export_acvv_sla_excel(recon_q, email_q)

    # 3. Aggregate Totals (Reconciliations)
    status_breakdown = ReconciliationWorksheet.objects.filter(recon_q).values(
        'reconciled_status'
    ).annotate(total=Count('id'))

    user_breakdown = ReconciliationWorksheet.objects.filter(recon_q).values(
        'updated_by__username'
    ).annotate(total=Count('id')).order_by('-total')

    # 4. Aggregate Totals (Email Delegations) - UPDATED FOR MULTI-TYPE LOGS
    # Grouping by both status and communication_type to capture 'Claim Sent', 'Two Pot Email', 'Reply'
    email_status_breakdown = EmailDelegation.objects.filter(email_q).values(
        'status', 'communication_type'
    ).annotate(total=Count('id'))

    email_user_breakdown = EmailDelegation.objects.filter(email_q).values(
        'assigned_user__username'
    ).annotate(total=Count('id')).order_by('-total')

    # Calculate Totals
    total_recon = sum(item['total'] for item in status_breakdown)
    total_emails = sum(item['total'] for item in email_status_breakdown)

    context = {
        'status_breakdown': status_breakdown,
        'user_breakdown': user_breakdown,
        'email_status_breakdown': email_status_breakdown, # Added
        'email_user_breakdown': email_user_breakdown,     # Added
        'total_actions': total_recon + total_emails,     # Grand Total
        'total_recon': total_recon,
        'total_emails': total_emails,
        'start_date': start_str,
        'end_date': end_str,
    }
    
    return render(request, 'acvv_app/acvv_sla_report.html', context)

def export_acvv_sla_excel(recon_q, email_q):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = openpyxl.Workbook()
    
    # --- Styles ---
    header_font = Font(bold=True, color="FFFFFF")
    recon_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    email_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    center_align = Alignment(horizontal="center")

    # ==========================================
    # SHEET 1: RECONCILIATIONS (Original)
    # ==========================================
    ws1 = wb.active
    ws1.title = "ACVV Reconciliation Audit"
    headers1 = ['MG Code', 'MG Name', 'Fiscal Month', 'Status', 'Agent', 'Last Updated', 'Contribution Amount']
    ws1.append(headers1)
    
    for cell in ws1[1]:
        cell.font, cell.fill, cell.alignment = header_font, recon_fill, center_align

    queryset_recon = ReconciliationWorksheet.objects.filter(recon_q).select_related('updated_by').order_by('-updated_at')
    for r in queryset_recon:
        ws1.append([
            r.mg_code, r.mg_name,
            r.fiscal_month.strftime('%B %Y') if r.fiscal_month else '',
            r.reconciled_status,
            r.updated_by.username if r.updated_by else "System",
            r.updated_at.strftime('%Y-%m-%d %H:%M') if r.updated_at else '',
            r.contribution_amount_reconciled
        ])

    # ==========================================
    # SHEET 2: EMAIL DELEGATIONS (Updated Columns)
    # ==========================================
    ws2 = wb.create_sheet(title="Email Audit Trail")
    # --- FIX: Appended 'Log Type' column into headers matrix layout array ---
    headers2 = ['Subject', 'Sender', 'Assigned To', 'Status', 'Log Type', 'Category', 'Delegated At']
    ws2.append(headers2)

    for cell in ws2[1]:
        cell.font, cell.fill, cell.alignment = header_font, email_fill, center_align

    queryset_email = EmailDelegation.objects.filter(email_q).select_related('assigned_user').order_by('-delegated_at')
    for e in queryset_email:
        # --- FIX: Ensure incoming 'communication_type' outputs cleanly alongside fallback values ---
        resolved_log_type = e.communication_type or "Email"

        ws2.append([
            e.subject,
            e.sender_address,
            e.assigned_user.username if e.assigned_user else "Unassigned",
            e.get_status_display(),
            resolved_log_type, # ➔ Displays 'Claim Sent', 'Two Pot Email', 'Reply', etc.
            e.email_category,
            e.delegated_at.strftime('%Y-%m-%d %H:%M') if e.delegated_at else ''
        ])

    # Auto-adjust column widths for both sheets
    for sheet in [ws1, ws2]:
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 25

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="ACVV_Full_SLA_Report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def import_acvv_csv(request):
    if request.method == "POST":
        file = request.FILES.get('csv_upload')
        
        if not file:
            messages.error(request, "Please select a file.")
            return redirect('import_acvv_csv')

        try:
            # 1. Define columns to force as strings to prevent .0 decimals
            string_cols = {
                'ACCOUNT NUMBER': str, 
                'MG Code': str, 
                'NPO CODE': str,
                'MG Contact Tel. 1': str,
                'MG Contact Tel. 2': str
            }

            # 2. Check file extension and load with dtype
            if file.name.endswith('.csv'):
                df = pd.read_csv(file, dtype=string_cols)
            elif file.name.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(file, dtype=string_cols)
            else:
                messages.error(request, "Unsupported file format. Please upload CSV or Excel.")
                return redirect('import_acvv_csv')

            # 3. Clean up any remaining .0 from pandas auto-conversion
            # This turns '53361366398.0' into '53361366398'
            for col in string_cols.keys():
                if col in df.columns:
                    df[col] = df[col].astype(str).str.replace(r'\.0$', '', regex=True)
                    # Replace 'nan' string (from empty cells) with empty string
                    df[col] = df[col].replace('nan', '')

            # Fill remaining NaN values with None for DB compatibility
            df = df.where(pd.notnull(df), None)
            
            records_created = 0
            records_updated = 0

            for _, row in df.iterrows():
                mip_name_val = row.get('Member Group Name')
                if not mip_name_val:
                    continue

                defaults_data = {
                    'branch_code': row.get('MG Code', ''),
                    'status': row.get('Company Status', ''),
                    'member': row.get('Member Count', ''),
                    'contribution_amount': row.get('Bill Amount', ''),
                    'mg_email_address': row.get('MG Contact Email', ''),
                    'tel': row.get('MG Contact Tel. 1', ''),
                    'tel_2': row.get('MG Contact Tel. 2', ''),
                    'npo_code': row.get('NPO CODE', ''),
                    'mg_bank_info': row.get('MG BANK INFO', ''),
                    'bank': row.get('BANK', ''),
                    'branch': row.get('BRANCH', ''),
                    'account_number': row.get('ACCOUNT NUMBER', ''),
                    'account_name': row.get('ACCOUNT NAME', ''),
                    'account_type': row.get('ACCOUNT TYPE', ''),
                    'employer_contacts': row.get('EMPLOYER CONTACTS', ''),
                    'schedule_date_received': row.get('Last Recon - Date', '')
                }

                obj, created = Globalacvv.objects.update_or_create(
                    mip_names=str(mip_name_val),
                    defaults=defaults_data
                )
                
                if created: records_created += 1
                else: records_updated += 1

            messages.success(request, f"Import complete! Created: {records_created}, Updated: {records_updated}.")
            
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
            
        return redirect('import_acvv_csv')

    return render(request, 'acvv_app/import_acvv.html')

from django.db import connection, transaction

@login_required
def import_reconciliation_csv(request):
    if request.method == "POST":
        file = request.FILES.get('csv_upload')
        if not file:
            messages.error(request, "Please select a file.")
            return redirect('import_reconciliation_csv')

        try:
            # 1. Load file
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            else:
<<<<<<< HEAD
                df = pd.read_excel(file)

            # Clean and strip whitespaces from column headers
            df.columns = df.columns.str.strip()
            available_columns = list(df.columns)

            # 🚀 ROBUST FLEXIBLE HEADER MATCHING
            # Look for Company Identifier Variations
            col_code = next((c for c in df.columns if c in ['MG Code', 'Company Code', 'Group Code', 'ACCOUNT NUMBER']), None)
            
            # Look for Fiscal Month Variations
            col_fiscal = next((c for c in df.columns if "Current Fiscal" in c or "Fiscal Month" in c or "Period" in c), None)
            
            # Look for Name Variations
            col_name = next((c for c in df.columns if c in ['Member Group Name', 'Company Name', 'MG Name']), None)

            # Other tracking dates
            col_lpi = next((c for c in df.columns if "LPI due date" in c or "Date Schedule Received" in c), None)
            col_step = next((c for c in df.columns if "DATE CONFIRM ON STeP" in c or "STeP" in c), None)
            col_debit = next((c for c in df.columns if "DEBIT ORDER DATE" in c or "CONFIRM BY EMPLOYER" in c), None)

            # Critical validation check: If key columns aren't found, tell the user exactly what headers were detected
            if not col_code or not col_fiscal:
                messages.error(
                    request, 
                    f"Import failed. Missing required columns. We need a Code column and a Fiscal column. "
                    f"Detected columns in your file: {', '.join(available_columns)}"
                )
                return redirect('import_reconciliation_csv')
=======
                df = pd.read_excel(file, engine='openpyxl')
>>>>>>> 7ce144f2b45bac80f36bccec2d6e05a25abe1b3b

            # 2. Helpers
            def clean_decimal(val):
                if pd.isna(val) or val == '' or val is None: 
                    return 0.00
                
                # Convert to string and clean out currency symbols, commas, and hidden spaces
                val_str = str(val).upper()
                for char in ['R', '$', 'ZAR', ',', ' ', '\xa0']:
                    val_str = val_str.replace(char, '')
                
                try:
                    return float(val_str)
                except ValueError:
                    return 0.00

            def clean_date_to_str(val):
<<<<<<< HEAD
=======
                if pd.isna(val) or val == '' or val is None:
                    return None
>>>>>>> 7ce144f2b45bac80f36bccec2d6e05a25abe1b3b
                dt = pd.to_datetime(val, errors='coerce')
                if pd.notna(dt):
                    return dt.strftime('%Y-%m-%d')
                return None

            records_updated = 0
            
            with transaction.atomic():
                with connection.cursor() as cursor:
<<<<<<< HEAD
                    for _, row in df.iterrows():
                        # Dynamically pull identifiers based on the columns found above
                        company_code = str(row.get(col_code, '')).strip()
                        fiscal_month = clean_date_to_str(row.get(col_fiscal))
                        
                        # Skip row safely if data is empty or invalid
                        if not company_code or company_code.lower() == 'nan' or not fiscal_month:
                            continue

                        # Clean fallback target rows
                        company_name = str(row.get(col_name, 'Unknown Company')).strip()
                        d1 = clean_date_to_str(row.get(col_lpi)) if col_lpi else None
                        d2 = clean_date_to_str(row.get(col_step)) if col_step else None
                        d3 = clean_date_to_str(row.get(col_debit)) if col_debit else None
=======
                    
                    # df.values converts data to list of rows, bypassing headers entirely
                    for row in df.values:
                        
                        # =========================================================
                        # EXACT MAPPING BASED ON YOUR EXCEL COLUMNS
                        # =========================================================
                        col_company_name    = 0  # MG Name
                        col_company_code    = 1  # MG Code
                        # col 2 is Email Address (Skipped)
                        col_company_status  = 3  # Company Status
                        col_payment_method  = 4  # Payment Method
                        col_fiscal_month    = 5  # Last Fiscal Reconciled
                        # col 6 is Arrears (Skipped)
                        col_members         = 7  # Member Count Reconciled
                        col_amount          = 8  # Contribution Amount Reconciled
                        col_current_status  = 9  # Reconciled Status
                        col_date_schedule   = 10 # Date Schedule Received
                        col_date_step       = 11 # Date Confirmed on Step
                        col_date_debit      = 12 # Debit order date
                        # =========================================================

                        # 1. Extract Identifiers First
                        company_code = str(row[col_company_code]).strip() if pd.notna(row[col_company_code]) else ''
                        fiscal_month = clean_date_to_str(row[col_fiscal_month])
                        
                        # Skip if essential identifiers are missing
                        if not company_code or company_code.lower() == 'nan' or not fiscal_month:
                            continue

                        # 2. Clean Dates
                        d1 = clean_date_to_str(row[col_date_schedule])
                        d2 = clean_date_to_str(row[col_date_step])
                        d3 = clean_date_to_str(row[col_date_debit])
>>>>>>> 7ce144f2b45bac80f36bccec2d6e05a25abe1b3b

                        # 3. Safely Extract Members (Handling missing or text data)
                        members_count = row[col_members]
                        if pd.isna(members_count) or members_count == '':
                            members_count = 0
                        else:
                            try:
                                members_count = int(float(members_count))
                            except ValueError:
                                members_count = 0

                        # 4. Safely Extract Amount
                        contribution_amount = clean_decimal(row[col_amount])
                        
                        # 5. Extract Basic Strings Safely
                        company_name = str(row[col_company_name]).strip() if pd.notna(row[col_company_name]) else ''
                        company_status = str(row[col_company_status]).strip() if pd.notna(row[col_company_status]) else 'Active'
                        payment_method = str(row[col_payment_method]).strip() if pd.notna(row[col_payment_method]) else 'Debit Order'
                        reconciled_status = str(row[col_current_status]).strip() if pd.notna(row[col_current_status]) else 'Unreconciled'

                        # 6. Database Insertion/Update
                        sql = """
                            INSERT INTO reconciliation_worksheet 
                            (mg_code, fiscal_month, mg_name, company_status, payment_method, 
                             member_count_reconciled, contribution_amount_reconciled, reconciled_status, 
                             date_schedule_received, date_confirmed_on_step, debit_order_date)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE
                            mg_name=VALUES(mg_name), company_status=VALUES(company_status), 
                            payment_method=VALUES(payment_method), member_count_reconciled=VALUES(member_count_reconciled),
                            contribution_amount_reconciled=VALUES(contribution_amount_reconciled),
                            reconciled_status=VALUES(reconciled_status),
                            date_schedule_received=VALUES(date_schedule_received),
                            date_confirmed_on_step=VALUES(date_confirmed_on_step),
                            debit_order_date=VALUES(debit_order_date)
                        """
                        
                        params = (
<<<<<<< HEAD
                            company_code, fiscal_month, company_name,
                            str(row.get('Company Status', 'Active')), str(row.get('Payment Method', 'Debit Order')),
                            int(row.get('MEMBERS', 0)), clean_decimal(row.get('CONTRIBUTION AMOUNT')),
                            str(row.get('Current Status', 'Unreconciled')),
=======
                            company_code, fiscal_month, company_name, company_status, payment_method,
                            members_count, contribution_amount, reconciled_status,
>>>>>>> 7ce144f2b45bac80f36bccec2d6e05a25abe1b3b
                            d1, d2, d3
                        )
                        
                        cursor.execute(sql, params)
                        records_updated += 1

            if records_updated == 0:
                messages.warning(
                    request, 
                    f"0 records matched formatting layouts. Verify your file data contains valid rows. "
                    f"Matched Identifier Column: '{col_code}', Matched Fiscal Month Column: '{col_fiscal}'"
                )
            else:
                messages.success(request, f"Successfully processed {records_updated} records.")
                
        except Exception as e:
            print(traceback.format_exc())
            messages.error(request, f"Import failed: {str(e)}")
            
        return redirect('import_reconciliation_csv')
    
    return render(request, 'acvv_app/import_reconciliation.html')

import re

@login_required
def import_withdrawals(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']
        
        # Helper to clean date strings that might contain extra text
        def clean_and_parse_date(val):
            if not val: return None
            val_str = str(val).strip()
            # Regex finds DD.MM.YYYY format
            match = re.search(r'\d{2}\.\d{2}\.\d{4}', val_str)
            if match:
                return pd.to_datetime(match.group(0), dayfirst=True).date()
            return None

        try:
            df = pd.read_excel(file)
            df.columns = df.columns.str.strip()
            df = df.where(pd.notnull(df), None)

            for index, row in df.iterrows():
                # Primary Key Lookup
                id_no = str(row.get('ID_Passport', '')).strip()
                if not id_no: continue
                
                # Strict Mapping based on your provided table
                defaults = {
                    'company_code': row.get('FUDN CODE', 'UNKNOWN'),
                    'mip_number': row.get('Member Nr', ''),
                    'member_name': row.get('Member Name', ''),
                    'member_surname': row.get('Member Surname', ''),
                    'claim_created_date': clean_and_parse_date(row.get('Claim Created')),
                    'claim_type': row.get('Claim Type', 'Two Pot'),
                    'exit_reason': row.get('Exit Reason', ''),
                    'claim_allocation': row.get('Claim Allocation', 'New Claim'),
                    'claim_status': row.get('Claim Status', 'Submitted'),
                    'date_paid': clean_and_parse_date(row.get('DATE CLAIM PAID')),
                    'last_contribution_date': clean_and_parse_date(row.get('J_Last_Contribution - EXIT DATE')),
                    'date_submitted': clean_and_parse_date(row.get('Date Submitted')),
                    'vested_pot_paid_date': clean_and_parse_date(row.get('Vested Pot Date Paid')),
                    'savings_pot_paid_date': clean_and_parse_date(row.get('Savigns Pot Date Paid')),
                    'infund_cert_date': clean_and_parse_date(row.get('Paid Up Certificate')),
                    'payment_option': row.get('Pay Full Benefit / Retirement', ''),
                    
                    # Calculated Booleans
                    'vested_pot_available': bool(row.get('Vested Pot Date Paid')),
                    'savings_pot_available': bool(row.get('Savigns Pot Date Paid')),
                }

                # Update existing or create new
                claim_obj, created = AcvvClaim.objects.update_or_create(
                    id_number=id_no,
                    defaults=defaults
                )

                # Add Note from the 'NOTES' Excel Column
                note_content = row.get('NOTES')
                if note_content:
                    ClaimNote.objects.create(
                        claim=claim_obj,
                        note_selection="IMPORT",
                        note_description=f"Excel Import Note: {note_content}",
                        created_by=request.user
                    )

            messages.success(request, "Withdrawals imported successfully.")
        except Exception as e:
            print(f"IMPORT ERROR: {e}") 
            messages.error(request, f"Import Failed: {str(e)}")
            
        return redirect('global_two_pot')
    
    return render(request, 'acvv_app/import_withdrawals.html')
