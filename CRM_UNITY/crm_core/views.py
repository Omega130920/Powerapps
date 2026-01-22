import base64
import csv
import datetime
import json
import os
import mimetypes
import requests
import openpyxl
from base64 import urlsafe_b64decode, urlsafe_b64encode
from decimal import Decimal
from datetime import datetime, date, timezone as py_timezone
from dateutil import parser
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import timedelta
from django.utils.dateparse import parse_datetime

from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count, Sum, Max
from django.http import FileResponse, Http404, HttpResponse, HttpRequest, JsonResponse
from django.utils import timezone
from django.utils.html import strip_tags
from django.views.decorators.clickjacking import xframe_options_exempt
from django.core.paginator import Paginator
from django.db import connection, transaction
from django.forms.models import model_to_dict
from django.urls import reverse

# Model Imports
from .models import (
    ComplaintLog, CrmUnityOutlookToken, DirectEmail, DirectEmailLog, GlobalFundContact, Cbc, CbcAdminPerson, CbcConsultancyPerson,
    Cfa, CfaAdminPerson, Cfa2, Cfa3, CommunicationsPerson,
    HumanResources, Section13a, ClientNotes,
    MemberDocument, CrmInbox, CrmDelegateTo, CrmDelegateAction, DelegationReport
)

# Form Imports
from .forms import (
    DocumentUploadForm, ComplaintLogForm, GlobalFundContactForm, 
    CbcForm, CbcAdminPersonForm, CbcConsultancyPersonForm,
    CfaForm, CfaAdminPersonForm, Cfa2Form, Cfa3Form, 
    CommunicationsPersonForm, HumanResourcesForm, Section13aForm
)

# Service Imports
from .services.outlook_graph_service import OutlookGraphService
from .services.delegation_service import delegate_email_task, add_action_note

User = get_user_model()
ZERO_DECIMAL = Decimal('0.00')


# --- GLOBAL MAPPINGS ---
CATEGORY_NAMES = {
    "1": "1. Reconciliation", "2": "2. Claims", "3": "3. Section 13A",
    "4": "4. New Business", "5": "5. Amendments", "6": "6. Section 28",
    "7": "7. Section 14", "8": "8. Section 27", "9": "9. Complaints",
    "10": "10. Two Pot", "11": "11. Report Back", "12": "12. Broker Investigations",
    "13": "13. General", "14": "14. Other"
}

ENQUIRY_GROUPING_MAP = {
    "1.1 Schedule Query": "Reconciliation (Queries, Sec 13A notices, Recon Queries)",
    "1.2 Member Data Query": "Reconciliation (Queries, Sec 13A notices, Recon Queries)",
    "1.3 Payment Query": "Reconciliation (Queries, Sec 13A notices, Recon Queries)",
    "1.4 Arrears Query": "Reconciliation (Queries, Sec 13A notices, Recon Queries)",
    "1.5 Online client RFW training Reconciliation": "RFW Training, SC Digital queries, New User Training",
    "1.6 Online Employer Amendment Data Query": "Reconciliation (Queries, Sec 13A notices, Recon Queries)",
    "1.7 Employer Data Referral for Overpayment": "Reconciliation (Queries, Sec 13A notices, Recon Queries)",
    "1.8 Employer Data Referral for Duplicate records and Merge queries": "Reconciliation (Queries, Sec 13A notices, Recon Queries)",
    "1.9 Member reinstatement": "Reconciliation (Queries, Sec 13A notices, Recon Queries)",
    "1.10 Member record Merge": "Reconciliation (Queries, Sec 13A notices, Recon Queries)",
    "2.1 Claim Form Query Active Withdrawal": "Claims (Two Pot included)",
    "2.2 Claim Form Receipt": "Claims (Two Pot included)",
    "2.3 Claim Form submission Active": "Claims (Two Pot included)",
    "2.4 Claim outstanding details": "Claims (Two Pot included)",
    "2.5 Claim Paid Confirmation": "Claims (Two Pot included)",
    "2.6 Forward Claim payment letter and tax documents": "Claims (Two Pot included)",
    "2.7 Claim Preservation Query": "Claims (Two Pot included)",
    "2.8 Claim ROT query": "Claims (Two Pot included)",
    "2.9 Claim Sec 14 query": "Claims (Two Pot included)",
    "2.10 Claim Paid Up Query": "Claims (Two Pot included)",
    "2.11  Emergency Two Pot Claim query": "Claims (Two Pot included)",
    "2.12 Retirement Claim query": "Claims (Two Pot included)",
    "2.13 Retirement Claim preservation Query": "Claims (Two Pot included)",
    "2.14 Retirement Claim Ill Health": "Claims (Two Pot included)",
    "2.15 Claim Paid Up Submission": "Claims (Two Pot included)",
    "2.16 Paid Up claim finalised (paid)": "Claims (Two Pot included)",
    "3.1  LPI  Section 13A Letter": "Reconciliation (Queries, Sec 13A notices, Recon Queries)",
    "3.2  Section 13A Client Query": "Reconciliation (Queries, Sec 13A notices, Recon Queries)",
    "3.3  Section 13A Termination": "Reconciliation (Queries, Sec 13A notices, Recon Queries)",
    "3.4 Section 13A Council Action": "Reconciliation (Queries, Sec 13A notices, Recon Queries)",
    "3.5 Section 13A Member Communication": "Reconciliation (Queries, Sec 13A notices, Recon Queries)",
    "3.6 Section 13A CBC Action": "Reconciliation (Queries, Sec 13A notices, Recon Queries)",
    "4.1 New Business Quote request": "New Business (Quotes, Installations and Queries)",
    "4.2 New Business COA and Accepted Quote submission": "New Business (Quotes, Installations and Queries)",
    "4.3 Welcome Email": "New Business (Quotes, Installations and Queries)",
    "4.4 Starter Pack Email": "New Business (Quotes, Installations and Queries)",
    "4.5 Starter Pack Meeting": "New Business (Quotes, Installations and Queries)",
    "4.6 Implementation confirmation": "New Business (Quotes, Installations and Queries)",
    "4.7 Employer Training": "RFW Training, SC Digital queries, New User Training",
    "5.1 Amendment Documentation": "Amendments",
    "5.2 Amendment to Acceptances": "Amendments",
    "5.3 Outstanding Amendment Documents": "Amendments",
    "5.4 Amendment Handover": "Amendments",
    "5.5 Amendment Rule to client broker": "Amendments",
    "5.6 Amendment COP to client broker": "Amendments",
    "5.7 Amendment RFW (Broker/Employer/3rd Party)": "Amendments",
    "6.1 Termination Query": "Terminations (Sec 14, 28 and 27)",
    "6.2 Termination Documentation Notice Letter": "Terminations (Sec 14, 28 and 27)",
    "6.3 Termination Banking Details for refund": "Terminations (Sec 14, 28 and 27)",
    "6.4 Termination Checklist": "Terminations (Sec 14, 28 and 27)",
    "6.5 Termination Withdrawal Forms": "Terminations (Sec 14, 28 and 27)",
    "6.6 Council confirmation of termination": "Terminations (Sec 14, 28 and 27)",
    "6.7 Member Termination Communication": "Terminations (Sec 14, 28 and 27)",
    "6.8 Section 14 Compulsory Form": "Terminations (Sec 14, 28 and 27)",
    "6.9 Section 14 Checklist": "Terminations (Sec 14, 28 and 27)",
    "6.10 Termination Fee Recovery": "Terminations (Sec 14, 28 and 27)",
    "7.1 Section 14 Query": "Terminations (Sec 14, 28 and 27)",
    "7.2 Section 14 Documentation": "Terminations (Sec 14, 28 and 27)",
    "7.3 Section 14 Outstanding requirements": "Terminations (Sec 14, 28 and 27)",
    "7.4 Section 14 Fund Query": "Terminations (Sec 14, 28 and 27)",
    "7.5 Section 14 Bank POP": "Terminations (Sec 14, 28 and 27)",
    "7.6 Section 14 Closure": "Terminations (Sec 14, 28 and 27)",
    "7.7 Section 14 Compulsory Form": "Terminations (Sec 14, 28 and 27)",
    "7.8 Section 14 Checklist": "Terminations (Sec 14, 28 and 27)",
    "7.9 Section 14 Member and Employer communication": "Terminations (Sec 14, 28 and 27)",
    "8.1 Termination Query": "Terminations (Sec 14, 28 and 27)",
    "8.2 Termination Documentation Notice Letter": "Terminations (Sec 14, 28 and 27)",
    "8.3 Termination Banking Details for refund": "Terminations (Sec 14, 28 and 27)",
    "8.4 Termination Checklist": "Terminations (Sec 14, 28 and 27)",
    "8.5 Termination Withdrawal Forms": "Terminations (Sec 14, 28 and 27)",
    "8.6 Council confirmation of termination": "Terminations (Sec 14, 28 and 27)",
    "8.7 Member Termination Communication (Retrenchment)": "Terminations (Sec 14, 28 and 27)",
    "8.8 Section 27 Compulsory Form": "Terminations (Sec 14, 28 and 27)",
    "8.9 Termination Member and Employer communication": "Terminations (Sec 14, 28 and 27)",
    "9.1 Complaint Detail Request": "Complaints (new 4)",
    "9.2 Complaint determination PFA": "Complaints (new 4)",
    "9.3 Member Complaint receipt and referral": "Complaints (new 4)",
    "9.4 Employer Complaint receipt and referral": "Complaints (new 4)",
    "9.5 Complaint Claim Payment": "Complaints (new 4)",
    "10.1 Employer Information Bulk Bank details": "Claims (Two Pot included)",
    "10.2 Employer Information Bulk Contact Details": "Claims (Two Pot included)",
    "10.3 Employer information Bulk Tax Detail changes": "Claims (Two Pot included)",
    "10.4 Training Employer Two Pot System changes": "RFW Training, SC Digital queries, New User Training",
    "10.5 Training Broker Two Pot Changes to Online": "RFW Training, SC Digital queries, New User Training",
    "10.6 Claim forms process query": "Claims (Two Pot included)",
    "10.7 Two Pot Process Claim Paid reports": "Claims (Two Pot included)",
    "10.8 Two Pot Member Claim Query": "Claims (Two Pot included)",
    "10.9  Two Pot Employer Query Reports for Qualifying Members": "Claims (Two Pot included)",
    "10.11 Two Pot Communication Employer Member Broker": "Claims (Two Pot included)",
    "11.1 Council Report Back": "Report Back (Broker and Employer Reports, Bulk Two Pot Communication)",
    "11.2 Broker Reports": "Report Back (Broker and Employer Reports, Bulk Two Pot Communication)",
    "11.3 Employer Reports": "Report Back (Broker and Employer Reports, Bulk Two Pot Communication)",
    "11.4 Sanlam Reports": "Report Back (Broker and Employer Reports, Bulk Two Pot Communication)",
    "11.5 Member Benefit Statement Monthly Reports": "General (Benefit statement CRM emails, New Claim forms generated etc, Commission Queries)",
    "11.6 Termination Report Back - Liquidations": "Report Back (Broker and Employer Reports, Bulk Two Pot Communication)",
    "11.7 Termination Report Back Clean Up Projects": "Report Back (Broker and Employer Reports, Bulk Two Pot Communication)",
    "11.8 Termination Report Back Dormant/Zero Members 2019 - 2024": "Report Back (Broker and Employer Reports, Bulk Two Pot Communication)",
    "12.1 Broker Investigation Letter": "Report Back (Broker and Employer Reports, Bulk Two Pot Communication)",
    "12.2 Broker Data and report Back Investigation Letter": "Report Back (Broker and Employer Reports, Bulk Two Pot Communication)",
    "12.3 Broker Client Reports for Claims Experience (Risk)": "Report Back (Broker and Employer Reports, Bulk Two Pot Communication)",
    "12.4 Broker Reports for Member Data, Underwriting and Claims Unresolved": "Report Back (Broker and Employer Reports, Bulk Two Pot Communication)",
    "13.1 Underwriting Queries": "Rate Revisions (CRM distribution of Rate letters, Queries)",
    "13.2 Death Claims Query": "Claims (Two Pot included)",
    "13.3 Funeral Claim Queries and notifications": "Claims (Two Pot included)",
    "13.4 Funeral Claim Outstanding Documents": "Claims (Two Pot included)",
    "13.5 Death Claim Outstanding Documents": "Claims (Two Pot included)",
    "13.6 Death Claim Beneficiary Fund Queries": "Claims (Two Pot included)",
    "13.7 Death Claim Distribution Approval to PE and Broker": "Claims (Two Pot included)",
    "13.8 AGRI Paid Up Funeral Claims": "Claims (Two Pot included)",
    "13.9 AGRI Paid Up Certificate Requests SGR": "Claims (Two Pot included)",
    "13.10 Repatriation Assistance to PE Family Broker": "Claims (Two Pot included)",
    "13.11 Funeral Claim Submission": "Claims (Two Pot included)",
    "13.12 Disability Query": "Claims (Two Pot included)",
    "13.13 Disability Application Submission": "Claims (Two Pot included)",
    "13.14 Disability Outstanding Reports": "Claims (Two Pot included)",
    "13.15 Disability Approval": "Claims (Two Pot included)",
    "13.16 Disability Declined": "Claims (Two Pot included)",
    "14.1 Bulk SMS to Members for documents, tax numbers": "General (Benefit statement CRM emails, New Claim forms generated etc, Commission Queries)",
    "14.2 Investment Reports for Brokers": "Report Back (Broker and Employer Reports, Bulk Two Pot Communication)",
    "14.3 Joint Forum Client Reports for PE and CBC": "Report Back (Broker and Employer Reports, Bulk Two Pot Communication)",
    "14.4 Divorce Orders - submission flagging": "General (Benefit statement CRM emails, New Claim forms generated etc, Commission Queries)",
    "14.5 Amendment to Member Guides - benefit changes": "General (Benefit statement CRM emails, New Claim forms generated etc, Commission Queries)",
    "14.7 Maintenance Orders - submission flagging": "General (Benefit statement CRM emails, New Claim forms generated etc, Commission Queries)",
    "14.8 Divorce order claim forms non member spouse": "General (Benefit statement CRM emails, New Claim forms generated etc, Commission Queries)",
    "14.9 Annual Rate Revision Data Preparation": "Rate Revisions (CRM distribution of Rate letters, Queries)",
    "14.10 Annual Rate Revision Amendments (AGR/Council etc)": "Rate Revisions (CRM distribution of Rate letters, Queries)",
    "14.11 Annual Rate Revision Distribution - PE Broker": "Rate Revisions (CRM distribution of Rate letters, Queries)",
    "14.12 Global List Monthly Checks and Amendments": "General (Benefit statement CRM emails, New Claim forms generated etc, Commission Queries)",
    "14.13 0101 Account Monthly Checks and Action": "General (Benefit statement CRM emails, New Claim forms generated etc, Commission Queries)",
    "14.14 Benefit Statement Query (member/employer)": "General (Benefit statement CRM emails, New Claim forms generated etc, Commission Queries)",
    "14.15 Bulk Benefit Statements (employer/member) ad-hoc": "General (Benefit statement CRM emails, New Claim forms generated etc, Commission Queries)",
}

# ==============================================================================
# AUTHENTICATION & DASHBOARD
# ==============================================================================

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('dashboard')
    form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})

@login_required
def dashboard(request):
    """Displays the CRM dashboard with specific notification badges."""
    username = request.user.username
    
    # Initialize counts
    new_emails_count = 0
    recycled_count = 0
    my_pending_tasks_count = 0

    # 1. Manager Logic (Only for omega)
    if username.lower() == 'omega' or request.user.is_superuser:
        # New emails in CrmInbox that haven't been processed yet
        new_emails_count = CrmInbox.objects.filter(status='Pending').count()
        
        # Recycle Bin: Based on your recycle_bin_view logic (status='Recycled')
        recycled_count = CrmDelegateTo.objects.filter(status='Recycled').count()

    # 2. Agent Logic: My Assigned Tasks
    # Based on your tasks_view logic (delegated_to = username)
    my_pending_tasks_count = CrmDelegateTo.objects.filter(
        delegated_to=request.user.username, 
        status='Delegated'
    ).count()

    context = {
        'username': username,
        'new_emails_count': new_emails_count,           # For 'Delegations' button
        'my_pending_delegations_count': my_pending_tasks_count, # For 'Dashboard' button
        'recycled_count': recycled_count,               # For 'Recycle Bin' button
    }
    return render(request, 'dashboard.html', context)

def logout_view(request):
    logout(request)
    messages.info(request, "You have successfully logged out.")
    return redirect('login')

# ==============================================================================
# MICROSOFT GRAPH EMAIL (UNITY_INTERNAL AUTOMATIC LOGIC)
# ==============================================================================

@login_required
def fetch_emails_view(request):
    """
    MANAGER (omega) ONLY: Live Inbox shows ONLY new emails not yet in the system.
    """
    if request.user.username.lower() != 'omega' and not request.user.is_superuser:
        messages.error(request, "Access restricted.")
        return redirect('tasks')

    inbox_data = OutlookGraphService.fetch_inbox_messages(top_count=100)
    
    if 'error' in inbox_data:
        messages.error(request, f"Outlook Error: {inbox_data['error']}")
        return render(request, 'inbox.html', {'email_list': []})

    emails = inbox_data.get('value', [])
    processed_emails = []

    # Get IDs of everything already in the delegation table to hide them from the Live Inbox
    delegated_ids = CrmDelegateTo.objects.values_list('email_id', flat=True)

    for msg in emails:
        email_id = msg['id']
        
        # Filter: Only show if it hasn't been delegated yet
        if email_id not in delegated_ids:
            # Parse the MS Graph string (e.g., "2026-01-22T08:00:00Z") into a Python datetime
            raw_date = msg.get('receivedDateTime')
            parsed_date = parse_datetime(raw_date) if raw_date else None

            local_entry, created = CrmInbox.objects.get_or_create(
                email_id=email_id,
                defaults={
                    'subject': msg.get('subject', 'No Subject'),
                    'sender': msg.get('from', {}).get('emailAddress', {}).get('address', 'Unknown'),
                    'snippet': msg.get('bodyPreview', ''),
                    'received_timestamp': parsed_date, # Saved to DB
                    'status': 'Pending'
                }
            )
            
            # Attach 'received_at' dynamically so the HTML template can find it
            local_entry.received_at = local_entry.received_timestamp
            processed_emails.append(local_entry)

    return render(request, 'inbox.html', {'email_list': processed_emails})

@login_required
def email_registry_view(request):
    """
    MANAGER (omega) ONLY: Email List shows ALL emails and their current statuses.
    """
    if request.user.username.lower() != 'omega':
        return redirect('dashboard')
    
    all_emails = CrmDelegateTo.objects.all().order_by('-received_timestamp')
    return render(request, 'email_registry.html', {'emails': all_emails})

@login_required
def recycle_bin_view(request):
    """
    MANAGER (omega) ONLY: Recycle Bin shows tasks marked as recycled.
    """
    if request.user.username.lower() != 'omega':
        return redirect('dashboard')
        
    tasks = CrmDelegateTo.objects.filter(is_recycled=True)
    return render(request, 'recycle_bin.html', {'tasks': tasks})

@login_required
def get_email_content_view(request, email_id):
    """Fetches raw HTML email content via Service Account."""
    endpoint = f"messages/{email_id}"
    email_data = OutlookGraphService._make_graph_request(endpoint, method='GET')
    
    if 'error' in email_data:
        return HttpResponse(f"<h1>Outlook Error</h1><p>{email_data['error']}</p>", status=500)

    body_data = email_data.get('body', {})
    content = body_data.get('content', 'No content found.')
    
    if body_data.get('contentType', '').lower() != 'html':
        content = f'<pre style="white-space: pre-wrap; font-family: sans-serif;">{content}</pre>'

    wrapped_content = f"<!DOCTYPE html><html><body style='font-family: sans-serif;'>{content}</body></html>"
    return HttpResponse(wrapped_content, content_type='text/html')

@login_required
def delegate_email_view(request, email_id):
    """
    Handles classification and delegation for CRM_UNITY.
    FIXED: Removed target_email from _make_graph_request to resolve TypeError.
    FIXED: Added validation for Graph API response data types.
    """
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    inbox_item = get_object_or_404(CrmInbox, email_id=email_id)
    available_agents = User.objects.filter(is_active=True).exclude(pk=request.user.pk)

    # ==========================================
    # 1. HANDLE POST (DELEGATION LOGIC)
    # ==========================================
    if request.method == 'POST':
        agent_name = request.POST.get('agent_name')
        work_related = request.POST.get('work_related') 
        member_code = request.POST.get('mip_number') 
        
        form_data = {
            'member_group_code': member_code,
            'category': request.POST.get('email_category'),
            'work_related': work_related,
            'type': request.POST.get('email_type'),
        }

        if work_related == 'No':
            success, message = delegate_email_task(email_id, None, request.user, form_data, is_recycle=True)
            if success:
                try:
                    move_payload = {"destinationId": "deleteditems"}
                    # Removed target_email from positional arguments
                    OutlookGraphService._make_graph_request(f"messages/{email_id}/move", method='POST', data=move_payload)
                    messages.info(request, "Email moved to Recycle Bin locally and in Outlook.")
                except Exception as e:
                    messages.warning(request, f"Task recycled locally, but Outlook move failed: {str(e)}")
                return redirect('fetch_emails')
            else:
                messages.error(request, f"Error: {message}")
                return redirect('fetch_emails')

        else:
            if not agent_name or agent_name == "__Select Agent__":
                messages.error(request, "Please select an agent for work-related tasks.")
            else:
                success, message = delegate_email_task(email_id, agent_name, request.user, form_data)
                
                if success:
                    try:
                        assignee = User.objects.get(pk=agent_name) if agent_name.isdigit() else User.objects.get(username=agent_name)
                        reply_payload = {
                            "comment": f"Dear Sender,\n\nThis has been delegated to: {assignee.username}.\nRef: {member_code}"
                        }
                        
                        # Removed target_email from positional arguments
                        draft = OutlookGraphService._make_graph_request(f"messages/{email_id}/createReply", method='POST', data=reply_payload)
                        if draft and isinstance(draft, dict) and 'id' in draft:
                            OutlookGraphService._make_graph_request(f"messages/{draft['id']}/send", method='POST')
                        
                        messages.success(request, f"Task delegated to {assignee.username} and reply sent.")
                    except Exception as e:
                        messages.warning(request, f"Delegated locally, Graph error: {str(e)}")
                
                    return redirect('fetch_emails')

    # ==========================================
    # 2. HANDLE GET (DATA FETCHING)
    # ==========================================
    endpoint = f"messages/{email_id}" 
    # Removed target_email to prevent "multiple values for argument 'method'"
    email_data = OutlookGraphService._make_graph_request(endpoint, method='GET') 

    # Safety check: Ensure email_data is a dictionary before accessing .get()
    if not isinstance(email_data, dict) or 'error' in email_data:
        email_body = inbox_item.snippet
        email_subject_final = inbox_item.subject
    else:
        email_body = email_data.get('body', {}).get('content', inbox_item.snippet)
        email_subject_final = email_data.get('subject', inbox_item.subject)

    # Fetching attachments (Service method uses settings internally)
    attachments_list = OutlookGraphService.fetch_attachments(target_email, email_id)
    
    # Process image contentBytes for previews
    if attachments_list and isinstance(attachments_list, list):
        for att in attachments_list:
            content_type = att.get('contentType', '').lower()
            if 'image' in content_type:
                raw_att = OutlookGraphService.get_attachment_raw(target_email, email_id, att['id'])
                if raw_att and isinstance(raw_att, dict) and 'contentBytes' in raw_att:
                    att['contentBytes'] = raw_att['contentBytes']

    return render(request, 'delegate_to.html', {
        'email_id': email_id, 
        'email_subject': email_subject_final, 
        'available_agents': available_agents, 
        'email_body': email_body,
        'attachments': attachments_list, 
        'inbox_item': inbox_item
    })
    
@login_required
def send_task_email_view(request, email_id):
    if request.method == 'POST':
        # Get the task record
        task = get_object_or_404(CrmDelegateTo, email_id=email_id)
        
        recipient = request.POST.get('recipient_email')
        subject = request.POST.get('email_subject_reply') 
        message_body = request.POST.get('email_body_reply')
        
        if not message_body or message_body == "<p><br></p>":
             messages.error(request, "Email body cannot be empty.")
             return redirect('delegate_action', email_id=email_id)

        # 1. Send the actual email via Outlook
        response = OutlookGraphService.send_outlook_email(
            recipient=recipient, 
            subject=subject, 
            body_html=message_body
        )
        
        if response.get('success'):
            # 2. LOG TO DELEGATE ACTIONS (The Thread Table)
            # This makes it show up in the "View Thread" timeline
            CrmDelegateAction.objects.create(
                task_email_id=email_id,
                action_type='REPLY_SENT', # Specific type to distinguish from notes
                action_user=request.user.username,
                note_content=message_body, # Stores the actual email content
                related_subject=subject     # Stores the reply subject
            )
            
            messages.success(request, f"Reply sent and logged in thread.")
        else:
            messages.error(request, f"Failed: {response.get('error')}")
            
        return redirect('delegate_action', email_id=email_id)
    
    return redirect('tasks')

# ==============================================================================
# CRM MEMBER MANAGEMENT
# ==============================================================================

@login_required
def global_members_list(request):
    # 1. Get all members initially
    members = GlobalFundContact.objects.all().order_by('member_group_code')
    
    # 2. Get the unique statuses for the dropdown menu
    # This ensures your dropdown always matches what is actually in the database
    fund_statuses = GlobalFundContact.objects.values_list('fund_status', flat=True).distinct().order_by('fund_status')

    # 3. Get filter values from the URL
    search_query = request.GET.get('search_query')
    status_filter = request.GET.get('fund_status_filter')

    # 4. Apply Search Filter (Text box)
    if search_query:
        members = members.filter(
            Q(member_group_code__icontains=search_query) |
            Q(member_group_name__icontains=search_query)
        )

    # 5. Apply Fund Status Filter (Dropdown)
    # We only filter if a status is selected and it's not "all"
    if status_filter and status_filter != 'all':
        members = members.filter(fund_status=status_filter)

    context = {
        'members': members,
        'fund_statuses': fund_statuses,
        'search_query': search_query,
        'fund_status_filter': status_filter,
    }

    return render(request, 'global_members_list.html', context)

RELATED_FORMS = [
    ('cbc_form', CbcForm, Cbc), 
    ('cbc_admin_form', CbcAdminPersonForm, CbcAdminPerson),
    ('cbc_consultancy_form', CbcConsultancyPersonForm, CbcConsultancyPerson),
    ('cfa_form', CfaForm, Cfa), 
    ('cfa_admin_form', CfaAdminPersonForm, CfaAdminPerson),
    ('cfa2_form', Cfa2Form, Cfa2), 
    ('cfa3_form', Cfa3Form, Cfa3),
    ('communications_form', CommunicationsPersonForm, CommunicationsPerson),
    ('hr_form', HumanResourcesForm, HumanResources), 
    ('section13a_form', Section13aForm, Section13a),
]

from django.core.mail import EmailMessage

@login_required
def member_information(request, member_group_code):
    contact_info = get_object_or_404(GlobalFundContact, member_group_code=member_group_code)
    
    # --- 0. HANDLE FILE DOWNLOADS ---
    # Handle original MemberDocument downloads
    if request.method == 'GET' and 'download_id' in request.GET:
        doc_id = request.GET.get('download_id')
        document = get_object_or_404(MemberDocument, id=doc_id, related_member_group_code=member_group_code)
        if document.document_file:
            try:
                return FileResponse(document.document_file.open('rb'), as_attachment=True, filename=document.document_file.name)
            except FileNotFoundError:
                messages.error(request, "File not found on server storage.")
        else:
            messages.error(request, "Database record exists, but no file is attached.")

    # Handle Note Attachment downloads via Filename
    if request.method == 'GET' and 'download_note_file' in request.GET:
        target_filename = request.GET.get('download_note_file')
        document = MemberDocument.objects.filter(
            related_member_group_code=member_group_code, 
            document_file__icontains=target_filename
        ).first()
        
        if document and document.document_file:
            try:
                return FileResponse(document.document_file.open('rb'), as_attachment=True)
            except FileNotFoundError:
                messages.error(request, "Attachment file not found.")
        else:
            messages.error(request, "Could not locate the specific attachment.")

    # --- START OF POST LOGIC ---
    if request.method == 'POST':
        action = request.POST.get('action')
        user_display = request.user.username 
        
        if action == 'send_direct_email':
            recipient = request.POST.get('recipient_email')
            subject = request.POST.get('subject')
            body_content = request.POST.get('email_body_html_content')
            attachments = request.FILES.getlist('attachments')

            response = OutlookGraphService.send_outlook_email(
                recipient=recipient,
                subject=subject,
                body_html=body_content,
                attachments=attachments
            )

            if response.get('success'):
                DirectEmailLog.objects.create(
                    member_group_code=member_group_code,
                    subject=subject,
                    recipient_email=recipient,
                    body_content=body_content,
                    sent_by_user=request.user,
                    outlook_message_id=response.get('outlook_id'), 
                    sent_at=timezone.now()
                )
                messages.success(request, f"Email sent successfully to {recipient}")
            else:
                messages.error(request, f"Microsoft Error: {response.get('error')}")
            
            # Redirect back to the Communications tab
            return redirect(f"/global-members/{member_group_code}/#communications")

        elif action == 'save_member_note':
            attached_filename = None
            
            # 1. Process File Upload First
            if 'note_file' in request.FILES:
                uploaded_file = request.FILES['note_file']
                attached_filename = uploaded_file.name 
                
                MemberDocument.objects.create(
                    related_member_group_code=member_group_code,
                    document_file=uploaded_file,
                    title=f"Note Attachment: {uploaded_file.name}",
                    uploaded_by=user_display,
                    uploaded_at=timezone.now()
                )

            # 2. Create Note with reference to filename
            ClientNotes.objects.create(
                related_member_group_code=member_group_code,
                notes=request.POST.get('note_content'),
                communication_type=request.POST.get('communication_type'),
                action_notes=request.POST.get('action_notes'),
                attached_file_name=attached_filename, 
                user=user_display, 
            )
            messages.success(request, 'Note saved.')
            
            # Redirect back to the Communications tab
            return redirect(f"/global-members/{member_group_code}/#communications")

        return redirect('member_information', member_group_code=member_group_code)

    # --- DATA RETRIEVAL ---
    personnel_data = {
        'cbc_info': Cbc.objects.filter(member_group_code=member_group_code).first(),
        'cbc_admin_info': CbcAdminPerson.objects.filter(member_group_code=member_group_code).first(),
        'cbc_consultancy_info': CbcConsultancyPerson.objects.filter(member_group_code=member_group_code).first(),
        'cfa_info': Cfa.objects.filter(member_group_code=member_group_code).first(),
        'cfa_admin_info': CfaAdminPerson.objects.filter(member_group_code=member_group_code).first(),
        'cfa2_info': Cfa2.objects.filter(member_group_code=member_group_code).first(),
        'cfa3_info': Cfa3.objects.filter(member_group_code=member_group_code).first(),
        'communications_info': CommunicationsPerson.objects.filter(member_group_code=member_group_code).first(),
        'hr_info': HumanResources.objects.filter(member_group_code=member_group_code).first(),
        'section13a_info': Section13a.objects.filter(member_group_code=member_group_code).first(),
    }

    # --- MERGED COMMUNICATION HISTORY ---
    combined_email_log = []
    delegated_items = CrmDelegateTo.objects.filter(member_group_code=member_group_code)
    related_email_ids = [item.email_id for item in delegated_items]
    thread_status_map = {item.email_id: item.status for item in delegated_items}
    
    inbox_records = CrmInbox.objects.filter(email_id__in=related_email_ids)
    inbox_map = {email.email_id: email.received_timestamp for email in inbox_records}

    for item in delegated_items:
        combined_email_log.append({
            'timestamp': item.received_timestamp,
            'arrival_timestamp': inbox_map.get(item.email_id), 
            'delegation_timestamp': item.received_timestamp,
            'type': 'Original',
            'subject': item.subject,
            'assigned_to': item.delegated_to,
            'status': item.status,
            'email_id': item.email_id,
            'action_user': item.delegated_by
        })

    threaded_replies = CrmDelegateAction.objects.filter(
        task_email_id__in=related_email_ids, 
        action_type='REPLY_SENT'
    )

    for reply in threaded_replies:
        parent_status = thread_status_map.get(reply.task_email_id, 'Replied')
        combined_email_log.append({
            'timestamp': reply.action_timestamp,
            'arrival_timestamp': None,
            'delegation_timestamp': None,
            'type': 'Reply',
            'subject': reply.related_subject or "Reply to Task",
            'assigned_to': reply.action_user,
            'status': parent_status, 
            'email_id': reply.task_email_id,
            'action_user': reply.action_user
        })

    combined_email_log.sort(key=lambda x: x['timestamp'], reverse=True)

    notes = ClientNotes.objects.filter(related_member_group_code=member_group_code).order_by('-date')
    direct_emails = DirectEmailLog.objects.filter(member_group_code=member_group_code).order_by('-sent_at')
    documents = MemberDocument.objects.filter(related_member_group_code=member_group_code).order_by('-uploaded_at')

    context = {
        'contact_info': contact_info,
        **personnel_data,
        'notes': notes,
        'documents': documents,
        'combined_email_log': combined_email_log, 
        'direct_emails': direct_emails,
    }

    return render(request, 'member_information.html', context)

@login_required
def add_member(request):
    if request.method == 'POST':
        member_form = GlobalFundContactForm(request.POST)
        related_forms = {name: cls(request.POST) for name, cls, _ in RELATED_FORMS}
        if member_form.is_valid():
            try:
                with transaction.atomic():
                    new_mem = member_form.save()
                    code = new_mem.member_group_code
                    if 'save_all' in request.POST:
                        for name, _, model_cls in RELATED_FORMS:
                            inst = related_forms[name].save(commit=False)
                            inst.member_group_code = code
                            inst.save()
                    messages.success(request, f"Member {code} created.")
                    return redirect('member_information', member_group_code=code)
            except Exception as e:
                messages.error(request, f"Error: {e}")
    else:
        member_form = GlobalFundContactForm()
        related_forms = {name: cls() for name, cls, _ in RELATED_FORMS}
    return render(request, 'add_member.html', {'member_form': member_form, **related_forms})

# ==============================================================================
# EXCEL IMPORT (RESTORED)
# ==============================================================================

@login_required
def import_global_data(request):
    """Excel master file import handler."""
    if request.user.username.lower() != 'omega':
        return redirect('dashboard')
    if request.method == 'POST' and 'master_file' in request.FILES:
        try:
            workbook = openpyxl.load_workbook(request.FILES['master_file'], read_only=True)
            messages.success(request, "File read successfully. Import processing initiated.")
        except Exception as e:
            messages.error(request, f"Import Error: {e}")
    return render(request, 'import_global_data.html')

# ==============================================================================
# AGENT WORKFLOW & REPORTS
# ==============================================================================

@login_required
def tasks_view(request):
    """
    EMAIL DASHBOARD: 
    Manager (omega) sees ALL delegated tasks (Delegated AND Completed).
    Agents see ONLY tasks assigned to them.
    """
    # 1. Define the statuses we want to see
    VISIBLE_STATUSES = ['Delegated', 'Completed']

    # 2. Fetch the tasks based on user role
    if request.user.username.lower() == 'omega':
        tasks_qs = CrmDelegateTo.objects.filter(status__in=VISIBLE_STATUSES).order_by('-received_timestamp')
    else:
        tasks_qs = CrmDelegateTo.objects.filter(
            delegated_to=request.user.username, 
            status__in=VISIBLE_STATUSES
        ).order_by('-received_timestamp')
    
    # 3. Get list of related Email IDs for the Lookup Map
    related_email_ids = [t.email_id for t in tasks_qs]
    
    # 4. Fetch the original Inbox records (Map both Timestamp AND Sender)
    inbox_records = CrmInbox.objects.filter(email_id__in=related_email_ids)
    
    # We create a dictionary where the key is the email_id and the value is the whole record
    inbox_map = {email.email_id: email for email in inbox_records}

    # 5. Build the display list
    display_tasks = []
    for t in tasks_qs:
        # Get the corresponding inbox record if it exists
        inbox_item = inbox_map.get(t.email_id)
        
        display_tasks.append({
            'subject': t.subject,
            'email_id': t.email_id,
            'member_group_code': t.member_group_code,
            
            # Timestamp A: When it originally arrived (from Inbox)
            'arrival_timestamp': inbox_item.received_timestamp if inbox_item else None,
            
            # NEW: Sender Email Address (from Inbox)
            'sender': inbox_item.sender if inbox_item else "Unknown Sender",
            
            # Timestamp B: When it was delegated (from DelegateTo table)
            'delegation_timestamp': t.received_timestamp,
            
            'category': t.category,
            'enquiry_type': t.type, 
            'status': t.status,
            'delegated_to': t.delegated_to
        })

    return render(request, 'tasks.html', {'delegated_tasks': display_tasks})

@login_required
def delegate_action_view(request, email_id):
    """
    Detailed view for a specific task in CRM_UNITY.
    Captures 'source' to handle conditional 'Back' button logic.
    """
    task = get_object_or_404(CrmDelegateTo, email_id=email_id)
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    
    # --- 🟢 CAPTURE SOURCE FOR CONDITIONAL BACK BUTTON ---
    source = request.GET.get('from') # Capture ?from=member
    
    action_history = CrmDelegateAction.objects.filter(task_email_id=email_id).order_by('-action_timestamp')

    # --- 🚀 FETCH ATTACHMENT METADATA & IMAGE BYTES ---
    attachments_list = []
    try:
        attachments_list = OutlookGraphService.fetch_attachments(target_email, email_id)
        if attachments_list and isinstance(attachments_list, list):
            for att in attachments_list:
                content_type = att.get('contentType', '').lower()
                if 'image' in content_type:
                    raw_att = OutlookGraphService.get_attachment_raw(target_email, email_id, att['id'])
                    if raw_att and isinstance(raw_att, dict) and 'contentBytes' in raw_att:
                        att['contentBytes'] = raw_att['contentBytes']
    except Exception as e:
        print(f"Error fetching CRM attachment data in delegate_action: {e}")

    if request.method == 'POST':
        action_type = request.POST.get('action_type')
        user_display = request.user.username
        
        with transaction.atomic():
            # --- 1. RESTORE TO MAIN INBOX LOGIC ---
            if action_type == 'restore_to_inbox':
                restore_note = request.POST.get('restore_note', 'Restored from Recycle Bin to Main Inbox.')
                is_from_recycle_bin = (task.status == 'Recycled')

                try:
                    inbox_item = CrmInbox.objects.get(email_id=email_id)
                    inbox_item.status = 'Pending'    
                    inbox_item.is_processed = False  
                    inbox_item.delegated_to = None   
                    inbox_item.save()
                except CrmInbox.DoesNotExist:
                    pass 

                CrmDelegateAction.objects.create(
                    task_email_id=email_id,
                    action_type='restore_to_inbox',
                    action_user=user_display,
                    note_content=restore_note
                )
                
                task.delete()
                messages.success(request, "Task successfully moved back to the original Inbox queue.")
                
                if is_from_recycle_bin:
                    return redirect('recycle_bin')
                return redirect('fetch_emails') 

            # --- 2. UPDATE METADATA ---
            elif action_type == 'update_metadata':
                form_mip = request.POST.get('mip_number')
                task.member_group_code = form_mip  
                task.id_passport = request.POST.get('id_passport')
                task.category = request.POST.get('email_category')
                task.type = request.POST.get('email_type')
                task.method = request.POST.get('email_method')
                task.save()
                
                CrmDelegateAction.objects.create(
                    task_email_id=email_id,
                    action_type='update_metadata',
                    action_user=user_display,
                    note_content=f"Updated metadata: Member Group Code {form_mip}"
                )
                messages.success(request, "Task metadata updated successfully.")

            # --- 3. ADD INTERNAL NOTE ---
            elif action_type == 'add_note':
                note_text = request.POST.get('internal_note')
                comm_type = request.POST.get('communication_type_note')
                action_note_val = request.POST.get('action_notes_note')
                
                current_notes_raw = task.internal_notes
                
                if not current_notes_raw or current_notes_raw == "":
                    current_notes = []
                else:
                    try:
                        current_notes = json.loads(current_notes_raw)
                    except json.JSONDecodeError:
                        current_notes = [{"user": "System", "note": current_notes_raw, "timestamp": ""}]

                new_note_entry = {
                    'user': user_display,
                    'note': f"[{comm_type} - {action_note_val}] {note_text}",
                    'timestamp': timezone.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                current_notes.append(new_note_entry)
                
                task.internal_notes = json.dumps(current_notes)
                task.save()
                
                CrmDelegateAction.objects.create(
                    task_email_id=email_id,
                    action_type='add_note',
                    action_user=user_display,
                    note_content=note_text
                )
                messages.success(request, "Internal note saved.")

            # --- 4. FINALIZE / COMPLETE ---
            elif action_type == 'complete':
                final_note = request.POST.get('completion_notes', 'Task completed.')
                task.status = 'Completed'
                task.save()
                
                CrmDelegateAction.objects.create(
                    task_email_id=email_id,
                    action_type='complete',
                    action_user=user_display,
                    note_content=final_note
                )
                messages.success(request, "Task marked as Completed.")
                return redirect('tasks')

        # Redirect back to the same view, preserving the 'from' parameter if it exists
        redirect_url = reverse('delegate_action', kwargs={'email_id': email_id})
        if source:
            redirect_url += f"?from={source}"
        return redirect(redirect_url)

    return render(request, 'delegate_action.html', {
        'task': task,
        'action_history': action_history,
        'email_id': email_id,
        'attachments': attachments_list,
        'source': source,  # --- 🟢 PASS TO TEMPLATE ---
    })

from django.db.models import Count

@login_required
def delegation_report_view(request):
    if request.user.username.lower() != 'omega':
        return redirect('dashboard')

    # 1. Get Date Filters
    date_query, start_date, end_date = get_date_filters(request)
    valid_entries_filter = Q(category__isnull=False) & ~Q(category='') & \
                           Q(type__isnull=False) & ~Q(type='') & ~Q(type='None')
    
    final_filter = date_query & valid_entries_filter

    # --- SECTION 1: OVERALL TOTALS ---
    all_emails_total = CrmDelegateTo.objects.filter(date_query).count()
    work_stats = CrmDelegateTo.objects.filter(date_query).values('work_related').annotate(count=Count('email_id'))
    
    work_related_data = {'Yes': 0, 'No': 0}
    for item in work_stats:
        status = item['work_related']
        if status in work_related_data:
            work_related_data[status] = item['count']

    # --- SECTION 2: BREAKDOWN & SUMMARY (Existing Logic) ---
    raw_breakdown = CrmDelegateTo.objects.filter(final_filter).values('category', 'type').annotate(count=Count('email_id')).order_by('category', 'type')
    type_selection_breakdown = []
    for item in raw_breakdown:
        type_selection_breakdown.append({
            'category': CATEGORY_NAMES.get(str(item['category']), item['category']),
            'type': item['type'],
            'grouping': ENQUIRY_GROUPING_MAP.get(item['type'], "Unmapped"),
            'count': item['count']
        })

    raw_summary = CrmDelegateTo.objects.filter(final_filter).values('category').annotate(subtotal=Count('email_id')).order_by('-subtotal')
    category_summary = []
    for item in raw_summary:
        category_summary.append({
            'category': CATEGORY_NAMES.get(str(item['category']), item['category']),
            'subtotal': item['subtotal']
        })

    # --- SECTION 4: DETAILED FLOW & SLA CALCULATION ---
    # This section feeds your new popup/modal
    detailed_records = []
    records_qs = CrmDelegateTo.objects.filter(date_query).order_by('-received_timestamp')

    for item in records_qs:
        # Get Original Arrival Time
        inbox_item = CrmInbox.objects.filter(email_id=item.email_id).first()
        received_date = inbox_item.received_timestamp if inbox_item else item.received_timestamp
        
        # Get the Manager/User who performed the action
        first_action = CrmDelegateAction.objects.filter(task_email_id=item.email_id).order_by('action_timestamp').first()
        actioned_by = first_action.action_user if first_action else "System"

        # SLA Logic: If (Delegation Date - Received Date) > 2 Days
        # Note: item.received_timestamp is the time it was delegated
        sla_flag = "Within SLA"
        if received_date:
            time_diff = item.received_timestamp - received_date
            if time_diff > timedelta(days=2):
                sla_flag = "Out of SLA"

        detailed_records.append({
            'received_date': received_date,
            'actioned_by': actioned_by,
            'action_date': item.received_timestamp,
            'delegated_to': item.delegated_to or "Unassigned",
            'member_code': item.member_group_code or "N/A",
            'sla_status': sla_flag
        })

    context = {
        'all_emails_total': all_emails_total,
        'work_related_data': work_related_data,
        'type_selection_breakdown': type_selection_breakdown,
        'category_summary': category_summary,
        'detailed_records': detailed_records, # Pass to popup
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'delegation_summary_report.html', context)

@login_required
def complaint_log_view(request):
    # --- 1. HANDLE FILTERS ---
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Helper to check for "None" strings from the URL
    def is_valid_dt(val):
        return val and val != "None" and val != ""

    # --- 2. CHECK FOR EDIT MODE ---
    edit_id = request.GET.get('edit_id')
    instance = None
    if edit_id:
        instance = get_object_or_404(ComplaintLog, pk=edit_id)

    # --- 3. INITIALIZE FORM ---
    form = ComplaintLogForm(request.POST or None, instance=instance)

    # --- 4. HANDLE POST (SAVE/UPDATE) ---
    if request.method == 'POST':
        if form.is_valid():
            comp = form.save(commit=False)
            if not instance:
                comp.created_by = request.user
            comp.save()
            
            msg = "Complaint updated successfully." if instance else "Complaint logged successfully."
            messages.success(request, msg)
            return redirect('complaint_log')

    # --- 5. DATA RETRIEVAL WITH FILTERS ---
    complaints = ComplaintLog.objects.all()

    if is_valid_dt(start_date):
        complaints = complaints.filter(created_at__date__gte=start_date)
    if is_valid_dt(end_date):
        complaints = complaints.filter(created_at__date__lte=end_date)

    complaints = complaints.order_by('-created_at')
    
    context = {
        'form': form, 
        'complaints': complaints,
        'is_editing': instance is not None,
        'title': 'Complaint Registry'
    }
    return render(request, 'complaint_log.html', context)


@login_required
def export_complaints_excel(request):
    """Exports filtered complaints to Excel."""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    def is_valid_dt(val):
        return val and val != "None" and val != ""

    # Fetch base records
    complaints = ComplaintLog.objects.all()

    # Apply same filters as the view
    if is_valid_dt(start_date):
        complaints = complaints.filter(created_at__date__gte=start_date)
    if is_valid_dt(end_date):
        complaints = complaints.filter(created_at__date__lte=end_date)

    complaints = complaints.order_by('-created_at')

    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Complaint Log Export"

    # Header Styling
    headers = [
        'ID', 'Complainant', 'Employer', 'Nature of Complaint', 
        'Status', 'Created By', 'Created Date', 'Resolved Date'
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo to match UI
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Fill Data
    for c in complaints:
        ws.append([
            c.id,
            c.complainant,
            c.employer or 'N/A',
            c.nature_of_complaint,
            c.current_status,
            c.created_by.username if c.created_by else 'System',
            c.created_at.replace(tzinfo=None) if c.created_at else 'N/A',
            c.resolved_date.replace(tzinfo=None) if c.resolved_date else 'Pending'
        ])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column_letter].width = max_length + 2

    # Return Excel File
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Complaint_Log_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response

from django.core.paginator import Paginator

@login_required
def email_registry_view(request):
    """
    MANAGER (omega) ONLY: Email Registry / Delegations Overview.
    Shows all emails, their statuses, and allows searching/pagination.
    """
    if request.user.username.lower() != 'omega':
        messages.error(request, "Access restricted to management.")
        return redirect('dashboard')
    
    # 1. Get all delegated tasks from the database
    # We order by received_timestamp descending so newest items are on top
    all_tasks = CrmDelegateTo.objects.all().order_by('-received_timestamp')

    # 2. Setup Pagination (matching your template's page_obj variable)
    # 10 tasks per page
    paginator = Paginator(all_tasks, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'delegations_overview.html', {
        'page_obj': page_obj
    })
    
@login_required
def recycle_bin_view(request):
    """
    MANAGER (omega) ONLY: View items marked for Recycle Bin.
    Filters by the status string 'Recycled'.
    """
    if request.user.username.lower() != 'omega':
        messages.error(request, "Access restricted.")
        return redirect('dashboard')

    # Status must match the 'final_status' variable in your service layer
    recycled_items = CrmDelegateTo.objects.filter(
        status='Recycled'
    ).order_by('-received_timestamp')

    return render(request, 'recycle_bin.html', {
        'recycled_tasks': recycled_items
    })

@login_required
def delete_recycled_tasks_view(request):
    """
    Handles the POST request from the 'Delete Selected' button.
    """
    if request.method == 'POST' and request.user.username.lower() == 'omega':
        email_ids = request.POST.getlist('email_ids')
        if email_ids:
            # Physically delete or mark as permanently purged
            count = CrmDelegateTo.objects.filter(email_id__in=email_ids).delete()
            messages.success(request, f"Successfully purged {count[0]} emails from the system.")
        else:
            messages.warning(request, "No items were selected for deletion.")
            
    return redirect('recycle_bin')

@login_required
def download_attachment_view(request, message_id, attachment_id):
    """Securely downloads email attachments via Microsoft Graph."""
    endpoint = f"messages/{message_id}/attachments/{attachment_id}"
    attachment_data = OutlookGraphService._make_graph_request(endpoint, method='GET')
    
    if 'error' in attachment_data:
        messages.error(request, f"Microsoft Error: {attachment_data['error']}")
        return redirect('delegate_action', email_id=message_id)

    try:
        file_content = base64.b64decode(attachment_data.get('contentBytes', ''))
        filename = attachment_data.get('name', 'attachment')
        content_type = attachment_data.get('contentType', 'application/octet-stream')
        
        response = HttpResponse(file_content, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        messages.error(request, f"Download failed: {str(e)}")
        return redirect('delegate_action', email_id=message_id)
    
@login_required
def view_email_thread(request, email_id):
    """A dedicated page to see the full history of an email thread including attachments."""
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    # 1. Get the delegation record
    task = get_object_or_404(CrmDelegateTo, email_id=email_id)
    
    # 2. Get the original inbox record
    inbox_item = CrmInbox.objects.filter(email_id=email_id).first()
    
    # 3. Get the full audit trail
    actions = CrmDelegateAction.objects.filter(task_email_id=email_id).order_by('action_timestamp')

    # 4. Fetch the full HTML body & Attachments
    email_body = "Could not retrieve content from Outlook."
    attachments_list = []
    
    try:
        # Fetch Body
        endpoint = f"messages/{email_id}"
        email_data = OutlookGraphService._make_graph_request(endpoint, method='GET')
        
        if isinstance(email_data, dict) and 'error' not in email_data:
            email_body = email_data.get('body', {}).get('content', "")
        elif inbox_item:
            email_body = inbox_item.snippet

        # 🚀 5. Fetch Attachments (Live from Outlook)
        attachments_list = OutlookGraphService.fetch_attachments(target_email, email_id)
        
        # Inject contentBytes for thumbnails
        if attachments_list and isinstance(attachments_list, list):
            for att in attachments_list:
                content_type = att.get('contentType', '').lower()
                if 'image' in content_type:
                    raw_att = OutlookGraphService.get_attachment_raw(target_email, email_id, att['id'])
                    if raw_att and isinstance(raw_att, dict) and 'contentBytes' in raw_att:
                        att['contentBytes'] = raw_att['contentBytes']

    except Exception as e:
        print(f"Error in view_email_thread: {e}")
        if inbox_item:
            email_body = inbox_item.snippet

    context = {
        'task': task,
        'inbox_item': inbox_item,
        'actions': actions,
        'email_body': email_body,
        'attachments': attachments_list, # Key for the HTML loop
        'email_id': email_id,
    }
    return render(request, 'email_thread.html', context)

@login_required
def download_attachment_view(request, message_id, attachment_id):
    """
    Fetches the raw attachment from Microsoft Graph and serves it as a download.
    """
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    
    # 1. Fetch raw data from your service
    raw_data = OutlookGraphService.get_attachment_raw(target_email, message_id, attachment_id)
    
    if 'error' in raw_data or 'contentBytes' not in raw_data:
        return HttpResponse("Error: Could not retrieve attachment.", status=404)

    # 2. Decode the Base64 content
    file_content = base64.b64decode(raw_data['contentBytes'])
    file_name = raw_data.get('name', 'attachment')
    content_type = raw_data.get('contentType', 'application/octet-stream')

    # 3. Create the HTTP response with file headers
    response = HttpResponse(file_content, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    
    return response

@login_required
def export_delegation_report_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from django.utils import timezone
    from django.db.models import Q

    # 1. Apply date filters using the helper
    date_query, start_str, end_str = get_date_filters(request)
    
    # Filter for entries that have been categorized/processed
    valid_entries_filter = Q(category__isnull=False) & ~Q(category='')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Delegation Flow Report"

    # 2. HEADER DEFINITION
    headers = [
        'Communication Type',       # From ClientNotes or Default 'Email'
        'Email Received Date',      # Original Arrival from Inbox
        'Actioned By',              # The Manager/User who delegated
        'Date of Action',           # When it was delegated
        'Delegated To',             # Assigned Agent
        'Member Group Code', 
        'Main Category', 
        'Enquiry Selection', 
        'Grouping', 
        'Status'
    ]
    ws.append(headers)

    # Apply Corporate Green Styling to Header Row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 3. Fetch Email Delegation Records
    email_records = CrmDelegateTo.objects.filter(
        date_query & valid_entries_filter
    ).order_by('-received_timestamp')

    # 4. Fill Excel Rows with Email Data
    for item in email_records:
        first_action = CrmDelegateAction.objects.filter(
            task_email_id=item.email_id
        ).order_by('action_timestamp').first()
        actioned_by_user = first_action.action_user if first_action else (item.delegated_by or "System")

        inbox_record = CrmInbox.objects.filter(email_id=item.email_id).first()
        received_date = inbox_record.received_timestamp.replace(tzinfo=None) if inbox_record and inbox_record.received_timestamp else item.received_timestamp.replace(tzinfo=None)

        category_display = CATEGORY_NAMES.get(str(item.category), str(item.category)) if item.category else "Unclassified"
        grouping_display = ENQUIRY_GROUPING_MAP.get(item.type, "Unmapped") if item.type else "N/A"

        ws.append([
            "Email",                                            # Communication Type
            received_date,                                      # Email Received Date
            actioned_by_user,                                   # Actioned By
            item.received_timestamp.replace(tzinfo=None),       # Date of Action
            item.delegated_to or "Unassigned",                  # Delegated To
            item.member_group_code or "N/A",                    # Member Group Code
            category_display,                                   # Main Category
            item.type or "None",                                # Enquiry Selection
            grouping_display,                                   # Grouping
            item.status                                         # Status
        ])

    # 5. Fetch and Append Client Notes (Calls/Teams/etc)
    # --- SAFETY CHECK FOR "None" STRINGS ---
    def is_valid_date_str(val):
        return val and val != "None" and val != ""

    notes_filter = Q()
    if is_valid_date_str(start_str):
        notes_filter &= Q(date__date__gte=start_str)
    if is_valid_date_str(end_str):
        notes_filter &= Q(date__date__lte=end_str)

    note_records = ClientNotes.objects.filter(notes_filter).order_by('-date')

    for note in note_records:
        ws.append([
            note.communication_type or "Note",                  # Communication Type
            "N/A",                                              # Email Received Date
            note.user,                                          # Actioned By
            note.date.replace(tzinfo=None),                     # Date of Action
            "N/A",                                              # Delegated To
            note.related_member_group_code or "N/A",            # Member Group Code
            "Communication Log",                                # Main Category
            note.action_notes or "N/A",                         # Enquiry Selection
            "N/A",                                              # Grouping
            "Completed"                                         # Status
        ])

    # 6. Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column_letter].width = max_length + 3

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Delegation_Flow_Report_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response

@login_required
def final_sla_report_view(request):
    if request.user.username.lower() != 'omega':
        return redirect('dashboard')

    start_str = request.GET.get('start_date')
    end_str = request.GET.get('end_date')
    
    # --- INTERNAL HELPER: Prevent "None" string from crashing filters ---
    def get_valid_date(date_val):
        if date_val and date_val != "None" and date_val != "":
            return parse_date(date_val)
        return None

    # 1. Setup Date Filters
    delegate_q = Q()
    notes_q = Q()
    email_log_q = Q()

    start_dt = get_valid_date(start_str)
    if start_dt:
        delegate_q &= Q(received_timestamp__date__gte=start_dt)
        notes_q &= Q(date__gte=start_dt)
        email_log_q &= Q(sent_at__date__gte=start_dt)

    end_dt = get_valid_date(end_str)
    if end_dt:
        delegate_q &= Q(received_timestamp__date__lte=end_dt)
        notes_q &= Q(date__lte=end_dt)
        email_log_q &= Q(sent_at__date__lte=end_dt)

    # 2. Handle Excel Export Request
    if request.GET.get('export') == 'excel':
        # Now passing clean Q objects that won't contain None values
        return export_sla_excel(delegate_q, notes_q, email_log_q)

    # 3. Aggregate Grouped Totals
    calls_breakdown = ClientNotes.objects.filter(
        notes_q, 
        communication_type__in=['Incoming Call', 'Outgoing Call']
    ).values('communication_type', 'action_notes').annotate(total=Count('id')).order_by('communication_type')

    inbox_breakdown = CrmDelegateTo.objects.filter(delegate_q).values(
        'type', 'status'
    ).annotate(total=Count('email_id')).order_by('type')

    queries_breakdown = ClientNotes.objects.filter(
        notes_q, 
        communication_type='Notes Log'
    ).values('action_notes').annotate(total=Count('id')).order_by('action_notes')

    sent_emails_breakdown = DirectEmailLog.objects.filter(email_log_q).values(
        'subject' 
    ).annotate(total=Count('id'))

    teams_breakdown = ClientNotes.objects.filter(
        notes_q, 
        communication_type='Teams Meeting'
    ).values('action_notes').annotate(total=Count('id')).order_by('action_notes')

    # 4. Final Data Structure with Badge Totals
    report_data = {
        'CALLS': calls_breakdown,
        'CALLS_TOTAL': sum(item['total'] for item in calls_breakdown),
        
        'INBOX_EMAILS': inbox_breakdown,
        'INBOX_TOTAL': sum(item['total'] for item in inbox_breakdown),
        
        'QUERIES': queries_breakdown,
        'QUERIES_TOTAL': sum(item['total'] for item in queries_breakdown),
        
        'SENT_EMAILS': sent_emails_breakdown,
        'SENT_TOTAL': sum(item['total'] for item in sent_emails_breakdown),
        
        'TEAMS': teams_breakdown,
        'TEAMS_TOTAL': sum(item['total'] for item in teams_breakdown),
    }

    grand_total = (
        report_data['CALLS_TOTAL'] + 
        report_data['INBOX_TOTAL'] + 
        report_data['QUERIES_TOTAL'] + 
        report_data['SENT_TOTAL'] + 
        report_data['TEAMS_TOTAL']
    )

    details = ClientNotes.objects.filter(notes_q).order_by('-date')[:50]

    context = {
        'report': report_data,
        'grand_total': grand_total,
        'start_date': start_str,
        'end_date': end_str,
        'details': details,
    }
    return render(request, 'final_sla_report.html', context)

def export_sla_excel(delegate_q, notes_q, email_log_q):
    """Enhanced helper to generate the detailed SLA Excel file matching live totals"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SLA Detailed Breakdown"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")

    headers = ['Date', 'Main Category', 'Enquiry Type', 'Action Type', 'User', 'Reference', 'Content Preview']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # 1. Add Communication Logs (Calls, Teams, Queries)
    for note in ClientNotes.objects.filter(notes_q):
        ws.append([
            note.date.replace(tzinfo=None),
            'Communication Log',
            note.communication_type,
            note.action_notes,
            note.user,
            note.related_member_group_code,
            note.notes[:100]
        ])

    # 2. ADD MISSING DATA: Inbox Emails & Task Queries
    # This pulls the 9 Inbox Emails and 2 Queries missing from your previous export
    for task in CrmDelegateTo.objects.filter(delegate_q):
        ws.append([
            task.received_timestamp.replace(tzinfo=None),
            'Inbox Delegation',
            task.type or 'Incoming Email',
            task.status,
            task.delegated_by,
            task.member_group_code,
            task.subject
        ])

    # 3. ADD MISSING DATA: Sent Email Replies (SLA Category: Sent Emails)
    # This pulls the 'REPLY_SENT' actions logged in the thread table
    for action in CrmDelegateAction.objects.filter(task_email_id__isnull=False, action_type='REPLY_SENT'):
        ws.append([
            action.action_timestamp.replace(tzinfo=None),
            'Sent Email (Thread Reply)',
            'Email Response',
            'REPLY_SENT',
            action.action_user,
            'N/A',
            action.related_subject
        ])

    # 4. Add Direct Emails
    for email in DirectEmailLog.objects.filter(email_log_q):
        ws.append([
            email.sent_at.replace(tzinfo=None),
            'Direct Email',
            'Outgoing Email',
            'Sent',
            email.sent_by_user.username,
            email.member_group_code,
            email.subject
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Detailed_SLA_Report_{date.today()}.xlsx"'
    wb.save(response)
    return response

from django.utils.dateparse import parse_date

def get_date_filters(request):
    """Helper function to extract dates safely from GET request"""
    start_str = request.GET.get('start_date')
    end_str = request.GET.get('end_date')
    
    # Check if the string is empty or literally the string "None"
    def is_valid(val):
        return val and val != "None" and val != ""

    filters = Q()
    
    if is_valid(start_str):
        parsed_start = parse_date(start_str)
        if parsed_start:
            filters &= Q(received_timestamp__date__gte=parsed_start)
            
    if is_valid(end_str):
        parsed_end = parse_date(end_str)
        if parsed_end:
            filters &= Q(received_timestamp__date__lte=parsed_end)
            
    return filters, start_str, end_str


def get_unified_email_data(request):
    """
    Helper function to merge CrmInbox (New) and CrmDelegateTo (Delegated) 
    data into a single list for the UI and CSV Export.
    Includes safety checks for 'None' strings and Agent lookups.
    """
    from django.db.models import Max, Q
    from datetime import datetime

    # 1. Action Lookup (Check for replies and original delegation actions)
    last_reply_map = CrmDelegateAction.objects.filter(
        action_type='REPLY_SENT'
    ).values('task_email_id').annotate(last_replied=Max('action_timestamp'))
    last_reply_dict = {item['task_email_id']: item['last_replied'] for item in last_reply_map}

    delegation_user_map = CrmDelegateAction.objects.filter(
        action_type='restore_to_inbox' 
    ).values('task_email_id', 'action_user')
    user_dict = {item['task_email_id']: item['action_user'] for item in delegation_user_map}

    # 2. Setup Filters
    date_filter = Q()
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    def is_valid_date(val):
        return val and val != "None" and val != ""

    if is_valid_date(start_date_str):
        date_filter &= Q(received_timestamp__date__gte=start_date_str)
    if is_valid_date(end_date_str):
        date_filter &= Q(received_timestamp__date__lte=end_date_str)

    # 3. Fetch Delegated Emails
    delegated_qs = CrmDelegateTo.objects.filter(date_filter).order_by('-received_timestamp')
    delegated_ids = {task.email_id for task in delegated_qs}

    # 4. Fetch New/Pending Emails
    new_emails_qs = CrmInbox.objects.filter(date_filter).exclude(
        email_id__in=delegated_ids
    ).order_by('-received_timestamp')

    unified_list = []

    # Process Delegated
    for task in delegated_qs:
        unified_list.append({
            'email_id': task.email_id,
            'subject': task.subject,
            'sender': task.sender or "Unknown",
            'status': task.status,
            'delegated_to': task.delegated_to or user_dict.get(task.email_id, "System"),
            'member_group_code': task.member_group_code,
            'category': CATEGORY_NAMES.get(str(task.category), task.category),
            'type': task.type,  # <--- CRITICAL: Now pulls Enquiry Selection from DB
            'received_timestamp': task.received_timestamp,
            'last_replied_timestamp': last_reply_dict.get(task.email_id),
            'is_delegated': True
        })

    # Process New
    for email in new_emails_qs:
        unified_list.append({
            'email_id': email.email_id,
            'subject': email.subject,
            'sender': email.sender,
            'status': 'New',
            'delegated_to': 'Inbox (Unassigned)',
            'member_group_code': 'N/A',
            'category': 'Unclassified',
            'type': 'Incoming Email', # Default for un-delegated items
            'received_timestamp': email.received_timestamp,
            'last_replied_timestamp': None,
            'is_delegated': False
        })

    # 5. Search Filter
    search_text = request.GET.get('search', '').lower()
    if search_text:
        unified_list = [
            row for row in unified_list 
            if search_text in row['subject'].lower() or 
               search_text in row['sender'].lower() or 
               search_text in str(row['member_group_code']).lower()
        ]

    # Final Sort
    unified_list.sort(key=lambda x: x['received_timestamp'] if x['received_timestamp'] else datetime.min, reverse=True)
    return unified_list

@login_required
def email_workflow_log_view(request):
    from django.core.paginator import Paginator

    # Uses the helper above which now includes the 'type' field
    data = get_unified_email_data(request)
    
    paginator = Paginator(data, 25) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'email_workflow_log.html', {'page_obj': page_obj})

@login_required
def export_email_workflow_csv(request):
    import csv
    from django.http import HttpResponse
    from django.utils import timezone

    data = get_unified_email_data(request)
    
    response = HttpResponse(content_type='text/csv')
    filename = f"Email_Workflow_{timezone.now().strftime('%Y-%m-%d')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    
    writer.writerow([
        'Received Date', 
        'Sender', 
        'Subject', 
        'Status', 
        'Agent (Assigned)', 
        'Member Code', 
        'Category',
        'Enquiry Selection', 
        'Date Replied'
    ])

    for row in data:
        received_dt = row['received_timestamp'].strftime('%Y-%m-%d %H:%M') if row['received_timestamp'] else 'N/A'
        
        reply_dt = 'No Reply'
        if row.get('last_replied_timestamp'):
            reply_dt = row['last_replied_timestamp'].strftime('%Y-%m-%d %H:%M')

        agent_name = row.get('delegated_to') or 'Inbox (Unassigned)'

        writer.writerow([
            received_dt,
            row.get('sender', 'Unknown'),
            row.get('subject', 'No Subject'),
            row.get('status', 'New'),
            agent_name,
            row.get('member_group_code', 'N/A'),
            row.get('category', 'Unclassified'),
            row.get('type', 'None'), # This will now pull the value you added to the helper
            reply_dt
        ])

    return response

@login_required
def download_actual_email(request, email_id):
    """
    Fetches the raw MIME content from Outlook and serves it as a .eml file.
    This can be called from the Thread view OR the Member Information Email Log.
    """
    try:
        # 1. Get raw bytes from our updated service
        raw_mime = OutlookGraphService.get_email_mime_content(email_id)

        if isinstance(raw_mime, dict) and 'error' in raw_mime:
            messages.error(request, f"Microsoft Error: {raw_mime['error']}")
            return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

        # 2. Try to find a subject for the filename
        # We check CrmDelegateTo first, then CrmInbox
        subject_source = CrmDelegateTo.objects.filter(email_id=email_id).first() or \
                         CrmInbox.objects.filter(email_id=email_id).first()
        
        filename = "email_record.eml"
        if subject_source and subject_source.subject:
            # Clean filename: remove non-alphanumeric characters
            clean_subject = "".join([c for c in subject_source.subject if c.isalnum() or c in (' ', '-', '_')]).strip()
            filename = f"{clean_subject[:50]}.eml"

        # 3. Return as a downloadable message/rfc822 file
        response = HttpResponse(raw_mime, content_type='message/rfc822')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        messages.error(request, f"Download failed: {str(e)}")
        return redirect(request.META.get('HTTP_REFERER', 'dashboard'))