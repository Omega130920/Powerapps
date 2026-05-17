import base64
from datetime import date
from email import parser
import random
import traceback
from dateutil import parser as date_parser
from time import timezone
from django.shortcuts import render, get_object_or_404, redirect
from django.conf import settings
from django.contrib import messages
from django.http import HttpResponse, Http404, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
import logging
from django.db.models import Q
import openpyxl
import pandas as pd
from django.db import transaction
from dateutil.relativedelta import relativedelta
from django.core.paginator import Paginator
from datetime import datetime
from .services.outlook_graph_service import OutlookGraphService

logger = logging.getLogger(__name__)

# Import your unmanaged models
from .models import AdHocList, ClaimAffordability, ClaimList, PssubfBeneficiary, PssubfDirectEmail, PssubfInbox, PssubfDelegate, PssubfAction, PssubfNote, PssubfProfileNote
# Import your verified services
from PSSUBF_APP.services.outlook_graph_service import OutlookGraphService
from PSSUBF_APP.services.delegation_service import delegate_pssubf_task

def clean_numeric(val):
    if not val or str(val).lower() == 'undefined' or str(val).strip() == '': 
        return 0
    return str(val).replace('R', '').replace(',', '').replace('%', '').strip()

@login_required
def pssubf_switchboard(request):
    """
    Main Menu / Switchboard.
    Mapped to path('')
    Renders the card-based navigation (dashboard.html).
    """
    return render(request, 'pssubf/dashboard.html')

from django.db.models import Q

@login_required
def outlook_dashboard_view(request):
    # Security Check
    if request.user.username.lower() != 'omega' and not request.user.is_superuser:
        messages.error(request, "Access restricted.")
        return redirect('pssubf_switchboard')

    target_email = request.GET.get('email', 'your_default_email@domain.com')
    search_query = request.GET.get('q', '').strip().lower()
    sort_order = request.GET.get('sort', 'newest')
    
    # 1. Fetch live data from Graph API
    inbox_data = OutlookGraphService.fetch_inbox_messages(target_email, top_count=50) 
    
    if 'error' in inbox_data:
        return render(request, 'pssubf/inbox_list.html', {'error': inbox_data['error']})

    all_emails = inbox_data.get('value', [])
    email_ids = [e['id'] for e in all_emails]

    # 2. Bulk fetch existing records to reduce DB hits
    local_inbox_map = PssubfInbox.objects.filter(email_id__in=email_ids).in_bulk(field_name='email_id')
    delegated_map = PssubfDelegate.objects.filter(email_id__in=email_ids).in_bulk(field_name='email_id')

    filtered_emails = []

    for email in all_emails:
        e_id = email['id']
        
        # Skip if already processed (anything not 'Assigned' is considered archived/done)
        if e_id in delegated_map and delegated_map[e_id].status != 'Assigned':
            continue

        # Sync PssubfInbox (The Archive)
        if e_id not in local_inbox_map:
            received_date = email.get('receivedDateTime')
            
            # FIX: Use 'or' to provide a fallback string if subject is None/Null from API
            safe_subject = email.get('subject') or '(No Subject)'
            
            local_record = PssubfInbox.objects.create(
                email_id=e_id,
                subject=safe_subject,
                sender=email.get('from', {}).get('emailAddress', {}).get('address', '').lower(),
                received_timestamp=date_parser.isoparse(received_date) if received_date else timezone.now(),
                snippet=email.get('bodyPreview', '') or '',
                status='Pending'
            )
        else:
            local_record = local_inbox_map[e_id]

        # Sync PssubfDelegate (The Active Task List)
        if e_id not in delegated_map:
            delegation = PssubfDelegate.objects.create(
                email_id=e_id,
                status='Assigned',
                subject=local_record.subject or '(No Subject)',
                sender=local_record.sender,
            )
        else:
            delegation = delegated_map[e_id]

        # Prepare for Template
        email_display = {
            'id': e_id,
            'subject': local_record.subject or '(No Subject)',
            'sender': local_record.sender,
            'received_at': local_record.received_timestamp,
            'snippet': local_record.snippet,
            'status': delegation.status,
        }

        # Search Filter
        if search_query:
            # Added safe handling for subject in search string
            subj_lower = (email_display['subject'] or '').lower()
            sender_lower = (email_display['sender'] or '').lower()
            content = f"{subj_lower} {sender_lower}"
            if search_query not in content:
                continue

        filtered_emails.append(email_display)

    # 3. Sort
    reverse_sort = (sort_order == 'newest')
    filtered_emails.sort(key=lambda x: x['received_at'] if x['received_at'] else timezone.now(), reverse=reverse_sort)

    return render(request, 'pssubf/inbox_list.html', {
        'messages': filtered_emails,
        'search_query': search_query,
        'sort_order': sort_order,
        'target_email': target_email
    })

@login_required
def pssubf_delegate_view(request, email_id):
    """View to fetch live email details and delegate."""
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    available_users = User.objects.filter(is_active=True)

    # We still keep a reference to local inbox if you need to update a status 
    # but the primary view is now live.
    inbox_item = PssubfInbox.objects.filter(email_id=email_id).first()

    if request.method == 'POST':
        agent_name = request.POST.get('assigned_agent')
        is_recycle = 'recycle' in request.POST
        
        success, message = delegate_pssubf_task(
            email_id=email_id,
            agent_name=agent_name,
            delegator_user=request.user,
            form_data=request.POST,
            is_recycle=is_recycle
        )
        
        if success:
            # Sync the local PssubfInbox status if the record exists
            if inbox_item:
                inbox_item.status = 'Recycled' if is_recycle else 'Delegated'
                inbox_item.save()

            messages.success(request, message)
            return redirect('pssubf_delegations_list')
        else:
            messages.error(request, f"Error: {message}")

    # Fetch live content for the detail view
    email_data = OutlookGraphService._make_graph_request(f"messages/{email_id}", method='GET')
    
    if isinstance(email_data, dict) and 'error' in email_data:
        email_subject = inbox_item.subject if inbox_item else "Error"
        email_content = inbox_item.snippet if inbox_item else "Content Unavailable"
        attachments = []
    else:
        email_subject = email_data.get('subject', '(No Subject)')
        email_content = email_data.get('body', {}).get('content', '')
        attachments = OutlookGraphService.fetch_attachments(target_email, email_id)
        
        for att in attachments:
            if att.get('isInline') and att.get('contentId'):
                cid = att.get('contentId')
                raw = OutlookGraphService.get_attachment_raw(target_email, email_id, att['id'])
                if raw and isinstance(raw, dict) and 'contentBytes' in raw:
                    data_url = f"data:{att.get('contentType', 'image/png')};base64,{raw['contentBytes']}"
                    email_content = email_content.replace(f"cid:{cid}", data_url)
                    att['contentBytes'] = raw['contentBytes']

    return render(request, 'pssubf/delegate.html', {
        'email_id': email_id,
        'email_subject': email_subject,
        'email_content': email_content,
        'attachments': attachments,
        'available_users': available_users,
        'inbox_item': inbox_item
    })

logger = logging.getLogger(__name__)

@login_required
def pssubf_action_view(request, email_id):
    """
    Agent Action View: Handles Notes, Metadata Updates, Completion, 
    and Email Replies while resolving broken inline images.
    Now handles both Delegated tasks and Direct Portal communications.
    """
    
    # 1. RESOLVE TASK SOURCE (Fixes the 404 for DIRECT_ IDs)
    # Check Delegations first
    task = PssubfDelegate.objects.filter(email_id=email_id).first()
    is_direct_entry = False

    if not task:
        # If not in delegates, check the Inbox (for Direct entries)
        task = PssubfInbox.objects.filter(email_id=email_id).first()
        is_direct_entry = True

    if not task:
        raise Http404("No record found for this ID in Delegations or Inbox.")

    target_email = settings.OUTLOOK_EMAIL_ADDRESS 

    # --- POST Logic ---
    if request.method == 'POST':
        action_type = request.POST.get('action_type')

        # 1. Update Metadata
        if action_type == 'update_metadata':
            task.member_group_code = request.POST.get('member_group_code')
            task.email_category = request.POST.get('email_category')
            task.status = request.POST.get('status')
            task.save()
            
            PssubfAction.objects.create(
                task_email_id=email_id,
                action_user=request.user.username,
                action_type="METADATA_UPDATE",
                note_content=f"Updated: Category={task.email_category}, Group={task.member_group_code}, Status={task.status}"
            )
            messages.success(request, "Task information updated successfully.")

        # 2. Add Internal Note
        elif action_type == 'add_note':
            note_text = request.POST.get('note_content')
            new_category = request.POST.get('email_category')
            new_status = request.POST.get('status')
            
            call_direction = request.POST.get('call_direction', 'Outbound')
            call_method = request.POST.get('call_method', 'Note')
            call_type = request.POST.get('call_type', 'General Note')
            
            if new_category:
                task.email_category = new_category
            if new_status:
                task.status = new_status
            task.save()

            audit_string = f"[{call_direction} | {call_method} | {call_type}]"
            
            PssubfNote.objects.create(
                task_email_id=email_id,
                agent_name=request.user.username,
                note_text=f"{audit_string} {note_text}",
                classification_at_time=getattr(task, 'email_category', 'N/A'),
                status_at_time=task.status
            )

            PssubfAction.objects.create(
                task_email_id=email_id,
                action_user=request.user.username,
                action_type="NOTE",
                note_content=f"{audit_string} {note_text}"
            )
            messages.success(request, "Internal note saved.")

        # 3. Handle External Email Reply
        elif action_type == 'send_reply':
            recipient = request.POST.get('reply_recipient')
            subject = request.POST.get('reply_subject')
            body_content = request.POST.get('reply_body')
            
            call_direction = "Outbound"
            call_method = "Emails"
            call_type = "Feedback to Beneficiary"

            uploaded_files = request.FILES.getlist('reply_attachments')
            attachments_payload = []
            
            for f in uploaded_files:
                try:
                    content_bytes = f.read()
                    encoded_content = base64.b64encode(content_bytes).decode('utf-8')
                    attachments_payload.append({
                        "@odata.type": "#microsoft.graph.fileAttachment",
                        "name": f.name,
                        "contentType": f.content_type,
                        "contentBytes": encoded_content
                    })
                except Exception as e:
                    logger.error(f"Attachment encoding error: {e}")

            response = OutlookGraphService.send_outlook_email(
                sender=target_email,
                recipient=recipient,
                subject=subject,
                body=body_content,
                attachments=attachments_payload
            )

            if isinstance(response, dict) and 'error' in response:
                messages.error(request, f"Email failed: {response.get('error')}")
            else:
                audit_string = f"[{call_direction} | {call_method} | {call_type}]"
                
                PssubfAction.objects.create(
                    task_email_id=email_id,
                    action_user=request.user.username,
                    action_type="EMAIL_REPLY",
                    note_content=f"{audit_string}\nTo: {recipient}\nSubject: {subject}\n\n{body_content}"
                )

                PssubfNote.objects.create(
                    task_email_id=email_id,
                    agent_name=request.user.username,
                    note_text=f"{audit_string} REPLY SENT TO {recipient}: {body_content}",
                    classification_at_time=getattr(task, 'email_category', 'N/A'),
                    status_at_time=task.status
                )
                messages.success(request, "Reply sent and logged.")

        # 4. Mark as Complete
        elif action_type == 'mark_complete':
            task.status = 'Completed'
            task.save()
            
            PssubfAction.objects.create(
                task_email_id=email_id,
                action_user=request.user.username,
                action_type="COMPLETED",
                note_content="Agent marked task as completed."
            )
            messages.success(request, "Task closed.")
            return redirect('pssubf_delegations_list')

        return redirect('pssubf_action', email_id=email_id)

    # --- GET Logic ---
    attachments = []
    email_content = ""
    email_subject = getattr(task, 'subject', '(No Subject)')

    # Only request from Graph if it's NOT a direct portal entry
    if not str(email_id).startswith("DIRECT_"):
        try:
            email_data = OutlookGraphService._make_graph_request(f"messages/{email_id}", method='GET')
            attachments = OutlookGraphService.fetch_attachments(target_email, email_id)
            email_content = email_data.get('body', {}).get('content', 'Content unavailable.')
            email_subject = email_data.get('subject', email_subject)

            # Inline Image Resolution
            for att in attachments:
                if att.get('isInline') and att.get('contentId'):
                    cid = att.get('contentId')
                    raw = OutlookGraphService.get_attachment_raw(target_email, email_id, att['id'])
                    if raw and isinstance(raw, dict) and 'contentBytes' in raw:
                        base64_data = raw['contentBytes']
                        content_type = att.get('contentType', 'image/png')
                        data_url = f"data:{content_type};base64,{base64_data}"
                        email_content = email_content.replace(f"cid:{cid}", data_url)
                        att['contentBytes'] = base64_data
        except Exception as e:
            logger.error(f"Graph API Error: {e}")
            email_content = "Could not retrieve email body from Outlook. This may be a local-only record."
    else:
        # Handle Direct Entry Display
        email_content = f"<p><strong>Internal Communication Record</strong></p><p>{getattr(task, 'snippet', 'No content available.')}</p>"

    # Fetch History
    history = PssubfAction.objects.filter(task_email_id=email_id).order_by('-action_timestamp')

    return render(request, 'pssubf/action_detail.html', {
        'task': task,
        'email_subject': email_subject,
        'email_content': email_content,
        'attachments': attachments,
        'history': history,
        'email_id': email_id,
        'is_direct': is_direct_entry
    })

import re # Add to imports

@login_required
def pssubf_view_thread(request, email_id):
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    email_data = OutlookGraphService._make_graph_request(f"messages/{email_id}", method='GET')
    
    email_body = email_data.get('body', {}).get('content', '')
    attachments = OutlookGraphService.fetch_attachments(target_email, email_id)

    # 🚀 FIX: Replace CID with Base64 for inline images
    for att in attachments:
        # Check if it's an inline attachment with a Content-ID
        if att.get('isInline') and att.get('contentId'):
            cid = att.get('contentId')
            # Fetch the raw content if not already present
            raw_data = OutlookGraphService.get_attachment_raw(target_email, email_id, att['id'])
            if raw_data and 'contentBytes' in raw_data:
                base64_data = raw_data['contentBytes']
                content_type = att.get('contentType', 'image/png')
                
                # Replace src="cid:..." with src="data:image/png;base64,..."
                data_url = f"data:{content_type};base64,{base64_data}"
                email_body = email_body.replace(f"cid:{cid}", data_url)

    actions = PssubfAction.objects.filter(task_email_id=email_id).order_by('-action_timestamp')

    return render(request, 'pssubf/thread_history.html', {
        'email_id': email_id,
        'email_subject': email_data.get('subject', 'No Subject'),
        'sender_name': email_data.get('from', {}).get('emailAddress', {}).get('name', 'Unknown'),
        'sender_email': email_data.get('from', {}).get('emailAddress', {}).get('address', ''),
        'email_body': email_body, # Now contains embedded images
        'attachments': attachments,
        'actions': actions
    })

@login_required
def download_pssubf_attachment(request, message_id, attachment_id):
    """Downloads raw file from Outlook Graph API."""
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    raw_data = OutlookGraphService.get_attachment_raw(target_email, message_id, attachment_id)
    
    if not raw_data or 'contentBytes' not in raw_data:
        raise Http404("Attachment not found.")
    
    file_content = base64.b64decode(raw_data['contentBytes'])
    response = HttpResponse(file_content, content_type=raw_data.get('contentType', 'application/octet-stream'))
    response['Content-Disposition'] = f'attachment; filename="{raw_data.get("name", "download")}"'
    return response

@login_required
def sync_pssubf_inbox(request):
    """Triggers the Outlook Graph API to fetch latest mail and save to MySQL."""
    try:
        # This calls your service logic to fetch and save to pssubf_inbox
        new_emails = OutlookGraphService.sync_latest_emails() 
        messages.success(request, f"Successfully synced {len(new_emails)} new emails.")
    except Exception as e:
        messages.error(request, f"Sync failed: {str(e)}")
    
    return redirect('pssubf_dashboard')

@login_required
def pssubf_delegations_list(request):
    """
    Displays the active queue, excluding completed and recycled items.
    """
    # Exclude both Recycled and Completed to keep the dashboard focused on active work
    delegations = PssubfDelegate.objects.exclude(
        status__in=['Recycled', 'Completed', 'Complete']
    ).order_by('-created_at')
    
    return render(request, 'pssubf/delegations_list.html', {
        'delegations': delegations
    })
    
@login_required
def pssubf_audit_logs(request):
    """The Master Archive / Audit Log view."""
    logs = PssubfAction.objects.all().order_by('-action_timestamp')
    return render(request, 'pssubf/audit_logs.html', {
        'logs': logs
    })
    
@login_required
def pssubf_recycle_bin(request):
    """
    Displays all items marked as 'Recycled'.
    Matches the status value found in the pssubf_delegate table.
    """
    # Changed filter from 'DLT' to 'Recycled' to match your DB screenshot
    recycled_tasks = PssubfDelegate.objects.filter(status='Recycled').order_by('-created_at')
    
    return render(request, 'pssubf/recycle_bin.html', {
        'recycled_tasks': recycled_tasks
    })

@login_required
def pssubf_restore_item(request, email_id):
    """
    Restores an item from the Recycle Bin back to the Triage Dashboard.
    Clears the status so the Dashboard filter no longer excludes it.
    """
    # 1. Get the recycled task record
    task = get_object_or_404(PssubfDelegate, email_id=email_id)
    
    # 2. Find the original Inbox record (The one the dashboard looks at)
    inbox_item = PssubfInbox.objects.filter(email_id=email_id).first()

    if inbox_item:
        # 🚀 CRITICAL: Clear the status. 
        # Setting it to '' or None makes it "New" again so it shows in the Dashboard.
        inbox_item.status = '' 
        inbox_item.save()

    # 3. Delete the 'Recycled' delegate entry so the system sees it as a fresh start
    task.delete() 
    
    # 4. Log the action
    PssubfAction.objects.create(
        task_email_id=email_id,
        action_user=request.user.username,
        action_type="RESTORE",
        note_content="Item restored. Status cleared to return to Triage Dashboard."
    )
    
    messages.success(request, "Email restored to Dashboard. You can now assign an agent.")
    
    # 5. Redirect to the Dashboard so you can see it's back
    return redirect('pssubf_dashboard')

@login_required
def pssubf_audit_logs(request):
    """
    Master Audit Log: Shows New, Delegated, and Completed actions.
    EXCLUDES all Recycle actions to keep the focus on productive workflows.
    """
    query = request.GET.get('q')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # 1. Start with all logs
    # 2. Exclude 'Recycle' action types and 'RESTORE' if you want a clean work log
    logs = PssubfAction.objects.exclude(action_type__in=['Recycle', 'RESTORE'])

    # Apply Search Filter (User, Type, or Content)
    if query:
        logs = logs.filter(
            Q(action_user__icontains=query) |
            Q(action_type__icontains=query) |
            Q(note_content__icontains=query)
        )

    # Apply Date Range Filters
    if start_date:
        logs = logs.filter(action_timestamp__date__gte=start_date)
    if end_date:
        logs = logs.filter(action_timestamp__date__lte=end_date)

    logs = logs.order_by('-action_timestamp')

    return render(request, 'pssubf/audit_logs.html', {
        'logs': logs,
        'query': query,
        'start_date': start_date,
        'end_date': end_date
    })
    
@login_required
def pssubf_recycle_view(request, email_id):
    """View to review recycled item details with fixed inline images before restoration."""
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    # Fetch from delegate table where status is Recycled
    task = get_object_or_404(PssubfDelegate, email_id=email_id, status='Recycled')

    # Fetch live content from Graph API
    email_data = OutlookGraphService._make_graph_request(f"messages/{email_id}", method='GET')
    
    if isinstance(email_data, dict) and 'error' in email_data:
        email_subject = task.subject or "Subject Unavailable"
        email_content = "Live content unavailable (Email may have been moved or deleted in Outlook)."
        attachments = []
    else:
        email_subject = email_data.get('subject', task.subject)
        email_content = email_data.get('body', {}).get('content', '')
        
        # Fetch Attachments
        attachments = OutlookGraphService.fetch_attachments(target_email, email_id)
        
        # 🚀 FIX: Resolve Inline Images (Signatures, etc.)
        for att in attachments:
            # Handle inline images for the email body
            if att.get('isInline') and att.get('contentId'):
                cid = att.get('contentId')
                raw = OutlookGraphService.get_attachment_raw(target_email, email_id, att['id'])
                if raw and isinstance(raw, dict) and 'contentBytes' in raw:
                    base64_data = raw['contentBytes']
                    content_type = att.get('contentType', 'image/png')
                    
                    # Replace CID link with Base64 Data URI
                    data_url = f"data:{content_type};base64,{base64_data}"
                    email_content = email_content.replace(f"cid:{cid}", data_url)
                    
                    # Attach bytes to object for display in the attachment list
                    att['contentBytes'] = base64_data
            
            # Fetch thumbnails for standard image attachments
            elif 'image' in att.get('contentType', '').lower() and not att.get('contentBytes'):
                raw = OutlookGraphService.get_attachment_raw(target_email, email_id, att['id'])
                if raw and isinstance(raw, dict) and 'contentBytes' in raw:
                    att['contentBytes'] = raw['contentBytes']

    return render(request, 'pssubf/recycle_detail.html', {
        'task': task,
        'email_id': email_id,
        'email_subject': email_subject,
        'email_content': email_content,
        'attachments': attachments
    })
    
@login_required
def pssubf_delete_permanent(request, email_id):
    """Permanently deletes a single record from the database."""
    task = get_object_or_404(PssubfDelegate, email_id=email_id)
    task.delete()
    
    # Also clean up the inbox status if needed, or just leave it
    PssubfInbox.objects.filter(email_id=email_id).update(status='DELETED')
    
    messages.error(request, "Record permanently deleted.")
    return redirect('pssubf_recycle_bin')

@login_required
def pssubf_bulk_delete(request):
    """Handles multiple deletions at once."""
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_ids')
        if selected_ids:
            # Delete from delegate table
            PssubfDelegate.objects.filter(email_id__in=selected_ids).delete()
            # Mark as deleted in inbox
            PssubfInbox.objects.filter(email_id__in=selected_ids).update(status='DELETED')
            messages.error(request, f"Permanently deleted {len(selected_ids)} records.")
            
    return redirect('pssubf_recycle_bin')

@login_required
def pssubf_history_preview(request, email_id):
    """
    AJAX View: Finds the sender of the current email and returns 
    all previous delegation records for that sender.
    """
    # 1. Get the current email from the inbox to identify the sender
    # We use .first() to avoid a 404 if the sync hasn't hit the DB yet
    current_mail = PssubfInbox.objects.filter(email_id=email_id).first()
    
    if not current_mail:
        # If not in local DB, we try to get the sender from the Graph API live
        email_data = OutlookGraphService._make_graph_request(f"messages/{email_id}", method='GET')
        sender_email = email_data.get('from', {}).get('emailAddress', {}).get('address', '')
    else:
        sender_email = current_mail.sender

    # 2. Find all previous delegations for this specific sender
    # We look in PssubfDelegate to see who worked on this person's files before
    previous_tasks = PssubfDelegate.objects.filter(
        sender=sender_email
    ).exclude(email_id=email_id).order_by('-created_at')

    context = {
        'sender_email': sender_email,
        'previous_tasks': previous_tasks,
    }

    # 3. Render the partial HTML that "pops up" in the modal
    return render(request, 'pssubf/partials/history_preview_content.html', context)

from django.db import connection # Import this at the top

@login_required
def beneficiary_import_view(request):
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        
        try:
            # Read as string to preserve IDs, then handle the "nan" values
            df = pd.read_excel(excel_file, dtype=str)
            df.columns = [c.strip() for c in df.columns]
            
            total_excel_rows = len(df)
            created_count = 0
            updated_count = 0
            import_errors = []

            with transaction.atomic():
                for index, row in df.iterrows():
                    
                    def to_date(val):
                        # Handle various "null" string versions from Excel/Pandas
                        s_val = str(val).strip().lower()
                        if pd.isna(val) or s_val in ['', 'nan', 'none', 'null']: 
                            return None
                        try: 
                            return pd.to_datetime(val).date()
                        except: 
                            return None

                    def to_float(val):
                        # FIXED: Specifically check for "nan" string before conversion
                        s_val = str(val).strip().lower()
                        if pd.isna(val) or s_val in ['', 'nan', 'none', 'null']:
                            return 0.00
                        try:
                            clean_val = s_val.replace(',', '').replace('r', '').strip()
                            return float(clean_val)
                        except: 
                            return 0.00

                    # Identification
                    m_no = str(row.get('Membership Number', '')).strip()
                    id_no = str(row.get('ID Number', '')).strip()

                    if not m_no or m_no.lower() == 'nan':
                        continue

                    try:
                        dob = to_date(row.get('DOB'))
                        cessation = dob + relativedelta(years=18) if dob else None

                        beneficiary_data = {
                            'old_membership_number': row.get('Old membership Number'),
                            'title': row.get('Title'),
                            'initials': row.get('Initials'),
                            'first_name': row.get('First Name'),
                            'second_name': row.get('Second Name'),
                            'last_name': row.get('Last Name'),
                            'id_number': id_no,
                            'dob': dob,
                            'cessation_date': cessation,
                            'gender': row.get('Gender'),
                            'employee_number': row.get('Employee Number'),
                            'fund_join_date': to_date(row.get('Fund Join Date')),
                            'stipened_frequency': row.get('Stipened Frequency'),
                            'stipened': to_float(row.get('Stipened')), # Now safe from "nan"
                            'mobile_1': row.get('Mobile 1'),
                            'email_1': row.get('Email 1'),
                            'mobile_2': row.get('Mobile 2'),
                            'email_2': row.get('Email 2'),
                            'mobile_3': row.get('Mobile 3'),
                            'email_3': row.get('Email 3'),
                            'guardian_mobile': row.get('Guardian Mobile'),
                            'guardian_email': row.get('Guaridan Email'),
                            'guardian_title': row.get('Guardian Title'),
                            'guardian_first_name': row.get('Guardian First Name'),
                            'guardian_second_name': row.get('Guardian Second Name'),
                            'guardian_last_name': row.get('Guardian Last Name'),
                            'guardian_initial': row.get('Guardian Initial'),
                            'guardian_dob': to_date(row.get('Guardian DOB')),
                            'guardian_id_number': row.get('Guardian ID Number'),
                            'guardian_id_type': row.get('Guardian ID Type'),
                            'guardian_gender': row.get('Guardian Gender'),
                            'guardian_address': row.get('Guardian Address'),
                        }

                        # Data Cleanup: Remove keys that are None if they are not allowed in DB
                        # (Membership Number and ID Number are handled separately)
                        beneficiary_data = {k: v for k, v in beneficiary_data.items() if pd.notnull(v) or v is None}

                        obj, created = PssubfBeneficiary.objects.update_or_create(
                            membership_number=m_no,
                            defaults=beneficiary_data
                        )
                        
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1

                    except Exception as e:
                        import_errors.append({"row": index + 2, "m_no": m_no, "reason": str(e)})

            # Session data for the list view
            request.session['import_stats'] = {
                'total': total_excel_rows,
                'created': created_count,
                'updated': updated_count,
                'failed': len(import_errors)
            }
            request.session['import_errors'] = import_errors
            
            return redirect('beneficiary_list')

        except Exception as e:
            messages.error(request, f"File Error: {str(e)}")

    return render(request, 'pssubf/beneficiary_import.html')

@login_required
def beneficiary_list_view(request):
    # 1. Get all records from the database
    queryset = PssubfBeneficiary.objects.all().order_by('last_name')

    # 2. Filter Logic (Status based on cessation_date)
    status_filter = request.GET.get('status')
    today = date.today()  # Live date for current calculation

    if status_filter == 'expired':
        # Members 18 and older (cessation date has passed or is today)
        queryset = queryset.filter(cessation_date__lte=today)
    elif status_filter == 'active':
        # Members under 18 (cessation date is in the future)
        queryset = queryset.filter(cessation_date__gt=today)

    # 3. Pagination (36 records per page)
    paginator = Paginator(queryset, 36)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 4. LIVE CALCULATION: Years and Months
    # This loop runs only for the 36 records on the current page for performance
    for member in page_obj:
        if member.dob:
            birth = member.dob
            
            # Calculate total months difference
            # Formula: (Years Diff * 12) + Months Diff
            total_months = (today.year - birth.year) * 12 + (today.month - birth.month)
            
            # Adjustment: If today's day of the month is earlier than the birth day, 
            # the current month hasn't fully completed yet.
            if today.day < birth.day:
                total_months -= 1
            
            # Prevent negative months for edge cases
            total_months = max(0, total_months)
            
            # Split total months into Years (Y) and remaining Months (M)
            years = total_months // 12
            months = total_months % 12
            
            # Format output: e.g., "19Y 03M"
            member.calculated_age = f"{years}Y {months:02d}M"
        else:
            member.calculated_age = "N/A"

    # 5. Handle Session Errors (from Excel imports)
    import_errors = request.session.pop('import_errors', [])
    
    return render(request, 'pssubf/beneficiary_list.html', {
        'page_obj': page_obj,
        'import_errors': import_errors,
        'current_status': status_filter,
        'total_count': queryset.count()
    })

import os
from django.core.files.storage import FileSystemStorage

@login_required
def beneficiary_details_view(request, membership_number):
    member = get_object_or_404(PssubfBeneficiary, membership_number=membership_number)
    
    # Handles South African formatting (spaces and commas) safely
    def clean_decimal(value):
        if not value or value == '': return 0.00
        cleaned = str(value).replace('R', '').replace(' ', '').replace(',', '.').replace('%', '').strip()
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return 0.00

    if request.method == 'POST':
        action = request.POST.get('action')

        # --- 1. HANDLE DIRECT EMAIL (COMPOSITION TAB) ---
        if action == 'send_direct_email':
            recipient = request.POST.get('to_email')
            subject = request.POST.get('subject')
            body_html = request.POST.get('email_html_content')

            if recipient and subject and body_html:
                result = OutlookGraphService.send_outlook_email(settings.OUTLOOK_EMAIL_ADDRESS, recipient, subject, body_html)
                
                if result.get('success') or result == {}:
                    direct_mail = PssubfDirectEmail.objects.create(
                        membership_number=membership_number,
                        agent_name=request.user.username,
                        recipient=recipient,
                        subject=subject,
                        body_html=body_html
                    )
                    
                    email_id = f"DIRECT_{membership_number}_{direct_mail.id}"
                    PssubfInbox.objects.create(
                        email_id=email_id,
                        subject=subject,
                        sender=settings.OUTLOOK_EMAIL_ADDRESS,
                        snippet=f"Direct Email to {recipient}",
                        status='Sent',
                        received_timestamp=timezone.now(),
                        member_group_code=membership_number
                    )

                    PssubfAction.objects.create(
                        task_email_id=email_id,
                        action_type="Direct Email Sent",
                        action_user=request.user.username,
                        note_content=f"Sent to: {recipient} | Subject: {subject}",
                        action_timestamp=timezone.now()
                    )
                    messages.success(request, f"Direct email sent successfully to {recipient}.")
                else:
                    messages.error(request, f"Email failed to send: {result.get('error')}")
            return redirect('beneficiary_details', membership_number=membership_number)

        # --- 2. HANDLE GENERAL NOTES (NOTES TAB) ---
        elif action == 'save_note' or 'save_note' in request.POST:
            note_text = request.POST.get('note_text')
            if note_text:
                current_status = "Expired" if member.is_expired else "Active"
                PssubfNote.objects.create(
                    task_email_id=f"PROFILE_{membership_number}",
                    agent_name=request.user.username,
                    note_text=note_text,
                    classification_at_time="Profile Detail Note",
                    status_at_time=current_status
                )
                PssubfProfileNote.objects.create(
                    membership_number=membership_number,
                    agent_name=request.user.username,
                    note_content=note_text
                )
                PssubfAction.objects.create(
                    task_email_id=f"NOTE_{membership_number}",
                    action_type="Internal Note",
                    action_user=request.user.username,
                    note_content=f"Added profile note: {note_text[:50]}...",
                    action_timestamp=timezone.now()
                )
                messages.success(request, "Internal note added to profile.")
            return redirect('beneficiary_details', membership_number=membership_number)

        # --- 3. HANDLE CORE PROFILE UPDATES ---
        elif action == 'update_profile':
            try:
                member.old_membership_number = request.POST.get('old_membership_number', member.old_membership_number)
                member.title = request.POST.get('title', member.title)
                member.initials = request.POST.get('initials', member.initials)
                member.first_name = request.POST.get('first_name', member.first_name)
                member.second_name = request.POST.get('second_name', member.second_name)
                member.last_name = request.POST.get('last_name', member.last_name)
                member.id_number = request.POST.get('id_number', member.id_number)
                
                dob_str = request.POST.get('dob')
                if dob_str:
                    dob_date = datetime.strptime(dob_str, '%Y-%m-%d').date()
                    member.dob = dob_date
                    member.cessation_date = dob_date + relativedelta(years=18)
                
                member.employee_number = request.POST.get('employee_number', member.employee_number)
                member.stipened_frequency = request.POST.get('stipened_frequency', member.stipened_frequency)
                
                if 'stipened' in request.POST:
                    member.stipened = clean_decimal(request.POST.get('stipened'))
                if 'total_fund_value' in request.POST:
                    member.total_fund_value = clean_decimal(request.POST.get('total_fund_value'))
                
                port_date_str = request.POST.get('portfolio_date')
                if port_date_str:
                    member.portfolio_date = datetime.strptime(port_date_str, '%Y-%m-%d').date()

                join_date_str = request.POST.get('fund_join_date')
                if join_date_str:
                    member.fund_join_date = datetime.strptime(join_date_str, '%Y-%m-%d').date()

                member.mobile_1 = request.POST.get('mobile_1', member.mobile_1)
                member.email_1 = request.POST.get('email_1', member.email_1)
                member.mobile_2 = request.POST.get('mobile_2', member.mobile_2)
                member.email_2 = request.POST.get('email_2', member.email_2)
                member.mobile_3 = request.POST.get('mobile_3', member.mobile_3)
                member.email_3 = request.POST.get('email_3', member.email_3)

                member.guardian_title = request.POST.get('guardian_title', member.guardian_title)
                member.guardian_first_name = request.POST.get('guardian_first_name', member.guardian_first_name)
                member.guardian_last_name = request.POST.get('guardian_last_name', member.guardian_last_name)
                member.guardian_mobile = request.POST.get('guardian_mobile', member.guardian_mobile)
                member.guardian_email = request.POST.get('guardian_email', member.guardian_email)
                member.guardian_address = request.POST.get('guardian_address', member.guardian_address)
                
                member.save()

                PssubfAction.objects.create(
                    task_email_id=f"PROFILE_MOD_{membership_number}",
                    action_type="Profile Update",
                    action_user=request.user.username,
                    note_content="Modified beneficiary personal/financial details and fund values.",
                    action_timestamp=timezone.now()
                )
                messages.success(request, f"Changes saved successfully for Member {member.membership_number}.")
                return redirect('beneficiary_details', membership_number=member.membership_number)
            except Exception as e:
                messages.error(request, f"Error updating record: {str(e)}")

        # --- 4. HANDLE NEW CLAIM ---
        elif action == 'add_claim_entry':
            try:
                date_logged_str = request.POST.get('date_logged')
                claim_date = datetime.strptime(date_logged_str, '%Y-%m-%d').date() if date_logged_str else date.today()
                
                calculated_age_str = "---"
                if member.dob:
                    diff_age = relativedelta(claim_date, member.dob)
                    total_m = round((diff_age.years * 12) + diff_age.months + (diff_age.days / 30.44))
                    calculated_age_str = f"{total_m // 12}Y {str(total_m % 12).zfill(2)}M"

                file_payload = request.FILES.get('supporting_document')
                if not file_payload and request.FILES:
                    file_payload = list(request.FILES.values())[0]

                file_status_label = "No"
                file_saved_path = None
                
                if file_payload:
                    fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT))
                    saved_filename = fs.save(file_payload.name, file_payload)
                    file_saved_path = saved_filename
                    file_status_label = "Yes"

                timestamp_string = datetime.now().strftime('%Y%m%d')
                generated_reference = f"CLM-{timestamp_string}{random.randint(1000, 9999)}"

                ClaimList.objects.create(
                    beneficiary=member,
                    reference_no=generated_reference,
                    claim_type=request.POST.get('claim_type'),
                    date_logged=claim_date,
                    amount_requested=clean_decimal(request.POST.get('amount_requested')),
                    status='Pending',
                    description=request.POST.get('description'),
                    guardian_name=f"{member.guardian_first_name} {member.guardian_last_name}".strip(),
                    beneficiary_name=f"{member.first_name} {member.last_name}".strip(),
                    beneficiary_dob=member.dob,
                    termination_date=member.cessation_date,
                    portfolio_value=clean_decimal(request.POST.get('portfolio_value', member.total_fund_value)),
                    portfolio_date=request.POST.get('portfolio_date') or member.portfolio_date,
                    age_at_claim=calculated_age_str,
                    supporting_docs_attached=file_status_label,
                    loaded_by_agent=request.user.username,
                    attachment_path=file_saved_path
                )
                messages.success(request, f"New claim registered successfully. Reference Code: {generated_reference}")
            except Exception as e:
                messages.error(request, f"Claim Error: {str(e)}")
            return redirect('beneficiary_details', membership_number=membership_number)

        # --- 5. HANDLE NEW AD HOC ---
        elif action == 'add_adhoc_entry':
            try:
                claim_date_str = request.POST.get('claim_form_date')
                claim_date = datetime.strptime(claim_date_str, '%Y-%m-%d').date() if claim_date_str else date.today()
                
                years_val = 0.0
                if member.cessation_date and claim_date < member.cessation_date:
                    diff = relativedelta(member.cessation_date, claim_date)
                    total_m = round((diff.years * 12) + diff.months + (diff.days / 30.44))
                    years_val = total_m / 12

                # 🟢 ROBUST FILE STREAM FALLBACK CHECKER
                file_payload = request.FILES.get('supporting_document')
                if not file_payload and request.FILES:
                    file_payload = list(request.FILES.values())[0]

                file_status_label = "No"
                file_saved_path = None
                
                if file_payload:
                    fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT))
                    saved_filename = fs.save(file_payload.name, file_payload)
                    file_saved_path = saved_filename
                    file_status_label = "Yes"

                requested_amt = clean_decimal(request.POST.get('amount_requested'))
                portfolio_val = clean_decimal(request.POST.get('portfolio_value', member.total_fund_value))
                stipend_val = float(member.stipened or 0)
                
                total_months = round(years_val * 12)
                calc_liability = total_months * stipend_val
                surplus_calc = portfolio_val - (calc_liability + requested_amt)
                
                # 🟢 TRUNCATION SECURITY GUARANTEE: Trim value strings to fit the varchar(20) target safely
                afford_string = f"R {surplus_calc:.2f}".strip()[:20]

                AdHocList.objects.create(
                    beneficiary=member,
                    title=request.POST.get('title'),
                    comments=request.POST.get('comments'),
                    claim_form_date=claim_date,
                    status='Pending',
                    supporting_docs_attached=file_status_label,
                    attachment_path=file_saved_path, 
                    years_to_maturity=years_val, 
                    affordability_calculation=afford_string,
                    portfolio_value=portfolio_val,
                    portfolio_date=member.portfolio_date or date.today(),
                    amount_requested=requested_amt
                )
                messages.success(request, "Ad Hoc entry saved successfully.")
            except Exception as e:
                print(traceback.format_exc())
                messages.error(request, f"Ad Hoc Error: {str(e)}")
            return redirect('beneficiary_details', membership_number=membership_number)

    # --- FETCH DATA FOR TABS ---
    claims = ClaimList.objects.filter(beneficiary__membership_number=membership_number).order_by('-date_logged')
    adhoc_records = AdHocList.objects.filter(beneficiary=member).order_by('-claim_form_date')

    incoming_emails = PssubfDelegate.objects.filter(member_group_code=membership_number)
    outgoing_emails = PssubfDirectEmail.objects.filter(membership_number=membership_number)

    combined_emails = []
    for e in incoming_emails:
        combined_emails.append({
            'email_id': e.email_id, 
            'agent': e.assigned_agent or "Unassigned",
            'subject': e.subject,
            'date': e.created_at,
            'status': e.status,
            'type': 'INCOMING'
        })

    for e in outgoing_emails:
        combined_emails.append({
            'email_id': f"DIRECT_{membership_number}_{e.id}", 
            'agent': e.agent_name,
            'subject': e.subject,
            'date': e.sent_at,
            'status': 'Sent',
            'type': 'OUTGOING'
        })

    combined_emails.sort(key=lambda x: x['date'] if x['date'] else timezone.now(), reverse=True)
    
    internal_notes = PssubfNote.objects.filter(task_email_id__icontains=membership_number).order_by('-created_at')
    pssubf_actions = PssubfAction.objects.filter(Q(task_email_id__icontains=membership_number)).order_by('-action_timestamp')

    # --- DYNAMICALLY CALCULATE DISPLAY STRINGS ---
    for c in claims:
        if member.dob and c.date_logged:
            diff = relativedelta(c.date_logged, member.dob)
            total_m = round((diff.years * 12) + diff.months + (diff.days / 30.44))
            c.age_display = f"{total_m // 12}Y {str(total_m % 12).zfill(2)}M"
        else:
            c.age_display = "---"

    for a in adhoc_records:
        if member.cessation_date and a.claim_form_date:
            diff = relativedelta(member.cessation_date, a.claim_form_date)
            total_m = round((diff.years * 12) + diff.months + (diff.days / 30.44))
            
            a.maturity_display = f"{total_m // 12}Y {str(total_m % 12).zfill(2)}M"
            
            stipend_value = float(member.stipened or 0)
            total_liability = total_m * stipend_value
            portfolio_val = float(a.amount_requested or member.total_fund_value or 0)
            
            surplus_val = portfolio_val - total_liability
            a.calculated_surplus = f"R {surplus_val:,.2f} ({'Surplus' if surplus_val >= 0 else 'Deficit'})"
        else:
            a.maturity_display = "---"
            a.calculated_surplus = "N/A"

    context = {
        'member': member,
        'claims': claims,
        'adhoc_records': adhoc_records,
        'email_logs': combined_emails,
        'internal_notes': internal_notes,
        'pssubf_actions': pssubf_actions,
        'title': f"Member Profile - {membership_number}"
    }
    return render(request, 'pssubf/beneficiary_details.html', context)

from datetime import date
from django.utils import timezone
from django.http import HttpResponse
import pandas as pd

@login_required
def export_beneficiaries_excel(request):
    """Exports the complete filtered beneficiary list with all columns to Excel."""
    status_filter = request.GET.get('status')
    queryset = PssubfBeneficiary.objects.all().order_by('last_name')

    # Apply same filter logic as the list view
    today = date.today()
    if status_filter == 'expired':
        queryset = queryset.filter(cessation_date__lte=today)
    elif status_filter == 'active':
        queryset = queryset.filter(cessation_date__gt=today)

    # Fetching EVERY column from the schema provided
    data = list(queryset.values(
        'membership_number', 'old_membership_number', 'title', 'initials', 
        'first_name', 'second_name', 'last_name', 'id_number', 'dob', 'gender', 
        'employee_number', 'fund_join_date', 'cessation_date', 
        'stipened_frequency', 'stipened', 
        'mobile_1', 'email_1', 'mobile_2', 'email_2', 'mobile_3', 'email_3',
        'guardian_title', 'guardian_initial', 'guardian_first_name', 
        'guardian_second_name', 'guardian_last_name', 'guardian_dob',
        'guardian_id_number', 'guardian_id_type', 'guardian_gender',
        'guardian_mobile', 'guardian_email', 'guardian_address',
        'created_at', 'updated_at'
    ))

    # Create DataFrame
    df = pd.DataFrame(data)

    # --- FIX FOR TIMEZONE ERROR ---
    # Convert timezone-aware datetimes to naive datetimes
    for col in df.columns:
        if pd.api.types.is_datetime64tz_dtype(df[col]):
            df[col] = df[col].dt.tz_localize(None)
    
    # Optional: Ensure Decimals are floats for Excel compatibility
    if 'stipened' in df.columns:
        df['stipened'] = df['stipened'].apply(lambda x: float(x) if x is not None else 0.0)

    # Mapping internal field names to readable Excel Headers
    column_mapping = {
        'membership_number': 'Membership Number',
        'old_membership_number': 'Old Membership Number',
        'title': 'Title',
        'initials': 'Initials',
        'first_name': 'First Name',
        'second_name': 'Second Name',
        'last_name': 'Last Name',
        'id_number': 'ID Number',
        'dob': 'Date of Birth',
        'gender': 'Gender',
        'employee_number': 'Employee Number',
        'fund_join_date': 'Fund Join Date',
        'cessation_date': 'Cessation Date',
        'stipened_frequency': 'Stipend Frequency',
        'stipened': 'Stipend Amount',
        'mobile_1': 'Mobile 1',
        'email_1': 'Email 1',
        'mobile_2': 'Mobile 2',
        'email_2': 'Email 2',
        'mobile_3': 'Mobile 3',
        'email_3': 'Email 3',
        'guardian_title': 'Guardian Title',
        'guardian_initial': 'Guardian Initials',
        'guardian_first_name': 'Guardian First Name',
        'guardian_second_name': 'Guardian Second Name',
        'guardian_last_name': 'Guardian Last Name',
        'guardian_dob': 'Guardian DOB',
        'guardian_id_number': 'Guardian ID Number',
        'guardian_id_type': 'Guardian ID Type',
        'guardian_gender': 'Guardian Gender',
        'guardian_mobile': 'Guardian Mobile',
        'guardian_email': 'Guardian Email',
        'guardian_address': 'Guardian Address',
        'created_at': 'Created At',
        'updated_at': 'Updated At'
    }
    
    df.rename(columns=column_mapping, inplace=True)

    # Set up the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=PSSUBF_Master_Export_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'

    # Write the dataframe to the response
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Beneficiaries')
        
        # Access the openpyxl worksheet object to adjust column widths
        worksheet = writer.sheets['Beneficiaries']
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if cell.value:
                        val_len = len(str(cell.value))
                        if val_len > max_length:
                            max_length = val_len
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = min(adjusted_width, 50) # Cap width at 50 for readability

    return response

@login_required
def get_beneficiary_data(request, membership_number):
    """API endpoint to auto-populate the New Claim popup"""
    member = get_object_or_404(PssubfBeneficiary, membership_number=membership_number)
    
    # Calculate age for the age_at_claim field (initial calculation)
    today = timezone.now().date()
    age = today.year - member.dob.year - ((today.month, today.day) < (member.dob.month, member.dob.day))
    
    data = {
        'guardian_name': f"{member.guardian_first_name or ''} {member.guardian_last_name or ''}".strip(),
        'beneficiary_name': f"{member.first_name} {member.last_name}",
        'dob': member.dob.strftime('%Y-%m-%d'),
        'termination_date': member.cessation_date.strftime('%Y-%m-%d') if member.cessation_date else '',
        'stipened': float(member.stipened or 0),
        'age': age
    }
    return JsonResponse(data)

@login_required
def claim_list_view(request):
    """Main view for the full Claim Registry with Dynamic Age Calculation"""
    
    def clean_numeric(val):
        if not val or str(val).lower() == 'undefined' or str(val).strip() == '':
            return 0
        return str(val).replace('R', '').replace(',', '').replace('%', '').strip()

    if request.method == 'POST':
        action = request.POST.get('action')
        m_num = request.POST.get('membership_number')
        
        try:
            member = PssubfBeneficiary.objects.filter(membership_number=m_num).first()
            uploaded_file = request.FILES.get('supporting_document')
            file_name = uploaded_file.name if uploaded_file else None
            timestamp = timezone.now().strftime('%Y-%m-%d %H:%M')
            agent_stamp = f"\n\n--- Managed by {request.user.username} on {timestamp} ---"

            if action == 'add_claim_entry':
                new_claim = ClaimList(
                    beneficiary=member,
                    reference_no=f"CLM-{timezone.now().strftime('%Y%m%d%H%M')}",
                    guardian_name=request.POST.get('guardian_name'),
                    beneficiary_name=request.POST.get('beneficiary_name'),
                    beneficiary_dob=request.POST.get('beneficiary_dob') or None,
                    termination_date=request.POST.get('termination_date') or None,
                    claim_type=request.POST.get('claim_type'),
                    description=(request.POST.get('description') or "") + agent_stamp,
                    date_logged=request.POST.get('date_logged') or None,
                    status=request.POST.get('status', 'Created'),
                    portfolio_value=clean_numeric(request.POST.get('portfolio_value')),
                    portfolio_date=request.POST.get('portfolio_date') or None,
                    amount_requested=clean_numeric(request.POST.get('amount_requested')),
                    supporting_docs_attached=request.POST.get('supporting_docs_attached'),
                    monthly_income_payment=clean_numeric(request.POST.get('monthly_income')),
                    date_paid=request.POST.get('date_paid') or None,
                    loaded_by_agent=request.user.username,
                    attachment_path=file_name
                )
                new_claim.save()
                messages.success(request, f"Claim for Member {m_num} logged.")

            elif action == 'update_claim_entry':
                claim_id = request.POST.get('claim_id')
                claim = get_object_or_404(ClaimList, id=claim_id)
                
                if file_name:
                    claim.attachment_path = file_name
                elif request.POST.get('remove_attachment') == 'true':
                    claim.attachment_path = None

                claim.guardian_name = request.POST.get('guardian_name')
                claim.beneficiary_name = request.POST.get('beneficiary_name')
                claim.beneficiary_dob = request.POST.get('beneficiary_dob') or None
                claim.termination_date = request.POST.get('termination_date') or None
                claim.date_paid = request.POST.get('date_paid') or None
                claim.claim_type = request.POST.get('claim_type')
                claim.description = (request.POST.get('description') or "") + agent_stamp
                claim.date_logged = request.POST.get('date_logged') or None
                claim.status = request.POST.get('status')
                claim.portfolio_value = clean_numeric(request.POST.get('portfolio_value'))
                claim.portfolio_date = request.POST.get('portfolio_date') or None
                claim.amount_requested = clean_numeric(request.POST.get('amount_requested'))
                claim.supporting_docs_attached = request.POST.get('supporting_docs_attached')
                claim.monthly_income_payment = clean_numeric(request.POST.get('monthly_income'))
                claim.save()
                messages.success(request, f"Claim {claim_id} updated.")

            PssubfAction.objects.create(
                task_email_id=f"CLAIM_{m_num}",
                action_type="Claim Record Managed",
                action_user=request.user.username,
                note_content=f"Action: {action} | Status: {request.POST.get('status')}",
                action_timestamp=timezone.now()
            )
            return redirect('claim_list')
            
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

    # Fetching logic with select_related optimization
    membership_number = request.GET.get('membership_number')
    claims = ClaimList.objects.all().select_related('beneficiary').order_by('-date_logged')
    
    if membership_number:
        claims = claims.filter(beneficiary__membership_number=membership_number)

    # 🟢 DYNAMIC DISPLAY CALCULATION: Age at Claim
    for c in claims:
        if c.beneficiary and c.beneficiary.dob and c.date_logged:
            diff = relativedelta(c.date_logged, c.beneficiary.dob)
            total_m = round((diff.years * 12) + diff.months + (diff.days / 30.44))
            c.age_display = f"{total_m // 12}Y {str(total_m % 12).zfill(2)}M"
        else:
            c.age_display = "---"

    return render(request, 'claim_list.html', {'claims': claims, 'title': 'Claims Registry'})


@login_required
def ad_hoc_list_view(request):
    """Main view for the Ad Hoc Registry with Dynamic Maturity Calculation"""
    
    def clean_numeric(val):
        if not val or str(val).lower() == 'undefined' or str(val).strip() == '':
            return 0
        return str(val).replace('R', '').replace(',', '').replace('%', '').strip()

    if request.method == 'POST':
        action = request.POST.get('action')
        m_num = request.POST.get('membership_number')
        
        try:
            member = get_object_or_404(PssubfBeneficiary, membership_number=m_num)
            uploaded_file = request.FILES.get('supporting_document')
            file_name = uploaded_file.name if uploaded_file else None
            timestamp = timezone.now().strftime('%Y-%m-%d %H:%M')
            agent_stamp = f"\n\n--- Managed by {request.user.username} on {timestamp} ---"

            if action == 'add_adhoc_entry':
                AdHocList.objects.create(
                    beneficiary=member,
                    title=request.POST.get('title'),
                    comments=(request.POST.get('comments') or "") + agent_stamp,
                    claim_form_date=request.POST.get('claim_form_date') or None,
                    date_paid=request.POST.get('date_paid') or None,
                    status=request.POST.get('status', 'Created'),
                    supporting_docs_attached=request.POST.get('supporting_docs_attached', 'No'),
                    attachment_path=file_name,
                    portfolio_value=clean_numeric(request.POST.get('portfolio_value')),
                    portfolio_date=request.POST.get('portfolio_date') or None,
                    amount_requested=clean_numeric(request.POST.get('amount_requested')),
                )
                messages.success(request, f"New Ad Hoc claim for Member {m_num} successfully logged.")

            elif action == 'update_adhoc_entry':
                record_id = request.POST.get('record_id')
                record = get_object_or_404(AdHocList, id=record_id)
                
                if file_name:
                    record.attachment_path = file_name
                elif request.POST.get('remove_attachment') == 'true':
                    record.attachment_path = None

                record.title = request.POST.get('title')
                record.status = request.POST.get('status')
                record.claim_form_date = request.POST.get('claim_form_date') or None
                record.date_paid = request.POST.get('date_paid') or None
                record.supporting_docs_attached = request.POST.get('supporting_docs_attached')
                record.portfolio_value = clean_numeric(request.POST.get('portfolio_value'))
                record.portfolio_date = request.POST.get('portfolio_date') or None
                record.amount_requested = clean_numeric(request.POST.get('amount_requested'))

                user_comments = request.POST.get('comments') or ""
                if agent_stamp not in (record.comments or ""):
                    record.comments = user_comments + agent_stamp
                else:
                    record.comments = user_comments

                record.save()
                messages.success(request, f"Ad Hoc Record {record_id} has been updated.")

            return redirect('adhoc_list')
            
        except Exception as e:
            messages.error(request, f"Process Error: {str(e)}")

    membership_number = request.GET.get('membership_number')
    adhoc_records = AdHocList.objects.all().select_related('beneficiary').order_by('-date_created')

    if membership_number:
        adhoc_records = adhoc_records.filter(beneficiary__membership_number=membership_number)

    # 🟢 DYNAMIC DISPLAY CALCULATION: Years to Maturity
    for a in adhoc_records:
        if a.beneficiary and a.beneficiary.cessation_date and a.claim_form_date:
            diff = relativedelta(a.beneficiary.cessation_date, a.claim_form_date)
            total_m = round((diff.years * 12) + diff.months + (diff.days / 30.44))
            a.maturity_display = f"{total_m // 12}Y {str(total_m % 12).zfill(2)}M"
        else:
            a.maturity_display = "---"

    context = {
        'adhoc_list': adhoc_records,
        'title': 'Ad Hoc Registry'
    }
    return render(request, 'Ad_hoc_list.html', context)

@login_required
def get_claim_details(request, claim_id):
    """Fetches full claim data and cleans the description for editing"""
    claim = get_object_or_404(ClaimList, id=claim_id)
    member = claim.beneficiary
    
    # Clean the description so the agent doesn't edit the old tracking stamps
    clean_description = claim.description or ""
    if "--- Managed by" in clean_description:
        clean_description = clean_description.split("---")[0].strip()

    data = {
        'membership_number': member.membership_number,
        'guardian_name': f"{member.guardian_first_name or ''} {member.guardian_last_name or ''}".strip(),
        'beneficiary_name': f"{member.first_name} {member.last_name}",
        'dob': member.dob.strftime('%Y-%m-%d') if member.dob else '',
        'term_date': member.cessation_date.strftime('%Y-%m-%d') if member.cessation_date else '',
        
        'claim_type': claim.claim_type,
        'age_at_claim': claim.age_at_claim,
        'monthly_income': float(claim.monthly_income_payment or 0),
        'date_logged': claim.date_logged.strftime('%Y-%m-%d') if claim.date_logged else '',
        'portfolio_value': float(claim.portfolio_value or 0),
        'portfolio_date': claim.portfolio_date.strftime('%Y-%m-%d') if claim.portfolio_date else '',
        'amount_requested': float(claim.amount_requested or 0),
        'status': claim.status,
        'date_paid': claim.date_paid.strftime('%Y-%m-%d') if hasattr(claim, 'date_paid') and claim.date_paid else '',
        'description': clean_description,
        'docs_attached': claim.supporting_docs_attached
    }
    return JsonResponse(data)

@login_required
def ad_hoc_list_view(request):
    """Main view for the Ad Hoc Registry with Dynamic Maturity Calculation"""
    
    def clean_numeric(val):
        if not val or str(val).lower() == 'undefined' or str(val).strip() == '':
            return 0
        return str(val).replace('R', '').replace(',', '').replace('%', '').strip()

    if request.method == 'POST':
        action = request.POST.get('action')
        m_num = request.POST.get('membership_number')
        
        try:
            member = get_object_or_404(PssubfBeneficiary, membership_number=m_num)
            uploaded_file = request.FILES.get('supporting_document')
            file_name = uploaded_file.name if uploaded_file else None
            timestamp = timezone.now().strftime('%Y-%m-%d %H:%M')
            agent_stamp = f"\n\n--- Managed by {request.user.username} on {timestamp} ---"

            if action == 'add_adhoc_entry':
                AdHocList.objects.create(
                    beneficiary=member,
                    title=request.POST.get('title'),
                    comments=(request.POST.get('comments') or "") + agent_stamp,
                    claim_form_date=request.POST.get('claim_form_date') or None,
                    date_paid=request.POST.get('date_paid') or None,
                    status=request.POST.get('status', 'Created'),
                    supporting_docs_attached=request.POST.get('supporting_docs_attached', 'No'),
                    attachment_path=file_name,
                    portfolio_value=clean_numeric(request.POST.get('portfolio_value')),
                    portfolio_date=request.POST.get('portfolio_date') or None,
                    amount_requested=clean_numeric(request.POST.get('amount_requested')),
                )
                messages.success(request, f"New Ad Hoc claim for Member {m_num} successfully logged.")

            elif action == 'update_adhoc_entry':
                record_id = request.POST.get('record_id')
                record = get_object_or_404(AdHocList, id=record_id)
                
                if file_name:
                    record.attachment_path = file_name
                elif request.POST.get('remove_attachment') == 'true':
                    record.attachment_path = None

                record.title = request.POST.get('title')
                record.status = request.POST.get('status')
                record.claim_form_date = request.POST.get('claim_form_date') or None
                record.date_paid = request.POST.get('date_paid') or None
                record.supporting_docs_attached = request.POST.get('supporting_docs_attached')
                record.portfolio_value = clean_numeric(request.POST.get('portfolio_value'))
                record.portfolio_date = request.POST.get('portfolio_date') or None
                record.amount_requested = clean_numeric(request.POST.get('amount_requested'))

                user_comments = request.POST.get('comments') or ""
                if agent_stamp not in (record.comments or ""):
                    record.comments = user_comments + agent_stamp
                else:
                    record.comments = user_comments

                record.save()
                messages.success(request, f"Ad Hoc Record {record_id} has been updated.")

            return redirect('adhoc_list')
            
        except Exception as e:
            messages.error(request, f"Process Error: {str(e)}")

    membership_number = request.GET.get('membership_number')
    adhoc_records = AdHocList.objects.all().select_related('beneficiary').order_by('-date_created')

    if membership_number:
        adhoc_records = adhoc_records.filter(beneficiary__membership_number=membership_number)

    # 🟢 DYNAMIC DISPLAY CALCULATION: Years to Maturity
    for a in adhoc_records:
        if a.beneficiary and a.beneficiary.cessation_date and a.claim_form_date:
            diff = relativedelta(a.beneficiary.cessation_date, a.claim_form_date)
            total_m = round((diff.years * 12) + diff.months + (diff.days / 30.44))
            a.maturity_display = f"{total_m // 12}Y {str(total_m % 12).zfill(2)}M"
        else:
            a.maturity_display = "---"

    context = {
        'adhoc_list': adhoc_records,
        'title': 'Ad Hoc Registry'
    }
    return render(request, 'Ad_hoc_list.html', context)

from openpyxl.styles import Font, PatternFill, Alignment
@login_required
def export_adhoc_excel(request):
    """
    Exports the Ad Hoc Registry to Excel with columns A through R.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ad Hoc Claims"

    headers = [
        'Member Number', 'Guardian Name & Surname', 'Beneficiary Name & Surname', 
        'Beneficiary Date Of Birth', 'Age at Claim Date', 'Termination Date', 
        'Years to Maturity', 'Original Claim Form Date', 'Portfolio Value', 
        'Portfolio Value Date', 'Monthly Income Payment', 'Amount Requested', 
        'Reason', 'Supporting Documents Attached', 'Affordability Calc', 
        'Note', 'Date Paid', 'Loaded by Agent'
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="8FCE7F", end_color="8FCE7F", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    records = AdHocList.objects.all().select_related('beneficiary').order_by('-date_created')

    for rec in records:
        age_display = ""
        if rec.beneficiary.dob and rec.claim_form_date:
            dob = rec.beneficiary.dob
            cfd = rec.claim_form_date
            age = cfd.year - dob.year - ((cfd.month, cfd.day) < (dob.month, dob.day))
            age_display = str(age)

        agent_info = "System"
        if rec.comments and "Managed by" in rec.comments:
            try:
                agent_info = rec.comments.split("Managed by")[-1].split("on")[0].strip()
            except:
                agent_info = "Unknown"

        clean_note = rec.comments.split("---")[0].strip() if rec.comments else ""

        ws.append([
            rec.beneficiary.membership_number,                       # A
            f"{rec.beneficiary.guardian_first_name} {rec.beneficiary.guardian_last_name}", # B
            f"{rec.beneficiary.first_name} {rec.beneficiary.last_name}", # C
            rec.beneficiary.dob.strftime('%Y-%m-%d') if rec.beneficiary.dob else "", # D
            age_display,                                              # E
            rec.beneficiary.cessation_date.strftime('%Y-%m-%d') if rec.beneficiary.cessation_date else "", # F
            rec.years_to_maturity,                                    # G
            rec.claim_form_date.strftime('%Y-%m-%d') if rec.claim_form_date else "", # H
            float(rec.portfolio_value or 0),                          # I
            rec.portfolio_date.strftime('%Y-%m-%d') if rec.portfolio_date else "", # J
            float(rec.beneficiary.stipened or 0),                     # K
            float(rec.amount_requested or 0),                          # L
            rec.title,                                                # M
            rec.supporting_docs_attached,                             # N
            rec.affordability_calculation,                            # O
            clean_note,                                               # P
            rec.date_paid.strftime('%Y-%m-%d') if rec.date_paid else "", # Q
            agent_info                                                # R
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=AdHoc_Export_{date.today()}.xlsx'
    wb.save(response)
    return response

@login_required
def export_claims_excel(request):
    """
    Exports the Claim Registry (ClaimList model) to Excel.
    Sequence: A-Member Number to M-Supporting Docs.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Claims Registry Export"

    # Define exact header mapping A-M + specific metadata
    headers = [
        'Member Number',                # A
        'Guardian Name & Surname',      # B
        'Beneficiary Name & Surname',   # C
        'Beneficiary Date Of Birth',    # D
        'Age at Claim Date',            # E
        'Termination Date',             # F
        'Original Claim Form Date',     # G
        'Portfolio Value',              # H
        'Portfolio Value Date',         # I
        'Monthly Income Payment',       # J
        'Amount Requested',             # K
        'Reason',                       # L
        'Date Paid',                    
        'Loaded by Agent',              
        'Supporting Documents Attached' # M
    ]
    ws.append(headers)

    # Styling for Green Registry Theme
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Get records with select_related to pull Beneficiary data efficiently
    records = ClaimList.objects.all().select_related('beneficiary').order_by('-date_logged')

    for rec in records:
        # 🛡️ SAFETY CHECK: Initialize variables with defaults in case beneficiary is missing
        m_num = "N/A"
        guardian = "N/A"
        beneficiary_full_name = "N/A"
        b_dob = ""
        c_date = ""
        age_display = ""

        # Only extract data if the beneficiary relationship exists
        if rec.beneficiary:
            m_num = rec.beneficiary.membership_number
            guardian = f"{rec.beneficiary.guardian_first_name} {rec.beneficiary.guardian_last_name}"
            beneficiary_full_name = f"{rec.beneficiary.first_name} {rec.beneficiary.last_name}"
            b_dob = rec.beneficiary.dob.strftime('%Y-%m-%d') if rec.beneficiary.dob else ""
            c_date = rec.beneficiary.cessation_date.strftime('%Y-%m-%d') if rec.beneficiary.cessation_date else ""
            
            # Age Calculation for Column E: Claim Date (date_logged) vs Beneficiary DOB
            if rec.beneficiary.dob and rec.date_logged:
                dob = rec.beneficiary.dob
                dl = rec.date_logged
                age = dl.year - dob.year - ((dl.month, dl.day) < (dob.month, dob.day))
                age_display = str(age)

        # Agent Extraction logic from Description stamp
        agent_info = "N/A"
        if rec.description and "Managed by" in rec.description:
            try:
                agent_info = rec.description.split("Managed by")[-1].split("on")[0].strip()
            except:
                pass

        ws.append([
            m_num,                                   # A
            guardian,                                # B
            beneficiary_full_name,                   # C
            b_dob,                                   # D
            age_display,                             # E
            c_date,                                  # F
            rec.date_logged.strftime('%Y-%m-%d') if rec.date_logged else "", # G
            rec.portfolio_value,                     # H
            rec.portfolio_date.strftime('%Y-%m-%d') if rec.portfolio_date else "", # I
            rec.monthly_income_payment,              # J
            rec.amount_requested,                    # K
            rec.claim_type,                          # L (Reason)
            rec.date_paid.strftime('%Y-%m-%d') if hasattr(rec, 'date_paid') and rec.date_paid else "", # Date Paid
            agent_info,                              # Loaded by Agent
            rec.supporting_docs_attached             # M
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Claims_List_{date.today()}.xlsx'
    wb.save(response)
    return response

from dateutil.relativedelta import relativedelta

@login_required
def affordability_dashboard(request):
    """View the calculator and history"""
    results = ClaimAffordability.objects.all().order_by('-calculated_at')
    return render(request, 'affordability_dashboard.html', {'results': results})

from django.db.models import Sum

@login_required
def run_manual_calc(request):
    if request.method == 'POST':
        m_num = request.POST.get('membership_number')
        member = get_object_or_404(PssubfBeneficiary, membership_number=m_num)
        
        # Use the stored fund value from the database
        fund_val = float(member.total_fund_value or 0)
        stipend = float(member.stipened or 0)
        
        if member.dob:
            today = date.today()
            majority_date = member.dob + relativedelta(years=18)
            
            # Calculate months remaining
            if majority_date <= today:
                total_months = 0
            else:
                diff = relativedelta(majority_date, today)
                total_months = (diff.years * 12) + diff.months

            # CORE MATH: Only Stipends
            total_stipend_commitment = stipend * total_months
            
            # Final Balance is Fund minus what we OWE in stipends
            final_bal = fund_val - total_stipend_commitment

            ClaimAffordability.objects.update_or_create(
                membership_number=m_num,
                defaults={
                    'majority_date': majority_date,
                    'years_to_majority': round(total_months / 12, 2),
                    'months_to_majority': total_months,
                    'total_stipend_commitment': total_stipend_commitment,
                    'fund_after_stipend': final_bal, 
                    'final_projected_balance': final_bal,
                    'requires_letter': final_bal < 0
                }
            )
            
    return redirect('affordability_dashboard')

@login_required
def get_beneficiary_data(request, membership_number):
    """
    API endpoint to auto-populate the New Claim popup.
    Matches your DB Schema: 'stipened' and 'total_fund_value'.
    """
    member = get_object_or_404(PssubfBeneficiary, membership_number=membership_number)
    
    # Calculate current age for the initial view
    today = timezone.now().date()
    age = 0
    if member.dob:
        age = today.year - member.dob.year - ((today.month, today.day) < (member.dob.month, member.dob.day))
    
    data = {
        'success': True,
        'membership_number': member.membership_number,
        'guardian_name': f"{member.guardian_first_name or ''} {member.guardian_last_name or ''}".strip(),
        'beneficiary_name': f"{member.first_name or ''} {member.last_name or ''}".strip(),
        # Date inputs in HTML require YYYY-MM-DD format
        'dob': member.dob.strftime('%Y-%m-%d') if member.dob else '',
        'term_date': member.cessation_date.strftime('%Y-%m-%d') if member.cessation_date else '',
        # Use exact DB column names from your table definition
        'stipened': float(member.stipened or 0),
        'total_fund_value': float(member.total_fund_value or 0),
        'age': age
    }
    return JsonResponse(data)

import io
from django.http import HttpResponse
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

@login_required
def download_email_eml(request, email_id):
    """
    Generates and downloads a .eml file for both Delegated and Direct emails.
    """
    msg = MIMEMultipart()
    
    if email_id.startswith("DIRECT_"):
        # Fetch from local Direct Email table
        local_id = email_id.replace("DIRECT_", "")
        email_obj = get_object_or_404(PssubfDirectEmail, id=local_id)
        
        msg['Subject'] = email_obj.subject
        msg['From'] = settings.OUTLOOK_EMAIL_ADDRESS
        msg['To'] = email_obj.recipient
        body_content = email_obj.body_html
    else:
        # Fetch from Delegated incoming table
        task = get_object_or_404(PssubfDelegate, email_id=email_id)
        msg['Subject'] = task.subject or "(No Subject)"
        msg['From'] = task.sender or "System"
        msg['To'] = settings.OUTLOOK_EMAIL_ADDRESS
        body_content = f"Source: PSSUBF Portal Delegation\nSender: {task.sender}\nDate: {task.created_at}"

    msg.attach(MIMEText(body_content, 'html'))

    buf = io.BytesIO()
    buf.write(msg.as_bytes())
    buf.seek(0)
    
    response = HttpResponse(buf.read(), content_type='message/rfc822')
    response['Content-Disposition'] = f'attachment; filename="Email_Record.eml"'
    return response

import openpyxl
from datetime import date
from django.db.models import Q, Count
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from openpyxl.styles import Font, PatternFill, Alignment
# 🚀 FIX: Imported PssubfDirectEmail instead of DirectEmailLog
from .models import ClaimList, PssubfAction, PssubfBeneficiary, PssubfDirectEmail

@login_required
def claim_sla_report_view(request):
    if request.user.username.lower() not in ['omega', 'manager']:
        return redirect('dashboard')

    start_str = request.GET.get('start_date')
    end_str = request.GET.get('end_date')
    
    def get_valid_date(date_val):
        if date_val and date_val != "None" and date_val != "":
            return parse_date(date_val)
        return None

    # 1. Setup Filters
    claims_q = Q()
    actions_q = Q()
    emails_q = Q()

    start_dt = get_valid_date(start_str)
    if start_dt:
        claims_q &= Q(date_logged__gte=start_dt)
        actions_q &= Q(action_timestamp__date__gte=start_dt)
        emails_q &= Q(sent_at__date__gte=start_dt)

    end_dt = get_valid_date(end_str)
    if end_dt:
        claims_q &= Q(date_logged__lte=end_dt)
        actions_q &= Q(action_timestamp__date__lte=end_dt)
        emails_q &= Q(sent_at__date__lte=end_dt)

    # 2. Handle Excel Export
    if request.GET.get('export') == 'excel':
        return export_claims_sla_excel(claims_q, actions_q, emails_q)

    # 3. Aggregate Data for Dashboard
    type_breakdown = ClaimList.objects.filter(claims_q).values('claim_type', 'status').annotate(total=Count('id')).order_by('claim_type')
    action_breakdown = PssubfAction.objects.filter(actions_q).values('action_type', 'action_user').annotate(total=Count('id')).order_by('action_type')
    
    # 🚀 FIX: Using agent_name instead of sent_by_user
    email_breakdown = PssubfDirectEmail.objects.filter(emails_q).values(
        'subject', 'agent_name'
    ).annotate(total=Count('id')).order_by('-total')

    report_data = {
        'TYPES': type_breakdown,
        'TYPES_TOTAL': sum(item['total'] for item in type_breakdown),
        
        'ACTIONS': action_breakdown,
        'ACTIONS_TOTAL': sum(item['total'] for item in action_breakdown),
        
        'SENT_EMAILS': email_breakdown,
        'SENT_TOTAL': sum(item['total'] for item in email_breakdown),
    }

    grand_total = report_data['TYPES_TOTAL'] + report_data['ACTIONS_TOTAL'] + report_data['SENT_TOTAL']
    recent_details = ClaimList.objects.filter(claims_q).order_by('-date_logged')[:50]

    context = {
        'report': report_data,
        'grand_total': grand_total,
        'start_date': start_str,
        'end_date': end_str,
        'details': recent_details,
        'title': 'SLA Performance Report'
    }
    return render(request, 'claim_sla_report.html', context)

def export_claims_sla_excel(claims_q, actions_q, emails_q):
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    
    def clean_date(dt):
        if dt:
            try:
                return dt.replace(tzinfo=None)
            except AttributeError:
                return dt 
        return "N/A"

    # --- SHEET 1: CLAIMS LOG ---
    ws1 = wb.active
    ws1.title = "Claims Audit Trail"
    headers = ['Reference', 'Member No', 'Beneficiary', 'Claim Type', 'Date Logged', 'Status', 'Agent', 'Portfolio Value']
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font, cell.fill, cell.alignment = header_font, header_fill, Alignment(horizontal="center")

    for c in ClaimList.objects.filter(claims_q).order_by('-date_logged'):
        ws1.append([c.reference_no, c.beneficiary_id, c.beneficiary_name, c.claim_type, c.date_logged, c.status, c.loaded_by_agent, float(c.portfolio_value or 0)])

    # --- SHEET 2: AGENT ACTIVITY ---
    ws2 = wb.create_sheet(title="Agent Performance")
    ws2.append(['Timestamp', 'Agent', 'Action', 'Target Record', 'Notes'])
    for cell in ws2[1]:
        cell.font, cell.fill, cell.alignment = header_font, header_fill, Alignment(horizontal="center")

    for a in PssubfAction.objects.filter(actions_q).order_by('-action_timestamp'):
        ws2.append([clean_date(a.action_timestamp), a.action_user, a.action_type, a.task_email_id, a.note_content])

    # --- SHEET 3: EMAIL LOGS ---
    ws3 = wb.create_sheet(title="Correspondence Log")
    ws3.append(['Sent At', 'User', 'Recipient', 'Subject', 'Member No'])
    for cell in ws3[1]:
        cell.font, cell.fill, cell.alignment = header_font, header_fill, Alignment(horizontal="center")

    # 🚀 FIX: Using PssubfDirectEmail and correct field names (agent_name, recipient)
    for e in PssubfDirectEmail.objects.filter(emails_q).order_by('-sent_at'):
        ws3.append([
            clean_date(e.sent_at), 
            e.agent_name, 
            e.recipient, 
            e.subject, 
            e.membership_number
        ])

    for sheet in [ws1, ws2, ws3]:
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 25

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="PSSUBF_SLA_{date.today()}.xlsx"'
    wb.save(response)
    return response