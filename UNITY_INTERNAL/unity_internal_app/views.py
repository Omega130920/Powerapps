from base64 import urlsafe_b64decode, urlsafe_b64encode
import base64
from collections import defaultdict
import csv
from functools import cache
import io
import json
import mimetypes
import os
import pickle
import time
from tkinter.font import Font
from django.forms import DecimalField, model_to_dict
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import connection, transaction
from django.db.models import Sum, F
from django.urls import reverse
import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
import openpyxl
import pandas as pd
import numpy as np
from reportlab.lib.styles import ParagraphStyle
from django.utils import timezone
from decimal import Decimal
from django.db import models
from django.utils import timezone
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.db.models import Q
from django.conf import settings
from django.core.mail import EmailMessage
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl.styles import Font, Alignment, PatternFill
from django.db.models import Q, Max
import datetime as dt_mod
from datetime import datetime, time

# Import the new Graph API service functions
from .services.outlook_graph_service import OutlookGraphService

# Import Delegation Service functions
from .services.delegation_service import (
    get_or_create_delegation_status, 
    delegate_email_task, 
    add_delegation_note, 
    get_delegated_emails_for_user,
    log_delegation_transaction
)

import logging

# Initialize the logger for this module
logger = logging.getLogger(__name__)

from dateutil import parser

from django.utils.safestring import mark_safe # for the email body & signature

from django.http import HttpResponse

# Import all models and forms
from .models import BillSettlement, CreditNote, DelegationNote, DelegationTransactionLog, EmailDelegation, ImportBank, JournalEntry, OutlookInbox, ReconnedBank, ScheduleSurplus, UnityBill, UnityClaimNote, UnityMgListing, ClientNotes, InternalFunds, UnityNotes, UnityClaim
from .forms import AddMemberForm, FiscalDateAssignmentForm, PreBillForm, UnityClaimForm

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph
from django.db.models import DateField, DateTimeField
from datetime import datetime, date

# --- Global Definitions ---
REVIEW_NOTES_OPTIONS = [
    "AWAIT 0101 REPORT",
    "AWAIT COVID APPROVAL",
    "BILLING FISCAL NOT AVAILABLE ON RFW",
    "DEBIT ADJUSTMENT",
    "GL0101 BALANCE NOT UTILISED",
    "HISTORIC CONTRIBUTIONS",
    "HISTORIC DEBIT - EMPLOYER TO PAY SHORTFALL",
    "NO SCHEDULE RECEIVED",
    "OVERS CREDIT LINE",
    "PARTIALLY RECONCILED",
    "RECONCILED",
    "REQUESTED SUPPORTING DOCUMENTS",
    "SALARY DOES NOT MATCH CONTRIBUTION RATE",
    "SCHEDULE DOES NOT MATCH PAYMENT",
]
ZERO_DECIMAL = Decimal('0.00')
# --------------------------

# --- Authentication Views ---
def login_view(request):
    """Handles user login."""
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('dashboard')
            else:
                messages.error(request, "Invalid username or password.")
    
    form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})

@login_required
def dashboard(request):
    """Displays the user dashboard with notification badges."""
    from .models import CreditNote, EmailDelegation, ReconnedBank
    
    username = request.user.username
    
    # Initialize counts
    pending_approval_count = 0
    new_emails_count = 0
    recycled_count = 0
    
    # 1. Manager Logic (Omega only)
    if username.lower() == 'omega' or request.user.is_superuser:
        # Financial Approvals (Managerial level)
        pending_approval_count = CreditNote.objects.filter(credit_link_status='Pending').count()
        
        # Inbox emails not yet assigned (Strictly NEW and work-related)
        new_emails_count = EmailDelegation.objects.filter(
            status='NEW', 
            assigned_user__isnull=True, 
            work_related=True
        ).count()
        
        # Recycle Bin: FIXED to match outlook_recycle_bin_view
        # Only count items explicitly marked with the 'DLT' status
        recycled_count = EmailDelegation.objects.filter(status='DLT').count()
    
    # 2. User Logic: My Assigned Tasks (Items explicitly marked as Delegated to current user)
    my_pending_delegations_count = EmailDelegation.objects.filter(
        assigned_user=request.user,
        status='DEL', # Standard status for delegated tasks in the agent's queue
        work_related=True
    ).count()

    # 3. Bank & Credit Note Logic
    # Count bank lines that require action (exclude Matched)
    bank_lines_count = ReconnedBank.objects.exclude(recon_status='Matched').count() 
    
    # Count pending credit notes for the badge indicator
    credit_notes_count = CreditNote.objects.filter(credit_link_status='Pending').count()
    
    context = {
        'username': username,
        'pending_approval_count': pending_approval_count,
        'new_emails_count': new_emails_count,
        'recycled_count': recycled_count,
        'my_pending_delegations_count': my_pending_delegations_count,
        'bank_lines_count': bank_lines_count,
        'credit_notes_count': credit_notes_count,
    }
    return render(request, 'dashboard.html', context)

def logout_view(request):
    """Logs the user out."""
    logout(request)
    messages.info(request, "You have successfully logged out.")
    return redirect('login')

def index(request):
    """Handles the root URL, redirecting to the login page."""
    return redirect('login')

# --- Unity Listing Views ---
@login_required
def unity_list(request):
    """
    Displays a list combining InternalFunds and UnityMgListing.
    Calculates the 'Current Status' and 'Current Fiscal' based on the latest UnityBill.
    """
    
    # 1. Fetch Base Records
    internal_funds_records = InternalFunds.objects.all()
    unity_listing_map = {
        record.a_company_code: record for record in UnityMgListing.objects.all()
    }
    
    # 2. Manual Calculation Setup
    bill_map = dict(UnityBill.objects.values_list('id', 'C_Company_Code'))
    
    # Aggregate Surplus & Allocation (Keep your existing logic here)
    surplus_map = defaultdict(Decimal)
    for s in ScheduleSurplus.objects.values('unity_bill_source_id', 'surplus_amount'):
        b_id = s['unity_bill_source_id']
        if b_id in bill_map:
            surplus_map[bill_map[b_id]] += (s['surplus_amount'] or Decimal('0.00'))

    allocation_map = defaultdict(Decimal)
    for a in JournalEntry.objects.values('target_bill_id', 'amount'):
        b_id = a['target_bill_id']
        if b_id in bill_map:
            allocation_map[bill_map[b_id]] += (a['amount'] or Decimal('0.00'))

    # --- NEW: Calculate "Current Billing Status" AND "Fiscal Date" Map ---
    billing_status_map = {}
    fiscal_date_map = {} # This will store the latest A_CCDatesMonth
    
    # Order by date so the latest bill for each company is processed last 
    # and overwrites previous entries in our maps.
    all_bills = UnityBill.objects.all().order_by('A_CCDatesMonth')
    
    for b in all_bills:
        code = b.C_Company_Code
        total_covered = JournalEntry.objects.filter(target_bill=b).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        
        # Priority Logic for Status
        if b.is_reconciled:
            status = "RECON COMPLETE"
        elif b.G_Schedule_Date and (b.H_Schedule_Amount or 0) > 0 and total_covered == 0:
            status = "SCHEDULED"
        elif (total_covered or 0) > 0:
            status = "OPEN"
        elif b.F_Pre_Bill_Date:
            status = "PRE-BILL"
        elif (b.H_Schedule_Amount or 0) == 0 and total_covered == 0:
            status = "AWAITING SCHEDULE"
        else:
            status = "OPEN"
            
        billing_status_map[code] = status
        
        # UPDATE: Capture the Fiscal Date from the Bill
        if b.A_CCDatesMonth:
            # Storing as string to match the Varchar type of the fallback column
            fiscal_date_map[code] = b.A_CCDatesMonth.strftime('%Y-%m-%d')

    # 3. Build Combined List
    combined_records = []
    
    # Phase 1: InternalFunds
    for fund_record in internal_funds_records:
        company_code = fund_record.A_Company_Code
        detail_record = unity_listing_map.pop(company_code, None)
        
        active_surplus_value = surplus_map.get(company_code, Decimal('0.00')) - allocation_map.get(company_code, Decimal('0.00'))
        
        # PRIORITY LOGIC FOR FISCAL DATE: Bill Date -> Listing Date -> "N/A"
        final_fiscal = fiscal_date_map.get(company_code)
        if not final_fiscal and detail_record:
            final_fiscal = detail_record.g_current_fiscal

        combined_records.append({
            'A_Company_Code': fund_record.A_Company_Code,
            'B_Company_Name': fund_record.B_Company_Name,
            'Source': fund_record.Source,
            'D_Company_Status': fund_record.D_Company_Status,
            'c_agent': detail_record.c_agent if detail_record else None,
            'e_payment_method': detail_record.e_payment_method if detail_record else None,
            'f_billing_method': detail_record.f_billing_method if detail_record else None,
            'g_current_fiscal': final_fiscal or "N/A",
            'h_current_status': billing_status_map.get(company_code, "N/A"),
            'j_arrears': detail_record.j_arrears if detail_record else None,
            'has_details': bool(detail_record),
            'active_surplus': active_surplus_value,
        })

    # Phase 2: Remaining UnityMgListing
    for company_code, detail_record in unity_listing_map.items():
        active_surplus_value = surplus_map.get(company_code, Decimal('0.00')) - allocation_map.get(company_code, Decimal('0.00'))
        
        # PRIORITY LOGIC FOR FISCAL DATE: Bill Date -> Listing Date
        final_fiscal = fiscal_date_map.get(company_code) or detail_record.g_current_fiscal

        combined_records.append({
            'A_Company_Code': detail_record.a_company_code,
            'B_Company_Name': detail_record.b_company_name,
            'Source': 'System Only (New)',
            'D_Company_Status': detail_record.d_company_status,
            'c_agent': detail_record.c_agent,
            'e_payment_method': detail_record.e_payment_method,
            'f_billing_method': detail_record.f_billing_method,
            'g_current_fiscal': final_fiscal,
            'h_current_status': billing_status_map.get(company_code, "NO BILLING"),
            'j_arrears': detail_record.j_arrears,
            'has_details': True,
            'active_surplus': active_surplus_value,
        })
        
    # 4. Context for Rendering
    context = {
        'unity_records': combined_records,
        'distinct_source': InternalFunds.objects.values_list('Source', flat=True).distinct().exclude(Source__isnull=True),
        'distinct_company_status': InternalFunds.objects.values_list('D_Company_Status', flat=True).distinct(),
        'distinct_agent': UnityMgListing.objects.values_list('c_agent', flat=True).distinct(),
        'distinct_payment': UnityMgListing.objects.values_list('e_payment_method', flat=True).distinct(),
        'distinct_billing': UnityMgListing.objects.values_list('f_billing_method', flat=True).distinct(),
        'distinct_fiscal': UnityMgListing.objects.values_list('g_current_fiscal', flat=True).distinct(),
        'distinct_current_status': ["RECON COMPLETE", "SCHEDULED", "OPEN", "PRE-BILL", "AWAITING SCHEDULE", "NO BILLING"]
    }
    return render(request, 'unity_internal_app/unity_list.html', context)

from django.db.models import Sum, Max, Q # Ensure Max is imported

@login_required
def unity_information(request: HttpRequest, company_code):
    """
    Displays detailed information for a single record.
    UPDATED: Integrated Bankline Logic to subtract Credit Note allocations from balance.
    UPDATED: Applied CRM_UNITY mapping pattern for robust email history retrieval.
    UPDATED: Fixed status filtering to ensure 'DEL' (Delegated) items are visible.
    UPDATED: Integrated DelegationTransactionLog to show sent replies in the history.
    UPDATED: Status display logic now matches unity_list behavior.
    UPDATED: surplus_created now includes "Overs" moved to CreditNote table.
    UPDATED: Added bank_assigned_total to show the original transaction amount.
    UPDATED: h_current_status now dynamically calculated to match unity_list priority.
    UPDATED: Added File Upload handler for PDF attachments in UnityNotes.
    UPDATED: Dynamic Fiscal Date priority logic (Bill Date -> Listing Table).
    UPDATED: Added applied_credit_total calculation for recon history.
    """
    from .models import (
        EmailDelegation, DelegationTransactionLog, UnityNotes, 
        OutlookInbox, CreditNote, BillSettlement, ReconnedBank,
        UnityMgListing, InternalFunds, ClientNotes, UnityBill,
        ScheduleSurplus, JournalEntry, UnityClaim
    )
    from django.db.models import Q, Sum
    from decimal import Decimal
    from django.utils import timezone
    from django.urls import reverse
    from django.contrib import messages
    from django.shortcuts import render, redirect
    from django.conf import settings

    # =========================================================
    # 0. DOWNLOAD HANDLER (Matches CRM_UNITY Logic)
    # =========================================================
    if request.method == 'GET' and 'download_email_id' in request.GET:
        email_id = request.GET.get('download_email_id')
        if not email_id or email_id == "MANUAL_SEND_SUCCESS":
            messages.warning(request, "The physical .eml file for this older record is not available.")
            return redirect('unity_information', company_code=company_code)
        return redirect('download_email_file', email_id=email_id)

    # --- 1. Fetch Main Unity Record ---
    try:
        unity_record = UnityMgListing.objects.filter(a_company_code=company_code).first()
    except Exception:
        unity_record = None

    is_fallback = False
    lookup_code = company_code 
    
    if not unity_record:
        unity_record = InternalFunds.objects.filter(A_Company_Code=company_code).first()
        if not unity_record:
            messages.error(request, f"Error: Record {company_code} not found.")
            return redirect('unity_list')
        is_fallback = True
        messages.warning(request, f"Full detail information is not available for {company_code}.")

    # --- NEW: Dynamic Status & Fiscal Calculation ---
    latest_bill = UnityBill.objects.filter(C_Company_Code=lookup_code).order_by('-A_CCDatesMonth').first()
    calculated_status = "NO BILLING"
    calculated_fiscal = None
    
    if latest_bill:
        total_covered = JournalEntry.objects.filter(target_bill=latest_bill).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        
        if latest_bill.is_reconciled:
            calculated_status = "RECON COMPLETE"
        elif latest_bill.G_Schedule_Date and (latest_bill.H_Schedule_Amount or 0) > 0 and total_covered == 0:
            calculated_status = "SCHEDULED"
        elif (total_covered or 0) > 0:
            calculated_status = "OPEN"
        elif latest_bill.F_Pre_Bill_Date:
            calculated_status = "PRE-BILL"
        elif (latest_bill.H_Schedule_Amount or 0) == 0 and total_covered == 0:
            calculated_status = "AWAITING SCHEDULE"
        else:
            calculated_status = "OPEN"

        if latest_bill.A_CCDatesMonth:
            calculated_fiscal = latest_bill.A_CCDatesMonth.strftime('%Y-%m-%d')
    
    if unity_record:
        unity_record.h_current_status = calculated_status
        if calculated_fiscal:
            unity_record.g_current_fiscal = calculated_fiscal

    # --- 2. Fetch Related Data ---
    notes = ClientNotes.objects.filter(a_company_code=company_code).order_by('-date')

    company_bill_ids = UnityBill.objects.filter(C_Company_Code=lookup_code).values_list('id', flat=True)
    if company_bill_ids:
        total_created = ScheduleSurplus.objects.filter(
            unity_bill_source_id__in=company_bill_ids
        ).aggregate(total=Sum('surplus_amount'))['total'] or Decimal('0.00')

        surplus_ids = ScheduleSurplus.objects.filter(
            unity_bill_source_id__in=company_bill_ids
        ).values_list('id', flat=True)
        
        total_allocated = JournalEntry.objects.filter(
            surplus_source_id__in=surplus_ids
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        available_surplus_value = total_created - total_allocated
    else:
        available_surplus_value = Decimal('0.00')
    
    bank_lines_assigned = ReconnedBank.objects.filter(company_code=company_code).select_related('bank_line').order_by('-transaction_date')
    for line in bank_lines_assigned:
        line.actual_bill_usage = BillSettlement.objects.filter(reconned_bank_line_id=line.id).aggregate(total=Sum('settled_amount'))['total'] or Decimal('0.00')
        line.credit_amount = CreditNote.objects.filter(source_bank_line=line).aggregate(total=Sum('schedule_amount'))['total'] or Decimal('0.00')
        line.true_remaining_balance = line.transaction_amount - line.actual_bill_usage - line.credit_amount
        line.is_fully_consumed = (line.true_remaining_balance <= Decimal('0.009'))
        line.original_assigned_amount = line.transaction_amount 
    
    bank_lines = bank_lines_assigned
    credit_notes = CreditNote.objects.filter(member_group_code=company_code).order_by('-ccdates_month')

    try:
        company_claims = UnityClaim.objects.filter(company_code=company_code).prefetch_related('notes').order_by('-claim_created_date', 'member_surname')
        delegation_pks = [c.linked_email_id for c in company_claims if c.linked_email_id]
        if delegation_pks:
            delegations_map = EmailDelegation.objects.in_bulk(delegation_pks)
            outlook_string_ids = [d.email_id for d in delegations_map.values()]
            inbox_map = OutlookInbox.objects.in_bulk(outlook_string_ids)
            for claim in company_claims:
                if claim.linked_email_id:
                    del_obj = delegations_map.get(int(claim.linked_email_id))
                    if del_obj:
                        inbox_item = inbox_map.get(del_obj.email_id)
                        if inbox_item:
                            claim.email_preview_subject, claim.email_preview_sender, claim.email_preview_body, claim.email_preview_date = inbox_item.subject, inbox_item.sender_address, inbox_item.body_content, inbox_item.received_at
    except Exception:
        company_claims = []
    
    try:
        communication_logs = UnityNotes.objects.filter(member_group_code=lookup_code).filter(~Q(communication_type='Sent Email')).order_by('-date')
    except Exception:
        communication_logs = []
    
    # --- 4. Build Unified Email History ---
    combined_email_log = []
    clean_lookup = str(lookup_code).strip()
    all_delegations = EmailDelegation.objects.filter(company_code__iexact=clean_lookup)
    all_del_ids = [item.id for item in all_delegations]
    related_string_ids = [item.email_id for item in all_delegations]
    thread_status_map = {item.id: item.status for item in all_delegations}
    thread_email_id_map = {item.id: item.email_id for item in all_delegations}
    inbox_records = OutlookInbox.objects.filter(email_id__in=related_string_ids)
    inbox_map = {email.email_id: email.received_at for email in inbox_records}

    for item in all_delegations:
        if item.status in ['DLT', 'DELETED', 'TRASH']: continue
        is_completed = item.status in ['COMP', 'DONE', 'CLS']
        combined_email_log.append({'timestamp': item.received_at, 'arrival_timestamp': inbox_map.get(item.email_id, item.received_at), 'delegation_timestamp': item.delegated_at, 'type': 'Original', 'display_type': 'Completed' if is_completed else 'Delegated', 'subject': item.email_category or f"Outlook Task: {item.email_id[:12]}...", 'assigned_to': item.assigned_user.username if item.assigned_user else 'UNASSIGNED', 'status': item.status, 'email_id': item.email_id, 'action_user': 'System', 'badge_color': '#3f51b5', 'icon': '📥'})

    threaded_replies = DelegationTransactionLog.objects.filter(delegation_id__in=all_del_ids, action_type='REPLIED').select_related('user')
    for reply in threaded_replies:
        combined_email_log.append({'timestamp': reply.timestamp, 'arrival_timestamp': None, 'delegation_timestamp': None, 'type': 'Reply', 'display_type': 'Reply Sent', 'subject': reply.subject or "Reply to Task", 'assigned_to': reply.recipient_email, 'status': thread_status_map.get(reply.delegation_id, "SENT"), 'email_id': thread_email_id_map.get(reply.delegation_id), 'action_user': reply.user.username if reply.user else 'System', 'badge_color': '#673ab7', 'icon': '📤'})

    direct_emails = UnityNotes.objects.filter(member_group_code__iexact=clean_lookup, communication_type='Sent Email')
    for email in direct_emails:
        outlook_id = email.attached_email_id or (email.action_notes.replace("OUTLOOK_ID:", "").strip() if email.action_notes and "OUTLOOK_ID:" in email.action_notes else None)
        combined_email_log.append({'timestamp': email.date, 'arrival_timestamp': None, 'delegation_timestamp': None, 'type': 'Direct', 'display_type': 'Email sent', 'subject': email.action_notes or 'Email Sent', 'assigned_to': email.notes.split('\n')[0][:50] if email.notes else 'Recipient', 'status': 'Direct', 'email_id': outlook_id, 'action_user': email.user, 'badge_color': '#4CAF50', 'icon': '📤'})

    combined_email_log.sort(key=lambda x: x['timestamp'], reverse=True)

    # --- 5. Billing Logic ---
    billing_queryset = UnityBill.objects.filter(C_Company_Code=lookup_code).order_by('-A_CCDatesMonth')
    open_bills, settled_bills = [], []
    for bill in list(billing_queryset):
        bill.total_covered = BillSettlement.objects.filter(unity_bill_source_id=bill.id).aggregate(total=Sum('settled_amount'))['total'] or Decimal('0.00')
        if bill.is_reconciled:
            bill.display_status = 'RECON COMPLETE'
            bill.bankline_total = BillSettlement.objects.filter(unity_bill_source_id=bill.id, reconned_bank_line_id__isnull=False).aggregate(total=Sum('settled_amount'))['total'] or Decimal('0.00')
            
            # --- Credit Notes ---
            bill.credit_allocated = BillSettlement.objects.filter(unity_bill_source_id=bill.id, source_credit_note_id__isnull=False).aggregate(total=Sum('settled_amount'))['total'] or Decimal('0.00')
            
            # --- Journal Surplus Transfers ---
            bill.surplus_allocated_from_journals = JournalEntry.objects.filter(target_bill_id=bill.id).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            
            # 🚀 NEW: Combined Applied Credit (Manual Credits + Surplus Journals)
            bill.applied_credit_total = bill.credit_allocated + bill.surplus_allocated_from_journals
            
            used_line_ids = BillSettlement.objects.filter(
                unity_bill_source_id=bill.id, 
                reconned_bank_line_id__isnull=False
            ).values_list('reconned_bank_line_id', flat=True)
            
            bill.bank_assigned_total = ReconnedBank.objects.filter(
                id__in=used_line_ids
            ).aggregate(total=Sum('transaction_amount'))['total'] or Decimal('0.00')
            
            overs_created = CreditNote.objects.filter(
                source_bank_line_id__in=used_line_ids,
                note_selection='OVERS',
                ccdates_month=bill.A_CCDatesMonth
            ).aggregate(total=Sum('schedule_amount'))['total'] or Decimal('0.00')
            
            legacy_surplus = ScheduleSurplus.objects.filter(unity_bill_source_id=bill.id).aggregate(total=Sum('surplus_amount'))['total'] or Decimal('0.00')
            
            bill.surplus_created = overs_created + legacy_surplus
            settled_bills.append(bill)
        else:
            bill.display_status = 'OPEN' if bill.total_covered > Decimal('0.00') else 'SCHEDULED'
            open_bills.append(bill)

    my_delegated_emails = EmailDelegation.objects.filter(assigned_user=request.user).exclude(status__in=['COMP', 'CLS', 'DONE']).order_by('-received_at')

    # --- 6. HANDLE POST REQUESTS ---
    if request.method == 'POST':
        if request.POST.get('email_submission_action') == 'send_email_and_log' or request.POST.get('action') == 'send_outgoing_member_note':
            subject, recipient, email_body_html, action_note_val = request.POST.get('member_email_subject_reply', 'Claim Update'), request.POST.get('member_recipient_email'), request.POST.get('email_body_html_content'), request.POST.get('action_notes', 'Email Composed')
            if recipient and email_body_html:
                from .services import OutlookGraphService 
                # Ensure attachment=None is passed or handled if updated previously
                response = OutlookGraphService.send_outlook_email(settings.OUTLOOK_EMAIL_ADDRESS, recipient, subject, email_body_html, 'HTML')
                if response.get('success'):
                    UnityNotes.objects.create(member_group_code=company_code, user=request.user.username, date=timezone.now(), communication_type='Sent Email', action_notes=action_note_val[:90], attached_email_id=response.get('id', ''), notes=f"To: {recipient}\nSubject: {subject}\n{email_body_html}")
                    messages.success(request, f"Email sent to {recipient}!")
                else: messages.error(request, f"Graph API Error: {response.get('error')}")
            return redirect(f"{reverse('unity_information', kwargs={'company_code': company_code})}#email-log")

        if 'update_general_info' in request.POST and unity_record and not is_fallback:
            try:
                unity_record.recon_contact_1_name, unity_record.recon_contact_1_email, unity_record.recon_contact_2_name, unity_record.recon_contact_2_email = request.POST.get('recon_contact_1_name'), request.POST.get('recon_contact_1_email'), request.POST.get('recon_contact_2_name'), request.POST.get('recon_contact_2_email')
                unity_record.commencement_date = request.POST.get('commencement_date') or None
                unity_record.fund_status_date = request.POST.get('fund_status_date') or None
                unity_record.fund_status, unity_record.i_last_recon, unity_record.c_agent, unity_record.d_company_status, unity_record.e_payment_method, unity_record.f_billing_method, unity_record.g_current_fiscal, unity_record.h_current_status, unity_record.j_arrears = request.POST.get('fund_status'), request.POST.get('last_recon_note'), request.POST.get('agent'), request.POST.get('company_status'), request.POST.get('payment_method'), request.POST.get('billing_method'), request.POST.get('current_fiscal'), request.POST.get('current_status'), request.POST.get('arrears')
                unity_record.save()
                messages.success(request, "General Information updated.")
            except Exception as e: messages.error(request, f"Error saving: {e}")
            return redirect('unity_information', company_code=company_code)
        
        elif request.POST.get('note_content') or request.POST.get('action_notes'):
            pdf_file = request.FILES.get('note_pdf_attachment')
            UnityNotes.objects.create(
                member_group_code=company_code, 
                user=request.user.username, 
                date=timezone.now(), 
                communication_type=request.POST.get('communication_type') or 'Notes Log', 
                action_notes=request.POST.get('action_notes'), 
                notes=request.POST.get('note_content'),
                attached_file=pdf_file 
            )
            messages.success(request, "Note and attachment added.")
            return redirect(f"{reverse('unity_information', kwargs={'company_code': company_code})}#notes-log")

    context = {'unity_record': unity_record, 'notes': notes, 'communication_logs': communication_logs, 'combined_email_log': combined_email_log, 'is_fallback': is_fallback, 'bank_lines': bank_lines, 'credit_notes': credit_notes, 'open_bills': open_bills, 'settled_bills': settled_bills, 'company_claims': company_claims, 'available_surplus': available_surplus_value, 'my_delegated_emails': my_delegated_emails}
    return render(request, 'unity_internal_app/unity_information.html', context)

@login_required
def unity_billing_history(request, company_code):
    """
    Displays the billing history.
    UPDATED: Status logic now relies on G_Pre_Bill_Date and G_Schedule_Date.
    UPDATED: Bank Line balance now accounts for Credit Note allocations.
    """
    from django.db.models import Sum
    from decimal import Decimal
    from django.contrib import messages
    from django.shortcuts import redirect, render
    from .models import (
        UnityMgListing, InternalFunds, UnityBill, BillSettlement, 
        ReconnedBank, CreditNote
    )

    ZERO_DECIMAL = Decimal('0.00')
    
    unity_record = UnityMgListing.objects.filter(a_company_code=company_code).first()
    is_fallback = False

    if not unity_record:
        unity_record = InternalFunds.objects.filter(A_Company_Code=company_code).first()
        is_fallback = True
    
    if not unity_record:
        messages.error(request, f"Company {company_code} not found.")
        return redirect('unity_list')

    # --- 2. FETCH BILLS ---
    billing_queryset = UnityBill.objects.filter(C_Company_Code=company_code).order_by('-A_CCDatesMonth')
    billing_records = list(billing_queryset)

    # --- 3. MANUAL CALCULATION LOOP & LIST SPLIT ---
    open_bills = []
    settled_bills = []
    
    for bill in billing_records:
        # A. Calculate Settlements
        settled_sum_agg = BillSettlement.objects.filter(
            unity_bill_source_id=bill.id
        ).aggregate(total=Sum('settled_amount'))['total'] or ZERO_DECIMAL
        
        total_covered = settled_sum_agg
        scheduled_amount = bill.H_Schedule_Amount or ZERO_DECIMAL
        remaining_balance = scheduled_amount - total_covered
        
        # --- NEW & UPDATED STATUS LOGIC ---
        pre_bill_date = getattr(bill, 'F_Pre_Bill_Date', None)
        schedule_date = getattr(bill, 'G_Schedule_Date', None) 
        
        if bill.is_reconciled:
            display_status = 'RECON COMPLETE'
        elif schedule_date and scheduled_amount > ZERO_DECIMAL:
            if total_covered > ZERO_DECIMAL:
                display_status = 'OPEN' 
            else:
                display_status = 'SCHEDULED' 
        elif pre_bill_date and not schedule_date:
            display_status = 'PRE-BILL' 
        else:
            if scheduled_amount > ZERO_DECIMAL:
                display_status = 'OPEN'
            else:
                display_status = 'Pre-Bill' 
                
        bill.temp_remaining = remaining_balance
        bill.total_covered = total_covered
        bill.display_status = display_status

        if bill.display_status == 'RECON COMPLETE':
            settled_bills.append(bill)
        else:
            open_bills.append(bill)

    # --- 4. FETCH DATA FOR THE 'BANK LINES & CREDIT' TAB (ENHANCED LOGIC) ---
    # UPDATED: Now matches unity_information logic to subtract Credit Notes from balance
    bank_lines_assigned = ReconnedBank.objects.filter(
        company_code=company_code
    ).select_related('bank_line').order_by('-transaction_date')

    for line in bank_lines_assigned:
        # A. Sum of usage on Unity Bills
        actual_bill_usage = BillSettlement.objects.filter(
            reconned_bank_line_id=line.id
        ).aggregate(total=Sum('settled_amount'))['total'] or ZERO_DECIMAL

        # B. Sum of usage moved to Credit Notes (Overs)
        credit_amount = CreditNote.objects.filter(
            source_bank_line=line
        ).aggregate(total=Sum('schedule_amount'))['total'] or ZERO_DECIMAL

        # C. Calculate True Remaining Balance
        line.true_remaining_balance = line.transaction_amount - actual_bill_usage - credit_amount
        
        # Safety epsilon for floating point math
        line.is_fully_consumed = (line.true_remaining_balance <= Decimal('0.009'))
        line.original_assigned_amount = line.transaction_amount 

    bank_lines_data = bank_lines_assigned
    credit_notes_data = CreditNote.objects.filter(member_group_code=company_code).order_by('-ccdates_month')

    context = {
        'company_code': company_code,
        'unity_record': unity_record,
        'is_fallback': is_fallback,
        'bank_lines': bank_lines_data,
        'credit_notes': credit_notes_data,
        'open_bills': open_bills,
        'settled_bills': settled_bills,
        'billing_records': billing_records,
        'default_tab_override': '#recon',
    }
    
    return render(request, 'unity_internal_app/unity_information.html', context)

# --- PRE-BILL CREATION VIEW ---
@login_required
@transaction.atomic
def create_pre_bill(request, company_code):
    """
    Handles the creation of a new UnityBill record, setting the Pre-Bill Date.
    """
    # Requires: from django.db.models import Sum, F
    
    company_name = f"Company Code {company_code}"
    
    # 1. Fetch Company Info
    try:
        company_info = InternalFunds.objects.get(A_Company_Code=company_code)
        company_name = company_info.B_Company_Name
    except InternalFunds.DoesNotExist:
        messages.error(request, f"Cannot find company details for code {company_code}.")
        return redirect('unity_information', company_code=company_code)

    calculated_debt_for_prefill = ZERO_DECIMAL
    
    # 2. Handle Form Submission (POST)
    if request.method == 'POST':
        form = PreBillForm(request.POST)
        bill_date_str = request.POST.get('A_CCDatesMonth')
        bill_date = None
        
        if bill_date_str:
            try:
                bill_date = datetime.strptime(bill_date_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Invalid date format submitted.")

        if form.is_valid() and bill_date:
            
            # --- NEW LOGIC: Calculate Debt (All available) ---
            # NOTE: We fetch ALL available debt, regardless of fiscal date, for pre-fill purposes.
            # This calculation is for guidance only, as it's not restricted by fiscal month anymore.
            debt_queryset = ReconnedBank.objects.filter(
                company_code=company_code,
                # Only unsettled lines
                amount_settled__lt=F('transaction_amount'),
            ).annotate(
                remaining_debt=F('transaction_amount') - F('amount_settled')
            )
            calculated_debt_for_prefill = debt_queryset.aggregate(
                total_schedule_amount=Sum('remaining_debt')
            )['total_schedule_amount'] or ZERO_DECIMAL
            # --- END NEW LOGIC ---
            
            bill_record = form.save(commit=False)
            bill_record.C_Company_Code = company_code
            
            # --- NEW LOGIC: Set Pre-Bill Date upon creation ---
            # This triggers the 'PRE-BILL' status as G_Schedule_Date will be None
            bill_record.F_Pre_Bill_Date = timezone.now().date()
            
            # Check the actual schedule amount being saved by the user
            scheduled_amount = bill_record.H_Schedule_Amount or ZERO_DECIMAL
            
            # --- CRITICAL FIX: Prevent premature closure ---
            if scheduled_amount <= ZERO_DECIMAL:
                # If R0.00 is scheduled, force the bill to remain OPEN/PRE-BILL by clearing final dates.
                bill_record.J_Final_Date = None
                # I_Submitted_Date is cleared here, which is fine since G_Schedule_Date controls the status now.
                bill_record.I_Submitted_Date = None
                messages.warning(request, "Bill created with R0.00 scheduled amount. It will remain in Pre-Bill status until updated.")
            # --- END CRITICAL FIX ---
            
            try:
                bill_record.save()
                
                messages.success(request, f"New Pre-Bill record created for {company_code} (Date: {bill_record.A_CCDatesMonth}). Scheduled Amount: R{bill_record.H_Schedule_Amount}")
                
                # 🛑 CRITICAL FIX: Add cache-busting timestamp to the redirect URL
                timestamp = timezone.now().timestamp()
                return redirect(f"{reverse('unity_billing_history', kwargs={'company_code': company_code})}?cache={timestamp}")
                
            except Exception as e:
                messages.error(request, f"Error saving new bill: {e}")
                
        else:
            messages.error(request, "Please correct the errors in the form and ensure the Bill Date is valid.")
    
    # 3. Handle GET Request (Initial Form Display)
    else:
        # 4. Final Context Construction for GET
        initial_data = {
            'C_Company_Code': company_code,
            'D_Company_Name': company_name,
            # Pre-fill with the calculated debt
            'H_Schedule_Amount': calculated_debt_for_prefill
        }
        form = PreBillForm(initial=initial_data)

    context = {
        'form': form,
        'company_code': company_code,
        'company_name': company_name,
        'is_editing': False,
    }
    return render(request, 'unity_internal_app/bill_form.html', context)

@login_required
def add_member_view(request):
    """Handles adding a new UnityMgListing member."""
    # NOTE: AddMemberForm must be imported or mocked above for runtime testing
    if request.method == 'POST':
        form = AddMemberForm(request.POST)
        if form.is_valid():
            try:
                # This form saves to the UnityMgListing table (internal_mg_list)
                form.save()
                messages.success(request, f"New member '{form.cleaned_data['b_company_name']}' added successfully!")
                return redirect('unity_list')
            except Exception as e:
                messages.error(request, f"Error saving member: {e}")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = AddMemberForm()

    context = {
        'form': form,
    }
    return render(request, 'unity_internal_app/add_member.html', context)

# --- Bank Reconciliation Views ---
@login_required
def import_excel_view(request):
    """Handles the upload and import of Excel data."""
    from django.db import connection
    from .models import ImportBank
    import pandas as pd
    import numpy as np
    
    if request.method == 'POST':
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, "Invalid file format. Please upload an Excel (.xlsx or .xls) file.")
                return redirect('import_data')
            
            try:
                df = pd.read_excel(excel_file, keep_default_na=False, header=None)
                
                db_columns = [
                    'Bank_account_name', 'Account_number', 'Statement_reference',
                    'DATE', 'Balance', 'Transaction_amount', 'Transaction_description',
                    'INTERNAL_IDENTIFICATION', 'Specialist', 'Date_identified',
                    'Fiscal', 'Comments', 'Interim_fiscal'
                ]
                df.columns = db_columns
                
                df = df.astype(str).replace({'nan': '', 'NaT': ''})
                
                df['DATE'] = pd.to_datetime(df['DATE'], errors='coerce')
                df['Date_identified'] = pd.to_datetime(df['Date_identified'], errors='coerce').dt.date
                df['Balance'] = pd.to_numeric(df['Balance'], errors='coerce')
                df['Transaction_amount'] = pd.to_numeric(df['Transaction_amount'], errors='coerce')
                
                initial_count = len(df)
                df.dropna(subset=['DATE'], inplace=True)
                dropped_count = initial_count - len(df)
                
                if dropped_count > 0:
                    messages.warning(request, f"Skipped {dropped_count} row(s) due to missing or invalid required data (Date).")

                df['DATE'] = df['DATE'].dt.date
                df = df.replace(r'^\s*$', np.nan, regex=True)
                df = df.where(pd.notna(df), None)

                with transaction.atomic():
                    columns_sql = ', '.join([f'`{col}`' for col in db_columns])
                    placeholders = ', '.join(['%s'] * len(db_columns))
                    sql = f"INSERT INTO {ImportBank._meta.db_table} ({columns_sql}) VALUES ({placeholders})"
                    
                    data_to_insert = []
                    for row in df[db_columns].values:
                        cleaned_row = [None if isinstance(item, float) and np.isnan(item) else item for item in row]
                        data_to_insert.append(tuple(cleaned_row))

                    with connection.cursor() as cursor:
                        cursor.executemany(sql, data_to_insert)
                        
                messages.success(request, f"Successfully imported {len(df)} records into the 'importbank' table. Data was appended.")
                
            except Exception as e:
                messages.error(request, f"An error occurred during import: {e}")
            
            return redirect('import_data')

    return render(request, 'unity_internal_app/import_excel.html', {})

@login_required
def bank_list(request):
    """
    Displays a list of active bank transaction segments and Approved Credit Note "Overs".
    
    CRITICAL FIX: Injects 'Approved' credit notes into the bank list as virtual deposits.
    This ensures that credits verified by the manager are not forgotten.
    """
    from django.db.models import F, Q
    from django.urls import reverse

    # --- Step 1: Automatic Segment Creation for New Imports ---
    
    assigned_import_ids = ReconnedBank.objects.values_list('bank_line_id', flat=True).distinct()
    
    new_imports_to_segment = ImportBank.objects.exclude(
        id__in=assigned_import_ids
    ).filter(
        transaction_amount__gt=ZERO_DECIMAL 
    )
    
    with transaction.atomic():
        for import_line in new_imports_to_segment:
            ReconnedBank.objects.create(
                bank_line_id=import_line.id,
                company_code=None, 
                transaction_amount=import_line.transaction_amount,
                transaction_date=import_line.date,
                fiscal_date=None, 
                recon_status='Unidentified - New Import',
                amount_settled=ZERO_DECIMAL,
            )
            
    # --- Step 2: Query ALL Active Segments (Physical Cash) ---
    
    active_segments_queryset = ReconnedBank.objects.annotate(
        remaining_balance=F('transaction_amount') - F('amount_settled')
    ).filter(
        remaining_balance__gt=ZERO_DECIMAL 
    ).select_related('bank_line').order_by('-transaction_date', 'id')

    combined_records = []
    
    # Process physical bank lines
    for recon_data in active_segments_queryset:
        allocated_code = recon_data.company_code
        is_unassigned = allocated_code is None 
        
        company_name = "-"
        agent_name = "-"
        if allocated_code and allocated_code != "N/A":
            mg_entry = UnityMgListing.objects.filter(a_company_code=allocated_code).first()
            if mg_entry:
                company_name = mg_entry.b_company_name
                agent_name = mg_entry.c_agent
            else:
                fund_entry = InternalFunds.objects.filter(A_Company_Code=allocated_code).first()
                if fund_entry:
                    company_name = fund_entry.B_Company_Name
                    agent_name = "Internal"
        
        action_url = reverse('bankline_recon', args=[recon_data.id]) if is_unassigned else reverse('display_bankline_review', args=[recon_data.id])
        action_text = "Assign" if is_unassigned else "View"
        
        combined_records.append({
            'id': recon_data.id, 
            'source_type': 'BANK', # Flag for UI
            'deposit_amount': recon_data.transaction_amount, 
            'transaction_date': recon_data.transaction_date,
            'transaction_description': recon_data.bank_line.transaction_description,
            'remaining_balance': recon_data.remaining_balance,
            'recon_status': recon_data.recon_status or "Unidentified",
            'allocated_code': allocated_code,
            'company_name': company_name,
            'agent': agent_name,
            'review_note': recon_data.review_note, 
            'action_url': action_url,
            'action_text': action_text,
        })

    # --- Step 3: FAIL-SAFE - Injected Approved Credit Notes (Virtual Cash) ---
    # Fetch credits that were approved by the manager but not yet linked/used
    approved_overs = CreditNote.objects.filter(
        credit_link_status='Approved',
        note_selection='OVERS',
        schedule_amount__gt=ZERO_DECIMAL
    )

    for credit in approved_overs:
        # Action URL for credit sends a 'source=credit' parameter to the recon view
        credit_action_url = f"{reverse('bankline_recon', args=[credit.id])}?source=credit"
        
        combined_records.append({
            'id': credit.id,
            'source_type': 'CREDIT', # Flag for UI to highlight as a Credit Note
            'deposit_amount': credit.schedule_amount,
            'transaction_date': credit.authorized_at.date() if credit.authorized_at else credit.processed_date,
            'transaction_description': f"APPROVED OVERS: {credit.comment or 'Manager Verified'}",
            'remaining_balance': credit.schedule_amount,
            'recon_status': 'APPROVED OVERS LINE', # High-visibility label
            'allocated_code': credit.member_group_code,
            'company_name': credit.member_group_name or "Verified Credit",
            'agent': credit.authorized_by or "Manager",
            'review_note': f"Authorized at {credit.authorized_at}",
            'action_url': credit_action_url,
            'action_text': "Use Credit",
        })

    # Sort everything together by date (Cash and Virtual Credits mixed)
    combined_records.sort(key=lambda x: x['transaction_date'], reverse=True)
    
    context = {
        'bank_records': combined_records,
    }
    return render(request, 'unity_internal_app/bank_list.html', context)

@login_required
def bankline_recon(request, record_id):
    """
    Handles the reconciliation/assignment of bank lines OR Approved Credits.
    UPDATED: Returns to bank_list.html after successful assignment.
    UPDATED: Allows saving without a company code if 'is_bulk' is ticked.
    """
    from django.urls import reverse
    
    # Check if we are dealing with a Credit Note instead of a Bank Line
    is_credit = request.GET.get('source') == 'credit'
    
    if is_credit:
        credit_note = get_object_or_404(CreditNote, id=record_id)
        
        open_bill = UnityBill.objects.filter(
            C_Company_Code=credit_note.member_group_code,
            is_reconciled=False
        ).order_by('A_CCDatesMonth').first()
        
        if open_bill:
            messages.info(request, f"Using Approved Credit of R{credit_note.schedule_amount} for {credit_note.member_group_code}.")
            return redirect('pre_bill_reconciliation_summary', 
                            company_code=credit_note.member_group_code, 
                            bill_id=open_bill.id)
        else:
            messages.warning(request, f"Credit is ready, but no open bills were found for {credit_note.member_group_code}.")
            # Note: Decided to keep this redirect to unity_info as a fallback for credits
            return redirect('unity_information', company_code=credit_note.member_group_code)

    # --- STANDARD BANK LINE LOGIC ---
    try:
        recon_segment = get_object_or_404(
            ReconnedBank.objects.select_related('bank_line'), 
            id=record_id
        )
    except Exception:
        messages.error(request, f"Error retrieving Recon Segment ID: {record_id}")
        return redirect('bank_list')

    if recon_segment.company_code is not None:
        messages.warning(request, f"Bank line {record_id} is already assigned to {recon_segment.company_code}.")
        return redirect('display_bankline_review', recon_id=record_id)
        
    company_codes = InternalFunds.objects.values_list('A_Company_Code', flat=True).distinct().order_by('A_Company_Code')

    if request.method == 'POST':
        allocated_company_code_value = request.POST.get('company_code')
        # Capture Bulk Tick Box
        is_bulk_ticked = request.POST.get('is_bulk') == 'on'
        
        # --- FIX: Only require Company Code if NOT a bulk split ---
        if not is_bulk_ticked:
            if not allocated_company_code_value or allocated_company_code_value == 'None':
                messages.error(request, "You must select a Company Code for reconciliation OR mark as Bulk Split.")
                return redirect('bankline_recon', record_id=record_id)

        try:
            # Only validate the company code if one was provided
            if allocated_company_code_value and allocated_company_code_value != 'None':
                code_exists = InternalFunds.objects.filter(A_Company_Code=allocated_company_code_value).exists() or \
                             UnityMgListing.objects.filter(a_company_code=allocated_company_code_value).exists()

                if not code_exists:
                    messages.error(request, f"Company code '{allocated_company_code_value}' is not recognized.")
                    return redirect('bankline_recon', record_id=record_id)

            # --- ASSIGNMENT ---
            # Set company_code to the provided value, or None if it's a Bulk Split without a code
            recon_segment.company_code = allocated_company_code_value if allocated_company_code_value != 'None' else None
            recon_segment.agent = request.user.get_full_name() or request.user.username
            
            # Apply Bulk Logic if ticked
            if is_bulk_ticked:
                recon_segment.recon_status = 'Unreconciled - Bulk Split'
                recon_segment.review_note = 'BULK'
            else:
                recon_segment.recon_status = 'Unreconciled - Assigned'
                
            recon_segment.save()
            
            success_msg = f"Bank line segment {record_id} processed successfully."
            if is_bulk_ticked:
                success_msg += " (Marked for Bulk Processing)"
            elif allocated_company_code_value:
                success_msg += f" (Assigned to Code: {allocated_company_code_value})"
                
            messages.success(request, success_msg)
            
            # REDIRECT TO BANK LIST
            return redirect('bank_list')

        except Exception as e:
            messages.error(request, f"Error saving assignment: {e}")
            return redirect('bankline_recon', record_id=record_id)

    # GET Request context for standard bank lines
    context = {
        'bank_record': recon_segment.bank_line, 
        'company_codes': company_codes,
        'current_recon': recon_segment,
    }
    return render(request, 'unity_internal_app/bankline_recon.html', context)

@login_required
def generate_recon_statement(request, recon_id):
    """Generates a PDF statement for a single reconciled bank line."""
    try:
        recon_record = get_object_or_404(ReconnedBank, pk=recon_id)
        company_listing = get_object_or_404(
            InternalFunds,
            A_Company_Code=recon_record.company_code
        )
    except Exception as e:
        messages.error(request, f"Error fetching PDF data: {e}")
        return redirect('bank_list')
    
    response = HttpResponse(content_type='application/pdf')
    filename = f"Recon_Statement_{recon_record.company_code}_{recon_id}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    try:
        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        styles = getSampleStyleSheet()

        x_margin = inch
        y_cursor = height - inch

        p.setFont("Helvetica-Bold", 16)
        p.drawString(x_margin, y_cursor, "Unity Management Reconciliation Statement")
        y_cursor -= 0.3 * inch

        p.setFont("Helvetica-Bold", 10)
        p.drawString(x_margin, y_cursor, "Company Details:")
        y_cursor -= 0.2 * inch

        data_rows = [
            ("Company Name:", company_listing.B_Company_Name),
            ("Company Code:", recon_record.company_code),
            ("Statement Date:", timezone.now().strftime("%B %d, %Y")),
        ]

        p.setFont("Helvetica", 10)
        for label, value in data_rows:
            p.drawString(x_margin, y_cursor, label)
            p.drawString(x_margin + 2.5 * inch, y_cursor, str(value))
            y_cursor -= 0.2 * inch
        
        y_cursor -= 0.3 * inch

        p.setFont("Helvetica-Bold", 10)
        p.drawString(x_margin, y_cursor, "Transaction Details:")
        y_cursor -= 0.2 * inch
        
        transaction_data = [
            ("Reconciliation ID:", recon_id),
            ("Date Received:", recon_record.transaction_date.strftime("%Y-%m-%d")),
            ("Amount:", f"R{recon_record.transaction_amount}"),
            ("Status:", recon_record.recon_status),
            ("Fiscal Review:", recon_record.fiscal_date.strftime("%Y-%m-%d") if recon_record.fiscal_date else "N/A"),
            ("Note:", recon_record.review_note if recon_record.review_note else "None"),
        ]
        
        p.setFont("Helvetica", 10)
        for label, value in transaction_data:
            p.drawString(x_margin, y_cursor, label)
            p.drawString(x_margin + 2.5 * inch, y_cursor, str(value))
            y_cursor -= 0.2 * inch

        p.setFont("Helvetica-Bold", 10)
        p.drawString(x_margin, y_cursor, "Source Description:")
        y_cursor -= 0.2 * inch
        
        p.setFont("Helvetica", 10)
        description_text = str(recon_record.bank_line.transaction_description)
        if 'Normal' not in styles:
            styles.add(ParagraphStyle(name='Normal'))
        
        description_paragraph = Paragraph(description_text, styles['Normal'])
        
        description_paragraph.wrapOn(p, width - 2 * x_margin, height)
        
        description_paragraph.drawOn(p, x_margin, y_cursor - description_paragraph.height)
        y_cursor -= description_paragraph.height + 0.1 * inch

        p.showPage()
        p.save()
        return response

    except Exception as e:
        messages.error(request, f"PDF Drawing/ReportLab Failure: {e}")
        return redirect('bank_list')
    
# --- BANKLINE REVIEW VIEWS ---
@login_required
def display_bankline_review(request, recon_id):
    """Displays a single reconciled bank line for review and includes bulk allocation breakdown."""
    from .models import ReconnedBank, InternalFunds, BillSettlement
    
    # --- CAPTURE NAVIGATION SOURCE ---
    source_param = request.GET.get('from')
    is_from_unity_info = request.GET.get('source') == 'unity' or source_param == 'unity'
    
    # Fetch the recon record with the original bank line details
    recon_record = get_object_or_404(ReconnedBank.objects.select_related('bank_line'), pk=recon_id)
    
    # --- FETCH BULK SETTLEMENTS ---
    settlements = BillSettlement.objects.filter(
        reconned_bank_line=recon_record
    ).select_related('unity_bill_source')

    # Fetch unique company codes for the selection dropdown
    company_codes = InternalFunds.objects.values_list('A_Company_Code', flat=True).distinct().order_by('A_Company_Code')

    context = {
        'recon_record': recon_record,
        'bank_record': recon_record.bank_line,
        'settlements': settlements,
        'company_codes': company_codes,
        'review_notes': REVIEW_NOTES_OPTIONS,
        'current_category': recon_record.review_note,      
        'is_from_unity_info': is_from_unity_info,
        'source': source_param,                               
    }
    
    return render(request, 'unity_internal_app/display_bankline_review.html', context)

@login_required
@transaction.atomic
def update_bankline_details(request, recon_id):
    """Updates ReconnedBank using separate columns for category and detailed notes."""
    from .models import ReconnedBank
    recon_record = get_object_or_404(ReconnedBank.objects.select_related('bank_line'), pk=recon_id)
    
    if request.method == 'POST':
        new_company_code = request.POST.get('company_code_select')
        new_fiscal_date = request.POST.get('fiscal_date')
        source_param = request.POST.get('source_param')
        
        # Capture separate inputs from the modal
        category = request.POST.get('review_note', '').strip()
        custom_text = request.POST.get('review_note_text', '').strip()

        allocation_cleared = (new_company_code in [None, '', 'None'])
        
        # --- SAVE TO INDIVIDUAL COLUMNS ---
        recon_record.company_code = new_company_code if new_company_code else None
        recon_record.fiscal_date = new_fiscal_date if new_fiscal_date else None
        recon_record.review_note = category            # Category dropdown
        recon_record.review_note_text = custom_text    # Detailed textarea (Saves to the new TEXT column)
        
        # Status Logic
        new_status = recon_record.recon_status
        if allocation_cleared:
            new_status = 'Unassigned'
        elif recon_record.company_code:
            new_status = 'Unreconciled - Allocated' if recon_record.fiscal_date else 'Unreconciled - Assigned'
        
        if category and "Query required" in category:
            new_status = 'Review Pending'
        
        recon_record.recon_status = new_status
        recon_record.save()

        # Sync to bank line comments for audit history
        bank_line = recon_record.bank_line
        bank_line.comments = f"Reviewed: {category} - {custom_text} (Code: {recon_record.company_code or 'N/A'})"
        bank_line.save()

        # --- EMAIL NOTIFICATION LOGIC ---
        if request.POST.get('send_email_toggle') == 'on':
            recipient = request.POST.get('recipient_email')
            cc_email = request.POST.get('cc_email')
            subject = request.POST.get('email_subject')
            full_recipients = f"{recipient},{cc_email}" if cc_email else recipient

            email_body_html = f"""
            <html>
            <body style="font-family: sans-serif; line-height: 1.6; color: #333;">
                <div style="background-color: #1b5e20; color: white; padding: 15px; border-radius: 8px 8px 0 0;">
                    <h2 style="margin:0;">Bank Line Review Update</h2>
                </div>
                <div style="padding: 20px; border: 1px solid #aed581; border-top: none; border-radius: 0 0 8px 8px;">
                    <p>Hello,</p>
                    <p>The following bank line has been updated by <strong>{request.user.username}</strong>:</p>
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr><td style="padding: 8px; border-bottom: 1px solid #eee;"><strong>Company:</strong></td><td style="padding: 8px; border-bottom: 1px solid #eee;">{recon_record.company_code or 'N/A'}</td></tr>
                        <tr><td style="padding: 8px; border-bottom: 1px solid #eee;"><strong>Amount:</strong></td><td style="padding: 8px; border-bottom: 1px solid #eee;">R {recon_record.transaction_amount}</td></tr>
                        <tr><td style="padding: 8px; border-bottom: 1px solid #eee;"><strong>Category:</strong></td><td style="padding: 8px; border-bottom: 1px solid #eee;">{category}</td></tr>
                        <tr><td style="padding: 8px; border-bottom: 1px solid #eee;"><strong>Internal Note:</strong></td><td style="padding: 8px; border-bottom: 1px solid #eee;">{custom_text}</td></tr>
                    </table>
                    <p style="margin-top: 20px; font-size: 12px; color: #777;">System: Unity_Internal | Ref: {recon_id}</p>
                </div>
            </body>
            </html>
            """
            
            result = OutlookGraphService.send_outlook_email(
                settings.OUTLOOK_EMAIL_ADDRESS, full_recipients, subject, email_body_html, 'HTML'
            )

            if result.get('success'):
                messages.success(request, f"Details saved and email sent.")
            else:
                messages.error(request, f"Saved, but Email failed: {result.get('error')}")
        else:
            messages.success(request, f"Bank Line {recon_id} details saved.")

        if source_param == 'global':
            return redirect('global_bank')
        return redirect('bank_list')
    
    return redirect('display_bankline_review', recon_id=recon_id)

# --- NEW HELPER FUNCTION FOR FLEXIBLE ALLOCATION (Replaces old calculate_bill_debt logic) ---
def get_available_bank_lines(company_code):
    """
    Returns a queryset of ReconnedBank lines that have remaining unsettled amounts 
    for the specific company code, regardless of fiscal date.
    """
    return ReconnedBank.objects.filter(
        company_code=company_code,
        # Check if the line has money remaining
        amount_settled__lt=F('transaction_amount'),
    ).select_related('bank_line').annotate(
        remaining_debt=F('transaction_amount') - F('amount_settled')
    ).order_by('transaction_date')

# --- BILLING RECONCILIATION VIEWS (UPDATED FOR FLEXIBLE ALLOCATION) ---
from decimal import Decimal
# NOTE: Ensure necessary imports like Decimal, Sum, timezone, etc. are at the top of your views.py

@login_required
def pre_bill_reconciliation_summary(request, company_code, bill_id):
    # Define a small tolerance for Decimal comparisons
    SAFETY_TOLERANCE = Decimal('0.0001') 
    ZERO_DECIMAL = Decimal('0.00')
    
    bill_record = get_object_or_404(UnityBill, id=bill_id, C_Company_Code=company_code)
    
    # --- CALCULATE AVAILABLE DEBT (Cash) ---
    available_debt_lines = get_available_bank_lines(company_code)
    total_debt = available_debt_lines.aggregate(total=Sum('remaining_debt'))['total'] or ZERO_DECIMAL
    
    # --- AGGREGATIONS (Already Applied) ---
    # 1. Total settled amount from all sources
    total_applied = BillSettlement.objects.filter(
        unity_bill_source_id=bill_record.pk
    ).aggregate(total=Sum('settled_amount'))['total'] or ZERO_DECIMAL
    
    # 2. Separate Cash applied
    total_cash_applied = BillSettlement.objects.filter(
        unity_bill_source_id=bill_record.pk,
        reconned_bank_line_id__isnull=False 
    ).aggregate(total_cash=Sum('settled_amount'))['total_cash'] or ZERO_DECIMAL

    # 3. Total Credit Notes already applied
    total_credit_notes_assigned = BillSettlement.objects.filter(
        unity_bill_source_id=bill_record.pk,
        source_credit_note_id__isnull=False
    ).aggregate(total_credit=Sum('settled_amount'))['total_credit'] or ZERO_DECIMAL

    # 4. Total Journals/Surplus applied
    total_surplus_applied_to_bill = JournalEntry.objects.filter(
        target_bill=bill_record
    ).aggregate(total_footer=Sum('amount'))['total_footer'] or ZERO_DECIMAL

    # --- MATH UPDATES ---
    scheduled_amount = bill_record.H_Schedule_Amount or ZERO_DECIMAL
    remaining_scheduled_amount = scheduled_amount - total_credit_notes_assigned - total_surplus_applied_to_bill - total_cash_applied
    current_outstanding = max(ZERO_DECIMAL, remaining_scheduled_amount)

    # --- NEW FAIL-SAFE: FETCH AVAILABLE APPROVED CREDITS ---
    # These are credits the manager approved (Overs or previous remainders) 
    # that haven't been applied to a bill yet.
    available_credits = CreditNote.objects.filter(
        member_group_code=company_code,
        credit_link_status='Approved',
        schedule_amount__gt=ZERO_DECIMAL
    )
    
    total_available_credit_value = available_credits.aggregate(t=Sum('schedule_amount'))['t'] or ZERO_DECIMAL

    # --- SURPLUS & JOURNALS (Existing Logic) ---
    company_bill_ids = UnityBill.objects.filter(C_Company_Code=company_code).values_list('id', flat=True)
    potential_surpluses = ScheduleSurplus.objects.filter(
        unity_bill_source_id__in=company_bill_ids
    ).exclude(status='FULLY_APPLIED')
    
    available_surpluses = []
    total_available_surplus_value = ZERO_DECIMAL
    for s in potential_surpluses:
        used = JournalEntry.objects.filter(surplus_source=s).aggregate(t=Sum('amount'))['t'] or ZERO_DECIMAL
        remaining = s.surplus_amount - used
        if remaining > ZERO_DECIMAL:
            s.temp_available = remaining
            total_available_surplus_value += remaining
            available_surpluses.append(s)

    applied_journals = JournalEntry.objects.filter(target_bill=bill_record).select_related('surplus_source')

    # --- FETCH ASSIGNED CREDIT NOTES (For audit trail display) ---
    assigned_note_ids = BillSettlement.objects.filter(
        unity_bill_source_id=bill_record.pk,
        source_credit_note_id__isnull=False
    ).values_list('source_credit_note_id', flat=True)
    credit_notes_history = CreditNote.objects.filter(id__in=assigned_note_ids)
    
    # --- UPDATED ACTION MESSAGE LOGIC ---
    # Include available credits in the "coverage" calculation
    total_coverage_available = total_debt + total_available_credit_value + total_available_surplus_value
    is_bill_fully_covered = total_applied >= (scheduled_amount - SAFETY_TOLERANCE)
    
    if is_bill_fully_covered:
        is_proceed_enabled = True
        action_message = "FULLY COVERED. Ready to Finalize."
    elif total_coverage_available >= current_outstanding:
        is_proceed_enabled = True
        action_message = f"FULL COVERAGE AVAILABLE: R{total_coverage_available:.2f} available to clear R{current_outstanding:.2f} balance."
    else:
        is_proceed_enabled = False
        action_message = f"Action REQUIRED: R{current_outstanding:.2f} liability remains."

    # --- FINAL CONTEXT MAPPING ---
    context = {
        'bill_record': bill_record,
        'company_code': company_code,
        'total_debt': total_debt, # Cash
        'total_available_credit': total_available_credit_value, # Approved Credits (Fail-safe)
        'scheduled_amount': scheduled_amount,
        'total_credit_notes_assigned': total_credit_notes_assigned, # Already used
        'total_available_surplus': total_available_surplus_value,
        'total_cash_applied': total_cash_applied,
        'remaining_schedule_amount': remaining_scheduled_amount,
        'current_outstanding': current_outstanding, 
        'all_lines': available_debt_lines.all().order_by('transaction_date'),
        
        'available_credits': available_credits, # NEW: For "Apply Credit" buttons in HTML
        'credit_notes': credit_notes_history, # Already applied list
        
        'available_surpluses': available_surpluses,
        'applied_journals': applied_journals,
        'action_message': action_message,
        'is_proceed_enabled': is_proceed_enabled,
    }
    
    return render(request, 'unity_internal_app/pre_bill_summary.html', context)

@login_required
@transaction.atomic
def process_bill_settlement(request, company_code, bill_id):
    """***FUNCTION DISABLED***"""
    messages.error(request, "Settlement processing is disabled due to the exclusion of Deposit Amount logic.")
    return redirect('pre_bill_reconciliation_summary', company_code=company_code, bill_id=bill_id)

def get_bank_lines_used_in_settlement(bill_record):
    """
    Retrieves the unique ReconnedBank lines that were applied to the given UnityBill 
    for the final summary table.
    """
    # 1. Get the primary keys (IDs) of all ReconnedBank lines found in the BillSettlement audit trail for this bill.
    used_line_ids = BillSettlement.objects.filter(
        unity_bill_source_id=bill_record.id
    ).values_list('reconned_bank_line_id', flat=True).distinct()
    
    # 2. Retrieve the ReconnedBank objects corresponding to those IDs.
    # Assumes ReconnedBank has a ForeignKey 'bank_line' for descriptions.
    used_lines = ReconnedBank.objects.filter(
        id__in=used_line_ids
    ).select_related('bank_line').order_by('transaction_date')
    
    return used_lines

# Note: The other two helpers (calculate_total_settled_for_display and 
# calculate_total_credit_assigned) are now calculated directly inside 
# reconciliation_success_view, making explicit helper functions unnecessary.

# --- CORE VIEW FUNCTIONS ---

# --- HELPER FUNCTIONS DEFINITIONS (To resolve Pylance errors and ensure data integrity) ---

def get_bank_lines_used_in_settlement(bill_record):
    """
    Retrieves the unique ReconnedBank lines that were applied to the given UnityBill 
    for the final summary table.
    """
    # 1. Get the primary keys (IDs) of all ReconnedBank lines found in the BillSettlement audit trail for this bill.
    used_line_ids = BillSettlement.objects.filter(
        unity_bill_source_id=bill_record.id
    ).values_list('reconned_bank_line_id', flat=True).distinct()
    
    # 2. Retrieve the ReconnedBank objects corresponding to those IDs.
    # We use select_related to efficiently fetch the description from the related BankLine table.
    used_lines = ReconnedBank.objects.filter(
        id__in=used_line_ids
    ).select_related('bank_line').order_by('transaction_date')
    
    return used_lines

# --- CORE VIEW FUNCTIONS ---

@login_required
@transaction.atomic
def process_cash_allocation(request, company_code, bill_id):
    """
    Handles cash allocation for a single selected bank line.
    PRESERVES the original bank transaction amount (e.g., R2500) for audit.
    Any remainder is moved to a CreditNote or New Segment, and the parent is marked as 'Exhausted'.
    """
    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect('pre_bill_reconciliation_summary', company_code=company_code, bill_id=bill_id)

    aware_dt = timezone.now()
    selected_recon_id = request.POST.get('selected_recon_id')
    amount_to_apply_str = request.POST.get('amount_to_apply')
    should_split_and_reallocate = request.POST.get('split_and_reallocate') == 'True' 

    try:
        amount_to_apply = Decimal(amount_to_apply_str) 
        recon_line = ReconnedBank.objects.select_for_update().get(pk=selected_recon_id, company_code=company_code)
        bill_record = UnityBill.objects.select_for_update().get(pk=bill_id, C_Company_Code=company_code)
        
        # 1. Audit Check: The 'Truth' is the original transaction amount.
        original_bank_amount = recon_line.transaction_amount
        line_unsettled = original_bank_amount - recon_line.amount_settled

        if amount_to_apply > (line_unsettled + Decimal('0.0001')):
            messages.error(request, f"Allocation failed: Only R{line_unsettled:.2f} remains.")
            return redirect('pre_bill_reconciliation_summary', company_code=company_code, bill_id=bill_id)
            
        # 2. Bill Coverage Check
        bill_settled_agg = BillSettlement.objects.filter(unity_bill_source_id=bill_record.pk).aggregate(total=Sum('settled_amount'))['total'] or ZERO_DECIMAL
        journal_total = JournalEntry.objects.filter(target_bill=bill_record).aggregate(total=Sum('amount'))['total'] or ZERO_DECIMAL
        bill_remaining_liability = bill_record.H_Schedule_Amount - (bill_settled_agg + journal_total)
        
        # 3. Final Application Amount
        final_amount_applied = min(amount_to_apply, bill_remaining_liability)
        
        # 4. Create Bill Settlement (The portion used for THIS bill)
        BillSettlement.objects.create(
            reconned_bank_line=recon_line,
            unity_bill_source=bill_record,
            settled_amount=final_amount_applied,
            settlement_date=aware_dt,
            confirmed_by=request.user,
            original_import_bank_id=recon_line.bank_line_id,
        )
        
        # 5. Handle the Remainder (The "Overs")
        amount_left_on_source = line_unsettled - final_amount_applied
        
        if amount_left_on_source > Decimal('0.009'):
            if should_split_and_reallocate:
                # Option A: Split to a new unassigned bank line segment
                ReconnedBank.objects.create(
                    bank_line_id=recon_line.bank_line_id,
                    company_code=None,
                    transaction_amount=amount_left_on_source,
                    transaction_date=recon_line.transaction_date,
                    recon_status='Unreconciled - Remainder',
                    amount_settled=ZERO_DECIMAL,
                )
            else:
                # Option B: Move to CreditNote for Manager Approval
                CreditNote.objects.create(
                    member_group_code=company_code,
                    schedule_amount=amount_left_on_source,
                    credit_link_status='Pending',
                    link_request_reason="Overs credit line",
                    source_bank_line=recon_line,
                    comment=f"Overs generated from R{original_bank_amount} deposit. R{final_amount_applied} used for Bill {bill_id}",
                    processed_by=request.user.username,
                    processed_date=aware_dt,
                    ccdates_month=bill_record.A_CCDatesMonth,
                    bank_stmt_date=recon_line.transaction_date,
                    note_selection="OVERS" 
                )
                messages.warning(request, f"R{amount_left_on_source:.2f} moved to Manager Approval.")

        # 6. EXHAUST THE PARENT LINE
        # By setting settled to the original amount, (transaction - settled) = 0.
        # But we NEVER change the transaction_amount itself.
        recon_line.amount_settled = original_bank_amount
        recon_line.recon_status = 'Reconciled'
        recon_line.save()
        
        messages.success(request, f"Applied R{final_amount_applied:.2f}. Source R{original_bank_amount} preserved.")
        return redirect('pre_bill_reconciliation_summary', company_code=company_code, bill_id=bill_id)
        
    except Exception as e:
        messages.error(request, f"Allocation Error: {str(e)}")
        return redirect('pre_bill_reconciliation_summary', company_code=company_code, bill_id=bill_id)
    
@login_required
@transaction.atomic
def finalize_reconciliation(request, company_code, bill_id):
    """
    Processes all selected bank line allocations in bulk, then
    finalizes the bill if the balance requirements are met.
    """
    try:
        bill_record = UnityBill.objects.select_for_update().get(pk=bill_id, C_Company_Code=company_code)
        aware_dt = timezone.now()

        # 1. PROCESS BULK ALLOCATIONS (If any checkboxes were checked)
        if request.method == 'POST':
            selected_ids = request.POST.getlist('selected_recon_ids')
            
            for recon_id in selected_ids:
                amount_str = request.POST.get(f'amount_to_apply_{recon_id}')
                if not amount_str:
                    continue
                
                amount_to_apply = Decimal(amount_str)
                if amount_to_apply <= ZERO_DECIMAL:
                    continue

                recon_line = ReconnedBank.objects.select_for_update().get(pk=recon_id)
                
                line_unsettled = recon_line.transaction_amount - recon_line.amount_settled
                applied_amount = min(amount_to_apply, line_unsettled)

                # Create Audit Trail
                BillSettlement.objects.create(
                    reconned_bank_line=recon_line,
                    unity_bill_source=bill_record,
                    settled_amount=applied_amount,
                    settlement_date=aware_dt,
                    confirmed_by=request.user,
                    original_import_bank_id=recon_line.bank_line_id,
                )

                # Update Bank Line
                recon_line.amount_settled += applied_amount
                if recon_line.amount_settled >= (recon_line.transaction_amount - Decimal('0.0001')):
                    recon_line.recon_status = 'Reconciled'
                else:
                    recon_line.recon_status = 'Partially Reconciled'
                recon_line.save()

        # 2. FINALIZATION CHECK (Verify if bill is now balanced)
        bill_settled_agg = BillSettlement.objects.filter(unity_bill_source_id=bill_record.pk).aggregate(total=Sum('settled_amount'))['total'] or ZERO_DECIMAL
        
        if bill_settled_agg >= (bill_record.H_Schedule_Amount - Decimal('0.0001')):
            if not bill_record.is_reconciled:
                bill_record.is_reconciled = True
                bill_record.save()
                messages.success(request, f"Bill #{bill_id} successfully processed and marked as **RECONCILED**.")
                
                # FIX: Redirect to the success view name used in urls.py
                return redirect('reconciliation_success_view', company_code=company_code, bill_id=bill_id)
            else:
                messages.info(request, "Bill is already reconciled.")
        else:
            remaining_liability = bill_record.H_Schedule_Amount - bill_settled_agg
            messages.error(request, f"Processed selected lines, but R{remaining_liability:.2f} still remains. Bill cannot be closed yet.")

        return redirect('pre_bill_reconciliation_summary', company_code=company_code, bill_id=bill_id)

    except Exception as e:
        messages.error(request, f"Error during finalization: {e}")
        return redirect('pre_bill_reconciliation_summary', company_code=company_code, bill_id=bill_id)
    
@login_required
def reconciliation_success_view(request, company_code, bill_id):
    """
    Renders the confirmation/final summary page after successful reconciliation.
    Gathers all required context for finalize_reconciliation.html.
    """
    bill_record = get_object_or_404(UnityBill, id=bill_id, C_Company_Code=company_code)
    
    # --- Data Aggregation ---
    
    # 1. Settled Totals (Cash + Credit + Journal)
    total_settled_against_bill = BillSettlement.objects.filter(
        unity_bill_source_id=bill_record.pk
    ).aggregate(total=Sum('settled_amount'))['total'] or ZERO_DECIMAL
    
    total_credit_assigned = BillSettlement.objects.filter(
        unity_bill_source_id=bill_record.pk, 
        source_credit_note_id__isnull=False
    ).aggregate(total=Sum('settled_amount'))['total'] or ZERO_DECIMAL
    
    # 2. Bank Lines used for settlement display
    lines_to_settle = get_bank_lines_used_in_settlement(bill_record) 
    
    # 3. Fiscal Dates calculation
    bill_date = bill_record.A_CCDatesMonth
    month_start_date = bill_date.replace(day=1)
    fiscal_starting_date = month_start_date 
    fiscal_closing_date = month_start_date + relativedelta(months=1) - relativedelta(days=1)
    
    # 4. Final Math for Template Display
    total_scheduled_amount_initial = bill_record.H_Schedule_Amount
    # Remaining liability should be 0, but we use max() to avoid negative display
    scheduled_amount = max(ZERO_DECIMAL, total_scheduled_amount_initial - total_settled_against_bill)
    
    # total_debt represents the total funds applied (total_settled_against_bill)
    total_debt = total_settled_against_bill
    
    # --- Context Setup (Matches finalize_reconciliation.html variables) ---
    context = {
        'company_code': company_code,
        'bill_record': bill_record,
        'lines_to_settle': lines_to_settle, 
        'total_debt': total_debt,
        'total_scheduled_amount_initial': total_scheduled_amount_initial,
        'total_settled_against_bill': total_settled_against_bill,
        'total_credit_assigned': total_credit_assigned,
        'scheduled_amount': scheduled_amount,
        
        'fiscal_starting_date': fiscal_starting_date,
        'fiscal_closing_date': fiscal_closing_date,

        'warning_message': '✅ RECONCILIATION SUCCESSFUL: This bill is permanently closed.', 
        'settle_button_text': 'Reconciliation Complete',
    }
    
    return render(request, 'unity_internal_app/finalize_reconciliation.html', context)
    
@login_required
@transaction.atomic
def edit_bill(request, company_code, bill_id):
    """
    Loads an existing UnityBill record for editing and handles form submission.
    CRITICAL FIX: Prevents R0.00 scheduled bills from being finalized when edited.
    FIX: Corrected redirection logic to manually handle the URL anchor.
    """
    # 1. Fetch the existing bill record
    # Assuming UnityBill is correctly imported
    bill_record = get_object_or_404(
        UnityBill,
        id=bill_id,
        C_Company_Code=company_code
    )
    
    if request.method == 'POST':
        # 2. Bind the form data to the instance
        # Assuming PreBillForm is imported
        form = PreBillForm(request.POST, instance=bill_record)
        
        if form.is_valid():
            # 3. Save the updated instance, but commit=False first to apply custom logic
            edited_bill = form.save(commit=False)
            
            # --- CRITICAL R0.00 CHECK ---
            scheduled_amount = edited_bill.H_Schedule_Amount or ZERO_DECIMAL
            
            if scheduled_amount <= ZERO_DECIMAL:
                # If the scheduled amount is zero, prevent database closure by clearing final dates.
                edited_bill.J_Final_Date = None
                edited_bill.I_Submitted_Date = None
                messages.warning(request, f"Bill #{bill_id} saved, but R0.00 scheduled amount prevents closure. Status remains open.")
            
            # Now save the record with the potentially updated dates
            edited_bill.save()
            
            messages.success(request, f"Bill for {company_code} (ID: {bill_id}) successfully updated.")
            
            # --- FIX: MANUALLY APPEND ANCHOR ---
            # 1. Reverse the URL using only the accepted keyword arguments ('company_code').
            url = reverse('unity_information', kwargs={'company_code': company_code})
            
            # 🛑 CRITICAL FIX: Add cache-busting timestamp to the redirect URL
            timestamp = timezone.now().timestamp()
            
            # 2. Append the anchor string ('#recon') manually.
            return redirect(f"{url}?cache={timestamp}#recon")

        else:
            messages.error(request, "Please correct the errors in the form.")
            
    else:
        # 4. For GET requests, load the form with the existing instance data
        form = PreBillForm(instance=bill_record)

    context = {
        'company_code': company_code,
        'form': form,
        'bill_id': bill_id,
        'is_editing': True,
    }
    
    # Use the generic template
    return render(request, 'unity_internal_app/bill_form.html', context)

# --- Define ALL required Excel headers (Remains the same) ---
EXCEL_FIELD_MAPPING = {
    'CCDates Month': 'ccdates_month',
    'Fund Code': 'fund_code',
    'Member Group Code': 'member_group_code',
    'Member Group Name': 'member_group_name',
    'Active Members - (Info from FuturaSA & NOT checked by Sanlam)': 'active_members',
    'Schedule Date': 'schedule_date',
    'Final Data Received Date': 'final_data_received_date',
    'Schedule Amount': 'schedule_amount',
    'Confirmation Date': 'confirmation_date',
    'Bank Stmt Date': 'bank_stmt_date',
    'Bank Deposit Amount': 'bank_deposit_amount',
    'Allocated Amount (For Front Office use & not to be checked by Sanlam)': 'allocated_amount',
    'Comment': 'comment',
    'Receipt In Live': 'receipt_in_live',
    'Receipting done by': 'receipting_done_by',
    '0101 Balance Sufficnt: Yes/No': 'balance_sufficient_flag',
    'Date & Letter checked:': 'date_letter_checked',
    'Done by:': 'done_by',
}
# ----------------------------------------------------

# --- HELPER FUNCTIONS (MUST BE DEFINED BEFORE import_credit) ---
# ----------------------------------------------------
def clean_value(value):
    """Cleans and strips white space from a value."""
    if value is None:
        return ''
    return str(value).strip()

# CORRECTED parse_date function (Around line 1060)
def parse_date(date_obj):
    """
    Handles parsing date strings or datetime/date objects.
    Returns a Python datetime.date object or None.
    """
    if date_obj is None or date_obj == '':
        return None
    
    # 1. Handle Python datetime/date objects
    if isinstance(date_obj, datetime):
        return date_obj.date()
    if isinstance(date_obj, date): # <--- Checks against the imported 'date' class
        return date_obj
    
    # 2. Convert value to clean string
    date_str = str(date_obj).strip()
    date_str = date_str.split(' ')[0] # Strip off any time component
    
    # 3. Try common formats
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y%m%d'):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
            
    return None

def parse_decimal(amount_str):
    """Cleans and converts string to Decimal."""
    if amount_str is None or amount_str == '':
        return None
    cleaned = str(amount_str).replace('R', '').replace(',', '').strip()
    try:
        if not cleaned: return None
        return Decimal(cleaned)
    except Exception:
        return None

# ----------------------------------------------------
# --- IMPORT_CREDIT FUNCTION (Calls the helpers above) ---
# ----------------------------------------------------
@login_required
@transaction.atomic
def import_credit(request):
    if request.method != 'POST' or 'credit_file' not in request.FILES:
        return render(request, 'unity_internal_app/import_credit.html', {'expected_headers': EXCEL_FIELD_MAPPING.keys()})
        
    credit_file = request.FILES['credit_file']
    filename = credit_file.name
    rows = []
    
    # 1. DETERMINE READER METHOD AND READ FILE DATA (Logic unchanged)
    if filename.endswith(('.xlsx', '.xls')):
        # XLSX/XLS: Use openpyxl to read
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(credit_file)
            sheet = workbook.active
            rows = [[clean_value(cell.value) for cell in row] for row in sheet.iter_rows()]
            
            if len(rows) > 4:
                rows = rows[4:]
            else:
                messages.error(request, "Excel file too short. Cannot skip the first 4 rows.")
                return redirect('import_credit')
                
        except Exception as e:
            messages.error(request, f"Error reading Excel file: {e}")
            return redirect('import_credit')
    
    elif filename.endswith('.csv'):
        # CSV: Use standard CSV reader
        try:
            file_data = credit_file.read().decode('utf-8')
            io_string = io.StringIO(file_data)
            
            try:
                import csv
                dialect = csv.Sniffer().sniff(io_string.readline())
                io_string.seek(0)
                reader = csv.reader(io_string, dialect)
            except Exception:
                io_string.seek(0)
                reader = csv.reader(io_string, delimiter=',')
                
            all_csv_rows = list(reader)

            if len(all_csv_rows) > 4:
                rows = all_csv_rows[4:]
            else:
                messages.error(request, "CSV file too short. Cannot skip the first 4 rows.")
                return redirect('import_credit')
                
        except UnicodeDecodeError:
            messages.error(request, "Unicode Decode Error: The CSV file is not encoded in UTF-8. Try saving it as CSV (Comma delimited) with UTF-8 encoding.")
            return redirect('import_credit')
        except Exception as e:
            messages.error(request, f"Error reading CSV file: {e}")
            return redirect('import_credit')
    
    else:
        messages.error(request, "Invalid file format. Please upload an Excel (.xlsx/.xls) or CSV file.")
        return redirect('import_credit')
    
    if not rows:
        messages.error(request, "No data rows found after skipping template headers.")
        return redirect('import_credit')

    # 2. READ HEADER (Logic unchanged)
    try:
        header = rows.pop(0)
        column_indices = {}
        
        for excel_header, target_field in EXCEL_FIELD_MAPPING.items():
            try:
                index = header.index(excel_header)
                column_indices[target_field] = index
            except ValueError:
                column_indices[target_field] = None
        
        if column_indices.get('member_group_code') is None:
            messages.error(request, f"Import failed: Mandatory column 'Member Group Code' is missing from the effective header row (Row 5 in the file).")
            return redirect('import_credit')
    except Exception as e:
        messages.error(request, f"Error during header processing: {e}")
        return redirect('import_credit')


    # 3. PROCESS ROWS (Using single saves and strict field type checking)
    rows_processed = 0
    
    # Get the CreditNote model's field map once for efficiency
    credit_note_fields = {f.name: f for f in CreditNote._meta.fields}

    try:
        with transaction.atomic():
            for row in rows:
                if not any(row) or len(row) < len(header):
                    continue
                
                row_data = {}
                has_error = False
                
                for target_field, index in column_indices.items():
                    if index is not None and index < len(row):
                        raw_value = row[index]
                        
                        # --- Type-specific parsing and ASSIGNMENT ---
                        if target_field in credit_note_fields:
                            field_instance = credit_note_fields[target_field]
                            
                            if isinstance(field_instance, DateField):
                                parsed_date = parse_date(raw_value)
                                
                                # Explicitly format to YYYY-MM-DD string only if parsed
                                if parsed_date:
                                    row_data[target_field] = parsed_date.strftime('%Y-%m-%d')
                                else:
                                    row_data[target_field] = None
                                    
                            elif isinstance(field_instance, DateTimeField):
                                # If you had DATETIME fields from Excel, parse them here
                                row_data[target_field] = None
                            
                            elif 'amount' in target_field or target_field == 'schedule_amount':
                                row_data[target_field] = parse_decimal(raw_value)
                            
                            elif target_field == 'active_members':
                                try:
                                    row_data[target_field] = int(clean_value(raw_value))
                                except ValueError:
                                    row_data[target_field] = None
                                    
                            else:
                                row_data[target_field] = clean_value(raw_value)
                        
                    else:
                        row_data[target_field] = None
                    
                    if target_field == 'member_group_code' and not row_data.get('member_group_code'):
                        messages.warning(request, f"Skipping row {rows_processed + 1}: Missing mandatory 'Member Group Code'.")
                        has_error = True
                        break
                
                if has_error:
                    continue

                # Handle processed_date (The only DATETIME field in your target table)
                # 🛠️ CRITICAL FIX APPLIED HERE: Using datetime.now() instead of datetime.datetime.now()
                row_data['processed_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                row_data['processed_by'] = request.user.username
                
                # --- Create the object ---
                CreditNote.objects.create(**row_data)
                
                rows_processed += 1
            
        if rows_processed > 0:
            messages.success(request, f"Successfully imported {rows_processed} Bill Detail Records into Credit_note.")
        else:
            messages.warning(request, "No valid records found to import.")
            
    except Exception as e:
        messages.error(request, f"A critical database error occurred during single insertion (Row {rows_processed + 1}): {e}")
        
    return redirect('import_credit')

@login_required
def credit_note_list(request):
    """
    Displays all imported and manually created CreditNote records.
    Updated to include Pending and Approved items so they can be 
    toggled in the UI.
    """
    # Use __in to fetch all three statuses
    credit_notes = CreditNote.objects.filter(
        credit_link_status__in=['Unlinked', 'Pending', 'Approved']
    ).order_by('-processed_date')
    
    context = {
        'page_title': 'Available Credit Imports',
        'credit_notes': credit_notes,
    }
    
    return render(request, 'unity_internal_app/credit_note_list.html', context)

@login_required
@transaction.atomic
def assign_fiscal_date_view(request, note_id):
    from django.contrib import messages
    from django.shortcuts import redirect, get_object_or_404, render
    from django.urls import reverse
    from django.utils import timezone
    from decimal import Decimal
    from .models import CreditNote, UnityBill
    from .forms import FiscalDateAssignmentForm

    note_record = get_object_or_404(CreditNote, id=note_id)
    current_company_code = note_record.member_group_code 

    if request.method == 'POST':
        form = FiscalDateAssignmentForm(request.POST, instance=note_record)
        
        # Ensure manually handled fields don't block validation
        if 'fiscal_date' in form.fields: form.fields['fiscal_date'].required = False
        if 'target_bill_id' in form.fields: form.fields['target_bill_id'].required = False

        new_group_code = request.POST.get('member_group_code')
        link_reason = request.POST.get('link_reason')
        requested_amt = request.POST.get('requested_link_amount')
        
        # Capture the new inputs from the updated HTML
        bank_fiscal_input = request.POST.get('fiscal_date')
        date_identified_input = request.POST.get('date_identified')

        if form.is_valid() and new_group_code:
            note = form.save(commit=False)
            
            # Update Core Workflow Fields
            note.member_group_code = new_group_code
            note.link_request_reason = link_reason
            
            # Update the specific date fields
            if bank_fiscal_input:
                note.fiscal_date = bank_fiscal_input
            if date_identified_input:
                note.date_identified = date_identified_input
            
            # Handle the Allocation Amount
            if requested_amt:
                try:
                    note.requested_amount = Decimal(requested_amt)
                except (ValueError, TypeError):
                    note.requested_amount = note.schedule_amount
            
            # Find an open bill for linking
            open_bill = UnityBill.objects.filter(
                C_Company_Code=new_group_code,
                is_reconciled=False
            ).order_by('A_CCDatesMonth').first()
            
            if open_bill:
                note.pending_linked_bill = open_bill
                note.credit_link_status = 'Pending'
                note.save()
                messages.success(request, f"Requested R{note.requested_amount} allocation for Bill #{open_bill.id}.")
            else:
                note.save()
                messages.warning(request, "Request saved, but no open bill was found for this code.")

            timestamp = timezone.now().timestamp()
            return redirect(f"{reverse('unity_billing_history', kwargs={'company_code': new_group_code})}?cache={timestamp}#credit")
    else:
        form = FiscalDateAssignmentForm(instance=note_record)

    context = {
        'page_title': 'Request Link Approval', 
        'note': note_record, 
        'form': form, 
        'company_code': current_company_code
    }
    return render(request, 'unity_internal_app/assign_fiscal_date.html', context)

@login_required
@transaction.atomic
def allocate_surplus_to_bill(request, bill_id):
    """
    Process the allocation of a Surplus to a Bill via a Journal Entry.
    UPDATED: Now triggers is_reconciled = True if the allocation completes the bill.
    """
    # Ensure local imports are available inside the function or globally
    from django.db.models import Sum
    from decimal import Decimal
    from django.utils import timezone
    from datetime import datetime

    if request.method != 'POST':
        return redirect('dashboard')

    # 1. Get Data from POST
    surplus_id = request.POST.get('surplus_id')
    amount_str = request.POST.get('amount')
    
    target_bill = get_object_or_404(UnityBill, pk=bill_id)
    company_code = target_bill.C_Company_Code
    
    try:
        amount_to_allocate = Decimal(amount_str)
        if amount_to_allocate <= ZERO_DECIMAL:
            raise ValueError("Amount must be positive.")

        # --- Fetch Surplus and Validate ---
        surplus = ScheduleSurplus.objects.select_for_update().get(pk=surplus_id)
        
        # Check if enough remains (DB level check)
        used_so_far = JournalEntry.objects.filter(surplus_source=surplus).aggregate(sum=Sum('amount'))['sum'] or ZERO_DECIMAL
        remaining_surplus = surplus.surplus_amount - used_so_far

        if amount_to_allocate > remaining_surplus:
            messages.error(request, f"Cannot allocate R{amount_to_allocate}. Only R{remaining_surplus} remains in this surplus.")
            return redirect(reverse('pre_bill_reconciliation_summary', kwargs={'company_code': company_code, 'bill_id': bill_id}))

        current_date_obj = timezone.now().date()
        naive_dt = datetime.combine(current_date_obj, datetime.min.time())
        aware_dt = timezone.make_aware(naive_dt)

        # 1. Create the Journal Entry
        journal_entry = JournalEntry.objects.create(
            surplus_source=surplus,
            target_bill=target_bill,
            amount=amount_to_allocate,
            created_by=request.user.username,
            allocation_date=current_date_obj
        )
        
        # 2. Create the BillSettlement record
        BillSettlement.objects.create(
            reconned_bank_line=None,
            unity_bill_source=target_bill,
            settled_amount=amount_to_allocate,
            settlement_date=aware_dt,
            source_credit_note_id=None,
            source_journal_entry_id=journal_entry.pk
        )
        
        # 3. Update Surplus Status
        new_used_amount = used_so_far + amount_to_allocate
        
        if new_used_amount >= surplus.surplus_amount:
            surplus.status = 'FULLY_APPLIED'
        else:
            surplus.status = 'PARTIALLY_APPLIED'
        surplus.save()

        # ============================================================
        # --- NEW LOGIC: CHECK FOR BILL COMPLETION (THE TRIGGER) ---
        # ============================================================
        
        # Calculate total paid so far (Cash + Credits + Journals)
        total_settled = BillSettlement.objects.filter(
            unity_bill_source_id=target_bill.pk
        ).aggregate(t=Sum('settled_amount'))['t'] or ZERO_DECIMAL
        
        # Check if bill is fully paid
        if total_settled >= target_bill.H_Schedule_Amount:
            target_bill.is_reconciled = True  # <--- FLIP THE SWITCH
            target_bill.save()
            messages.success(request, f"Allocation successful. Bill #{target_bill.id} is now FULLY RECONCILED.")
        else:
            messages.success(request, f"Journal Entry created! R{amount_to_allocate} allocated. Bill remains OPEN.")

    except Exception as e:
        messages.error(request, f"Allocation failed: {str(e)}")

    # Redirect
    timestamp = timezone.now().timestamp()
    return redirect(f"{reverse('pre_bill_reconciliation_summary', kwargs={'company_code': company_code, 'bill_id': bill_id})}?cache={timestamp}")

@login_required
def settle_bill_report(request, company_code, bill_id):
    """
    Read-only Audit Report for a settled bill.
    FIX: Added remaining_balance logic to prevent incorrect negative credit display.
    """
    from decimal import Decimal
    from django.db.models import Sum
    from django.shortcuts import get_object_or_404, render
    # Ensure ZERO_DECIMAL is defined if not global
    ZERO_DECIMAL = Decimal('0.00')
    
    from .models import UnityBill, BillSettlement, CreditNote, JournalEntry, ScheduleSurplus

    bill_record = get_object_or_404(UnityBill, id=bill_id, C_Company_Code=company_code)
    
    # 1. Fetch CASH Settlements (where reconned_bank_line is NOT NULL)
    settlements = BillSettlement.objects.filter(
        unity_bill_source_id=bill_record.id,
        reconned_bank_line_id__isnull=False,
    ).select_related('reconned_bank_line', 'reconned_bank_line__bank_line').order_by('settlement_date')
    
    settled_total = sum(s.settled_amount for s in settlements)

    # 2. Fetch CREDIT Settlements (where source_credit_note_id is NOT NULL)
    credit_settlements = BillSettlement.objects.filter(
        unity_bill_source_id=bill_record.id,
        source_credit_note_id__isnull=False
    ).order_by('settlement_date')
    
    credit_total = sum(c.settled_amount or ZERO_DECIMAL for c in credit_settlements)

    # 3. ATTACH CREDIT NOTE DETAILS
    credit_ids = [s.source_credit_note_id for s in credit_settlements if s.source_credit_note_id is not None]
    credit_note_map = {
        cn.id: cn for cn in CreditNote.objects.filter(id__in=credit_ids)
    }
    
    for settlement in credit_settlements:
        settlement.original_credit_note = credit_note_map.get(settlement.source_credit_note_id)

    # 4. Fetch SURPLUS Settlements (where source_journal_entry_id is NOT NULL)
    journal_settlements = BillSettlement.objects.filter(
        unity_bill_source_id=bill_record.id,
        source_journal_entry_id__isnull=False
    ).order_by('settlement_date')

    journal_total = sum(j.settled_amount or ZERO_DECIMAL for j in journal_settlements)

    # 5. Calculation Logic Fix
    total_paid = settled_total + credit_total + journal_total
    scheduled_amount = bill_record.H_Schedule_Amount or ZERO_DECIMAL
    
    # This is the fix: Baseline is the Schedule, not the Cash Received
    remaining_balance = scheduled_amount - total_paid
    
    # 6. Check for Surplus Generated
    generated_surplus = ScheduleSurplus.objects.filter(unity_bill_source_id=bill_record.id).first()

    context = {
        'bill': bill_record,
        'company_code': company_code,
        
        # Data Lists
        'settlements': settlements,
        'credit_settlements': credit_settlements,
        'journal_settlements': journal_settlements,
        'generated_surplus': generated_surplus,
        
        # Totals
        'settled_total': settled_total,    # This is your "Net Cash"
        'credit_total': credit_total,
        'journal_total': journal_total,
        'total_paid': total_paid,
        'remaining_balance': remaining_balance, # Use this in your template for the balance row
        'zero': ZERO_DECIMAL
    }
    
    return render(request, 'unity_internal_app/settle_bill_report.html', context)

from django.http import HttpResponse
@login_required
def export_settled_bill_csv(request, company_code, bill_id):
    """
    Exports the SINGLE settled bill details to a CSV file.
    """
    bill = get_object_or_404(UnityBill, id=bill_id, C_Company_Code=company_code)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="Settled_Bill_{company_code}_{bill_id}.csv"'

    writer = csv.writer(response)
    
    headers = [
        'A_CCDatesMonth', 'B_Fund_Code', 'C_Company_Code', 'D_Company_Name',
        'E_Active_Members', 'F_Pre-Bill_Date', 'G_Schedule_Date', 'H_Schedule_Amount',
        'I_Submitted_Date', 'J_Final_Date', 'K_Bank_Stmt_Date', 'L_Bank_Deposit_Amount'
    ]
    writer.writerow(headers)

    def write_row(date_val, amount_val):
        writer.writerow([
            bill.A_CCDatesMonth, bill.B_Fund_Code, bill.C_Company_Code, bill.D_Company_Name,
            bill.E_Active_Members, bill.F_Pre_Bill_Date, bill.G_Schedule_Date, bill.H_Schedule_Amount,
            bill.I_Submitted_Date, bill.J_Final_Date,
            date_val, amount_val
        ])

    # 1. Bank Settlements
    settlements = BillSettlement.objects.filter(unity_bill_source_id=bill.id).select_related('reconned_bank_line__bank_line')
    for s in settlements:
        # Need to check if reconned_bank_line is not None (credit/journal settlements won't have it)
        if s.reconned_bank_line:
            write_row(s.reconned_bank_line.bank_line.date if s.reconned_bank_line.bank_line else s.settlement_date.date(), s.settled_amount)

    # 2. Credit Notes (redundant if using BillSettlement, but retained for old CreditNote model lookup)
    credits = CreditNote.objects.filter(assigned_unity_bill=bill)
    for c in credits:
        write_row(c.fiscal_date or c.processed_date, c.schedule_amount)

    # 3. Journal Entries
    journals = JournalEntry.objects.filter(target_bill=bill)
    for j in journals:
        write_row(j.allocation_date, j.amount)

    return response

# Assuming UnityBill model and other necessary imports exist.
# MAX_DEPOSITS constant
MAX_DEPOSITS = 5
TWO_PLACES = Decimal('0.00') # Assuming this is correctly defined globally

@login_required
def export_global_history_csv(request):
    """
    Exports the payment history for Bills that had settlement activity 
    in a horizontal format (pivoted deposits).
    
    FIX: Unifies data fetching by querying BillSettlement, CreditNote, and 
          JournalEntry models separately and merging the results before pivoting.
    """
    # --- Constants ---
    MAX_DEPOSITS = 5
    TWO_PLACES = Decimal('0.00')
    
    # --- Date Filtering Logic (Unchanged) ---
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    filter_start_date = None
    filter_end_date = None
    
    try:
        if start_date_str:
            filter_start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        if end_date_str:
            filter_end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Invalid date format provided for filtering.", status=400)
    
    T1_TABLE = 'bill_settlement'
    
    # --- 1. Determine Bill IDs to Display (Filtered by Settlement Date) ---
    # This logic is kept concise, relying on BillSettlement for date filtering if requested.
    
    all_bills_queryset = UnityBill.objects.all()
    
    if filter_start_date or filter_end_date:
        # Get bill IDs from BillSettlement based on date range
        settlement_filter = BillSettlement.objects.all()
        if filter_start_date:
            settlement_filter = settlement_filter.filter(settlement_date__gte=filter_start_date)
        if filter_end_date:
            settlement_filter = settlement_filter.filter(settlement_date__lte=filter_end_date)
        
        filtered_bill_ids = settlement_filter.values_list('unity_bill_source_id', flat=True).distinct()
        
        if not filtered_bill_ids:
            all_bills_queryset = UnityBill.objects.none()
        else:
            all_bills_queryset = all_bills_queryset.filter(id__in=filtered_bill_ids)
            
    all_bills = list(all_bills_queryset.order_by('C_Company_Code', '-A_CCDatesMonth'))
    filtered_bill_ids = [bill.id for bill in all_bills]
    
    # --- 2. Fetch ALL Granular Settlements (Cash, Credit, Journal) ---

    deposits_by_bill = defaultdict(list)
    credits_map = defaultdict(Decimal)

    if filtered_bill_ids:
        # A. Fetch ALL BillSettlement records for the target bills
        all_settlements = BillSettlement.objects.filter(
            unity_bill_source_id__in=filtered_bill_ids
        ).select_related(
            'reconned_bank_line',
            'reconned_bank_line__bank_line'
        ).order_by('settlement_date')

        # B. Fetch all relevant Credit Notes and Journals for lookups
        credit_ids = all_settlements.values_list('source_credit_note_id', flat=True).distinct()
        journal_ids = all_settlements.values_list('source_journal_entry_id', flat=True).distinct()

        # Optimize: Pre-fetch source object maps (if necessary for rich detail)
        credit_note_details = {cn.id: cn for cn in CreditNote.objects.filter(id__in=credit_ids)}
        journal_entry_details = {je.id: je for je in JournalEntry.objects.filter(id__in=journal_ids)}
        
        # C. Map BillSettlement entries to deposits_by_bill list
        for s in all_settlements:
            deposit_amount = s.settled_amount or ZERO_DECIMAL
            source_type = 'Unknown'
            deposit_date = s.settlement_date.date() # Default date

            if s.reconned_bank_line_id:
                # 1. Cash Settlement (Primary Source is ReconnedBank/ImportBank)
                source_type = 'Cash'
                if s.reconned_bank_line and s.reconned_bank_line.bank_line:
                    deposit_date = s.reconned_bank_line.bank_line.date
            
            elif s.source_credit_note_id:
                # 2. Credit Settlement
                source_type = 'Credit'
                credits_map[s.unity_bill_source_id] += deposit_amount # Track total credit for status check
                
                cn = credit_note_details.get(s.source_credit_note_id)
                if cn and cn.fiscal_date:
                    deposit_date = cn.fiscal_date
            
            elif s.source_journal_entry_id:
                # 3. Journal/Surplus Settlement
                source_type = 'Journal'
                
                je = journal_entry_details.get(s.source_journal_entry_id)
                if je and je.allocation_date:
                    deposit_date = je.allocation_date

            deposits_by_bill[s.unity_bill_source_id].append({
                'date': deposit_date,
                'amount': deposit_amount,
                'type': source_type
            })

    # --- 3. Generate CSV Response ---

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="Global_Horizontal_Payments_DOWNLOAD.csv"'

    writer = csv.writer(response)
    
    # 1. Define Headers (Unchanged)
    base_headers = [
        'A_CCDatesMonth', 'B_Fund_Code', 'C_Company_Code', 'D_Company_Name',
        'E_Active_Members', 'F_Pre-Bill_Date', 'G_Schedule_Date', 'H_Schedule_Amount',
        'I_Submitted_Date', 'J_Final_Date',
    ]
    
    payment_headers = []
    for i in range(MAX_DEPOSITS):
        payment_headers.extend([
            f'{chr(75 + 2 * i)}_Bank_Stmt_Date',  # K, M, O, Q, S
            f'{chr(76 + 2 * i)}_Bank_Deposit_Amount'  # L, N, P, R, T
        ])
    
    writer.writerow(base_headers + payment_headers)
    
    # Date format for CSV output
    CSV_DATE_FORMAT = '%d/%m/%Y'

    for bill in all_bills:
        
        deposits = deposits_by_bill.get(bill.id, [])
        
        # Calculate settlement status (CRITICAL: Needs to rely on total paid vs schedule)
        total_settled = sum((d['amount'] for d in deposits), start=TWO_PLACES)
        
        is_settled = total_settled >= (bill.H_Schedule_Amount or TWO_PLACES)

        # CRITICAL FILTER: Skip row if the bill is not fully RECONCILED (as requested)
        if not is_settled:
            continue

        # A. Gather Base Bill Data (Columns A-J)
        row_data = [
            bill.A_CCDatesMonth.strftime(CSV_DATE_FORMAT) if bill.A_CCDatesMonth else '',
            bill.B_Fund_Code or '',
            bill.C_Company_Code or '',
            bill.D_Company_Name or '',
            bill.E_Active_Members or 0,
            bill.F_Pre_Bill_Date.strftime(CSV_DATE_FORMAT) if bill.F_Pre_Bill_Date else '',
            bill.G_Schedule_Date.strftime(CSV_DATE_FORMAT) if bill.G_Schedule_Date else '',
            str((bill.H_Schedule_Amount or ZERO_DECIMAL).quantize(TWO_PLACES)),
            bill.I_Submitted_Date.strftime(CSV_DATE_FORMAT) if bill.I_Submitted_Date else '',
            bill.J_Final_Date.strftime(CSV_DATE_FORMAT) if bill.J_Final_Date else '',
        ]
        
        # B. Prepare for Payment Data (Dynamic Columns K-T)
        payment_data = [''] * (MAX_DEPOSITS * 2)
        
        # Sort all deposits (Cash, Credit, Journal) by date
        deposits.sort(key=lambda d: d['date'])

        for i in range(MAX_DEPOSITS):
            if i < len(deposits):
                deposit = deposits[i]
                
                date_col_index = i * 2
                amount_col_index = i * 2 + 1
                
                # Fill payment data array
                payment_data[date_col_index] = deposit['date'].strftime(CSV_DATE_FORMAT)
                payment_data[amount_col_index] = str(deposit['amount'].quantize(TWO_PLACES))

        # C. Write the final row
        writer.writerow(row_data + payment_data)

    return response

# --- DEFINE CONSTANTS ---
# ZERO_DECIMAL already defined globally at the top
# Tolerance to handle floating point errors when comparing Decimal amounts to zero
TOLERANCE = Decimal('0.00001')

@login_required
def global_history_overview(request):
    """
    Renders a high-level overview of ALL Reconciled Bill History.
    UPDATED: 
    1. Only shows reconciled bills (is_reconciled=True).
    2. Excludes empty/placeholder bills (0 members or 0 amount).
    """
    from decimal import Decimal
    from collections import defaultdict
    
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    filter_start_date = None
    filter_end_date = None
    ZERO_DECIMAL = Decimal('0.00')
    
    try:
        if start_date_str:
            filter_start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        if end_date_str:
            filter_end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "Invalid date format provided for filtering.")
        
    T1_TABLE = 'bill_settlement'
    filtered_bill_ids = set()

    # --- 1. Base Query with Strict Filters ---
    # We apply the filters here first to ensure we only look for settled IDs 
    # belonging to "Valid" reconciled bills.
    base_bills = UnityBill.objects.filter(
        is_reconciled=True
    ).exclude(
        E_Active_Members=0
    ).exclude(
        H_Schedule_Amount=0
    )

    # --- 2. Determine Bill IDs to Display (Filtering by Date) ---
    if filter_start_date or filter_end_date:
        # 1A. Filter by Cash Settlement Date
        where_conditions_cash = []
        sql_args_cash = []
        if filter_start_date:
            where_conditions_cash.append("settlement_date >= %s")
            sql_args_cash.append(filter_start_date)
        if filter_end_date:
            where_conditions_cash.append("settlement_date <= %s")
            sql_args_cash.append(filter_end_date)
            
        where_clause_cash = "WHERE " + " AND ".join(where_conditions_cash) if where_conditions_cash else ""
        cash_filter_sql = f"SELECT DISTINCT unity_bill_source_id FROM {T1_TABLE} {where_clause_cash}"
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(cash_filter_sql, sql_args_cash)
                cash_ids = [row[0] for row in cursor.fetchall()]
                filtered_bill_ids.update(cash_ids)
        except Exception as e:
            messages.error(request, f"Database Error during cash filter: {e}")
            
        # 1B. Filter by Journal Entry Date
        journal_queryset = JournalEntry.objects.all()
        if filter_start_date:
            journal_queryset = journal_queryset.filter(allocation_date__gte=filter_start_date)
        if filter_end_date:
            journal_queryset = journal_queryset.filter(allocation_date__lte=filter_end_date)
        journal_ids = journal_queryset.values_list('target_bill_id', flat=True).distinct()
        filtered_bill_ids.update(journal_ids)
        
        # Apply the found IDs to our base strictly-filtered queryset
        all_bills_queryset = base_bills.filter(id__in=list(filtered_bill_ids))
    else:
        # No date filter: Show all strictly-filtered reconciled bills
        all_bills_queryset = base_bills

    # Prefetch data
    all_bills = list(all_bills_queryset.order_by('-A_CCDatesMonth', 'C_Company_Code'))
    final_bill_ids = [bill.id for bill in all_bills]
    
    # --- 3. Fetch Granular Settlements ---
    final_records = []
    if final_bill_ids:
        id_placeholders = ', '.join(['%s'] * len(final_bill_ids))
        T2_TABLE = 'reconned_bank'
        T3_TABLE = 'importbank'

        # 3A. Cash Deposits
        sql_query_cash = f"""
        SELECT T1.unity_bill_source_id, T3.DATE, T1.settled_amount
        FROM {T1_TABLE} T1
        LEFT JOIN {T2_TABLE} T2 ON T1.reconned_bank_line_id = T2.bank_line_id
        LEFT JOIN {T3_TABLE} T3 ON T2.bank_line_id = T3.id
        WHERE T1.unity_bill_source_id IN ({id_placeholders})
        """
        deposits_by_bill = defaultdict(list)
        
        with connection.cursor() as cursor:
            cursor.execute(sql_query_cash, final_bill_ids)
            for row in cursor.fetchall():
                if row[1]: # Only if deposit_date exists
                    deposits_by_bill[row[0]].append({'date': row[1], 'amount': row[2], 'type': 'Cash'})
            
        # 3B. Journal Entries
        journal_queryset = JournalEntry.objects.filter(target_bill__in=final_bill_ids)
        for je in journal_queryset:
            deposits_by_bill[je.target_bill_id].append({
                'date': je.allocation_date,
                'amount': je.amount,
                'type': 'Journal',
            })
            
        # 3C. Credits
        credit_notes_agg = BillSettlement.objects.filter(
            unity_bill_source_id__in=final_bill_ids,
            source_credit_note_id__isnull=False
        ).values('unity_bill_source_id').annotate(total_credit=Sum('settled_amount'))
        credits_map = {item['unity_bill_source_id']: item['total_credit'] for item in credit_notes_agg}
    
        # --- 4. Final Consolidation ---
        for bill in all_bills:
            deposits = deposits_by_bill.get(bill.id, [])
            cash_journal_settled = sum((d['amount'] for d in deposits), start=ZERO_DECIMAL)
            credit_settled = credits_map.get(bill.id, ZERO_DECIMAL)
            
            final_records.append({
                'bill': bill,
                'deposits': deposits,
                'status_name': 'RECON COMPLETE',
                'status_class': 'badge-success',
                'is_settled': True,
                'total_settled': cash_journal_settled + credit_settled,
            })

    context = {
        'bill_records': final_records,
        'filter_start_date': filter_start_date.strftime('%Y-%m-%d') if filter_start_date else '',
        'filter_end_date': filter_end_date.strftime('%Y-%m-%d') if filter_end_date else '',
    }
    return render(request, 'unity_internal_app/global_history_overview.html', context)

@login_required
def unallocate_surplus(request, bill_id):
    if request.method == 'POST':
        # Retrieve necessary data from the POST request
        journal_id = request.POST.get('journal_entry_id')
        company_code = request.POST.get('company_code') 
        
        # Default redirect in case of missing data
        if not journal_id or not company_code:
            messages.error(request, "Error: Journal Entry ID or Company Code is missing.")
            return redirect('pre_bill_reconciliation_summary', company_code=company_code or 'DEFAULT', bill_id=bill_id)

        try:
            with transaction.atomic():
                # 1. Fetch the Journal Entry (e.g., ID 16 or 17 in your example)
                journal_entry = get_object_or_404(
                    JournalEntry, 
                    pk=journal_id, 
                    target_bill_id=bill_id
                )
                amount = journal_entry.amount
                
                # 2. Find and delete the corresponding BillSettlement record.
                # This record reduces the bill's schedule/total_applied metric.
                settlement_record = get_object_or_404(
                    BillSettlement,
                    source_journal_entry_id=journal_entry.pk,
                    unity_bill_source_id=bill_id
                )
                settlement_record.delete()
                
                # 3. Delete the Journal Entry.
                # This releases the surplus funds back into the available pool.
                journal_entry.delete()
            
            messages.success(request, f"Journal Entry successfully reversed. R{amount:.2f} unallocated from Bill #{bill_id}.")

        except BillSettlement.DoesNotExist:
            messages.error(request, f"Error: BillSettlement record for Journal #{journal_id} not found. Metrics may be inconsistent.")
        except JournalEntry.DoesNotExist:
            messages.error(request, f"Error: Journal Entry #{journal_id} not found or does not belong to this bill.")
        except Exception as e:
            messages.error(request, f"An unexpected error occurred during reversal: {e}")

    # Redirect back to the summary page using the correct company code
    return redirect('pre_bill_reconciliation_summary', company_code=company_code, bill_id=bill_id)

@login_required
def confirmations_view(request):
    """
    Displays bills ready for daily confirmation review.
    
    UPDATED: 
    1. Only shows reconciled bills (is_reconciled=True).
    2. Filters out empty/placeholder bills (0 members or 0 amount).
    """
    from decimal import Decimal
    
    # 1. Date Filtering
    filter_start_date_str = request.GET.get('start_date')
    filter_end_date_str = request.GET.get('end_date')

    # Base Query: Order by 'Final Date' (Ascending), then 'Company Code'
    bills_queryset = UnityBill.objects.all().order_by('J_Final_Date', 'C_Company_Code')
    
    # --- NEW FILTERS ---
    # Only include bills that are fully reconciled/closed
    # AND exclude bills with 0 members or 0 schedule amount to hide "N/A" or empty data
    bills_queryset = bills_queryset.filter(
        is_reconciled=True
    ).exclude(
        E_Active_Members=0
    ).exclude(
        H_Schedule_Amount=0
    )

    if filter_start_date_str:
        try:
            start_dt = datetime.strptime(filter_start_date_str, '%Y-%m-%d').date()
            bills_queryset = bills_queryset.filter(J_Final_Date__gte=start_dt)
        except ValueError:
            pass

    if filter_end_date_str:
        try:
            end_dt = datetime.strptime(filter_end_date_str, '%Y-%m-%d').date()
            bills_queryset = bills_queryset.filter(J_Final_Date__lte=end_dt)
        except ValueError:
            pass
            
    # Limit to top 50 for performance
    review_bills = bills_queryset[:50]

    # 2. Data Consolidation
    confirmation_data = []
    
    for bill in review_bills:
        settlements = BillSettlement.objects.filter(unity_bill_source_id=bill.pk).order_by('settlement_date')
        source_details = []
        active_members = bill.E_Active_Members or 0 
        
        for settlement in settlements:
            source = {}
            source['amount'] = settlement.settled_amount 

            if settlement.reconned_bank_line:
                bank_line = settlement.reconned_bank_line
                source['date'] = bank_line.transaction_date
                source['type'] = 'Bank Line'
            elif settlement.source_credit_note_id:
                try:
                    credit_note = CreditNote.objects.get(id=settlement.source_credit_note_id)
                    source['date'] = credit_note.bank_stmt_date or settlement.settlement_date.date()
                    source['type'] = 'Credit Note'
                except Exception:
                    source['date'] = settlement.settlement_date.date()
                    source['type'] = 'Credit Note (Source Missing)'
            else:
                source['date'] = settlement.settlement_date.date()
                source['type'] = 'Other Source'

            source_details.append(source)
            
        source_details.sort(key=lambda x: x['date'] if x['date'] else datetime.date(1900, 1, 1))

        schedule_amount = bill.H_Schedule_Amount if bill.H_Schedule_Amount is not None else 0 

        confirmation_data.append({
            'bill_id': bill.id,
            'cc_dates_month': bill.A_CCDatesMonth,
            'company_code': bill.C_Company_Code,
            'active_members': active_members, 
            'schedule_date': bill.A_CCDatesMonth, 
            'final_date': bill.J_Final_Date or None, 
            'schedule_amount': schedule_amount,
            'confirmed_date': settlements.first().settlement_date.date() if settlements.exists() else None,
            'source_details': source_details,
        })

    # =========================================================
    # EXPORT TO EXCEL LOGIC (CLEAN GROUPED LAYOUT)
    # =========================================================
    if request.GET.get('export_excel'):
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Confirmations"

        headers = [
            'CC Dates Month', 'Company Code', 'Active Members', 
            'Schedule Date', 'Final Date', 'Schedule Amount', 
            'Confirmed Date', 'Bank Date', 'Bank Amount'
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for item in confirmation_data:
            bill_common = [
                item['cc_dates_month'],          
                item['company_code'],          
                item['active_members'],          
                item['schedule_date'],         
                item['final_date'],            
                item['schedule_amount'],         
                item['confirmed_date'],          
            ]

            empty_common = [''] * len(bill_common)
            sources = item['source_details']

            if not sources:
                ws.append(bill_common + ['', '']) 
            else:
                for index, source in enumerate(sources):
                    bank_cols = [
                        source['date'],
                        source['amount']
                    ]
                    
                    if index == 0:
                        ws.append(bill_common + bank_cols)
                    else:
                        ws.append(empty_common + bank_cols)

        filename = f"Daily_Confirmations_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    # =========================================================

    context = {
        'confirmation_records': confirmation_data,
        'filter_start_date': filter_start_date_str,
        'filter_end_date': filter_end_date_str,
    }
    
    return render(request, 'unity_internal_app/confirmations.html', context)


# --- Admin Billing View (Line-by-Line Display) ---

@login_required
def admin_billing_view(request):
    """
    Displays a raw, line-by-line list of bills ready for Admin Billing confirmation.
    Calculates a 0.3% Admin Fee per bill line based on monthly salary.
    
    UPDATED: 
    1. Only shows reconciled bills (is_reconciled=True).
    2. Strictly excludes bills with 0 members or R0.00 schedule amount.
    3. Updated fee calculation to 0.003 (0.3%).
    """
    from decimal import Decimal
    from datetime import datetime

    filter_start_date = request.GET.get('start_date')
    filter_end_date = request.GET.get('end_date')

    # Base Query: Order by CC Dates Month, then by Company Code
    bills_queryset = UnityBill.objects.all().order_by('-A_CCDatesMonth', 'C_Company_Code')
    
    # --- UPDATED FILTERS ---
    # 1. Must be Reconciled
    # 2. Exclude bills with 0 members (prevents N/A rows)
    # 3. Exclude bills with 0.00 schedule (prevents zero-fee rows)
    bills_queryset = bills_queryset.filter(
        is_reconciled=True
    ).exclude(
        E_Active_Members=0
    ).exclude(
        H_Schedule_Amount=0
    )

    if filter_start_date:
        try:
            start_dt = datetime.strptime(filter_start_date, '%Y-%m-%d').date()
            bills_queryset = bills_queryset.filter(A_CCDatesMonth__gte=start_dt)
        except ValueError:
            pass

    if filter_end_date:
        try:
            end_dt = datetime.strptime(filter_end_date, '%Y-%m-%d').date()
            bills_queryset = bills_queryset.filter(A_CCDatesMonth__lte=end_dt)
        except ValueError:
            pass
            
    final_bill_data = []
    
    for bill in bills_queryset:
        # Since we excluded 0.00 above, we know schedule_amount will be a valid positive number
        schedule_amount = bill.H_Schedule_Amount or Decimal('0.00')
        active_members = bill.E_Active_Members or 0 
        
        # 🚀 NEW: Calculate 0.3% Admin Fee (0.003) 🚀
        admin_fee = schedule_amount * Decimal('0.003')
        
        # Find the FIRST settlement record for metadata (Posted Date/User)
        first_settlement = BillSettlement.objects.filter(
            unity_bill_source_id=bill.pk
        ).order_by('settlement_date').first()
        
        posted_date = first_settlement.settlement_date if first_settlement else None
        
        # Get the username of the person who finalized the recon
        posted_user = "N/A"
        if first_settlement and first_settlement.confirmed_by:
            posted_user = first_settlement.confirmed_by.username
        
        fiscal_period_key = bill.A_CCDatesMonth.strftime("%Y-%m") if bill.A_CCDatesMonth else "N/A"

        final_bill_data.append({
            'fiscal_period': fiscal_period_key,
            'company_code': bill.C_Company_Code or "N/A",
            'company_name': bill.D_Company_Name or "N/A",
            'active_members': active_members, 
            'total_schedule_amount': schedule_amount,
            'total_admin_fee': admin_fee,
            'posted_date': posted_date,
            'posted_user': posted_user, 
        })

    context = {
        'bill_records': final_bill_data,
        'filter_start_date': filter_start_date,
        'filter_end_date': filter_end_date,
    }
    
    return render(request, 'unity_internal_app/admin_billing.html', context)

# ==============================================================================
# [UPDATED VIEWS] ADD THESE BELOW YOUR unity_information FUNCTION
# ==============================================================================

# In views.py

@login_required
def save_claim(request, company_code):
    if request.method == 'POST':
        claim_id = request.POST.get('claim_id')
        
        if claim_id:
            claim_instance = get_object_or_404(UnityClaim, pk=claim_id)
            form = UnityClaimForm(request.POST, instance=claim_instance)
        else:
            form = UnityClaimForm(request.POST)

        if form.is_valid():
            saved_claim = form.save(commit=False)

            # --- 1. CAPTURE NEW FIELDS (POTS & PRESERVATION) ---
            saved_claim.vested_pot_available = request.POST.get('vested_pot_available') == 'on'
            saved_claim.savings_pot_available = request.POST.get('savings_pot_available') == 'on'
            
            v_date = request.POST.get('vested_pot_paid_date')
            saved_claim.vested_pot_paid_date = v_date if v_date else None
            
            s_date = request.POST.get('savings_pot_paid_date')
            saved_claim.savings_pot_paid_date = s_date if s_date else None
            
            p_date = request.POST.get('infund_cert_date')
            saved_claim.infund_preservation_cert_received_date = p_date if p_date else None

            # --- 2. CAPTURE MIP & AMOUNT ---
            saved_claim.mip_number = request.POST.get('mip_number')
            
            amount_val = request.POST.get('claim_amount')
            if amount_val and amount_val.strip():
                try:
                    saved_claim.claim_amount = float(amount_val)
                except ValueError:
                    saved_claim.claim_amount = None
            else:
                saved_claim.claim_amount = None

            # --- 3. LINKED EMAIL LOGIC ---
            outlook_string_id = request.POST.get('linked_email_id')
            if outlook_string_id and outlook_string_id.strip():
                try:
                    email_obj = EmailDelegation.objects.get(email_id=outlook_string_id)
                    saved_claim.linked_email_id = email_obj.id
                    
                    UnityClaimNote.objects.create(
                        claim=saved_claim,
                        note_selection="SUBMITTED VIA E-MAIL",
                        note_description=f"System: Linked to Delegated Email Received at {email_obj.received_at}",
                        created_by=request.user
                    )
                except EmailDelegation.DoesNotExist:
                    saved_claim.linked_email_id = None

            saved_claim.save()

            # --- 4. Handle Manual Notes ---
            note_selection = request.POST.get('note_selection')
            note_description = request.POST.get('note_description')

            if note_selection or (note_description and note_description.strip()):
                UnityClaimNote.objects.create(
                    claim=saved_claim,
                    note_selection=note_selection,
                    note_description=note_description,
                    created_by=request.user
                )
                messages.success(request, "Claim saved and Notes updated.")
            else:
                messages.success(request, "Claim saved successfully.")
        else:
            messages.error(request, f"Error saving claim: {form.errors}")
            
    return redirect(f"{reverse('unity_information', kwargs={'company_code': company_code})}#company-claims")


@login_required
def global_claims_view(request):
    """
    Dashboard for all claims EXCEPT Two Pot.
    Integrated: Email Pre-loading for instant preview (No AJAX required).
    """
    query = request.GET.get('q')
    base_claims = UnityClaim.objects.exclude(claim_type='Two Pot')

    if query:
        claims = base_claims.filter(
            Q(id_number__icontains=query) | 
            Q(member_surname__icontains=query) | 
            Q(company_code__icontains=query)
        ).order_by('-claim_created_date')
    else:
        # Note: If using pagination later, apply it here. 
        # For now, we fetch the last 50 as per your original logic.
        claims = base_claims.order_by('-claim_created_date')[:50] 

    # --- 1. PRE-FETCH EMAIL CONTENT FOR TABLE ICONS ---
    # Collect IDs for any claims that have a linked email
    delegation_pks = [c.linked_email_id for c in claims if c.linked_email_id]
    
    if delegation_pks:
        # Map Delegation Primary Key -> Outlook String Email ID
        # Convert list to set to remove duplicates if multiple claims link to the same email
        delegations_map = EmailDelegation.objects.in_bulk(list(set(delegation_pks)))
        outlook_string_ids = [d.email_id for d in delegations_map.values()]
        
        # Map Outlook String Email ID -> Full Body/Subject Content
        inbox_map = OutlookInbox.objects.in_bulk(outlook_string_ids)

        for claim in claims:
            if claim.linked_email_id:
                try:
                    # Match the ID from the database to the pre-fetched map
                    del_obj = delegations_map.get(int(claim.linked_email_id))
                    if del_obj:
                        inbox_item = inbox_map.get(del_obj.email_id)
                        if inbox_item:
                            # Attach these temporarily to the object for the template <template>
                            claim.email_preview_subject = inbox_item.subject
                            claim.email_preview_sender = inbox_item.sender_address
                            claim.email_preview_body = inbox_item.body_content
                            claim.email_preview_date = inbox_item.received_at
                except (ValueError, TypeError):
                    continue

    all_companies = UnityMgListing.objects.values('a_company_code', 'b_company_name', 'c_agent')

    # --- 2. FETCH DELEGATIONS FOR MODAL DROPDOWN (Compose/Attach Logic) ---
    my_delegated_emails = EmailDelegation.objects.filter(
        assigned_user=request.user, 
        status='DEL'
    ).order_by('-received_at')

    # Attach basic info for dropdown display labels
    dropdown_email_ids = [d.email_id for d in my_delegated_emails]
    dropdown_inbox_items = OutlookInbox.objects.in_bulk(dropdown_email_ids)

    for delegation in my_delegated_emails:
        item = dropdown_inbox_items.get(delegation.email_id)
        if item:
            delegation.subject = item.subject
            delegation.sender = item.sender_address
        else:
            delegation.subject = "(Subject Unavailable)"
            delegation.sender = "Unknown"

    context = {
        'claims': claims,
        'all_companies': all_companies,
        'my_delegated_emails': my_delegated_emails,
        'is_two_pot_view': False
    }
    return render(request, 'unity_internal_app/global_claims.html', context)

@login_required
def global_two_pot_view(request):
    """
    Dedicated Dashboard for ONLY Two Pot claims with Date Range & Pagination.
    Integrated: Email Pre-loading for instant preview (No AJAX required).
    """
    query = request.GET.get('q')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    base_claims = UnityClaim.objects.filter(claim_type='Two Pot').order_by('-claim_created_date')

    # --- Filtering Logic ---
    if query:
        base_claims = base_claims.filter(
            Q(id_number__icontains=query) | 
            Q(member_surname__icontains=query) | 
            Q(company_code__icontains=query) |
            Q(mip_number__icontains=query)
        )

    if start_date and end_date:
        try:
            s_date = parse_date(start_date)
            e_date = parse_date(end_date)
            if s_date and e_date:
                base_claims = base_claims.filter(claim_created_date__range=[s_date, e_date])
        except ValueError:
            pass

    # --- Pagination ---
    paginator = Paginator(base_claims, 12) 
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # --- 1. PRE-FETCH EMAIL CONTENT FOR TABLE ICONS (Instant Preview Logic) ---
    # We collect all delegation IDs from the current page of claims to minimize DB hits
    delegation_pks = [c.linked_email_id for c in page_obj if c.linked_email_id]
    
    if delegation_pks:
        # Map Delegation Primary Key -> Outlook String Email ID
        delegations_map = EmailDelegation.objects.in_bulk(delegation_pks)
        outlook_string_ids = [d.email_id for d in delegations_map.values()]
        
        # Map Outlook String Email ID -> Full Body/Subject Content
        inbox_map = OutlookInbox.objects.in_bulk(outlook_string_ids)

        for claim in page_obj:
            if claim.linked_email_id:
                # Use int conversion to match in_bulk dictionary keys
                del_obj = delegations_map.get(int(claim.linked_email_id))
                if del_obj:
                    inbox_item = inbox_map.get(del_obj.email_id)
                    if inbox_item:
                        # Attach attributes directly to the claim object for the template <template>
                        claim.email_preview_subject = inbox_item.subject
                        claim.email_preview_sender = inbox_item.sender_address
                        claim.email_preview_body = inbox_item.body_content
                        claim.email_preview_date = inbox_item.received_at

    all_companies = UnityMgListing.objects.values('a_company_code', 'b_company_name', 'c_agent')

    # --- 2. FETCH DELEGATIONS FOR MODAL DROPDOWN (Compose/Attach Logic) ---
    my_delegated_emails = EmailDelegation.objects.filter(
        assigned_user=request.user, 
        status='DEL'
    ).order_by('-received_at')

    # Attach basic info for dropdown display labels
    dropdown_email_ids = [d.email_id for d in my_delegated_emails]
    dropdown_inbox_items = OutlookInbox.objects.in_bulk(dropdown_email_ids)

    for delegation in my_delegated_emails:
        item = dropdown_inbox_items.get(delegation.email_id)
        if item:
            delegation.subject = item.subject
            delegation.sender = item.sender_address
        else:
            delegation.subject = "(Subject Unavailable)"
            delegation.sender = "Unknown"

    context = {
        'page_obj': page_obj, 
        'all_companies': all_companies,
        'my_delegated_emails': my_delegated_emails,
        'is_two_pot_view': True,
        'search_query': query, 
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'unity_internal_app/global_two_pot.html', context)


@transaction.atomic
@login_required
def save_global_claim(request):
    if request.method == 'POST':
        # 1. Make a mutable copy of the POST data
        post_data = request.POST.copy()
        
        claim_type_input = post_data.get('claim_type', 'Standard')
        redirect_url = 'global_two_pot' if claim_type_input == 'Two Pot' else 'global_claims'
        
        claim_id = post_data.get('claim_id')
        old_linked_email_id = None
        
        if claim_id:
            claim_instance = get_object_or_404(UnityClaim, pk=claim_id)
            old_linked_email_id = claim_instance.linked_email_id
            form = UnityClaimForm(post_data, request.FILES, instance=claim_instance)
        else:
            form = UnityClaimForm(post_data, request.FILES)

        # --- THE FIX: Silence the allocation requirement ---
        # This prevents the "Select a valid choice" error from blocking the save.
        if 'claim_allocation' in form.fields:
            form.fields['claim_allocation'].required = False

        if form.is_valid():
            saved_claim = form.save(commit=False)
            
            # Ensure Company Code is set (Global view doesn't have it in URL)
            if not saved_claim.company_code:
                saved_claim.company_code = post_data.get('company_code')

            # --- Two Pot Specific Field Override ---
            if claim_type_input == 'Two Pot':
                saved_claim.claim_type = 'Two Pot'
                saved_claim.mip_number = post_data.get('mip_number')
                
                # We set the allocation here manually if it's missing
                if not saved_claim.claim_allocation:
                    saved_claim.claim_allocation = "Two Pot"
                
                # Checkboxes: 'on' if checked, else False
                saved_claim.vested_pot_available = post_data.get('vested_pot_available') == 'on'
                saved_claim.savings_pot_available = post_data.get('savings_pot_available') == 'on'
                
                # Dates: Ensure empty strings become None for the DB
                v_date = post_data.get('vested_pot_paid_date')
                saved_claim.vested_pot_paid_date = v_date if v_date else None
                
                s_date = post_data.get('savings_pot_paid_date')
                saved_claim.savings_pot_paid_date = s_date if s_date else None
                
                p_date = post_data.get('infund_cert_date')
                saved_claim.infund_preservation_cert_received_date = p_date if p_date else None

                # Handle Amount
                amount = post_data.get('claim_amount')
                try:
                    saved_claim.claim_amount = float(amount) if amount and amount.strip() else 0.00
                except ValueError:
                    saved_claim.claim_amount = 0.00

            # Handle Email Linking
            new_linked_email_id = post_data.get('linked_email_id')
            if new_linked_email_id and new_linked_email_id.strip():
                saved_claim.linked_email_id = new_linked_email_id

            if 'claim_attachment' in request.FILES:
                saved_claim.claim_attachment = request.FILES['claim_attachment']

            saved_claim.save()

            # Logging & Notes
            # 1. Email Link Note
            if str(new_linked_email_id or "") != str(old_linked_email_id or ""):
                UnityClaimNote.objects.create(
                    claim=saved_claim,
                    note_selection="SUBMITTED VIA E-MAIL",
                    note_description=f"System: Attached Email ID #{new_linked_email_id}",
                    created_by=request.user
                )

            # 2. Manual Note
            note_desc = post_data.get('note_description')
            if note_desc and note_desc.strip():
                UnityClaimNote.objects.create(
                    claim=saved_claim,
                    note_selection=post_data.get('note_selection') or "GENERAL NOTE",
                    note_description=note_desc,
                    created_by=request.user
                )

            messages.success(request, f"Claim for {saved_claim.member_surname} saved successfully.")
            return redirect(redirect_url)
            
        else:
            # If any other errors remain, they will show up in the messages block
            messages.error(request, f"Could not save claim: {form.errors}")
            return redirect(redirect_url)
            
    return redirect('global_claims')
# --------------------------------------------------------------------- #
# OUTLOOK DELEGATOR VIEWS (Inbox & Assignment)
# --------------------------------------------------------------------- #

from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse
from django.http import HttpResponse # Import HttpResponse for the content view
from dateutil import parser # Ensure parser is imported
from django.contrib.auth import get_user_model # Use get_user_model for clarity

# Services and Models (Assuming these imports are correct based on our prior conversation)
from .models import EmailDelegation
from .services.delegation_service import (
    get_or_create_delegation_status, delegate_email_task, 
    add_delegation_note, get_delegated_emails_for_user, 
    log_delegation_transaction
)
# Correct import: only import the class, and use the class name to call methods
from .services.outlook_graph_service import OutlookGraphService 

User = get_user_model() # Alias for the User model

# Assuming there are other required imports here (e.g., from unity_internal_app.models import ...)


@login_required
def outlook_dashboard_view(request):
    if request.user.username.lower() != 'omega' and not request.user.is_superuser:
        messages.error(request, "Access restricted.")
        return redirect('outlook_delegated_box')

    target_email = request.GET.get('email', settings.OUTLOOK_EMAIL_ADDRESS)
    
    # --- Capture Search and Sort Params ---
    search_query = request.GET.get('q', '').strip().lower()
    sort_order = request.GET.get('sort', 'newest') # Default to Newest
    
    context = {
        'target_email': target_email, 
        'messages': [],
        'search_query': search_query,
        'sort_order': sort_order
    }
    
    inbox_data = OutlookGraphService.fetch_inbox_messages(target_email, top_count=50) 
    
    if 'error' not in inbox_data:
        all_emails = inbox_data.get('value', [])
        email_ids = [e['id'] for e in all_emails]
        
        local_inbox_map = OutlookInbox.objects.filter(email_id__in=email_ids).in_bulk(field_name='email_id')
        delegated_or_recycled_ids = EmailDelegation.objects.filter(
            email_id__in=email_ids
        ).exclude(status='NEW').values_list('email_id', flat=True)
        
        filtered_emails = []
        for email in all_emails:
            email_id = email['id']
            if email_id in delegated_or_recycled_ids:
                continue
                
            received_date_str = email.get('receivedDateTime') 
            delegation, created = EmailDelegation.objects.get_or_create(
                email_id=email_id,
                defaults={'received_at': received_date_str, 'status': 'NEW'}
            )
            
            local_record = local_inbox_map.get(email_id)
            # Standardize date for sorting logic
            sort_date = local_record.received_at if local_record else received_date_str
            
            email['internal_received_at'] = sort_date
            email['delegation_status'] = delegation.get_status_display()
            email['delegation_id'] = delegation.pk 

            # --- 2. WILDCARD SEARCH LOGIC ---
            if search_query:
                subject = email.get('subject', '').lower()
                sender = email.get('from', {}).get('emailAddress', {}).get('address', '').lower()
                sender_name = email.get('from', {}).get('emailAddress', {}).get('name', '').lower()
                
                # If search term isn't in subject or sender, skip this email
                if search_query not in subject and search_query not in sender and search_query not in sender_name:
                    continue

            filtered_emails.append(email)
            
        # --- 3. DATE SORTER LOGIC ---
        # Note: local_record.received_at is a datetime object, received_date_str is a string.
        # Python handles string ISO dates or datetime objects well in sorted()
        reverse_sort = (sort_order == 'newest')
        filtered_emails.sort(key=lambda x: x['internal_received_at'], reverse=reverse_sort)
            
        context['messages'] = filtered_emails
    else:
        context['error'] = inbox_data['error']

    return render(request, 'unity_internal_app/outlook_dashboard.html', context)


@login_required
def send_email_view(request):
    """
    Handles displaying the email form and processing the email submission 
    to the Microsoft Graph API. The email is sent FROM the mailbox specified by target_email.
    """
    # DELEGATION AWARENESS: Get target email from URL or settings default 
    target_email = request.GET.get('email', settings.OUTLOOK_EMAIL_ADDRESS)
    
    if request.method == 'POST':
        recipient = request.POST.get('recipient')
        subject = request.POST.get('subject')
        # 🛑 FIX 1: Define body_html (using 'body' content, assuming HTML) 🛑
        body_html = request.POST.get('body') 
        # 🛑 FIX 2: Define sender (the delegated mailbox) 🛑
        sender = target_email 

        # Simple validation
        if not all([recipient, subject, body_html]):
            messages.error(request, "All fields are required.")
            return render(request, 'unity_internal_app/send_email_form.html', {'target_email': target_email})
        
        # Call the service function, passing the target_email as the sender mailbox
        # Line 3212 (Approximate):
        response = OutlookGraphService.send_outlook_email(sender, recipient, subject, body_html, 'HTML')
        
        # 🛑 FIX 3: Change 'result' to 'response' (Line 3214 Approx.) 🛑
        if response.get('success'): 
            messages.success(request, f"Email sent successfully from {target_email} to {recipient}.")
            # Redirect back to the dashboard, preserving the target email
            return redirect(f"{reverse('outlook_dashboard')}?email={target_email}")
        else:
            error_message = f"Email failed to send from {target_email}. {response.get('error', 'Unknown API Error')}"
            
            # Extract details if they exist in the nested error structure
            details = response.get('details', {})
            if isinstance(details, dict) and 'error' in details and 'message' in details['error']:
                error_message += f" Details: {details['error']['message']}"
            
            messages.error(request, error_message)
            # Render the form again with the failure message
            return render(request, 'unity_internal_app/send_email_form.html', {
                'recipient': recipient,
                'subject': subject,
                'body': body_html, # Pass the correct variable back to the form
                'target_email': target_email
            })

    # Render the empty form on GET request
    return render(request, 'unity_internal_app/send_email_form.html', {'target_email': target_email})


# --------------------------------------------------------------------- #
# OUTLOOK DELEGATED VIEWS (Assigned User Workflow)
# --------------------------------------------------------------------- #

@login_required
def outlook_delegated_box(request):
    """
    Displays the list of tasks specifically assigned to the current user.
    Includes additional DB fields and handles Search/Sort.
    """
    # --- Capture Search and Sort Params ---
    search_query = request.GET.get('q', '').strip().lower()
    sort_order = request.GET.get('sort', 'newest')

    # 1. Fetch delegations for this user
    delegations = EmailDelegation.objects.filter(
        assigned_user=request.user, 
        status='DEL'
    )

    # 2. Map local Inbox details for received_at (using in_bulk for efficiency)
    email_ids = delegations.values_list('email_id', flat=True)
    local_inbox_map = OutlookInbox.objects.filter(email_id__in=email_ids).in_bulk(field_name='email_id')

    tasks = []
    target_email = settings.OUTLOOK_EMAIL_ADDRESS 

    for delegation in delegations:
        # Get received_at from local_inbox_map (Table: unity_internal_inbox)
        local_inbox = local_inbox_map.get(delegation.email_id)
        db_received_at = local_inbox.received_at if local_inbox else delegation.received_at
        
        # --- WILDCARD SEARCH LOGIC ---
        # Search against Subject, Category, Company Code, or From
        # (Note: For subject/from we still use Graph API data, for codes we use DB)
        
        # We fetch Graph details first to allow searching against Subject/From
        endpoint = f"messages/{delegation.email_id}?$select=subject,from"
        try:
            response = OutlookGraphService._make_graph_request(endpoint, target_email, method='GET')
            graph_subject = response.get('subject', '(No Subject)') if 'error' not in response else "N/A"
            graph_from = response.get('from', {}).get('emailAddress', {}).get('name', 'Unknown') if 'error' not in response else "N/A"
            
            if search_query:
                match = any([
                    search_query in graph_subject.lower(),
                    search_query in graph_from.lower(),
                    search_query in (delegation.email_category or '').lower(),
                    search_query in (delegation.company_code or '').lower()
                ])
                if not match:
                    continue

            tasks.append({
                'delegation_id': delegation.pk,
                'status': delegation.get_status_display(),
                'subject': graph_subject,
                'from': graph_from,
                'received': db_received_at, # From local inbox table
                'company_code': delegation.company_code, # From delegation table
                'email_type': delegation.email_category, # From delegation table
            })
        except Exception:
            continue

    # --- DATE SORTER LOGIC ---
    reverse_sort = (sort_order == 'newest')
    tasks.sort(key=lambda x: x['received'] if x['received'] else timezone.now(), reverse=reverse_sort)

    context = {
        'tasks': tasks,
        'task_count': len(tasks),
        'search_query': search_query,
        'sort_order': sort_order
    }
    return render(request, 'unity_internal_app/outlook_delegated_box.html', context)


@login_required
def outlook_delegated_action(request, delegation_id):
    """
    Handles Notes, Replies, Metadata Updates, RESTORATION, and COMPLETION.
    Now supports sending file attachments with replies.
    """
    delegation = get_object_or_404(EmailDelegation, pk=delegation_id)
    
    # --- ROLE-BASED ACCESS CONTROL ---
    is_manager = request.user.username.lower() == 'omega' or request.user.is_superuser
    
    if not is_manager and delegation.assigned_user != request.user:
        messages.error(request, "Access restricted. You are not assigned to this task.")
        return redirect('outlook_delegated_box')

    target_email = settings.OUTLOOK_EMAIL_ADDRESS 

    if request.method == 'POST':
        action_type = request.POST.get('action_type')

        # --- 1. HANDLE RESTORE ---
        if action_type == 'restore_to_inbox':
            delegation.work_related = True
            delegation.status = 'NEW'
            delegation.assigned_user = None 
            delegation.save()

            add_delegation_note(
                delegation_id, 
                request.user, 
                "ACTION: Restored task from Recycle Bin to Main Inbox queue."
            )
            messages.success(request, "Email successfully restored to the Live Inbox.")
            return redirect('outlook_recycle_bin')

        # --- 2. HANDLE COMPLETE ---
        elif action_type == 'mark_complete':
            delegation.status = 'COM'
            delegation.save()

            add_delegation_note(
                delegation_id, 
                request.user, 
                "ACTION: Task marked as COMPLETED. Removed from active queue."
            )
            messages.success(request, "Task successfully marked as Completed.")
            return redirect('outlook_delegated_box')

        # --- 3. HANDLE METADATA UPDATES ---
        elif action_type == 'update_metadata':
            delegation.company_code = request.POST.get('company_code')
            delegation.email_category = request.POST.get('email_category')
            delegation.status = request.POST.get('status')
            delegation.save()

            add_delegation_note(
                delegation_id, 
                request.user, 
                f"SYSTEM: Metadata Updated. Status: [{delegation.status}], Category: [{delegation.email_category}]"
            )
            messages.success(request, "Task metadata updated successfully.")
            return redirect('outlook_delegated_action', delegation_id=delegation_id)

        # --- 4. HANDLE NOTE SUBMISSION ---
        elif action_type == 'add_note':
            note_content = request.POST.get('internal_note')
            success, message = add_delegation_note(delegation_id, request.user, note_content)
            if success: 
                messages.success(request, "Internal note saved.")
            else:
                messages.error(request, message)
            return redirect('outlook_delegated_action', delegation_id=delegation_id)
        
        # --- 5. HANDLE REPLY/SEND EMAIL ---
        elif 'reply_recipient' in request.POST:
            recipient = request.POST.get('reply_recipient')
            raw_subject = request.POST.get('reply_subject')
            subject = raw_subject if raw_subject else f"Reply: {delegation.email_category or 'Task Action'}"
            body_html = request.POST.get('reply_body')
            action_destination = request.POST.get('action_notes', 'EMAIL_REPLY')
            
            log_type = request.POST.get('email_log_type', 'DIRECT') 
            
            # 🚀 NEW: Grab the file from the request
            reply_file = request.FILES.get('reply_file')

            # Pass the file to the service
            response = OutlookGraphService.send_outlook_email(
                target_email, 
                recipient, 
                subject, 
                body_html,
                attachment=reply_file
            )
            
            if response.get('success'):
                final_action_type = 'REPLIED' if log_type == 'REPLY' else action_destination
                
                log_delegation_transaction(
                    delegation_id, 
                    request.user, 
                    subject, 
                    recipient, 
                    action_type=final_action_type 
                )
                
                messages.success(request, f"Reply sent successfully with attachment (if provided).")
            else:
                messages.error(request, f"Failed to send email: {response.get('error')}")
                
            return redirect('outlook_delegated_action', delegation_id=delegation_id)

    # --- GET DATA ---
    endpoint = f"messages/{delegation.email_id}"
    email_data = OutlookGraphService._make_graph_request(endpoint, target_email, method='GET')
    
    attachments = OutlookGraphService.fetch_attachments(target_email, delegation.email_id)
    
    # Loop through attachments for image previews
    for att in attachments:
        content_type = att.get('contentType', '').lower()
        if 'image' in content_type:
            raw_att = OutlookGraphService.get_attachment_raw(target_email, delegation.email_id, att['id'])
            if isinstance(raw_att, dict) and 'contentBytes' in raw_att:
                att['contentBytes'] = raw_att['contentBytes']
    
    if isinstance(email_data, dict) and 'error' in email_data:
        messages.warning(request, "Could not fetch live email content.")
        email_display = {'subject': delegation.email_id, 'body': {'content': 'Live content unavailable.'}}
    else:
        email_display = email_data

    context = {
        'delegation': delegation,
        'email': email_display,
        'attachments': attachments,
        'notes': delegation.notes.all().order_by('-created_at'),
        'target_email': target_email,
        'is_manager': is_manager,
    }
    return render(request, 'unity_internal_app/outlook_delegated_action.html', context)

@login_required
def outlook_delegate_to(request, email_id):
    """
    Handles delegation and re-delegation of Outlook tasks.
    UPDATED: All successful POST actions now redirect back to 'outlook_dashboard'.
    """
    from .models import EmailDelegation, DelegationTransactionLog
    
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    # Available users for assignment (excluding the current user performing the action)
    available_users = User.objects.filter(is_active=True).exclude(pk=request.user.pk)
    
    # 1. Fetch current delegation status if it exists
    current_delegation = EmailDelegation.objects.filter(email_id=email_id).first()

    if request.method == 'POST':
        work_related_input = request.POST.get('work_related')
        assignee_pk = request.POST.get('agent_name')
        
        data_for_delegation = {
            'company_code': request.POST.get('company_code'),
            'email_category': request.POST.get('email_category'),
            'work_related': True if work_related_input == 'Yes' else False,
            'comm_type': request.POST.get('comm_type') or 'Email',
        }

        # --- HANDLE ARCHIVE/RECYCLE ---
        if work_related_input == 'No':
            EmailDelegation.objects.update_or_create(
                email_id=email_id,
                defaults={
                    'work_related': False,
                    'status': 'DLT',  
                    'assigned_user': None,
                    'company_code': data_for_delegation['company_code'],
                    'email_category': data_for_delegation['email_category'],
                    'communication_type': data_for_delegation['comm_type'],
                }
            )
            messages.error(request, "Email moved to Recycle Bin (Status: DLT).")
            # REDIRECT: Back to Dashboard
            return redirect('outlook_dashboard')

        # --- HANDLE DELEGATION / RE-DELEGATION ---
        else:
            if not assignee_pk or assignee_pk in ['', '__Select Agent__']:
                messages.error(request, "Please select an agent for delegation.")
            else:
                # Fetch new assignee object
                assignee = User.objects.get(pk=assignee_pk)

                # Check if this is a re-assignment of an existing record
                if current_delegation:
                    old_agent_name = current_delegation.assigned_user.username if current_delegation.assigned_user else "Unassigned"
                    
                    # Update existing record fields
                    current_delegation.assigned_user = assignee
                    current_delegation.company_code = data_for_delegation['company_code']
                    current_delegation.email_category = data_for_delegation['email_category']
                    current_delegation.status = 'DEL'
                    current_delegation.delegated_at = timezone.now()
                    current_delegation.save()

                    # Audit Log for the Re-delegation
                    DelegationTransactionLog.objects.create(
                        delegation=current_delegation,
                        user=request.user,
                        action_type='RE_DELEGATED',
                        subject=f"Task moved from {old_agent_name} to {assignee.username}",
                        recipient_email=assignee.email or "N/A"
                    )
                    messages.success(request, f"Task successfully re-assigned to {assignee.username}.")
                    # REDIRECT: Back to Dashboard
                    return redirect('outlook_dashboard')

                else:
                    # Logic for NEW DELEGATION
                    success, message = delegate_email_task(
                        email_id, 
                        assignee_pk, 
                        request.user, 
                        classification_data=data_for_delegation
                    )
                    
                    if success:
                        # Auto-Reply Logic
                        reply_endpoint = f"messages/{email_id}/createReply"
                        reply_payload = {
                            "comment": f"Dear Sender,\n\nThis request has been successfully received and delegated to our agent: {assignee.username}.\n\nPlease use Reference: {data_for_delegation['company_code'] or 'N/A'} for future queries.\n\nRegards,\nMIP Support Team"
                        }
                        try:
                            draft_res = OutlookGraphService._make_graph_request(reply_endpoint, target_email, method='POST', data=reply_payload)
                            if 'id' in draft_res:
                                send_endpoint = f"messages/{draft_res['id']}/send"
                                OutlookGraphService._make_graph_request(send_endpoint, target_email, method='POST')
                                messages.success(request, f"Task assigned and confirmation email sent!")
                        except Exception as e:
                            messages.warning(request, f"Task assigned, but auto-reply failed: {str(e)}")
                        
                        # REDIRECT: Back to Dashboard
                        return redirect('outlook_dashboard')
                    else:
                        messages.error(request, message)

    # --- Fetch Data for GET Request ---
    endpoint = f"messages/{email_id}" 
    email_data = OutlookGraphService._make_graph_request(endpoint, target_email, method='GET') 

    if 'error' in email_data:
        messages.error(request, f"Error fetching email content: {email_data.get('error')}")
        # REDIRECT: Back to Dashboard on error
        return redirect('outlook_dashboard')

    # Fetch Attachments Metadata
    attachments_data = OutlookGraphService.fetch_attachments(target_email, email_id)
    
    for att in attachments_data:
        content_type = att.get('contentType', '').lower()
        if 'image' in content_type:
            raw_att = OutlookGraphService.get_attachment_raw(target_email, email_id, att['id'])
            if 'contentBytes' in raw_att:
                att['contentBytes'] = raw_att['contentBytes']
    
    raw_content = email_data.get('body', {}).get('content', '')
    received_date_str = email_data.get('receivedDateTime')
    
    if not current_delegation:
        current_delegation = get_or_create_delegation_status(email_id, received_date_str=received_date_str)
    
    context = {
        'email_id': email_id,
        'email_subject': email_data.get('subject', '(No Subject)'),
        'email_content': raw_content, 
        'attachments': attachments_data, 
        'available_users': available_users,
        'current_delegation': current_delegation,
    }
    return render(request, 'unity_internal_app/outlook_delegate_to.html', context)

def outlook_email_content(request, email_id):
    """
    Fetches the raw HTML content of an email and returns it as a response 
    to be loaded by an iframe's 'src' attribute.
    """
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    endpoint = f"messages/{email_id}" 
    
    # 🛑 FIX 11: Define 'method' and 'email_data' for this GET request 🛑
    method = 'GET'
    email_data = None
    
    email_data = OutlookGraphService._make_graph_request(endpoint, target_email, method=method, data=email_data)

    if 'error' in email_data:
        return HttpResponse("<h1>Error fetching email content.</h1>", status=500)

    body_data = email_data.get('body', {})
    
    if body_data.get('contentType', '').lower() == 'html':
        content = body_data.get('content', 'No HTML body found.')
    else:
        # If plain text, wrap in <pre> tags
        content = body_data.get('content', 'No email body found.')
        content = f'<pre style="white-space: pre-wrap; word-wrap: break-word; font-family: sans-serif;">{content}</pre>'

    # Wrap in a full HTML document (essential for iframe rendering)
    wrapped_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>body {{ font-family: sans-serif; margin: 15px; }}</style>
    </head>
    <body>
        {content}
    </body>
    </html>
    """
    
    # Return as plain HTML response (not marked safe, but the browser loads it securely)
    return HttpResponse(wrapped_content, content_type='text/html')

@login_required
def outlook_recycle_bin_view(request):
    """
    Displays all emails marked as DLT (Recycle Bin).
    Maps email_id to the actual subject stored in OutlookInbox for better readability.
    """
    # 1. Fetch delegations marked as DLT or not work-related
    recycled_tasks = EmailDelegation.objects.filter(status='DLT').order_by('-received_at')
    
    # 2. Extract email IDs to fetch subjects in bulk
    email_ids = [task.email_id for task in recycled_tasks]
    
    # 3. Create a map of {email_id: subject} from the OutlookInbox model
    inbox_details = OutlookInbox.objects.filter(email_id__in=email_ids).values('email_id', 'subject')
    subject_map = {item['email_id']: item['subject'] for item in inbox_details}
    
    # 4. Attach the subject to each task object
    for task in recycled_tasks:
        # Fallback to truncated ID if subject isn't found in local cache
        task.subject_display = subject_map.get(task.email_id, f"ID: {task.email_id[:15]}...")

    return render(request, 'unity_internal_app/recycle_bin.html', {
        'recycled_tasks': recycled_tasks
    })

@login_required
def outlook_restore_email(request, email_id):
    """Restores a task by setting work_related back to True."""
    task = get_object_or_404(EmailDelegation, email_id=email_id)
    task.work_related = True
    task.status = 'NEW'  
    task.save()
    
    DelegationNote.objects.create(
        delegation=task,
        user=request.user,
        content="Task restored from Recycle Bin."
    )
    
    messages.success(request, "Email restored to Live Inbox.")
    # FIX: Match the name in urls.py
    return redirect('outlook_recycle_bin')

@login_required
def outlook_delete_permanent(request):
    """
    Hides items from the Recycle Bin by moving them to ARC status.
    This avoids IntegrityErrors with foreign key notes while removing them from the UI.
    """
    if request.method == 'POST':
        email_ids = request.POST.getlist('email_ids')
        
        if email_ids:
            # Move from 'DLT' (Recycle Bin) to 'ARC' (Hidden/Archived)
            # This satisfies the DB constraint because the row still exists, 
            # but it will no longer show up in the filter(status='DLT') query.
            EmailDelegation.objects.filter(email_id__in=email_ids).update(status='ARC')
            
            messages.success(request, f"Successfully removed {len(email_ids)} items from the Recycle Bin.")
        else:
            messages.warning(request, "No items were selected for removal.")
            
        return redirect('outlook_recycle_bin')
    
@login_required
def view_email_thread(request, email_id):
    """
    Displays email content, attachments, and combined timeline of 
    automated transactions, delegation notes, and manual unity notes.
    """
    from .models import EmailDelegation, DelegationTransactionLog, UnityNotes, DelegationNote, OutlookInbox

    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    
    # 1. Fetch Live Content from Graph API
    endpoint = f"messages/{email_id}"
    email_data = OutlookGraphService._make_graph_request(endpoint, target_email, method='GET')
    
    # Ensure email_data is a dictionary for safe template lookup
    if not isinstance(email_data, dict):
        email_data = {}

    if 'error' in email_data:
        messages.error(request, f"Microsoft Graph Error: {email_data.get('error')}")
        return redirect(request.META.get('HTTP_REFERER', 'outlook_dashboard'))

    # 2. Handle Attachments
    attachments = OutlookGraphService.fetch_attachments(target_email, email_id)
    for att in attachments:
        if 'image' in att.get('contentType', '').lower():
            raw_att = OutlookGraphService.get_attachment_raw(target_email, email_id, att['id'])
            # Safeguard against raw_att not containing contentBytes
            if isinstance(raw_att, dict) and 'contentBytes' in raw_att:
                att['contentBytes'] = raw_att.get('contentBytes')

    email_body = email_data.get('body', {}).get('content', '')

    # 3. Fetch Local Data (Task, Transactions, and Manual Notes)
    # Using .first() prevents 404s on Direct Emails (Sent Items)
    task = EmailDelegation.objects.filter(email_id=email_id).first()
    
    actions = []
    delegation_notes = []
    if task:
        actions = list(DelegationTransactionLog.objects.filter(delegation=task).select_related('user'))
        delegation_notes = list(DelegationNote.objects.filter(delegation=task).select_related('user'))

    # 4. Fetch Manual Unity Notes associated with this Email ID
    manual_unity_notes = list(UnityNotes.objects.filter(attached_email_id=email_id))

    # 5. Build Combined Timeline
    combined_timeline = []

    # Map Action logs
    for act in actions:
        combined_timeline.append({'type': 'action', 'date': act.timestamp, 'obj': act})
    
    # Map Delegation/Internal notes
    for d_note in delegation_notes:
        combined_timeline.append({'type': 'del_note', 'date': d_note.created_at, 'obj': d_note})

    # Map Unity Management notes (Direct Emails & Manual Logs)
    for u_note in manual_unity_notes:
        combined_timeline.append({'type': 'unity_note', 'date': u_note.date, 'obj': u_note})

    # Sort timeline: Newest at the top
    combined_timeline.sort(key=lambda x: x['date'], reverse=True)

    # 6. Final Context Preparation
    context = {
        'task': task,
        'email': email_data,
        'email_body': email_body,
        'attachments': attachments,
        'combined_timeline': combined_timeline,
        # Try to find a local inbox item, otherwise use the task as fallback for metadata
        'inbox_item': OutlookInbox.objects.filter(email_id=email_id).first() or task, 
    }
    
    return render(request, 'unity_internal_app/view_email_thread.html', context)

@login_required
def email_list_view(request):
    """
    MASTER ARCHIVE: Displays 'NEW', 'DEL' (Delegated), and 'COM' (Completed) tasks.
    Supports filtering by Status, Agent, Type, Date, and Search Text.
    """
    # 1. Capture Filter Parameters
    filter_type = request.GET.get('type', 'all')
    search_query = request.GET.get('search', '').strip().lower()
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    
    # New Filter Parameters from HTML dropdowns
    status_filter = request.GET.get('status')
    agent_filter = request.GET.get('agent')
    type_filter = request.GET.get('email_type')
    
    # 2. Base Query: Start by excluding deleted/archived items
    delegations_qs = EmailDelegation.objects.exclude(
        status__in=['DLT', 'ARC']
    ).select_related('assigned_user').order_by('-received_at')

    # Apply Database-Level Filters
    if status_filter:
        delegations_qs = delegations_qs.filter(status=status_filter)
    else:
        # UPDATED: Default view now shows NEW and DEL (Delegated) 
        # to show active workload, or COM if explicitly requested via filter_type
        if filter_type == 'com':
            delegations_qs = delegations_qs.filter(status='COM')
        elif filter_type == 'new':
            delegations_qs = delegations_qs.filter(status='NEW')
        else:
            # Default: Show New and currently Delegated (Active Work)
            delegations_qs = delegations_qs.filter(status__in=['NEW', 'DEL'])

    if agent_filter:
        delegations_qs = delegations_qs.filter(assigned_user_id=agent_filter)
        
    if type_filter:
        delegations_qs = delegations_qs.filter(email_category=type_filter)

    # 3. Link to OutlookInbox for Subject and Sender Address
    email_ids = delegations_qs.values_list('email_id', flat=True)
    inbox_map = {obj.email_id: obj for obj in OutlookInbox.objects.filter(email_id__in=email_ids)}
    
    final_items = []
    for d in delegations_qs:
        inbox_detail = inbox_map.get(d.email_id)
        if inbox_detail:
            d.subject = inbox_detail.subject
            d.sender_address = inbox_detail.sender_address
        else:
            d.subject = f"[Details Missing: {d.email_id[:10]}]"
            d.sender_address = "Unknown"

        # 4. Apply Python-Side Filters (Search and Date)
        if search_query:
            if search_query not in d.subject.lower() and \
               search_query not in (d.company_code or '').lower() and \
               search_query not in (d.email_category or '').lower():
                continue

        if start_date and d.received_at:
            try:
                if d.received_at.date() < datetime.strptime(start_date, '%Y-%m-%d').date():
                    continue
            except ValueError: pass
            
        if end_date and d.received_at:
            try:
                if d.received_at.date() > datetime.strptime(end_date, '%Y-%m-%d').date():
                    continue
            except ValueError: pass

        final_items.append(d)

    # 5. Handle "Filter Category" backwards compatibility
    # (Already handled in Step 2 for query optimization, but keeping list logic safe)
    if filter_type == 'new':
        final_items = [item for item in final_items if item.status == 'NEW']
    elif filter_type == 'com':
        final_items = [item for item in final_items if item.status == 'COM']
    elif filter_type == 'delegated':
        final_items = [item for item in final_items if item.status == 'DEL']

    # 6. Pagination
    paginator = Paginator(final_items, 24)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 7. Provide context
    return render(request, 'unity_internal_app/email_list.html', {
        'page_obj': page_obj,
        'current_filter': filter_type,
        'search_query': search_query,
        'all_agents': User.objects.filter(is_active=True).order_by('username'),
        'status_filter': status_filter,
    })


from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
    
# ==============================================================================
# CORRECTED EXPORT FUNCTIONS
# ==============================================================================

# --- HULPFUNKSIE VIR TAK-NAME ---
def get_branch_map(claims_queryset):
    company_codes = claims_queryset.values_list('company_code', flat=True).distinct()
    return {
        item['a_company_code']: item['b_company_name'] 
        for item in UnityMgListing.objects.filter(
            a_company_code__in=company_codes
        ).values('a_company_code', 'b_company_name')
    }

# --- VERSLAG 1: TWO-POT CLAIMS INVOICE (Vir Cecile) ---
@login_required
def export_two_pot_invoice(request):
    """
    Report 1: Spesifieke faktuur-formaat vir Cecile.
    """
    query = request.GET.get('q')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    claims = UnityClaim.objects.filter(claim_type='Two Pot').order_by('claim_created_date')
    if query:
        claims = claims.filter(Q(id_number__icontains=query) | Q(member_surname__icontains=query))
    if start_date and end_date:
        claims = claims.filter(claim_created_date__range=[start_date, end_date])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Two-Pot Invoice"

    # Header Styl
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Kolomtitels (R1)
    headers = [
        "Date", "Initials", "Surname", "Member nr", "ID NUMBER", "Fund", 
        "Branch", "Admin From", "Qualified", "Date submitted online", 
        "Successful Loaded confirmation", "Admin Fee R33 + 15% Vat", "SUBMIT ONLINE"
    ]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, size=10)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    branch_map = get_branch_map(claims)

    for claim in claims:
        initials = "".join([n[0] for n in claim.member_name.split() if n]) if claim.member_name else ""
        is_paid = str(claim.claim_status).upper().strip() == "PAID"
        
        ws.append([
            claim.claim_created_date.strftime('%d/%m/%Y') if claim.claim_created_date else '',
            initials,
            claim.member_surname,
            claim.mip_number,
            claim.id_number,
            claim.company_code,
            branch_map.get(claim.company_code, ""),
            claim.agent or "",
            "yes" if is_paid else "no",
            claim.date_submitted.strftime('%d/%m/%Y') if claim.date_submitted else '',
            "YES" if is_paid else "NO",
            "R37.95",
            "SUBMIT ONLINE"
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Two_Pot_Invoice_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


# --- VERSLAG 2: TWO-POT FULL TRACKING (Geel Formaat) ---
@login_required
def export_two_pot_tracking(request):
    """
    Report 2: Volledige Tracking lys met Geel opskrifte en rooi teks vir 'NO'.
    """
    query = request.GET.get('q')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    claims_queryset = UnityClaim.objects.filter(claim_type='Two Pot').order_by('claim_created_date')

    if query:
        claims_queryset = claims_queryset.filter(
            Q(id_number__icontains=query) | Q(member_surname__icontains=query) | 
            Q(company_code__icontains=query) | Q(mip_number__icontains=query)
        )

    if start_date and end_date:
        claims_queryset = claims_queryset.filter(claim_created_date__range=[start_date, end_date])

    branch_map = get_branch_map(claims_queryset)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Two-Pot Tracking"

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    now = timezone.now()
    display_start = start_date if start_date else now.replace(day=1).strftime('%d.%m.%Y')
    display_end = end_date if end_date else now.strftime('%d.%m.%Y')
    
    # Row 4 Title
    ws.merge_cells('A4:O4')
    header_cell = ws['A4']
    header_cell.value = f"Billing - Member Emergency Savings Pot Withdrawal Requested - {display_start} to {display_end}"
    header_cell.font = Font(bold=True, size=11, underline="single")
    header_cell.fill = yellow_fill
    header_cell.border = thin_border

    # Row 5 Headers
    headers = [
        "DATE EXTRACT INFO / FORM FROM WEB", "Initials", "Surname", 
        "Member number", "ID NUMBER", "Fund", "Branch", "Query", "Claim", 
        "Qualified", "Date submitted/ online", "Succesfull Loaded confirm", 
        "Amount Apply for", "Admin Fee R33+15%", "Note"
    ]
    
    for _ in range(3): ws.append([]) # Empty rows to reach R5
    ws.append(headers)
    
    for cell in ws[5]:
        cell.font = Font(bold=True, size=9)
        cell.fill = yellow_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    ws.row_dimensions[5].height = 50

    for claim in claims_queryset:
        initials = "".join([n[0] for n in claim.member_name.split() if n]) if claim.member_name else ""
        is_paid = str(claim.claim_status).upper().strip() == "PAID"
        qualified_val = "YES" if is_paid else "NO"
        
        if is_paid:
            submit_date_label = claim.date_submitted.strftime('%d.%m.%Y') if claim.date_submitted else "Pending"
        else:
            submit_date_label = "Withdrawal Not Allowed"

        row = [
            claim.claim_created_date.strftime('%d/%m/%Y') if claim.claim_created_date else '',
            initials,
            claim.member_surname,
            claim.mip_number,
            claim.id_number,
            claim.company_code,
            branch_map.get(claim.company_code, "Unknown"),
            "Savings Form Request",
            "Savings Form Submitted" if is_paid else "Member Emergency Savings Pot Withdrawal Requested",
            qualified_val,
            submit_date_label,
            "YES" if is_paid else "",
            float(claim.claim_amount or 0),
            "37.95",
            claim.notes.last().note_description if claim.notes.exists() else ""
        ]
        ws.append(row)

        # Rooi teks vir "NO" status
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', horizontal='left')
            cell.font = Font(size=9, color="FF0000") if qualified_val == "NO" else Font(size=9)

    widths = [22, 8, 18, 14, 18, 10, 25, 20, 35, 10, 22, 18, 14, 14, 40]
    for i, width in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Two_Pot_Full_Tracking_{now.strftime("%Y_%m_%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def export_global_claims_excel(request):
    """
    STRICT: Excludes all 'Two Pot' claims.
    Format: Green Audit Spreadsheet matching the database-style layout.
    """
    query = request.GET.get('q')
    
    # 1. Filter: Exclude 'Two Pot' to keep this as a standard audit export
    claims_queryset = UnityClaim.objects.exclude(claim_type='Two Pot').order_by('-claim_created_date')

    if query:
        claims_queryset = claims_queryset.filter(
            Q(id_number__icontains=query) | 
            Q(member_surname__icontains=query) | 
            Q(company_code__icontains=query) |
            Q(mip_number__icontains=query)
        )

    # 2. Map Branch Names
    company_codes = claims_queryset.values_list('company_code', flat=True).distinct()
    mg_map = {
        item['a_company_code']: item['b_company_name'] 
        for item in UnityMgListing.objects.filter(a_company_code__in=company_codes).values('a_company_code', 'b_company_name')
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Withdrawal Audit"

    # --- Styles Definition ---
    # Light green fill matching the dashboard style
    green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True, size=11)

    # 3. Headers (Row 1) - Standard Audit Layout
    headers = [
        "Co Code", "Branch", "Agent", "MIP Number", "ID Number", 
        "Name", "Surname", "Type", "Status", "Exit Reason", 
        "Created", "Submitted", "Paid"
    ]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = green_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 4. Add Data
    for claim in claims_queryset:
        branch_name = mg_map.get(claim.company_code, "Unknown Group")
        
        # Safe Date Formatting to prevent NoneType errors
        created = claim.claim_created_date.strftime('%Y-%m-%d') if claim.claim_created_date else ''
        submitted = claim.date_submitted.strftime('%Y-%m-%d') if claim.date_submitted else ''
        paid = claim.date_paid.strftime('%Y-%m-%d') if claim.date_paid else ''

        row_data = [
            claim.company_code,
            branch_name,
            claim.agent if claim.agent else '',
            claim.mip_number if claim.mip_number else '',
            claim.id_number,
            claim.member_name,
            claim.member_surname,
            claim.claim_type,
            claim.claim_status,
            claim.exit_reason if claim.exit_reason else '',
            created,
            submitted,
            paid
        ]
        ws.append(row_data)

        # Apply borders to data rows
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # 5. Set column widths for professional spacing
    column_widths = [12, 35, 20, 15, 20, 20, 20, 15, 15, 20, 15, 15, 15]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = width

    # 6. Generate Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Global_Claims_Audit_Export.xlsx"'
    wb.save(response)
    return response

@login_required
def manager_approval_dashboard(request):
    """
    Lists all Credit Notes waiting for Linking Approval.
    """
    pending_credits = CreditNote.objects.filter(credit_link_status='Pending').order_by('-processed_date')
    
    context = {
        'pending_credits': pending_credits
    }
    return render(request, 'unity_internal_app/manager_approval_dashboard.html', context)

@login_required
@transaction.atomic
def approve_credit_link(request, note_id):
    """
    Manager Approval Logic.
    UPDATED: Handles "Overs credit line" as a fail-safe by marking as 'Approved'.
    This ensures the credit appears in the Bank List for the Clerk.
    """
    from decimal import Decimal
    from django.db.models import Sum
    from django.utils import timezone
    from django.contrib import messages
    from django.shortcuts import redirect, get_object_or_404
    from .models import CreditNote, BillSettlement

    note = get_object_or_404(CreditNote, id=note_id)
    ZERO_DECIMAL = Decimal('0.00')
    
    if note.credit_link_status != 'Pending':
        messages.error(request, "This credit is not pending approval.")
        return redirect('manager_approval_dashboard')

    # --- FAIL-SAFE LOGIC: HANDLE "OVERS CREDIT LINE" VERIFICATION ---
    # If the reason is 'Overs credit line', we change status to 'Approved'.
    # This acts as the trigger for bank_list.html to display this as a virtual deposit.
    if note.link_request_reason == "Overs credit line":
        note.credit_link_status = 'Approved' # <--- Changed from Unlinked
        note.pending_linked_bill = None
        note.authorized_by = request.user.username
        note.authorized_at = timezone.now()
        note.save()
        
        messages.success(request, f"Overs credit line of R{note.schedule_amount} for {note.member_group_code} approved. It is now visible in the Bank List for the Clerk.")
        return redirect('manager_approval_dashboard')

    # --- STANDARD LOGIC: HANDLE SPECIFIC BILL ALLOCATION REQUESTS ---
    target_bill = note.pending_linked_bill
    if not target_bill:
        note.credit_link_status = 'Unlinked'
        note.save()
        messages.warning(request, "Request reset: No target bill was attached to this link request.")
        return redirect('manager_approval_dashboard')

    # 1. Calculate Cap to prevent over-settling the bill
    total_already_settled = BillSettlement.objects.filter(
        unity_bill_source_id=target_bill.pk
    ).aggregate(total=Sum('settled_amount'))['total'] or ZERO_DECIMAL
    
    bill_remaining_debt = target_bill.H_Schedule_Amount - total_already_settled
    
    # Use the amount the agent requested
    current_temp_request = note.requested_amount or ZERO_DECIMAL
    final_allocation = min(current_temp_request, bill_remaining_debt)

    if final_allocation <= ZERO_DECIMAL:
        messages.warning(request, f"Bill #{target_bill.id} is already fully paid. Credit reset to Unlinked.")
        note.credit_link_status = 'Unlinked'
        note.pending_linked_bill = None
        note.requested_amount = ZERO_DECIMAL 
        note.save()
        return redirect('manager_approval_dashboard')

    # 2. Update Balances
    note.schedule_amount -= final_allocation 
    note.requested_amount = final_allocation 

    # If money remains on the credit note, it goes back to 'Unlinked' (Fail-safe for remainders)
    if note.schedule_amount <= Decimal('0.009'):
        note.assigned_unity_bill = target_bill
        note.credit_link_status = 'Approved' 
    else:
        note.assigned_unity_bill = None
        note.credit_link_status = 'Unlinked' 

    note.pending_linked_bill = None
    note.authorized_by = request.user.username
    note.authorized_at = timezone.now()
    note.save()

    # 3. Create Audit Trail (Allocates the money to the bill)
    BillSettlement.objects.create(
        unity_bill_source=target_bill,
        settled_amount=final_allocation,
        settlement_date=timezone.now(),
        source_credit_note_id=note.id,
        reconned_bank_line=None
    )

    # NOTE: We specifically DO NOT set target_bill.is_reconciled = True here.
    # The Clerk must still go to the summary page and click "Finalize".

    messages.success(request, f"Approved R{final_allocation} for Bill #{target_bill.id}. Clerk can now finalize this bill.")
    return redirect('manager_approval_dashboard')

@login_required
@transaction.atomic
def reject_credit_link(request, note_id):
    from django.contrib import messages
    from django.shortcuts import redirect, get_object_or_404
    from .models import CreditNote

    note = get_object_or_404(CreditNote, id=note_id)
    
    if note.credit_link_status != 'Pending':
        messages.error(request, "This credit is not in a pending state.")
        return redirect('manager_approval_dashboard')

    # --- UPDATED LOGIC ---
    # If the manager rejects it, we move it back to 'Unlinked'
    # This allows the clerk to try the request again or fix a mistake.
    note.credit_link_status = 'Unlinked'
    note.pending_linked_bill = None
    note.requested_amount = 0  
    
    # Track who rejected it for the audit trail
    note.review_note = f"Rejected by {request.user.username} on {timezone.now().date()}"
    note.save()

    messages.warning(request, f"Credit ID {note_id} was rejected and returned to the unlinked pool.")
    return redirect('manager_approval_dashboard')

@login_required
def global_bank_view(request):
    from django.db.models import F, Q, Sum
    from decimal import Decimal
    from django.core.paginator import Paginator
    # Import the necessary models for calculation
    from .models import (
        ReconnedBank, CreditNote, BillSettlement, UnityMgListing
    )

    query = request.GET.get('q')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # --- 1. Fetch Physical Bank Lines ---
    bank_qs = ReconnedBank.objects.select_related('bank_line').order_by('-transaction_date', '-id')
    
    if start_date and end_date:
        bank_qs = bank_qs.filter(transaction_date__range=[start_date, end_date])
    
    if query:
        # FIXED: Changed to lowercase fields to prevent "Unsupported lookup" FieldError
        bank_qs = bank_qs.filter(
            Q(company_code__icontains=query) |
            Q(bank_line__transaction_description__icontains=query) | 
            Q(bank_line__statement_reference__icontains=query)
        )

    # --- 2. Fetch Approved Virtual Credits (The Overs) ---
    credit_qs = CreditNote.objects.filter(
        Q(note_selection='OVERS') | Q(link_request_reason='Overs credit line'),
        credit_link_status='Approved',
        schedule_amount__gt=Decimal('0.00')
    )
    if start_date and end_date:
        credit_qs = credit_qs.filter(authorized_at__date__range=[start_date, end_date])
    if query:
        credit_qs = credit_qs.filter(Q(member_group_code__icontains=query))

    # --- 3. Pre-Calculate Bill Settlements for Bank Lines ---
    bank_line_ids = list(bank_qs.values_list('id', flat=True))
    
    # Map of actual bill usage (consumption) per bank line
    bill_usage_map = {
        item['reconned_bank_line_id']: item['total']
        for item in BillSettlement.objects.filter(reconned_bank_line_id__in=bank_line_ids)
        .values('reconned_bank_line_id')
        .annotate(total=Sum('settled_amount'))
    }

    # Map of overs/credits created from these bank lines
    overs_usage_map = {
        item['source_bank_line_id']: item['total']
        for item in CreditNote.objects.filter(source_bank_line_id__in=bank_line_ids, note_selection='OVERS')
        .values('source_bank_line_id')
        .annotate(total=Sum('schedule_amount'))
    }

    # --- 4. Build MG Map for Names and Agents ---
    all_codes = list(bank_qs.values_list('company_code', flat=True).distinct()) + \
                list(credit_qs.values_list('member_group_code', flat=True).distinct())
    
    mg_map = {
        item['a_company_code']: {'name': item['b_company_name'], 'agent': item['c_agent']}
        for item in UnityMgListing.objects.filter(a_company_code__in=all_codes).values('a_company_code', 'b_company_name', 'c_agent')
    }

    # --- 5. Unify the Data into the combined_list ---
    combined_list = []

    # Map Bank Lines
    for r in bank_qs:
        mg_info = mg_map.get(r.company_code, {})
        
        # Get actual bill-only usage (e.g., R 1200)
        bill_usage = bill_usage_map.get(r.id, Decimal('0.00'))
        credit_overs = overs_usage_map.get(r.id, Decimal('0.00'))

        raw_description = "No description found"
        if r.bank_line:
            # Handle potential casing variations in the object attributes
            raw_description = getattr(r.bank_line, 'transaction_description', 
                                     getattr(r.bank_line, 'Transaction_description', "No description found"))

        combined_list.append({
            'id': r.id,
            'source': 'BANK',
            'transaction_date': r.transaction_date,
            'amount': r.transaction_amount,
            # FIXED: "Used Amount" now shows only the bill consumption
            'settled_amount': bill_usage, 
            'credit_amount': credit_overs,
            # FIXED: "Remaining" now shows the surplus (amount - consumption)
            'remaining': r.transaction_amount - bill_usage,
            'status': r.recon_status or "Unidentified",
            'company_code': r.company_code or "—",
            'company_name': mg_info.get('name', "Unassigned") if r.company_code else "—",
            'review_note': r.review_note if r.review_note else "Bank Statement",
            'agent': mg_info.get('agent', "System") if r.company_code else "—",
            'transaction_description': raw_description,
        })

    # Map Approved Credits
    for c in credit_qs:
        combined_list.append({
            'id': c.id,
            'source': 'CREDIT',
            'transaction_date': c.authorized_at.date() if c.authorized_at else c.processed_date,
            'amount': c.schedule_amount,
            'settled_amount': Decimal('0.00'), 
            'credit_amount': Decimal('0.00'),
            'remaining': c.schedule_amount,
            'status': "APPROVED VIRTUAL",
            'company_code': c.member_group_code,
            'company_name': c.member_group_name or mg_map.get(c.member_group_code, {}).get('name', "Verified Credit"),
            'review_note': "APPROVED OVERS",
            'agent': c.authorized_by or "Manager",
            'transaction_description': f"Virtual Credit - {c.note_selection}",
        })

    # Sort combined list by date
    combined_list.sort(key=lambda x: x['transaction_date'], reverse=True)

    # Pagination
    paginator = Paginator(combined_list, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'global_records': page_obj, 
        'search_query': query,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'unity_internal_app/global_bank.html', context)

@login_required
def export_global_bank_excel(request):
    """
    Exports the Global Bank history to Excel.
    UPDATED: Matches dashboard logic for Correct Settled vs Remaining math.
    UPDATED: Splits review_note into 'Note Category' and 'Internal Note'.
    """
    import openpyxl
    from django.http import HttpResponse
    from django.db.models import Q, Sum
    from decimal import Decimal
    from .models import ReconnedBank, CreditNote, BillSettlement, UnityMgListing
    
    query = request.GET.get('q')
    
    # --- 1. Fetch Physical Bank Lines ---
    bank_records = ReconnedBank.objects.select_related('bank_line').order_by('-transaction_date', '-id')

    if query:
        bank_records = bank_records.filter(
            Q(company_code__icontains=query) |
            Q(bank_line__Transaction_description__icontains=query) |
            Q(bank_line__Statement_reference__icontains=query) |
            Q(recon_status__icontains=query) |
            Q(review_note__icontains=query)
        )

    # --- 2. Fetch Approved Virtual Credits (Overs) ---
    credit_records = CreditNote.objects.filter(
        Q(note_selection='OVERS') | Q(link_request_reason='Overs credit line'),
        credit_link_status='Approved',
        schedule_amount__gt=Decimal('0.00')
    )

    if query:
        credit_records = credit_records.filter(
            Q(member_group_code__icontains=query) |
            Q(member_group_name__icontains=query)
        )

    # --- 3. Pre-Calculate Bill Settlements for Bank Lines ---
    # This is the critical step to ensure math (Deposit - Settled = Remaining) is correct
    bank_line_ids = list(bank_records.values_list('id', flat=True))
    bill_usage_map = {
        item['reconned_bank_line_id']: item['total']
        for item in BillSettlement.objects.filter(reconned_bank_line_id__in=bank_line_ids)
        .values('reconned_bank_line_id')
        .annotate(total=Sum('settled_amount'))
    }

    # --- 4. Build Company Lookup Map ---
    bank_codes = list(bank_records.values_list('company_code', flat=True).distinct())
    credit_codes = list(credit_records.values_list('member_group_code', flat=True).distinct())
    all_codes = set(bank_codes + credit_codes)

    mg_map = {
        item['a_company_code']: {
            'name': item['b_company_name'], 
            'agent': item['c_agent']
        }
        for item in UnityMgListing.objects.filter(a_company_code__in=all_codes).values('a_company_code', 'b_company_name', 'c_agent')
    }

    # --- 5. Normalize and Merge Data ---
    export_rows = []

    # Process Bank Lines
    for r in bank_records:
        mg_info = mg_map.get(r.company_code, {})
        
        # Calculate actual bill usage vs remaining (Overs)
        bill_usage = bill_usage_map.get(r.id, Decimal('0.00'))
        
        # Robust Description handling
        raw_description = "No description found"
        if r.bank_line:
            raw_description = getattr(r.bank_line, 'Transaction_description', 
                                     getattr(r.bank_line, 'transaction_description', "No description found"))

        # Unpack notes
        note_category = ""
        internal_note_text = ""
        if r.review_note:
            if " | " in r.review_note:
                parts = r.review_note.split(" | ", 1)
                note_category = parts[0]
                internal_note_text = parts[1]
            else:
                note_category = r.review_note

        export_rows.append({
            'date': r.transaction_date,
            'description': raw_description,
            'code': r.company_code or "Unassigned",
            'name': mg_info.get('name', "—"),
            'deposit': r.transaction_amount,
            'settled': bill_usage, # The R 1200
            'remaining': r.transaction_amount - bill_usage, # The R 300
            'status': r.recon_status,
            'agent': mg_info.get('agent', "System"),
            'note_category': note_category,
            'internal_note': internal_note_text
        })

    # Process Credit Notes (Overs)
    for c in credit_records:
        c_name = c.member_group_name or mg_map.get(c.member_group_code, {}).get('name', "Verified Credit")
        c_date = c.authorized_at.date() if c.authorized_at else c.processed_date

        export_rows.append({
            'date': c_date,
            'description': f"APPROVED OVERS (Virtual Credit: {c.note_selection})",
            'code': c.member_group_code,
            'name': c_name,
            'deposit': c.schedule_amount,
            'settled': Decimal('0.00'),
            'remaining': c.schedule_amount,
            'status': "APPROVED VIRTUAL",
            'agent': c.authorized_by or "Manager",
            'note_category': "OVERS",
            'internal_note': "Virtual Credit Line created from surplus"
        })

    # Sort Combined List by Date
    export_rows.sort(key=lambda x: x['date'], reverse=True)

    # --- 6. Write to Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Global Bank Export"

    # HEADERS
    headers = [
        "Date", "Bank Description", "Company Code", "Company Name", 
        "Deposit Amount", "Settled (Bill Usage)", "Remaining (Surplus/Overs)", 
        "Status", "Agent", "Note Category", "Internal Note"
    ]
    ws.append(headers)

    # Bold headers
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    # Append rows
    for row in export_rows:
        ws.append([
            row['date'], row['description'], row['code'], row['name'],
            row['deposit'], row['settled'], row['remaining'],
            row['status'], row['agent'], 
            row['note_category'], row['internal_note']
        ])

    # Auto-adjust column widths for better look
    col_widths = {
        'A': 15, 'B': 45, 'C': 15, 'D': 30, 'E': 18, 
        'F': 18, 'G': 18, 'H': 15, 'I': 15, 'J': 25, 'K': 60
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Global_Bank_Export.xlsx"'
    wb.save(response)
    return response

@login_required
def download_attachment_view(request, message_id, attachment_id):
    """
    Acts as a proxy to download files from Microsoft Graph.
    Handles both FileAttachments (PDF, Excel, etc.) and ItemAttachments (nested emails).
    """
    target_email = settings.OUTLOOK_EMAIL_ADDRESS
    
    # 1. Fetch the metadata first to determine the attachment type
    attachment_data = OutlookGraphService.get_attachment_raw(target_email, message_id, attachment_id)
    
    if isinstance(attachment_data, dict) and 'error' in attachment_data:
        logger.error(f"Metadata retrieval failed for {attachment_id}: {attachment_data}")
        return HttpResponse("Error retrieving attachment info from Microsoft.", status=400)

    odata_type = attachment_data.get('@odata.type', '')
    file_name = attachment_data.get('name', 'attachment')

    # 2. Extract content based on type
    if odata_type == '#microsoft.graph.fileAttachment':
        # Standard file: content is already in the metadata as a base64 string
        if 'contentBytes' in attachment_data:
            file_content = base64.b64decode(attachment_data['contentBytes'])
            content_type = attachment_data.get('contentType', 'application/octet-stream')
        else:
            return HttpResponse("Attachment content is missing from metadata.", status=404)

    elif odata_type == '#microsoft.graph.itemAttachment':
        # Nested Email: must be fetched as raw MIME via the /$value endpoint
        file_content = OutlookGraphService.get_attachment_mime(target_email, message_id, attachment_id)
        
        # Check if binary content was returned or an error dict
        if not file_content or (isinstance(file_content, dict) and 'error' in file_content):
            return HttpResponse("Could not retrieve nested email content from Microsoft.", status=400)
            
        content_type = 'message/rfc822'
        # Force .eml extension so Windows handles the file correctly
        if not file_name.lower().endswith('.eml'):
            file_name += ".eml"
    else:
        return HttpResponse(f"Unsupported attachment type: {odata_type}", status=400)

    # 3. Build and return the final response
    response = HttpResponse(file_content, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    return response

@login_required
def export_email_list(request):
    """
    Exports the Email Archive to Excel.
    UPDATED: Now includes 'NEW' and 'DEL' (Delegated) by default.
    COLUMNS: Status, Subject, From, Received, Delegated, Agent, Company, Type.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    # 1. Capture Filter Parameters
    filter_type = request.GET.get('type', 'all')
    search_query = request.GET.get('search', '').strip().lower()
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    
    status_filter = request.GET.get('status')
    agent_filter = request.GET.get('agent')
    type_filter = request.GET.get('email_type')

    # 2. Fetch Data with Database-Level Filtering
    # Exclude Recycle Bin/Archived items
    delegations_qs = EmailDelegation.objects.exclude(status__in=['DLT', 'ARC']).select_related('assigned_user')

    # Apply Dropdown Filters at DB Level
    if status_filter:
        delegations_qs = delegations_qs.filter(status=status_filter)
    else:
        # UPDATED: Default restricted to NEW and DEL (Delegated) as requested
        delegations_qs = delegations_qs.filter(status__in=['NEW', 'DEL'])

    if agent_filter:
        delegations_qs = delegations_qs.filter(assigned_user_id=agent_filter)
        
    if type_filter:
        delegations_qs = delegations_qs.filter(email_category=type_filter)

    # 3. Map Inbox details (Subject and From)
    email_ids = delegations_qs.values_list('email_id', flat=True)
    inbox_map = {obj.email_id: obj for obj in OutlookInbox.objects.filter(email_id__in=email_ids)}

    final_items = []
    for d in delegations_qs:
        inbox_detail = inbox_map.get(d.email_id)
        d.subject = inbox_detail.subject if inbox_detail else "Unknown Subject"
        # ADDED: Sender Address field
        d.sender_address = inbox_detail.sender_address if inbox_detail else "Unknown"

        # Apply Python-Side Search Filter
        if search_query:
            if search_query not in d.subject.lower() and \
               search_query not in (d.company_code or '').lower() and \
               search_query not in (d.email_category or '').lower():
                continue

        # Apply Date Filters
        if start_date and d.received_at:
            try:
                if d.received_at.date() < datetime.strptime(start_date, '%Y-%m-%d').date():
                    continue
            except ValueError: pass
        if end_date and d.received_at:
            try:
                if d.received_at.date() > datetime.strptime(end_date, '%Y-%m-%d').date():
                    continue
            except ValueError: pass
        
        # Backward compatibility check (ensures we don't accidentally export COM if filter_type is 'new')
        if filter_type == 'new' and d.status != 'NEW':
            continue

        final_items.append(d)

    # Sort final list
    final_items.sort(key=lambda x: x.received_at if x.received_at else timezone.now(), reverse=True)

    # 4. Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Active Emails"

    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E5077", end_color="2E5077", fill_type="solid") # Changed to a professional blue
    
    # UPDATED: Columns as requested
    headers = [
        'Email Status', 'Subject', 'From', 'Date Received', 
        'Date Delegation', 'Agent', 'Company Code', 'Email Type'
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 5. Populate Rows
    for item in final_items:
        status_display = item.get_status_display()
        subject = item.subject
        sender = item.sender_address
        received_dt = item.received_at.strftime('%Y-%m-%d %H:%M') if item.received_at else ""
        delegated_dt = item.delegated_at.strftime('%Y-%m-%d %H:%M') if item.delegated_at else "---"
        user = item.assigned_user.username if item.assigned_user else "Unassigned"
        comp_code = item.company_code or "N/A"
        category = item.email_category or "General"

        ws.append([
            status_display, subject, sender, received_dt, 
            delegated_dt, user, comp_code, category
        ])

    # Adjust column widths for better readability
    # Status, Subject, From, RecDate, DelDate, Agent, CompCode, Type
    column_widths = [18, 55, 30, 20, 20, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 6. Return File
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Active_Emails_{timezone.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response

@login_required
def download_email_file(request, email_id):
    """
    Fetches the raw MIME (.eml) content from Microsoft Graph for a specific message.
    """
    try:
        # Use your existing OutlookGraphService to fetch raw content
        # The /$value endpoint returns the raw MIME content of the message
        endpoint = f"messages/{email_id}/$value"
        
        # Ensure your service has a method to handle raw/binary responses
        # If not, you can use a direct call or I can help you update the service
        raw_content = OutlookGraphService._make_graph_request(
            endpoint, 
            settings.OUTLOOK_EMAIL_ADDRESS, 
            method='GET', 
            is_raw=True # This tells the service NOT to parse as JSON
        )

        if not raw_content:
            messages.error(request, "Could not retrieve email content from Outlook.")
            return redirect(request.META.get('HTTP_REFERER', 'unity_list'))

        # Prepare the response as a file download
        response = HttpResponse(raw_content, content_type='message/rfc822')
        response['Content-Disposition'] = f'attachment; filename="Outlook_Email_{email_id[:10]}.eml"'
        return response

    except Exception as e:
        messages.error(request, f"Error downloading email: {str(e)}")
        return redirect(request.META.get('HTTP_REFERER', 'unity_list'))
    
@login_required
@transaction.atomic
def create_manual_credit(request):
    """
    Handles manual 'Overs' credit creation from the modal.
    Sets status to 'Pending' for Manager Approval.
    """
    if request.method == 'POST':
        # Capture form data from the modal
        company_code = request.POST.get('company_code')
        amount = request.POST.get('deposit_amount')
        deposit_date = request.POST.get('deposit_date')
        bank_fiscal = request.POST.get('bank_fiscal')
        date_identified = request.POST.get('date_identified')
        agent_input = request.POST.get('agent_name')

        if not all([company_code, amount, deposit_date]):
            messages.error(request, "Please provide Company Code, Amount, and Deposit Date.")
            return redirect('credit_note_list')

        # Create the record in credit_note table
        new_credit = CreditNote.objects.create(
            member_group_code=company_code,
            schedule_amount=Decimal(amount),          # Using schedule_amount as the balance
            bank_stmt_date=parse_date(deposit_date),  # Maps to 'Deposit Date'
            fiscal_date=parse_date(bank_fiscal),      # Maps to 'Bank Fiscal'
            date_identified=parse_date(date_identified),
            processed_by=agent_input or request.user.username,
            processed_date=timezone.now(),
            note_selection='OVERS',                   # Triggers yellow badge styling
            link_request_reason='Overs credit line',  # Specific flag for your fail-safe logic
            credit_link_status='Pending'              # Sends it to Manager Dashboard
        )

        # Manager Notification Logic
        # (This triggers the auto-notify flow for manager approval)
        try:
            subject = f"APPROVAL REQUIRED: Manual Overs Credit - {company_code}"
            body = f"A manual overs credit of R{amount} has been created for {company_code}."
            OutlookGraphService.send_outlook_email(settings.OUTLOOK_EMAIL_ADDRESS, "omega@example.com", subject, body, 'TEXT')
        except Exception:
            pass 

        messages.success(request, f"Manual credit for {company_code} created. Awaiting manager approval.")
        return redirect('credit_note_list')

    return redirect('credit_note_list')

@login_required
def export_unity_list_excel(request):
    """
    Exports the combined Unity List (InternalFunds + UnityMgListing) to Excel.
    Includes the calculated 'Active Surplus'.
    """
    from collections import defaultdict
    from decimal import Decimal
    import openpyxl
    from django.http import HttpResponse
    
    # --- REPLICATE DATA GATHERING LOGIC FROM unity_list ---
    
    # 1. Fetch Base Records
    internal_funds_records = InternalFunds.objects.all()
    
    # Create a mutable copy of the Unity listing map
    unity_listing_map = {
        record.a_company_code: record for record in UnityMgListing.objects.all()
    }
    
    # 2. Calculate Surplus/Allocation (Manual Calculation)
    bill_map = dict(UnityBill.objects.values_list('id', 'C_Company_Code'))
    ZERO_DECIMAL = Decimal('0.00')

    # Aggregate Total Surplus
    surplus_map = defaultdict(Decimal)
    surpluses = ScheduleSurplus.objects.values('unity_bill_source_id', 'surplus_amount')
    for s in surpluses:
        b_id = s['unity_bill_source_id']
        amount = s['surplus_amount'] or ZERO_DECIMAL
        if b_id in bill_map:
            surplus_map[bill_map[b_id]] += amount

    # Aggregate Total Allocation
    allocation_map = defaultdict(Decimal)
    allocations = JournalEntry.objects.values('target_bill_id', 'amount')
    for a in allocations:
        b_id = a['target_bill_id']
        amount = a['amount'] or ZERO_DECIMAL
        if b_id in bill_map:
            allocation_map[bill_map[b_id]] += amount

    # 3. Build Combined List
    combined_records = []

    # Phase 1: InternalFunds
    for fund_record in internal_funds_records:
        company_code = fund_record.A_Company_Code
        detail_record = unity_listing_map.pop(company_code, None)
        
        total_gained = surplus_map.get(company_code, ZERO_DECIMAL)
        total_used = allocation_map.get(company_code, ZERO_DECIMAL)
        active_surplus_value = total_gained - total_used
        
        combined_records.append({
            'code': fund_record.A_Company_Code,
            'name': fund_record.B_Company_Name,
            'source': fund_record.Source,
            'status': fund_record.D_Company_Status,
            'agent': detail_record.c_agent if detail_record else None,
            'payment': detail_record.e_payment_method if detail_record else None,
            'billing': detail_record.f_billing_method if detail_record else None,
            'fiscal': detail_record.g_current_fiscal if detail_record else None,
            'current_status': detail_record.h_current_status if detail_record else None,
            'last_recon': detail_record.i_last_recon if detail_record else None,
            'arrears': detail_record.j_arrears if detail_record else None,
            'email': detail_record.contact_email if detail_record else None,
            'surplus': active_surplus_value,
        })

    # Phase 2: Remaining UnityMgListing
    for company_code, detail_record in unity_listing_map.items():
        total_gained = surplus_map.get(company_code, ZERO_DECIMAL)
        total_used = allocation_map.get(company_code, ZERO_DECIMAL)
        active_surplus_value = total_gained - total_used
        
        combined_records.append({
            'code': detail_record.a_company_code,
            'name': detail_record.b_company_name,
            'source': 'System Only (New)',
            'status': detail_record.d_company_status,
            'agent': detail_record.c_agent,
            'payment': detail_record.e_payment_method,
            'billing': detail_record.f_billing_method,
            'fiscal': detail_record.g_current_fiscal,
            'current_status': detail_record.h_current_status,
            'last_recon': detail_record.i_last_recon,
            'arrears': detail_record.j_arrears,
            'email': detail_record.contact_email,
            'surplus': active_surplus_value,
        })

    # --- GENERATE EXCEL ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unity List Export"

    # Define Headers
    headers = [
        "Company Code", "Company Name", "Source", "Company Status", 
        "Agent", "Payment Method", "Billing Method", "Fiscal Year", 
        "Current Status", "Last Recon Note", "Arrears", "Contact Email", "Active Surplus"
    ]
    ws.append(headers)

    # Write Data
    for r in combined_records:
        ws.append([
            r['code'],
            r['name'],
            r['source'],
            r['status'],
            r['agent'],
            r['payment'],
            r['billing'],
            r['fiscal'],
            r['current_status'],
            r['last_recon'],
            r['arrears'],
            r['email'],
            r['surplus']
        ])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Unity_List_Export.xlsx"'
    wb.save(response)
    return response

@login_required
def apply_available_credit(request, bill_id):
    if request.method == 'POST':
        # 1. Get the Bill
        bill = get_object_or_404(UnityBill, pk=bill_id)
        
        # 2. Get Data from Form
        # 'credit_id' is the primary key of the CreditNote record
        credit_id = request.POST.get('credit_id') 
        amount_to_apply_str = request.POST.get('amount_to_apply')
        
        try:
            amount_to_apply = Decimal(amount_to_apply_str)
            
            # --- CRITICAL FIX: Use CreditNote, not ScheduleSurplus ---
            credit_item = get_object_or_404(CreditNote, pk=credit_id)
            
            # 3. Validation
            if amount_to_apply <= 0:
                messages.error(request, "Amount must be greater than 0.")
                return redirect('pre_bill_reconciliation_summary', company_code=bill.C_Company_Code, bill_id=bill.id)

            # Check available balance (using schedule_amount from CreditNote model)
            if amount_to_apply > credit_item.schedule_amount: 
                messages.error(request, "Cannot apply more than the available credit amount.")
                return redirect('pre_bill_reconciliation_summary', company_code=bill.C_Company_Code, bill_id=bill.id)

            # 4. Create the Allocation
            with transaction.atomic():
                # Note: We use 'source_credit_note_id' to match your BillSettlement audit trail
                BillSettlement.objects.create(
                    unity_bill_source=bill,
                    settled_amount=amount_to_apply,
                    settlement_date=timezone.now(),
                    source_credit_note_id=credit_item.id, # Link to the credit note
                    reconned_bank_line=None
                )
                
                # Deduct the amount from the Credit Note so it isn't used twice
                credit_item.schedule_amount -= amount_to_apply
                
                # If fully used, you might want to update status
                if credit_item.schedule_amount <= Decimal('0.009'):
                    credit_item.credit_link_status = 'Approved' # or 'Fully Used'
                
                credit_item.save()
                
                messages.success(request, f"Successfully applied R{amount_to_apply} from Credit Note.")

        except Exception as e:
            messages.error(request, f"Error applying credit: {str(e)}")
            
        return redirect('pre_bill_reconciliation_summary', company_code=bill.C_Company_Code, bill_id=bill.id)
    
    return redirect('dashboard')

@login_required
@transaction.atomic
def move_bank_line_to_credit(request, recon_id):
    """
    Manually moves a Bank Line's remaining balance into the CreditNote (Overs) pool.
    This sets the bank line to 'Reconciled' so it stops showing a balance.
    """
    recon_line = get_object_or_404(ReconnedBank, pk=recon_id)
    company_code = recon_line.company_code
    
    # Calculate what is left on this specific line
    remaining_balance = recon_line.transaction_amount - recon_line.amount_settled
    
    if remaining_balance <= Decimal('0.009'):
        messages.warning(request, "This line is already fully consumed.")
        return redirect('unity_information', company_code=company_code)

    # 1. Create the 'Overs' record in the CreditNote table
    CreditNote.objects.create(
        member_group_code=company_code,
        schedule_amount=remaining_balance,
        credit_link_status='Pending', # Sent to Manager Approval
        link_request_reason="Overs credit line",
        source_bank_line=recon_line, 
        comment=f"Manually moved to Credit Pool by {request.user.username}",
        processed_by=request.user.username,
        processed_date=timezone.now(),
        note_selection="OVERS"
    )

    # 2. Exhaust the bank line so it shows R0.00 remaining
    recon_line.amount_settled = recon_line.transaction_amount
    recon_line.recon_status = 'Reconciled'
    recon_line.save()

    messages.success(request, f"R{remaining_balance} moved to Credit Pool. Awaiting Manager Approval.")
    return redirect('unity_information', company_code=company_code)

@login_required
def bulk_billing_dashboard(request):
    from .models import ReconnedBank, UnityBill
    from django.db.models import Q, Sum, Min
    from django.shortcuts import render

    # 1. Fetch bank lines marked for bulk
    bulk_bank_lines = ReconnedBank.objects.filter(
        review_note__icontains="BULK", 
        recon_status__icontains="Unreconciled"
    ).select_related('bank_line').order_by('-transaction_date')

    # 2. FIXED: Added A_CCDatesMonth to .values() so bills for different dates don't consolidate
    bill_aggregates = UnityBill.objects.filter(
        is_reconciled=False
    ).values('C_Company_Code', 'A_CCDatesMonth').annotate(
        total_outstanding=Sum('H_Schedule_Amount')
    ).order_by('C_Company_Code', 'A_CCDatesMonth')

    # 3. Build the tuple (Code, Date, Amount)
    eligible_members = [
        (item['C_Company_Code'], item['A_CCDatesMonth'], item['total_outstanding']) 
        for item in bill_aggregates
    ]

    context = {
        'bulk_bank_lines': bulk_bank_lines,
        'eligible_members': eligible_members,
    }
    return render(request, 'unity_internal_app/bulk_billing.html', context)

from django.db import transaction

@login_required
def process_bulk_allocation(request):
    if request.method == 'POST':
        from .models import ReconnedBank, UnityBill, BillSettlement, JournalEntry
        from django.db.models import Sum
        from decimal import Decimal
        from django.contrib import messages
        from django.utils import timezone
        from django.db import transaction

        bank_line_id = request.POST.get('bank_line_id')
        member_codes = request.POST.getlist('member_codes[]')
        bill_dates = request.POST.getlist('bill_dates[]') # NEW: Capture dates from popup
        amounts = request.POST.getlist('amounts[]')

        try:
            with transaction.atomic():
                bank_line = ReconnedBank.objects.select_for_update().get(id=bank_line_id)
                bank_available = bank_line.transaction_amount - bank_line.amount_settled
                total_successfully_allocated = Decimal('0.00')

                # Updated zip to include bill_dates
                for code, bill_date, amount_str in zip(member_codes, bill_dates, amounts):
                    if not code or not amount_str: continue
                    
                    requested_amount = Decimal(amount_str)
                    if requested_amount <= 0: continue

                    if total_successfully_allocated + requested_amount > bank_available:
                        requested_amount = bank_available - total_successfully_allocated
                        if requested_amount <= 0: break 

                    # 2. UPDATED: Find the SPECIFIC bill for that code and date
                    target_bill = UnityBill.objects.filter(
                        C_Company_Code=code, 
                        A_CCDatesMonth=bill_date, # Filter by specific date
                        is_reconciled=False
                    ).first()

                    if not target_bill:
                        messages.warning(request, f"Skipped {code} for {bill_date}: Bill not found or already reconciled.")
                        continue

                    # 3. Calculate exactly what is still owed
                    already_paid = BillSettlement.objects.filter(
                        unity_bill_source=target_bill
                    ).aggregate(total=Sum('settled_amount'))['total'] or Decimal('0.00')
                    
                    journals_paid = JournalEntry.objects.filter(
                        target_bill=target_bill
                    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

                    remaining_on_bill = target_bill.H_Schedule_Amount - already_paid - journals_paid

                    actual_settle_amount = min(requested_amount, remaining_on_bill)

                    if actual_settle_amount <= 0:
                        continue

                    # 5. Create the Settlement
                    BillSettlement.objects.create(
                        unity_bill_source=target_bill,
                        reconned_bank_line=bank_line,
                        settled_amount=actual_settle_amount,
                        settlement_date=timezone.now(),
                        confirmed_by=request.user
                    )

                    bank_line.amount_settled += actual_settle_amount
                    total_successfully_allocated += actual_settle_amount

                    if (already_paid + journals_paid + actual_settle_amount) >= (target_bill.H_Schedule_Amount - Decimal('0.01')):
                        target_bill.is_reconciled = True
                        target_bill.save()

                if bank_line.amount_settled >= (bank_line.transaction_amount - Decimal('0.01')):
                    bank_line.recon_status = "Reconciled"
                else:
                    bank_line.recon_status = "Partially Reconciled"
                
                bank_line.save()
                messages.success(request, f"Bulk processing complete. Total Allocated: R {total_successfully_allocated}")

        except Exception as e:
            messages.error(request, f"Error during bulk processing: {str(e)}")

    return redirect('bulk_billing_dashboard')

from django.shortcuts import render, HttpResponse, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.utils import timezone
from django.db.models import Q, Sum
from datetime import timedelta
from collections import defaultdict
from django.contrib import messages
from django.conf import settings

# Ensure these models are imported correctly based on your app structure
# from .models import EmailDelegation, ReconnedBank, InternalFunds, BillSettlement

@login_required
def sla_report_view(request):
    """
    SLA Report: Tracks Email Delegation stats AND Bank Line Review Notes.
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # --- 1. EMAIL DELEGATION LOGIC ---
    delegations = EmailDelegation.objects.all().select_related('assigned_user').prefetch_related('transactions')

    if start_date and end_date and start_date not in ["None", ""] and end_date not in ["None", ""]:
        delegations = delegations.filter(received_at__range=[start_date, end_date])

    report_data = []
    agent_stats = defaultdict(lambda: {'count': 0, 'total_res_time': timedelta(), 'resolved_count': 0})

    for d in delegations:
        history_logs = d.transactions.all().order_by('timestamp')
        response_time = d.delegated_at - d.received_at if d.delegated_at and d.received_at else None
        resolution_time = None
        
        if d.status in ['COM', 'CLS', 'DONE']:
            last_act = history_logs.last()
            finish_date = last_act.timestamp if last_act else timezone.now()
            resolution_time = finish_date - d.received_at
            
            if d.assigned_user:
                agent_stats[d.assigned_user.username]['resolved_count'] += 1
                agent_stats[d.assigned_user.username]['total_res_time'] += resolution_time

        if d.assigned_user:
            agent_stats[d.assigned_user.username]['count'] += 1

        is_overdue = resolution_time > timedelta(hours=24) if resolution_time else False

        report_data.append({
            'delegation': d,
            'response_time': response_time,
            'resolution_time': resolution_time,
            'history_logs': history_logs,
            'is_overdue': is_overdue
        })

    final_agent_stats = []
    for agent, stats in agent_stats.items():
        avg_res = stats['total_res_time'] / stats['resolved_count'] if stats['resolved_count'] > 0 else None
        final_agent_stats.append({
            'agent': agent,
            'total_assigned': stats['count'],
            'total_resolved': stats['resolved_count'],
            'avg_resolution': str(avg_res).split('.')[0] if avg_res else "N/A"
        })

    # --- 2. BANK LINE NOTES LOGIC (Integrated) ---
    # Fetch bank segments that have either a category or internal text notes
    bank_notes_qs = ReconnedBank.objects.filter(
        Q(review_note__isnull=False) & ~Q(review_note='') | 
        Q(review_note_text__isnull=False) & ~Q(review_note_text='')
    ).select_related('bank_line')

    if start_date and end_date and start_date not in ["None", ""] and end_date not in ["None", ""]:
        bank_notes_qs = bank_notes_qs.filter(transaction_date__range=[start_date, end_date])

    return render(request, 'unity_internal_app/sla_report.html', {
        'report_data': report_data,
        'agent_stats': final_agent_stats,
        'bank_notes': bank_notes_qs, 
        'start_date': start_date if start_date != "None" else "",
        'end_date': end_date if end_date != "None" else "",
    })

@login_required
def export_sla_report_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    wb = openpyxl.Workbook()
    
    # --- SHEET 1: EMAIL DELEGATION ---
    ws1 = wb.active
    ws1.title = "Email SLA Report"
    
    header_fill = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)

    headers_email = [
        "Work Related", "Company Code", "Category", "Agent", 
        "Status", "Received At", "Delegated At", "Response Time", 
        "Resolution Time", "Action History"
    ]
    ws1.append(headers_email)
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    qs_emails = EmailDelegation.objects.all().select_related('assigned_user').prefetch_related('transactions')
    if start_date and end_date and start_date not in ["None", ""]:
        qs_emails = qs_emails.filter(received_at__range=[start_date, end_date])

    for d in qs_emails:
        notes_list = [f"[{t.timestamp.strftime('%H:%M')}] {t.action_type}" for t in d.transactions.all()]
        resp_str = str(d.delegated_at - d.received_at).split('.')[0] if d.delegated_at else "Pending"
        res_str = "Open"
        if d.status == 'COM':
            last_act = d.transactions.last()
            if last_act: res_str = str(last_act.timestamp - d.received_at).split('.')[0]

        ws1.append([
            "YES" if d.work_related else "NO",
            d.company_code or "N/A",
            d.email_category or "General",
            d.assigned_user.username if d.assigned_user else "Unassigned",
            d.get_status_display(),
            d.received_at.strftime('%Y-%m-%d %H:%M') if d.received_at else "",
            d.delegated_at.strftime('%Y-%m-%d %H:%M') if d.delegated_at else "",
            resp_str, res_str, " | ".join(notes_list)
        ])

    # --- SHEET 2: BANK LINE REVIEW NOTES ---
    ws2 = wb.create_sheet(title="Bank Review Notes")
    headers_bank = [
        "Date", "Company", "Amount", "Status", "Category", "Detailed Note", "Bank Reference"
    ]
    ws2.append(headers_bank)
    bank_header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    for cell in ws2[1]:
        cell.fill = bank_header_fill
        cell.font = white_font

    qs_banks = ReconnedBank.objects.filter(
        Q(review_note__isnull=False) & ~Q(review_note='') | 
        Q(review_note_text__isnull=False) & ~Q(review_note_text='')
    ).select_related('bank_line')

    if start_date and end_date and start_date not in ["None", ""]:
        qs_banks = qs_banks.filter(transaction_date__range=[start_date, end_date])

    for b in qs_banks:
        ws2.append([
            b.transaction_date.strftime('%Y-%m-%d') if b.transaction_date else "",
            b.company_code or "N/A",
            b.transaction_amount,
            b.recon_status,
            b.review_note or "",
            b.review_note_text or "",
            b.bank_line.statement_reference if b.bank_line else "N/A"
        ])

    # Final Formatting
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            sheet.column_dimensions[column].width = min(max_length + 2, 60)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="SLA_Log_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response